"""
Servidor Flask — API + frontend para exibir dados do StatArea
"""
from flask import Flask, jsonify, send_from_directory, abort, request
import json, os, glob, re, threading, time, sqlite3, itertools, math, traceback, queue, sys
import requests as http_req
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
import github_storage

# O console do Windows usa cp1252 por padrão, que não cobre nomes de time com
# caracteres como ş/ğ/č/đ (comuns em ligas turcas, balcânicas etc) — qualquer
# print(f"...{nome_do_time}...") com um desses derrubava a requisição inteira
# com UnicodeEncodeError (achado testando o Scanner: _find_uniscore_id e
# _process_momentum já tinham prints assim). Forçar UTF-8 aqui corrige de vez
# pra qualquer print futuro também, em vez de remendar um por um. Sem efeito
# em produção (Railway/Linux já usa UTF-8 por padrão).
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ── Sentry — captura erro do backend automaticamente (exceção não tratada em
# qualquer rota vira um evento no Sentry, com traceback completo), sem
# depender de print() que pode se perder no buffer do log do Railway. Só liga
# se SENTRY_DSN estiver configurado (variável de ambiente no Railway) — sem
# isso, roda normal, sem captura nenhuma (não trava nada em dev local).
SENTRY_DSN = os.environ.get("SENTRY_DSN", "")
if SENTRY_DSN:
    import sentry_sdk
    from sentry_sdk.integrations.flask import FlaskIntegration
    sentry_sdk.init(
        dsn=SENTRY_DSN,
        integrations=[FlaskIntegration()],
        traces_sample_rate=0.1,  # amostra 10% das requisições pra tracing de performance (fica dentro do free tier)
        environment=os.environ.get("RAILWAY_ENVIRONMENT_NAME", "production"),
    )

FOTMOB_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Referer": "https://www.fotmob.com/",
    "Accept":  "*/*",
}

app = Flask(__name__, static_folder="static", static_url_path="/static")
DATA_DIR     = os.path.dirname(__file__)
MOMENTUM_DIR = os.path.join(DATA_DIR, "momentum_history")
SHOTMAP_DIR  = os.path.join(DATA_DIR, "shotmap_history")
MAPA_CACHE_DIR = os.path.join(DATA_DIR, "mapa_cache")
FORCA_HISTORY_DIR = os.path.join(DATA_DIR, "forca_history")
os.makedirs(MOMENTUM_DIR, exist_ok=True)
os.makedirs(SHOTMAP_DIR,  exist_ok=True)
os.makedirs(MAPA_CACHE_DIR, exist_ok=True)
os.makedirs(FORCA_HISTORY_DIR, exist_ok=True)

# ── Cache em memória dos arquivos de momentum_history — evita reler e reparsear os
# +2000 arquivos do disco a cada busca de padrão (aba Análise/CS do Ao Vivo).
# Reaproveita o que já foi parseado; só relê arquivos novos ou modificados (por mtime).
_momentum_files_cache = {}   # fpath -> {mtime, pt_list, goals, casa, fora, liga, date, shotmap, score}
_momentum_files_lock  = threading.Lock()

def _get_momentum_files_cached():
    """Retorna a lista de dados já parseados de todos os arquivos de momentum_history,
    reutilizando o cache em memória sempre que possível."""
    with _momentum_files_lock:
        files = sorted(glob.glob(os.path.join(MOMENTUM_DIR, "*.json")))
        result = []
        for fpath in files:
            try:
                mtime = os.path.getmtime(fpath)
            except OSError:
                continue
            cached = _momentum_files_cache.get(fpath)
            if cached and cached["mtime"] == mtime:
                result.append(cached)
                continue
            try:
                with open(fpath, encoding="utf-8") as f:
                    d = json.load(f)
                pt_list = sorted(
                    [(float(p["minute"]), float(p["value"]))
                     for p in d.get("graphPoints", [])
                     if "minute" in p and "value" in p],
                    key=lambda x: x[0]
                )
                entry = {
                    "mtime":   mtime,
                    "pt_list": pt_list,
                    "goals":   d.get("goals", []),
                    "casa":    d.get("casa", "—"),
                    "fora":    d.get("fora", "—"),
                    "liga":    d.get("liga", ""),
                    "date":    d.get("date", ""),
                    "shotmap": d.get("shotmap", []),
                    "score":   d.get("score", {}),
                }
                _momentum_files_cache[fpath] = entry
                result.append(entry)
            except Exception:
                continue
        # Remove do cache arquivos que não existem mais
        stale = set(_momentum_files_cache) - set(files)
        for fpath in stale:
            _momentum_files_cache.pop(fpath, None)
        return result


def ajustar_hora(hora_str):
    """Subtrai 3 horas do horário vindo do StatArea (UTC → BRT)."""
    if not hora_str:
        return hora_str
    from datetime import datetime, timedelta
    try:
        t = datetime.strptime(hora_str.strip(), "%H:%M") - timedelta(hours=3)
        return t.strftime("%H:%M")
    except Exception:
        return hora_str


def _shift_hora(hora_str, delta_hours):
    """Soma delta_hours ao horário HH:MM. Retorna novo HH:MM ou original."""
    from datetime import datetime, timedelta
    if not hora_str:
        return hora_str
    try:
        t = datetime.strptime(hora_str.strip(), "%H:%M") + timedelta(hours=delta_hours)
        return t.strftime("%H:%M")
    except Exception:
        return hora_str


FLAG_FILE = os.path.join(DATA_DIR, ".times_adjusted")


def _flag_today():
    """Retorna a data gravada no flag, ou '' se não existir."""
    try:
        with open(FLAG_FILE, "r") as f:
            return f.read().strip()
    except Exception:
        return ""


@app.route("/api/fix-times/status")
def api_fix_times_status():
    """Verifica se o ajuste já foi aplicado hoje."""
    today = datetime.now().strftime("%Y-%m-%d")
    done  = _flag_today() == today
    return jsonify({"done": done, "date": today})


@app.route("/api/fix-times", methods=["POST"])
def api_fix_times():
    """Adiciona +3h nos horários do predictions JSON — só uma vez por dia."""
    today = datetime.now().strftime("%Y-%m-%d")

    # Bloqueia segunda execução no mesmo dia
    if _flag_today() == today:
        return jsonify({"ok": False, "already_done": True,
                        "msg": "Horários já foram ajustados hoje."})

    body  = request.get_json(force=True, silent=True) or {}
    delta = int(body.get("delta", 3))

    matches, source = load_predictions()
    if not matches:
        return jsonify({"ok": False, "error": "Nenhum arquivo de partidas encontrado"}), 404

    fname = source or "predictions_full.json"
    path  = os.path.join(DATA_DIR, fname)

    updated = 0
    for m in matches:
        if m.get("hora"):
            m["hora"] = _shift_hora(m["hora"], delta)
            updated += 1

    with open(path, "w", encoding="utf-8") as f:
        json.dump(matches, f, ensure_ascii=False, indent=2)

    # Grava flag com a data de hoje
    with open(FLAG_FILE, "w") as f:
        f.write(today)

    return jsonify({"ok": True, "updated": updated, "delta": delta})


def load_predictions():
    # Prefere o arquivo full (com detalhes), senão usa o simples
    for fname in ["predictions_full.json", "predictions.json"]:
        path = os.path.join(DATA_DIR, fname)
        if os.path.exists(path):
            with open(path, encoding="utf-8") as f:
                return json.load(f), fname
    return [], None


# ── CONFRONTO DIRETO (H2H) + SCORE DE CONVICÇÃO ──────────────────────────
# Reaproveitam só dados já coletados (m["detalhes"]), sem nenhuma chamada
# externa nova. Cálculo é uma soma ponderada simples sobre o objeto de cada
# partida já carregado em memória — sem impacto de performance perceptível.

CONVICCAO_PESOS = {
    "ofensivo": 0.25,
    "defensivo_adversario": 0.25,
    "forma_recente": 0.20,
    "casa_fora": 0.15,
    "h2h": 0.10,
    "contexto": 0.05,  # sem indicador de contexto implementado ainda — o peso é
                        # redistribuído proporcionalmente entre os presentes (ver abaixo)
}
CONVICCAO_PENALIDADE_DIVERGENCIA = 0.15  # TODO: recalibrar depois de acumular histórico real


def _convicao_pesos_efetivos(indicadores_presentes):
    """Remove 'contexto' (não implementado) e qualquer indicador ausente nessa
    partida, redistribuindo os pesos que sobrarem proporcionalmente entre os
    indicadores realmente presentes, mantendo a soma em 1.0."""
    pesos = {k: v for k, v in CONVICCAO_PESOS.items() if k != "contexto" and k in indicadores_presentes}
    soma = sum(pesos.values())
    if soma <= 0:
        return pesos
    falta = 1.0 - soma
    for k in pesos:
        pesos[k] += falta * (pesos[k] / soma)
    return pesos


def _h2h_norm_nome(s):
    import unicodedata
    s = (s or "").lower().strip()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def _h2h_confronto_direto(confrontos, home, fora):
    """Estatísticas do confronto direto entre os dois times de hoje, calculadas
    em cima de detalhes.confrontos_diretos (já coletado pelo scraper). Marca
    amostra insuficiente com menos de 3 jogos, conforme especificado."""
    confrontos = confrontos or []
    if len(confrontos) < 3:
        return {"amostra_suficiente": False, "total": len(confrontos)}

    home_n = _h2h_norm_nome(home).split(" ")[0] if _h2h_norm_nome(home) else ""
    vit_casa = vit_fora = empates = gols_totais = over25 = 0
    mm_vit_casa = mm_vit_fora = mm_empates = mm_total = 0

    for p in confrontos:
        try: gc = int(p.get("gols_casa") or 0)
        except (TypeError, ValueError): gc = 0
        try: gf = int(p.get("gols_fora") or 0)
        except (TypeError, ValueError): gf = 0
        p_casa_norm = _h2h_norm_nome(p.get("casa"))
        p_casa_eh_home = bool(home_n) and (home_n in p_casa_norm)
        gols_home = gc if p_casa_eh_home else gf
        gols_fora_time = gf if p_casa_eh_home else gc
        if gols_home > gols_fora_time: vit_casa += 1
        elif gols_home < gols_fora_time: vit_fora += 1
        else: empates += 1
        gols_totais += gc + gf
        if gc + gf > 2.5: over25 += 1
        if p_casa_eh_home:
            mm_total += 1
            if gols_home > gols_fora_time: mm_vit_casa += 1
            elif gols_home < gols_fora_time: mm_vit_fora += 1
            else: mm_empates += 1

    n = len(confrontos)
    resultado = {
        "amostra_suficiente": True,
        "total": n,
        "vitorias_casa": vit_casa,
        "vitorias_fora": vit_fora,
        "empates": empates,
        "media_gols_totais": round(gols_totais / n, 2),
        "pct_over_25": round(over25 / n * 100),
        "quem_abre_placar": None,  # dado não disponível no histórico coletado hoje
    }
    resultado["mesmo_mando"] = (
        {"total": mm_total, "vitorias_casa": mm_vit_casa, "vitorias_fora": mm_vit_fora, "empates": mm_empates}
        if mm_total >= 2 else None
    )
    return resultado


def _convicao_indicadores(m, is_home):
    """Indicadores normalizados 0-100, cada um alinhado a favor do time avaliado
    (is_home decide se 'vota' pelo mandante ou visitante) — mesma escala de
    pontos 0-100 já usada no Score do Time (frontend), não uma normalização nova."""
    d = m.get("detalhes") or {}
    team = m.get("casa") if is_home else m.get("fora")
    ind = {}

    last10 = d.get("ultimas_10_partidas") or {}
    team_data = last10.get(team)
    if team_data is None and len(last10) >= 2:
        keys = list(last10.keys())
        team_data = last10.get(keys[0] if is_home else keys[1])
    form = (team_data or {}).get("form") or ""
    if form:
        wins, draws = form.count("W"), form.count("D")
        ind["forma_recente"] = round((wins * 3 + draws) / (len(form) * 3) * 100)

    stats = d.get("estatisticas") or {}
    ts = stats.get(team)
    if ts is None and len(stats) >= 2:
        keys = list(stats.keys())
        ts = stats.get(keys[0] if is_home else keys[1])
    ts = ts or {}
    try: avg_scored = float(ts.get("Average scored goals per match") or 0)
    except (TypeError, ValueError): avg_scored = 0
    try: avg_conc = float(ts.get("Average conceded goals per match") or 0)
    except (TypeError, ValueError): avg_conc = 0
    if avg_scored > 0:
        ind["ofensivo"] = (100 if avg_scored >= 2.0 else 75 if avg_scored >= 1.5 else
                            50 if avg_scored >= 1.0 else 25 if avg_scored >= 0.5 else 10)
    if avg_conc > 0 or avg_scored > 0:
        ind["defensivo_adversario"] = (100 if avg_conc <= 0.7 else 75 if avg_conc <= 1.0 else
                                        50 if avg_conc <= 1.5 else 25 if avg_conc <= 2.0 else 10)

    # casa/fora — proxy: posição na tabela. O JSON de previsões não tem um
    # split casa/fora dedicado por time; classificação é o indicador de força
    # relativa mais próximo já calculado hoje (mesma lógica do Score do Time).
    standings = d.get("classificacao") or []
    hl = [r for r in standings if r.get("destacado")]
    if len(hl) >= 2:
        my_row = hl[0] if is_home else hl[1]
        try: pos = int(my_row.get("pos") or 0)
        except (TypeError, ValueError): pos = 0
        total = len(standings)
        if pos > 0 and total > 0:
            pct = pos / total
            ind["casa_fora"] = (100 if pct <= 0.20 else 80 if pct <= 0.40 else
                                 55 if pct <= 0.60 else 27 if pct <= 0.80 else 0)

    h2h = _h2h_confronto_direto(d.get("confrontos_diretos"), m.get("casa"), m.get("fora"))
    if h2h.get("amostra_suficiente"):
        vit = h2h["vitorias_casa"] if is_home else h2h["vitorias_fora"]
        ind["h2h"] = round(vit / h2h["total"] * 100)

    return ind, h2h


def _convicao_score(m, is_home):
    """Soma ponderada dos indicadores (pesos em CONVICCAO_PESOS), com penalização
    quando os indicadores divergem fortemente entre si. None se não houver
    nenhum indicador calculável pra essa partida ainda (detalhes incompletos)."""
    ind, h2h = _convicao_indicadores(m, is_home)
    if not ind:
        return None, h2h
    pesos = _convicao_pesos_efetivos(ind.keys())
    score = sum(pesos[k] * v for k, v in ind.items())
    if len(ind) >= 2:
        divergencia = max(ind.values()) - min(ind.values())
        score -= divergencia * CONVICCAO_PENALIDADE_DIVERGENCIA
    return round(min(100, max(0, score))), h2h


APP_VERSION = "2026-05-18-v9"

# ── PAINEL PRINCIPAL — jogos + odds (1X2 e Over/Under) do BetExplorer ──────────
# A listagem de jogos por liga do BetExplorer (homepage) já vem 100% renderizada
# em HTML no endpoint /gres/ajax/homepage-data.php (sem precisar de Selenium/JS).
# Cada chamada só traz UM tipo de aposta por vez (betType=1x2 ou betType=ou), então
# buscamos as duas e casamos os jogos pelo event-id pra ter 1/X/2 + Over/Under juntos.
BETEXPLORER_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Referer": "https://www.betexplorer.com/br/",
    "Accept": "text/html, */*; q=0.01",
    "X-Requested-With": "XMLHttpRequest",
}
BETEXPLORER_BASE = "https://www.betexplorer.com"
_PAINEL_CACHE_TTL = 60  # odds mudam, mas o BetExplorer passou a bloquear (429) com requests demais
_painel_cache = {}   # cache_key (data ou "today") -> {"ts":, "data":}
_painel_lock = threading.Lock()

# ── Rate limit — o BetExplorer começou a devolver 429 (Too Many Requests) quando
# batemos rápido demais (vários widgets abertos ao mesmo tempo, auto-refresh,
# múltiplos usuários). Serializa TODAS as chamadas ao BetExplorer com um espaço
# mínimo entre elas + retry com backoff quando toma 429, em vez de derrubar a
# aba na cara do usuário.
_be_rate_lock = threading.Lock()
_be_last_request_ts = 0.0
_BE_MIN_INTERVAL = 0.6  # segundos entre requests consecutivos ao BetExplorer


def _be_get(url, params=None, timeout=15, retries=3):
    """GET no BetExplorer com espaçamento mínimo entre chamadas e retry/backoff
    em cima de 429 — evita que um pico de acessos derrube a página pro usuário."""
    global _be_last_request_ts
    last_exc = None
    for attempt in range(retries):
        with _be_rate_lock:
            wait = _BE_MIN_INTERVAL - (time.time() - _be_last_request_ts)
            if wait > 0:
                time.sleep(wait)
            _be_last_request_ts = time.time()
        r = http_req.get(url, params=params, headers=BETEXPLORER_HEADERS, timeout=timeout)
        if r.status_code == 429:
            last_exc = RuntimeError("O BetExplorer está limitando as requisições no momento (429). Tente de novo em alguns segundos.")
            time.sleep(1.5 * (attempt + 1))
            continue
        r.raise_for_status()
        r.encoding = "utf-8"
        return r
    raise last_exc


def _be_fetch_bettype_html(bettype, date_params=None):
    """date_params, quando informado, é {"year":, "month":, "day":} — mesmo
    parâmetro que o calendário do BetExplorer usa pra navegar entre dias."""
    params = {"tab": "all", "betType": bettype, "lang": "br", "tz": "-3:00", "start": 0, "end": 300}
    if date_params:
        params.update(date_params)
    r = _be_get(f"{BETEXPLORER_BASE}/gres/ajax/homepage-data.php", params=params)
    return r.text


def _be_parse_bettype(html):
    """Parseia o fragment HTML do BetExplorer pra um tipo de aposta, retornando
    {event_id: {...}} e a lista de ligas na ordem em que aparecem na página."""
    soup = BeautifulSoup(html, "html.parser")
    matches = {}
    leagues_order = []
    for ul in soup.find_all("ul", class_="leagues-list"):
        country = ul.get("data-country", "")
        header_li = ul.find("li", class_="js-tournament")
        league_name, flag_url, ttid = "", "", None
        if header_li:
            name_tag = header_li.find("p", class_="leaguesNames")
            if name_tag:
                league_name = name_tag.get_text(strip=True)
            img_tag = header_li.find("img")
            if img_tag:
                flag_url = img_tag.get("data-src") or img_tag.get("src") or ""
            ttid = header_li.get("data-ttid")
        league_key = ttid or league_name
        leagues_order.append({"key": league_key, "country": country, "league_name": league_name, "flag_url": flag_url})

        for row in ul.find_all("li", class_="table-main__tournamentLiContent"):
            event_id = row.get("data-event-id")
            if not event_id:
                continue
            status_el = row.select_one(".matchDateStatus")
            status_text = status_el.get_text(strip=True) if status_el else ""

            participants = row.select(".table-main__truncate")
            home = participants[0].get_text(strip=True) if len(participants) > 0 else ""
            away = participants[1].get_text(strip=True) if len(participants) > 1 else ""

            logos = row.select(".table-main__participantLogo")
            home_logo = logos[0].get("data-src") or logos[0].get("src") if len(logos) > 0 else None
            away_logo = logos[1].get("data-src") or logos[1].get("src") if len(logos) > 1 else None
            if home_logo and home_logo.startswith("/"):
                home_logo = BETEXPLORER_BASE + home_logo
            if away_logo and away_logo.startswith("/"):
                away_logo = BETEXPLORER_BASE + away_logo

            score_home = score_away = None
            score_div = row.select_one(".mainResult.table-main__Bold.mobileHidden")
            if score_div:
                parts = [d.get_text(strip=True) for d in score_div.find_all("div")]
                digits = [p for p in parts if p and p not in ("-", ":")]
                if len(digits) >= 2:
                    score_home, score_away = digits[0], digits[1]

            link_tag = row.find("a", attrs={"data-live-cell": "matchlink"})
            match_url = link_tag.get("href") if link_tag else None

            odds_wrap = row.select_one(".oddsColumn")
            values, line = [], None
            if odds_wrap:
                line_div = odds_wrap.find("div", class_="table-main__oddOU")
                if line_div:
                    line = line_div.get_text(strip=True)
                for odd_div in odds_wrap.select(".table-main__odd"):
                    btn = odd_div.find(["button", "p"])
                    values.append(btn.get("data-odd") if btn else None)

            try:
                ts = int(row.get("data-ts") or 0)
            except (TypeError, ValueError):
                ts = 0

            matches[event_id] = {
                "event_id": event_id, "league_key": league_key, "ts": ts,
                "status_text": status_text, "home": home, "away": away,
                "home_logo": home_logo, "away_logo": away_logo,
                "score_home": score_home, "score_away": score_away,
                "match_url": match_url, "line": line, "odds": values,
            }
    return matches, leagues_order


# ── NowGoal — fonte alternativa dos jogos do dia (BetExplorer passou a bloquear
# com 429 com frequência demais). O NowGoal serve os jogos como um array JS puro
# (sem HTML pra parsear), só exige um cookie de sessão (LS_ACCESS_TOKEN) obtido
# visitando a home antes de acessar o feed de dados. Por ora só migramos jogos +
# placar + liga/país (odds e link de análise continuam vindo do BetExplorer, que
# ainda alimenta os widgets de Confronto Direto/Últimos Resultados/Classificações).
NOWGOAL_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Referer": "https://www.nowgoal.net/",
    "Accept": "*/*",
}
NOWGOAL_BASE = "https://www.nowgoal.net"
_ng_cookie_cache = {"ts": 0.0, "jar": None}
_ng_cookie_lock = threading.Lock()
_NG_COOKIE_TTL = 3600  # 1h — token de sessão simples, não expira rápido, mas renovamos por segurança

_NG_STATUS_MAP = {
    "0": "Agendado", "1": "1º Tempo", "2": "Intervalo", "3": "2º Tempo",
    "4": "Prorrogação", "5": "Pênaltis", "-1": "Encerrado", "7": "Adiado", "8": "Cancelado",
}


def _ng_get_cookie_jar():
    """Visita a home do NowGoal pra pegar o cookie de sessão (LS_ACCESS_TOKEN) exigido
    pelo feed de dados — sem ele o endpoint devolve {"code":100401} em vez do array JS."""
    with _ng_cookie_lock:
        now = time.time()
        if _ng_cookie_cache["jar"] is not None and (now - _ng_cookie_cache["ts"]) < _NG_COOKIE_TTL:
            return _ng_cookie_cache["jar"]
        r = http_req.get(f"{NOWGOAL_BASE}/", headers=NOWGOAL_HEADERS, timeout=15)
        r.raise_for_status()
        _ng_cookie_cache["jar"] = r.cookies
        _ng_cookie_cache["ts"] = now
        return r.cookies


def _ng_split_js_array(content):
    """Faz o split de um literal de array JS tipo `1,'a, b',,'c'` respeitando aspas
    simples e elementos vazios (vírgulas seguidas) — não dá pra usar json.loads porque
    o NowGoal usa aspas simples e omite elementos nulos em vez de usar `null`."""
    tokens = []
    buf = ""
    in_str = False
    escape = False
    for ch in content:
        if in_str:
            if escape:
                buf += ch
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == "'":
                in_str = False
            else:
                buf += ch
            continue
        if ch == "'":
            in_str = True
            continue
        if ch == ",":
            tokens.append(buf.strip())
            buf = ""
            continue
        buf += ch
    tokens.append(buf.strip())
    return tokens


_NG_ODDS_COMPANY = "8"  # bookmaker usado como fonte de odds (handicap asiático/1x2/O-U) — ver goal{id}.xml


def _ng_fetch_odds():
    """Busca o feed de odds do NowGoal (handicap asiático + 1X2 + over/under) pra
    um bookmaker fixo. Formato por linha (dentro de <m>...</m>):
    match_id, ah_provider_id, ah_line, ah_home_odd, ah_away_odd,
    x12_provider_id, odd_1, odd_x, odd_2,
    ou_provider_id, ou_line, odd_over, odd_under, ...flags
    Retorna {match_id: {...}} — se falhar, retorna {} (odds ficam vazias, sem quebrar a listagem)."""
    try:
        jar = _ng_get_cookie_jar()
        r = http_req.get(
            f"{NOWGOAL_BASE}/gf/data/odds/en/goal{_NG_ODDS_COMPANY}.xml",
            headers=NOWGOAL_HEADERS, cookies=jar, timeout=15,
        )
        r.raise_for_status()
        text = r.text
    except Exception:
        return {}

    odds = {}
    for m in re.finditer(r"<m>(.*?)</m>", text):
        fields = m.group(1).split(",")
        if len(fields) < 13:
            continue
        match_id = fields[0]
        odds[match_id] = {
            "ah_line": fields[2] or None, "ah_home": fields[3] or None, "ah_away": fields[4] or None,
            "odd_1": fields[6] or None, "odd_x": fields[7] or None, "odd_2": fields[8] or None,
            "ou_line": fields[10] or None, "odd_over": fields[11] or None, "odd_under": fields[12] or None,
        }
    return odds


def _ng_fetch_today_matches():
    """Busca e parseia o feed de jogos do dia do NowGoal (array JS puro em vez de
    HTML). Retorna (matches: list[dict], leagues_info: {league_index: {...}})."""
    jar = _ng_get_cookie_jar()
    r = http_req.get(f"{NOWGOAL_BASE}/gf/data/bf_en-idn1.js", headers=NOWGOAL_HEADERS, cookies=jar, timeout=15)
    r.raise_for_status()
    text = r.text
    odds_by_match = _ng_fetch_odds()

    countries = {}   # idx -> nome do país
    for m in re.finditer(r"C\[(\d+)\]=\[(.*?)\];", text):
        idx = int(m.group(1))
        parts = _ng_split_js_array(m.group(2))
        countries[idx] = parts[1] if len(parts) > 1 else ""

    leagues = {}     # idx -> {"name":, "country":}
    for m in re.finditer(r"B\[(\d+)\]=\[(.*?)\];", text):
        idx = int(m.group(1))
        parts = _ng_split_js_array(m.group(2))
        name = parts[2] if len(parts) > 2 else (parts[1] if len(parts) > 1 else "")
        country_idx = None
        try:
            country_idx = int(parts[10]) if len(parts) > 10 and parts[10] else None
        except ValueError:
            country_idx = None
        leagues[idx] = {"name": name, "country": countries.get(country_idx, "")}

    matches = []
    for m in re.finditer(r"A\[(\d+)\]=\[(.*?)\];", text):
        parts = _ng_split_js_array(m.group(2))
        if len(parts) < 11:
            continue
        try:
            match_id = parts[0]
            league_idx = int(parts[1]) if parts[1] else None
            home_name, away_name = parts[4], parts[5]
            kickoff = parts[6]
            status_code = parts[8]
            not_started = status_code == "0"  # antes do apito inicial o NowGoal já manda "0" em placar/HT/escanteio
            score_home = parts[9] if parts[9] != "" and not not_started else None
            score_away = parts[10] if parts[10] != "" and not not_started else None
            ht_home = parts[11] if len(parts) > 11 and parts[11] != "" and not not_started else None
            ht_away = parts[12] if len(parts) > 12 and parts[12] != "" and not not_started else None
            corner_home = parts[27] if len(parts) > 27 and parts[27] != "" and not not_started else None
            corner_away = parts[28] if len(parts) > 28 and parts[28] != "" and not not_started else None
        except (IndexError, ValueError):
            continue
        lg = leagues.get(league_idx, {"name": "", "country": ""})
        try:
            ts = int(datetime.strptime(kickoff, "%Y-%m-%d %H:%M:%S").timestamp())
        except ValueError:
            ts = 0

        minute = None
        if status_code in ("1", "3") and ts:
            elapsed = int((time.time() - ts) / 60)
            if status_code == "1":
                minute = max(0, min(elapsed, 45))
            else:  # 2º tempo — aproximado: desconta o intervalo (~15min) do tempo corrido
                minute = max(46, min(elapsed - 15, 90))

        odds = odds_by_match.get(match_id, {})
        matches.append({
            "event_id": match_id,
            "league_key": league_idx,
            "league_name": lg["name"], "country": lg["country"],
            "time": _NG_STATUS_MAP.get(status_code, status_code),
            "minute": minute,
            "home": home_name, "away": away_name,
            "home_logo": None, "away_logo": None,
            "score_home": score_home, "score_away": score_away,
            "ht_home": ht_home, "ht_away": ht_away,
            "corner_home": corner_home, "corner_away": corner_away,
            "match_url": None,
            "odd_1": odds.get("odd_1"), "odd_x": odds.get("odd_x"), "odd_2": odds.get("odd_2"),
            "ou_line": odds.get("ou_line"), "odd_over": odds.get("odd_over"), "odd_under": odds.get("odd_under"),
            "ah_line": odds.get("ah_line"), "ah_home": odds.get("ah_home"), "ah_away": odds.get("ah_away"),
            "ts": ts,
        })
    return matches


# ── "Comparação de força" (widget de análise pré-jogo do NowGoal) ─────────────
# O NowGoal já calcula tudo isso no client (grades/percentuais de H2H, Estado,
# Ataque, Defesa, Valor de mercado, Escanteios/Cartões/Faltas/Posse) a partir de
# dados embutidos na própria página (`battleData`, `lastMatchData`, `marketData`,
# `survayData`) e expõe o resultado pronto em `window._strength` depois que a
# página carrega. Em vez de reimplementar essa fórmula (script minificado de
# ~1MB, não vale o risco de divergir do site original), abrimos a página com
# Playwright (igual já fazemos pro contexto do BetExplorer) e lemos esse objeto
# já calculado direto do browser.
_ng_strength_cache = {}   # match_id -> {"ts":, "data": {...}}
_ng_strength_lock = threading.Lock()
_NG_STRENGTH_TTL = 900  # 15min — dado muda pouco entre atualizações
_ng_playwright_semaphore = threading.Semaphore(2)


def _ng_sanitize_nan(obj):
    """O NowGoal calcula alguns percentuais como 0/0 => NaN (ex.: time sem jogos
    ainda nesse recorte). Python aceita NaN como literal JSON na serialização,
    mas isso não é JSON válido de verdade — o `fetch().json()` do navegador
    rejeita com "Unexpected token 'N'". Troca por None (vira null, JSON válido)."""
    if isinstance(obj, float) and (obj != obj):  # NaN nunca é igual a si mesmo
        return None
    if isinstance(obj, dict):
        return {k: _ng_sanitize_nan(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_ng_sanitize_nan(v) for v in obj]
    return obj


_NG_MATCH_TABLE_JS = """
() => {
    function extract(n) {
        const q = (id) => { const el = document.getElementById(id); return el ? el.textContent.trim() : null; };
        const qv = (id) => { const el = document.querySelector('#' + id + ' .value'); return el ? el.textContent.trim() : null; };
        const table = document.getElementById('table_v' + n);
        if (!table) return null;
        const rows = [...table.querySelectorAll('tr[id^="tr' + n + '_"]')].map(tr => {
            const tds = tr.querySelectorAll('td');
            const g = (i) => tds[i] ? tds[i] : null;
            const scoreTd = g(3), cornerTd = g(5);
            const ft = scoreTd ? scoreTd.querySelector('[class^="fscore"]') : null;
            const ht = scoreTd ? scoreTd.querySelector('[class^="hscore"]') : null;
            const cFt = cornerTd ? cornerTd.querySelector('[class^="fcorner"]') : null;
            const cHt = cornerTd ? cornerTd.querySelector('[class^="hcorner"]') : null;
            const odd = (i) => { const td = g(i); return td ? td.getAttribute('data-o') : null; };
            return {
                league: g(0) ? (g(0).getAttribute('title') || g(0).textContent.trim()) : null,
                date: (() => { const s = g(1) ? g(1).querySelector('[data-t]') : null; return s ? s.getAttribute('data-t') : null; })(),
                home: g(2) ? g(2).textContent.trim() : null,
                score_ft: ft ? ft.textContent.trim() : null,
                score_ht: ht ? ht.textContent.replace(/[()]/g, '').trim() : null,
                away: g(4) ? g(4).textContent.trim() : null,
                corner_ft: cFt ? cFt.textContent.trim() : null,
                corner_ht: cHt ? cHt.textContent.replace(/[()]/g, '').trim() : null,
                odd_hw: odd(6), odd_d: odd(7), odd_aw: odd(8),
                wl_badge: g(9) ? g(9).textContent.trim() : null,
                odd_ah_home: odd(10), ah_line: odd(11), odd_ah_away: odd(12),
                ah_badge: g(13) ? g(13).textContent.trim() : null,
                ou_badge: g(14) ? g(14).textContent.trim() : null,
            };
        });
        return {
            rows,
            summary: {
                win: q('hW_v' + n), draw: q('d_v' + n), lose: q('gW_v' + n),
                goal_avg_home: q('hsAvg_v' + n), goal_avg_away: q('gsAvg_v' + n),
                ah_home_pct: qv('ahWBar_v' + n), ah_draw_pct: qv('ahDBar_v' + n), ah_away_pct: qv('ahLBar_v' + n),
                ah_count: q('ahCount_v' + n),
                ou_over_pct: qv('ouWBar_v' + n), ou_draw_pct: qv('ouDBar_v' + n), ou_under_pct: qv('ouLBar_v' + n),
                ou_count: q('ouCount_v' + n),
            },
        };
    }
    return { home: extract(1), away: extract(2), h2h: extract(3) };
}
"""

# "Estatísticas de probabilidades" (Win/Draw/Lose + Over/Draw/Under de todas as
# odds parecidas), "Distribuição de metas" (nº de gols / cronograma de gols /
# momento do 1º gol), "Meio período/Tempo integral" (matriz HT x FT) e
# "Diferença de gols HT x FT" — todos widgets prontos do NowGoal (ids oddsStat/
# goalStat/HFStat/GDStat), lidos direto do DOM já renderizado como os outros.
_NG_EXTRA_STATS_JS = """
() => {
    function readGroups(ul) {
        if (!ul) return null;
        // O NowGoal já embute TODAS as variantes de HT/HA-Igual no atributo "rate"
        // (por item, no oddsStat; no <ul> inteiro, no HFStat/goalStat) — os
        // checkboxes só trocam qual variante já calculada é exibida, sem nova
        // busca. Repassa o "rate" cru pro frontend poder alternar sem refazer
        // scraping nenhum.
        return {
            ul_rate: ul.getAttribute('rate'),
            groups: [...ul.querySelectorAll('li.group')].map(li => {
                const items = [...li.querySelectorAll('.item2')].map(it => ({
                    home_pct: it.querySelector('.home.bar') ? parseFloat(it.querySelector('.home.bar').style.height) : null,
                    away_pct: it.querySelector('.away.bar') ? parseFloat(it.querySelector('.away.bar').style.height) : null,
                    home_val: it.querySelector('.home .value') ? it.querySelector('.home .value').textContent.trim() : null,
                    away_val: it.querySelector('.away .value') ? it.querySelector('.away .value').textContent.trim() : null,
                    label: it.querySelector('.txt') ? it.querySelector('.txt').textContent.trim() : null,
                    rate: it.getAttribute('rate'),
                }));
                const titEl = li.querySelector('.tit');
                return { title: titEl ? titEl.textContent.replace(/\\s+/g, ' ').trim() : null, items };
            }),
        };
    }
    // "Momento do primeiro gol" (3ª sub-aba) usa o MESMO elemento #goalTimeStat
    // que "Cronograma de metas" — só troca via switchGoalStat(2), e o clique
    // síncrono não dá tempo do DOM re-renderizar antes da gente ler (a leitura
    // saía vazia). Em vez de depender desse timing, o rate de #goalTimeStat já
    // vem com as 4 variantes (time-normal, time-HA, primeiro gol-normal,
    // primeiro gol-HA) no mesmo atributo — só troca de sub-aba pra pegar os
    // rótulos certos (que são os mesmos nas duas abas) e decodifica o índice
    // 2/3 no frontend em vez de tentar reler o DOM depois do 3º clique.
    let goalNum = null, goalTime = null;
    if (typeof switchGoalStat === 'function' && document.getElementById('goalNumStat')) {
        switchGoalStat(0); goalNum = readGroups(document.getElementById('goalNumStat'));
        switchGoalStat(1); goalTime = readGroups(document.getElementById('goalTimeStat'));
    }
    return {
        odds_stat: readGroups(document.getElementById('panLuStat')),
        goal_num_stat: goalNum,
        goal_time_stat: goalTime,
        hf_stat: readGroups(document.getElementById('HFStat')),
        gd_stat: readGroups(document.getElementById('GDStat')),
    };
}
"""


def _ng_fetch_strength(match_id, _attempt=1):
    with _ng_strength_lock:
        cached = _ng_strength_cache.get(match_id)
        if cached and (time.time() - cached["ts"]) < _NG_STRENGTH_TTL:
            return cached["data"]

    from playwright.sync_api import sync_playwright

    data = None
    try:
        with _ng_playwright_semaphore:
            with sync_playwright() as p:
                browser = p.chromium.launch(
                    headless=True,
                    args=["--disable-blink-features=AutomationControlled"],
                )
                try:
                    page = browser.new_page(
                        user_agent=NOWGOAL_HEADERS["User-Agent"],
                        viewport={"width": 1280, "height": 900},
                    )
                    page.goto(f"{NOWGOAL_BASE}/match/h2h-{match_id}", wait_until="domcontentloaded", timeout=30000)
                    page.wait_for_function(
                        "() => window._strength && window._strength.count !== undefined",
                        timeout=20000,
                    )
                    data = _ng_sanitize_nan(page.evaluate("() => window._strength"))
                    # As 3 tabelas de histórico (últimos resultados de casa/fora + confronto
                    # direto — table_v1/v2/v3) são preenchidas pelo NowGoal via JS depois que a
                    # página carrega, igual o _strength — extrai direto do DOM já renderizado em
                    # vez de tentar achar/replicar o endpoint que as alimenta.
                    tables = _ng_sanitize_nan(page.evaluate(_NG_MATCH_TABLE_JS))
                    data["last_results_home"] = tables.get("home")
                    data["last_results_away"] = tables.get("away")
                    data["h2h_table"] = tables.get("h2h")
                    # "Partidas históricas com as mesmas probabilidades" (AH/1X2/O-U) — o
                    # NowGoal também calcula isso no client e expõe pronto em window._sameOdds,
                    # mas isso carrega um pouco depois (via ajax assíncrono próprio) do que o
                    # _strength — sem esperar por ele especificamente, `data` ainda vem vazio.
                    try:
                        page.wait_for_function(
                            "() => window._sameOdds && window._sameOdds.data && window._sameOdds.data.AHAllSclass",
                            timeout=10000,
                        )
                        data["same_odds"] = _ng_sanitize_nan(page.evaluate("() => window._sameOdds"))
                    except Exception:
                        data["same_odds"] = None
                    extra = _ng_sanitize_nan(page.evaluate(_NG_EXTRA_STATS_JS))
                    data.update(extra)
                finally:
                    browser.close()
    except Exception:
        if _attempt < 2:
            time.sleep(1.5)
            return _ng_fetch_strength(match_id, _attempt=_attempt + 1)
        raise RuntimeError("Não foi possível carregar a Comparação de Força dessa partida no NowGoal.")

    if data is None:
        raise RuntimeError("Não foi possível carregar a Comparação de Força dessa partida no NowGoal.")

    with _ng_strength_lock:
        _ng_strength_cache[match_id] = {"ts": time.time(), "data": data}
    return data


@app.route("/api/painel/ng_strength")
def api_painel_ng_strength():
    match_id = request.args.get("match_id", "")
    if not match_id or not match_id.isdigit():
        return jsonify({"error": "match_id inválido"}), 400
    try:
        data = _ng_fetch_strength(match_id)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return jsonify(data)


# Grades por eixo (H2H/Estado/Ataque/Defesa/Valor) — só o suficiente pra
# calcular a "diferença de força" mostrada na listagem do Painel Principal.
_PAINEL_FORCA_RADAR_AXES = ("battle", "state", "attack", "defend", "market")


@app.route("/api/painel/ng_strength_cached", methods=["POST"])
def api_painel_ng_strength_cached():
    """Devolve a força (só os 5 eixos usados pro grade geral) dos match_ids
    que JÁ estiverem em cache (de alguém ter aberto a Comparação de Força
    antes) — nunca dispara um Playwright novo aqui. É o que permite mostrar
    a diferença de força na listagem inteira sem custo extra: os jogos ainda
    não abertos simplesmente não aparecem na resposta, e a listagem mostra
    "—" pra eles até alguém abrir o modal daquele jogo alguma vez."""
    body = request.get_json(force=True, silent=True) or {}
    match_ids = body.get("match_ids") or []
    if not isinstance(match_ids, list):
        return jsonify({}), 400
    now = time.time()
    result = {}
    with _ng_strength_lock:
        for mid in match_ids:
            mid = str(mid)
            cached = _ng_strength_cache.get(mid)
            if cached and (now - cached["ts"]) < _NG_STRENGTH_TTL:
                d = cached["data"]
                result[mid] = {k: d[k] for k in _PAINEL_FORCA_RADAR_AXES if k in d}
    return jsonify(result)


# ── Metodologias — ranking de tipsters do tips.nowgoal.net ────────────────────
# Ao contrário dos widgets de análise (window._strength etc), o ranking e os
# palpites de cada usuário são JSON puro servido direto pelo backend deles —
# não precisa de Playwright, só requests normal.
TIPS_BASE = "https://tips.nowgoal.net"
TIPS_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Referer": "https://tips.nowgoal.net/",
}
_tips_ranking_cache = {}   # tipo -> {"ts":, "data":}
_tips_ranking_lock = threading.Lock()
_TIPS_RANKING_TTL = 600  # 10min

_tips_user_cache = {}   # user_id -> {"ts":, "data":}
_tips_user_lock = threading.Lock()
_TIPS_USER_TTL = 600


def _tips_fetch_ranking(kind):
    """kind: 1=Semana+Taxa de vitória, 2=Semana+ROI, 3=Mês+Taxa de vitória, 4=Mês+ROI
    (mapeamento confirmado testando os 2 seletores — Semana/Mês e Win Rate/ROI —
    no site original e comparando qual `type` cada combinação disparava)."""
    with _tips_ranking_lock:
        cached = _tips_ranking_cache.get(kind)
        if cached and (time.time() - cached["ts"]) < _TIPS_RANKING_TTL:
            return cached["data"]
    r = http_req.get(f"{TIPS_BASE}/home/getrankingjson", params={"type": kind}, headers=TIPS_HEADERS, timeout=15)
    r.raise_for_status()
    data = r.json()
    with _tips_ranking_lock:
        _tips_ranking_cache[kind] = {"ts": time.time(), "data": data}
    return data


_tips_article_cache = {}   # article_id -> {"ts":, "data":}
_tips_article_lock = threading.Lock()
_TIPS_ARTICLE_TTL = 3600  # 1h — o palpite de um artigo já publicado não muda mais


_TIPS_PICK_RE = re.compile(
    r"var odds1 = changeOdds\(([\d.]+).*?"
    r"var odds2 = changeOdds\(([\d.]+).*?"
    r"var odds3 = changeOdds\(([\d.]+).*?"
    r"data-kind='(\d)'>(Home|Over) .*?"
    r"data-kind='\d'>(Away|Under)",
    re.S,
)
_TIPS_FORMAT_RE = re.compile(r'dv\.format\(\s*"([^"]*)"\s*,\s*"([^"]*)"\s*,\s*"([^"]*)"\s*\)')


def _tips_fetch_article_pick(article_id):
    """A maioria das dicas do tips.nowgoal.net é paga (o texto da análise vem
    trocado por um parágrafo de marketing genérico e idêntico em todas), mas o
    palpite em si (que lado, com qual odd) fica visível de graça — só o texto
    de análise fica bloqueado. O detalhe é que essa parte NÃO vem pronta no
    HTML: o servidor gera um trechinho de JS que monta a div na hora (usando
    .format() estilo Python pra decidir qual lado leva a classe "on", que é o
    que marca o palpite de fato) — então em vez de rodar esse JS (precisaria
    de Playwright), extrai os valores literais desse script com regex."""
    with _tips_article_lock:
        cached = _tips_article_cache.get(article_id)
        if cached and (time.time() - cached["ts"]) < _TIPS_ARTICLE_TTL:
            return cached["data"]
    pick = None
    try:
        r = http_req.get(f"{TIPS_BASE}/article/{article_id}", headers=TIPS_HEADERS, timeout=10)
        r.raise_for_status()
        m = _TIPS_PICK_RE.search(r.text)
        fmt = _TIPS_FORMAT_RE.search(r.text)
        if m and fmt:
            home_odd, line_val, away_odd, kind_code, home_label, away_label = m.groups()
            home_on, _, away_on = fmt.groups()
            pick = {
                "kind": "AH" if kind_code == "2" else "OU",
                "line": line_val,
                "home_label": home_label, "home_odd": home_odd, "home_pick": home_on.strip() == "on",
                "away_label": away_label, "away_odd": away_odd, "away_pick": away_on.strip() == "on",
            }
    except Exception:
        pick = None
    with _tips_article_lock:
        _tips_article_cache[article_id] = {"ts": time.time(), "data": pick}
    return pick


def _tips_fetch_user_tips_raw(user_id):
    """Só a listagem básica (sem buscar o palpite de cada artigo) — usada tanto
    como base pra `_tips_fetch_user_tips` quanto pro "Previsões mais acertivas"
    (que só precisa de okind/isWin/isEnd, já vêm de graça nessa chamada)."""
    with _tips_user_lock:
        cached = _tips_user_cache.get(user_id)
        if cached and (time.time() - cached["ts"]) < _TIPS_USER_TTL:
            return cached["data"]
    params = {"userid": user_id, "kind": 0, "pre_page": 0, "req_page": 1, "endid": 0, "minid": 0, "type": 0}
    r = http_req.get(f"{TIPS_BASE}/user/getusertopiclist", params=params, headers=TIPS_HEADERS, timeout=15)
    r.raise_for_status()
    data = r.json()
    with _tips_user_lock:
        _tips_user_cache[user_id] = {"ts": time.time(), "data": data}
    return data


_TIPS_MARKET_LABELS = {2: "Handicap Asiático", 3: "Over/Under", 1: "1X2"}


def _tips_compute_best_market(user_id):
    """Agrupa as dicas encerradas do tipster por mercado (okind) e devolve o
    mercado onde ele mais acerta (com pelo menos 3 dicas encerradas nesse
    mercado, senão o recorte é pequeno demais pra significar algo)."""
    try:
        data = _tips_fetch_user_tips_raw(user_id)
    except Exception:
        return None
    tips = data.get("list") or []
    by_market = {}
    for t in tips:
        if not t.get("isEnd"):
            continue
        k = t.get("okind")
        m = by_market.setdefault(k, {"wins": 0, "total": 0})
        m["total"] += 1
        if t.get("isWin"):
            m["wins"] += 1
    best = None
    for k, m in by_market.items():
        if m["total"] < 3:
            continue
        pct = m["wins"] / m["total"]
        if best is None or pct > best["pct"]:
            best = {"kind": k, "label": _TIPS_MARKET_LABELS.get(k, "Outros"), "pct": pct, "wins": m["wins"], "total": m["total"]}
    return best


def _tips_fetch_user_tips(user_id):
    data = _tips_fetch_user_tips_raw(user_id)

    # Busca o palpite de cada dica em paralelo (a página do artigo é lenta,
    # ~2-3s cada — sequencial levaria mais de 1min pras 50 dicas de uma vez).
    tips = data.get("list") or []
    if tips:
        from concurrent.futures import ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=20) as ex:
            picks = list(ex.map(lambda t: _tips_fetch_article_pick(t["id"]), tips))
        for t, pick in zip(tips, picks):
            t["pick"] = pick

    with _tips_user_lock:
        _tips_user_cache[user_id] = {"ts": time.time(), "data": data}
    return data


@app.route("/api/painel/tips_ranking")
def api_painel_tips_ranking():
    kind = request.args.get("type", "1")
    if kind not in ("1", "2", "3", "4"):
        return jsonify({"error": "type inválido"}), 400
    try:
        data = _tips_fetch_ranking(kind)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return jsonify(data)


@app.route("/api/painel/tips_best_markets")
def api_painel_tips_best_markets():
    """Pro mesmo ranking (Semana/Mês × Taxa de Vitória/ROI já existente), busca
    em qual mercado (Handicap Asiático ou Over/Under) cada tipster mais acerta.
    Bem mais leve que /tips_user: só a listagem básica de cada um (sem os
    palpites por artigo), então dá pra buscar todo mundo do ranking em paralelo
    numa boa."""
    kind = request.args.get("type", "1")
    if kind not in ("1", "2", "3", "4"):
        return jsonify({"error": "type inválido"}), 400
    try:
        ranking = _tips_fetch_ranking(kind)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    users = ranking.get("list") or []
    from concurrent.futures import ThreadPoolExecutor
    with ThreadPoolExecutor(max_workers=20) as ex:
        best_markets = list(ex.map(lambda u: _tips_compute_best_market(u["uid"]), users))
    out = []
    for u, best in zip(users, best_markets):
        out.append({
            "uid": u["uid"], "uname": u["uname"], "uimg": u.get("uimg"), "rank": u["rank"], "rrc": u.get("rrc"),
            "best_market": best,
        })
    return jsonify({"list": out})


@app.route("/api/painel/tips_user")
def api_painel_tips_user():
    user_id = request.args.get("uid", "")
    if not user_id or not user_id.isdigit():
        return jsonify({"error": "uid inválido"}), 400
    try:
        data = _tips_fetch_user_tips(user_id)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return jsonify(data)


# ── Watchlist de tipsters — avisa no Telegram a cada prognóstico novo ─────────
TIPSTER_WATCH_DB_PATH = os.path.join(DATA_DIR, "tipster_watch.db")
_tipster_watch_db_lock = threading.Lock()
_TIPSTER_WATCH_POLL_INTERVAL = 5 * 60  # 5min — dicas novas valem a pena avisar rápido
TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID")


def _tipster_watch_db():
    conn = sqlite3.connect(TIPSTER_WATCH_DB_PATH, timeout=30)
    conn.execute("PRAGMA busy_timeout = 30000")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS watched_tipster (
            user_id TEXT PRIMARY KEY,
            user_name TEXT,
            added_ts INTEGER
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS seen_tip (
            user_id TEXT NOT NULL,
            tip_id TEXT NOT NULL,
            PRIMARY KEY (user_id, tip_id)
        )
    """)
    return conn


def _telegram_send(text):
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        print("[tipster-watch] TELEGRAM_BOT_TOKEN/TELEGRAM_CHAT_ID não configurados — aviso não enviado")
        return
    try:
        http_req.post(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage",
            data={"chat_id": TELEGRAM_CHAT_ID, "text": text, "parse_mode": "HTML"},
            timeout=10,
        )
    except Exception as e:
        print(f"[tipster-watch] Erro enviando pro Telegram: {e}")


@app.route("/api/painel/tips_watchlist")
def api_painel_tips_watchlist():
    with _tipster_watch_db_lock:
        conn = _tipster_watch_db()
        try:
            rows = conn.execute("SELECT user_id, user_name FROM watched_tipster").fetchall()
        finally:
            conn.close()
    return jsonify({"watched": [{"user_id": r[0], "user_name": r[1]} for r in rows]})


@app.route("/api/painel/tips_watch", methods=["POST"])
def api_painel_tips_watch():
    body = request.get_json(silent=True) or {}
    user_id = str(body.get("user_id", ""))
    user_name = body.get("user_name", "")
    if not user_id or not user_id.isdigit():
        return jsonify({"error": "user_id inválido"}), 400
    with _tipster_watch_db_lock:
        conn = _tipster_watch_db()
        try:
            conn.execute(
                "INSERT OR REPLACE INTO watched_tipster (user_id, user_name, added_ts) VALUES (?,?,?)",
                (user_id, user_name, int(time.time())),
            )
            # Marca as dicas já existentes como "vistas" — só quer aviso de dicas NOVAS
            # a partir de agora, não um spam retroativo de tudo que o tipster já postou.
            try:
                data = _tips_fetch_user_tips_raw(user_id)
                for t in (data.get("list") or []):
                    conn.execute("INSERT OR IGNORE INTO seen_tip (user_id, tip_id) VALUES (?,?)", (user_id, str(t["id"])))
            except Exception:
                pass
            conn.commit()
        finally:
            conn.close()
    github_storage.push_file_bg(TIPSTER_WATCH_DB_PATH, "tipster_watch.db")
    return jsonify({"ok": True})


@app.route("/api/painel/tips_unwatch", methods=["POST"])
def api_painel_tips_unwatch():
    body = request.get_json(silent=True) or {}
    user_id = str(body.get("user_id", ""))
    with _tipster_watch_db_lock:
        conn = _tipster_watch_db()
        try:
            conn.execute("DELETE FROM watched_tipster WHERE user_id=?", (user_id,))
            conn.execute("DELETE FROM seen_tip WHERE user_id=?", (user_id,))
            conn.commit()
        finally:
            conn.close()
    github_storage.push_file_bg(TIPSTER_WATCH_DB_PATH, "tipster_watch.db")
    return jsonify({"ok": True})


def _tipster_watch_poll_cycle():
    with _tipster_watch_db_lock:
        conn = _tipster_watch_db()
        try:
            watched = conn.execute("SELECT user_id, user_name FROM watched_tipster").fetchall()
        finally:
            conn.close()

    for user_id, user_name in watched:
        try:
            data = _tips_fetch_user_tips_raw(user_id)
        except Exception as e:
            print(f"[tipster-watch] Erro buscando dicas de {user_name} ({user_id}): {e}")
            continue
        tips = data.get("list") or []
        if not tips:
            continue

        with _tipster_watch_db_lock:
            conn = _tipster_watch_db()
            try:
                seen_ids = {row[0] for row in conn.execute("SELECT tip_id FROM seen_tip WHERE user_id=?", (user_id,)).fetchall()}
                new_tips = [t for t in tips if str(t["id"]) not in seen_ids]
                for t in new_tips:
                    conn.execute("INSERT OR IGNORE INTO seen_tip (user_id, tip_id) VALUES (?,?)", (user_id, str(t["id"])))
                conn.commit()
            finally:
                conn.close()

        for t in new_tips:
            try:
                pick = _tips_fetch_article_pick(t["id"])
            except Exception:
                pick = None
            match = f"{t.get('hname', '')} x {t.get('gname', '')}".strip()
            league = t.get("fname") or t.get("sname") or ""
            pick_line = "palpite indisponível"
            if pick:
                lado = pick["home_label"] if pick.get("home_pick") else pick["away_label"]
                odd = pick["home_odd"] if pick.get("home_pick") else pick["away_odd"]
                linha = f" {pick['line']}" if pick.get("line") else ""
                pick_line = f"{lado}{linha} @ {odd}"
            text = (
                f"🔔 <b>{user_name}</b> postou novo prognóstico\n"
                f"🏆 {league}\n"
                f"⚽ {match}\n"
                f"🎯 {pick_line}"
            )
            _telegram_send(text)


def _tipster_watch_loop():
    _github_sync_done.wait(timeout=120)  # espera a restauração do GitHub terminar antes do 1º ciclo
    while True:
        try:
            _tipster_watch_poll_cycle()
            github_storage.push_file_bg(TIPSTER_WATCH_DB_PATH, "tipster_watch.db")
        except Exception as e:
            print(f"[tipster-watch] Erro no ciclo: {e}")
        time.sleep(_TIPSTER_WATCH_POLL_INTERVAL)


def _painel_fetch_matches_nowgoal(force=False, date_str=None):
    """Versão NowGoal — só serve o dia atual por enquanto (o feed do NowGoal não
    tem parâmetro de data ainda descoberto; navegação por calendário fica limitada
    ao dia de hoje até isso ser mapeado)."""
    now = time.time()
    cache_key = date_str or "today"
    with _painel_lock:
        cached = _painel_cache.get(cache_key)
        if not force and cached is not None and (now - cached["ts"]) < _PAINEL_CACHE_TTL:
            return cached["data"]

        try:
            matches = _ng_fetch_today_matches()
        except Exception as e:
            if cached is not None:
                stale = dict(cached["data"])
                stale["stale"] = True
                stale["stale_error"] = str(e)
                return stale
            return {"error": str(e), "leagues": [], "updated_at": now, "date": date_str}

        # Anexa (quando encontrado) o link direto pra Betfair Exchange / Bolsa de
        # Aposta daquela partida específica, reaproveitando os links que o
        # RadarFutebol já resolve no feed público dele (casando por nome dos
        # times). Ver _find_radar_links — nunca derruba o carregamento do painel
        # se o RadarFutebol estiver fora do ar, só fica sem os links dessa vez.
        try:
            for m in matches:
                lb, lba, lr = _find_radar_links(m.get("home"), m.get("away"), m.get("ts"))
                m["link_betfair"] = lb
                m["link_bolsa"] = lba
                m["link_radar"] = lr
        except Exception as e:
            print(f"[radar-links] Erro anexando links: {e}")

        leagues_map = {}
        for m in matches:
            key = m["league_key"]
            lg = leagues_map.setdefault(key, {
                "key": key, "country": m["country"], "league_name": m["league_name"],
                "flag_url": "", "matches": [],
            })
            lg["matches"].append({k: v for k, v in m.items() if k not in ("league_key", "league_name", "country")})

        leagues = [lg for lg in leagues_map.values() if lg["matches"]]
        for lg in leagues:
            lg["matches"].sort(key=lambda x: x["ts"])

        # O NowGoal às vezes devolve uma página de bloqueio/verificação em vez do feed
        # de verdade — isso não estoura exceção (o request "funciona", só que o regex
        # não acha nenhum jogo pra extrair), e sem essa checagem o cache ficava com
        # "leagues: []" por 60s, mostrando "Nenhum jogo encontrado" com o app cheio de
        # jogos de verdade. Se veio vazio e já tínhamos dados bons antes, mantém os
        # dados antigos (marcados como stale) em vez de aceitar o vazio como válido.
        if not leagues and cached is not None and cached["data"].get("leagues"):
            stale = dict(cached["data"])
            stale["stale"] = True
            stale["stale_error"] = "NowGoal devolveu feed vazio (possível bloqueio temporário)"
            return stale

        data = {"leagues": leagues, "updated_at": now, "date": date_str}
        _painel_cache[cache_key] = {"ts": now, "data": data}
        return data


# ── Links diretos pra Betfair Exchange / Bolsa de Aposta ───────────────────────
# O RadarFutebol expõe publicamente (sem login) um feed SSE com os links já
# resolvidos pra cada partida — inclusive o "affid=radarfutebol" no link da
# Bolsa de Aposta (afiliado deles; usamos o mesmo código a pedido do usuário).
# Como não temos os IDs internos de cada partida nessas 2 plataformas, casar
# por nome dos times (mesmo _name_match usado pra SofaScore/Uniscore/FotMob)
# é o jeito de ligar nosso jogo ao link certo sem precisar integrar direto com
# Betfair/Bolsa (o Betfair, inclusive, bloqueia scraping direto por política).
_RADAR_LINKS_URL = (
    "https://www.radarfutebol.com/sse/home"
    "?idioma=pt-br&campoBusca=&somLigado=false&mostrarApenasJogosLive=false"
    "&mostrarApenasJogosFavoritos=false&countJogosMostrar=300"
    "&mostrarFiltroAcrescimo=false&filtroAcrescimoHt=1&filtroAcrescimoFt=1"
    "&filtroAcrescimoHtOperador=%3E%3D&filtroAcrescimoFtOperador=%3E%3D"
    "&filtroAcrescimoCondicao=ou&mostrarApenasJogosOraculo=false"
    "&mostrarApenasJogosBolsa=false&mostrarApenasJogosBetfair=false"
    "&mostrarApenasJogosOver=false&mostrarApenasJogosLayCs=false"
    "&favoritoVencendo=false&favoritoPerdendo=false&casaVencendo=false"
    "&visitanteVencendo=false&empatado=false&filtroAlertas=false"
    "&filtroDiferencaXg=false&ordemInicio=false"
)
_RADAR_LINKS_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Referer":    "https://www.radarfutebol.com/",
    "Accept":     "text/event-stream",
}
_RADAR_LINKS_TTL = 90  # feed muda pouco de um minuto pro outro, evita bater toda hora
_radar_links_cache = {"ts": 0.0, "events": []}
_radar_links_lock = threading.Lock()


def _get_radar_futebol_links():
    """Busca o feed público (SSE) do RadarFutebol e extrai, de cada partida,
    o link pronto pra Betfair Exchange e pra Bolsa de Aposta. Só lê a primeira
    linha 'data: {...}' do stream e fecha a conexão — não fica pendurado
    esperando os próximos eventos ao vivo do SSE."""
    with _radar_links_lock:
        if time.time() - _radar_links_cache["ts"] < _RADAR_LINKS_TTL and _radar_links_cache["events"]:
            return _radar_links_cache["events"]
    try:
        r = http_req.get(_RADAR_LINKS_URL, headers=_RADAR_LINKS_HEADERS, stream=True, timeout=15)
        # O servidor não declara charset no Content-Type do SSE, então o requests
        # cai no fallback ISO-8859-1 (padrão HTTP pra text/*) e decodifica errado
        # qualquer acento (ex: "Fenerbahçe" virava "FenerbahÃ§e") — isso quebrava
        # o casamento de nome pra times com acento/cedilha. Força UTF-8 (é o que
        # o feed realmente manda).
        r.encoding = "utf-8"
        payload = None
        for raw_line in r.iter_lines(decode_unicode=True):
            if raw_line and raw_line.startswith("data:"):
                payload = raw_line[len("data:"):].strip()
                break
        r.close()
        if not payload:
            return _radar_links_cache["events"]

        obj = json.loads(payload)
        events = []
        for camp in obj.get("campeonatos", []):
            for ev in (camp.get("eventos") or {}).values():
                link_betfair = ev.get("linkBetfair")
                link_bolsa = ev.get("linkBolsadeaposta")
                if not (link_betfair or link_bolsa):
                    continue
                ts = None
                inicio = ev.get("inicio")
                if inicio:
                    try:
                        ts = datetime.strptime(inicio, "%Y-%m-%d %H:%M:%S").timestamp()
                    except ValueError:
                        ts = None
                slug_evento = ev.get("slugEvento")
                id_evento = ev.get("idEvento")
                link_radar = f"https://www.radarfutebol.com/radar/{slug_evento}/{id_evento}" if slug_evento and id_evento else None
                events.append({
                    "home": ev.get("timeCasa") or "",
                    "away": ev.get("timeFora") or "",
                    "ts": ts,
                    "link_betfair": link_betfair,
                    "link_bolsa": link_bolsa,
                    "link_radar": link_radar,
                })

        with _radar_links_lock:
            _radar_links_cache["ts"] = time.time()
            _radar_links_cache["events"] = events
        print(f"[radar-links] {len(events)} jogos com link Betfair/Bolsa de Aposta")
        return events
    except Exception as e:
        print(f"[radar-links] Erro buscando feed do RadarFutebol: {e}")
        return _radar_links_cache["events"]


def _find_radar_links(home, away, ts=None):
    """Casa (home, away) do nosso feed com os eventos do RadarFutebol. Quando
    o nome bate em mais de uma partida (raro — 2 times com nome parecido
    jogando no mesmo dia), desempata pelo horário mais próximo; se mesmo assim
    a diferença passar de 3h, não arrisca linkar pro jogo errado."""
    if not home or not away:
        return None, None, None
    events = _get_radar_futebol_links()
    candidates = [ev for ev in events if _name_match(home, ev["home"]) and _name_match(away, ev["away"])]
    if not candidates:
        return None, None, None
    if len(candidates) > 1 and ts:
        candidates.sort(key=lambda ev: abs((ev["ts"] or 0) - ts))
        if abs((candidates[0]["ts"] or 0) - ts) > 3 * 3600:
            return None, None, None
    ev = candidates[0]
    return ev.get("link_betfair"), ev.get("link_bolsa"), ev.get("link_radar")


# ── Pré-carga de Força (versão leve) — enche o _ng_strength_cache sozinho, em
# segundo plano, pra coluna "Força" do Painel Principal não depender de
# alguém abrir a Comparação de força manualmente. Uma 1ª versão tentava cobrir
# TODOS os jogos do dia com 2 workers simultâneos e derrubou o app em produção
# (Chromium headless nas costas um do outro, sem pausa, estourou a memória do
# container do Railway). Essa versão é bem mais cautelosa: só 1 Playwright por
# vez (nunca mais que 1 rodando junto com o que alguém abrir na hora, então no
# pior caso são 2 simultâneos — dentro do limite já usado em todo o resto do
# sistema), com uma pausa entre cada partida, e só cobre uma JANELA de horário
# (jogos ao vivo + começando nas próximas horas) em vez do dia inteiro — jogo
# muito distante no futuro não interessa agora mesmo, e entra na janela
# conforme o horário dele se aproxima.
_NG_STRENGTH_PREFETCH_WINDOW_PAST = 2 * 3600    # cobre jogos que começaram até 2h atrás (ainda podem estar ao vivo)
_NG_STRENGTH_PREFETCH_WINDOW_FUTURE = 3 * 3600  # e que começam nas próximas 3h
_NG_STRENGTH_PREFETCH_DELAY = 6        # segundos de respiro entre uma partida e a próxima
_NG_STRENGTH_PREFETCH_CYCLE_GAP = 45   # segundos entre uma checagem "o que falta" e a próxima
_ng_strength_prefetch_queue = queue.Queue()
_ng_strength_prefetch_queued = set()  # dedup — evita enfileirar o mesmo jogo 2x antes dele ser processado
_ng_strength_prefetch_queued_lock = threading.Lock()


def _ng_strength_prefetch_worker():
    while True:
        mid = _ng_strength_prefetch_queue.get()
        try:
            with _ng_strength_lock:
                cached = _ng_strength_cache.get(mid)
                fresh = cached and (time.time() - cached["ts"]) < _NG_STRENGTH_TTL
            if not fresh:
                try:
                    _ng_fetch_strength(mid)
                except Exception as e:
                    print(f"[forca-prefetch] Erro no jogo {mid}: {e}")
        finally:
            with _ng_strength_prefetch_queued_lock:
                _ng_strength_prefetch_queued.discard(mid)
            _ng_strength_prefetch_queue.task_done()
        time.sleep(_NG_STRENGTH_PREFETCH_DELAY)


def _ng_strength_prefetch_filler_loop():
    _github_sync_done.wait(timeout=120)
    while True:
        try:
            matches_data = _painel_fetch_matches_nowgoal()
            now = time.time()
            added = 0
            for lg in matches_data.get("leagues", []):
                for m in lg["matches"]:
                    ts = m.get("ts")
                    if not ts or not (now - _NG_STRENGTH_PREFETCH_WINDOW_PAST <= ts <= now + _NG_STRENGTH_PREFETCH_WINDOW_FUTURE):
                        continue
                    mid = m.get("event_id")
                    if not mid:
                        continue
                    mid = str(mid)
                    with _ng_strength_lock:
                        cached = _ng_strength_cache.get(mid)
                        fresh = cached and (now - cached["ts"]) < _NG_STRENGTH_TTL
                    if fresh:
                        continue
                    with _ng_strength_prefetch_queued_lock:
                        if mid in _ng_strength_prefetch_queued:
                            continue
                        _ng_strength_prefetch_queued.add(mid)
                    _ng_strength_prefetch_queue.put(mid)
                    added += 1
            if added:
                print(f"[forca-prefetch] {added} jogo(s) da janela atual enfileirado(s)")
        except Exception as e:
            print(f"[forca-prefetch] Erro buscando jogos do dia: {e}")
        time.sleep(_NG_STRENGTH_PREFETCH_CYCLE_GAP)


# ── Backup de Força — arquiva em disco (JSON, um arquivo por jogo, sincronizado
# com o GitHub igual momentum_history/shotmap_history) os dados de H-T,
# escanteio, odds (1/X/2/Over/Under) e força (H2H/Estado/Ataque/Defesa/Valor)
# de todo jogo do dia que já ENCERROU. Diferente da pré-carga acima (que
# re-varria jogos ao vivo sem parar e derrubou o app em produção), aqui cada
# jogo só entra na fila UMA VEZ — placar de jogo encerrado não muda mais, e se
# o arquivo já existe no disco nem tenta de novo — então o volume total fica
# limitado a "quantos jogos terminam por dia", nunca cresce sem limite. Mesmo
# assim usa só 1 worker + pausa entre partidas, pela mesma cautela de sempre
# com o Playwright.
_FORCA_BACKUP_DELAY = 8           # segundos de respiro entre uma partida e a próxima
_FORCA_BACKUP_SCAN_INTERVAL = 120 # segundos entre uma varredura "quem terminou" e a próxima
_forca_backup_queue = queue.Queue()
_forca_backup_queued = set()   # event_id em fila — evita duplicar antes de processar
_forca_backup_queued_lock = threading.Lock()


def _forca_backup_path(date_str, event_id):
    return os.path.join(FORCA_HISTORY_DIR, f"{date_str}_{event_id}.json")


def _forca_backup_worker():
    while True:
        date_str, m = _forca_backup_queue.get()
        event_id = str(m.get("event_id"))
        try:
            path = _forca_backup_path(date_str, event_id)
            if not os.path.exists(path):
                strength = _ng_fetch_strength(event_id)
                axes = {k: strength[k] for k in ("battle", "state", "attack", "defend", "market") if k in strength}
                payload = {
                    "event_id": event_id, "date": date_str,
                    "league": m.get("league_name"), "country": m.get("country"),
                    "home": m.get("home"), "away": m.get("away"),
                    "score_home": m.get("score_home"), "score_away": m.get("score_away"),
                    "ht_home": m.get("ht_home"), "ht_away": m.get("ht_away"),
                    "corner_home": m.get("corner_home"), "corner_away": m.get("corner_away"),
                    "odd_1": m.get("odd_1"), "odd_x": m.get("odd_x"), "odd_2": m.get("odd_2"),
                    "odd_over": m.get("odd_over"), "odd_under": m.get("odd_under"),
                    "forca": axes,
                }
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(payload, f, ensure_ascii=False, indent=2)
                github_storage.push_file_bg(path, f"forca_history/{date_str}_{event_id}.json")
                print(f"[forca-backup] Salvo: {date_str}_{event_id}.json ({m.get('home')} x {m.get('away')})")
        except Exception as e:
            print(f"[forca-backup] Erro no jogo {event_id}: {e}")
        finally:
            with _forca_backup_queued_lock:
                _forca_backup_queued.discard(event_id)
            _forca_backup_queue.task_done()
        time.sleep(_FORCA_BACKUP_DELAY)


def _forca_backup_scan_loop():
    _github_sync_done.wait(timeout=120)
    while True:
        try:
            data = _painel_fetch_matches_nowgoal()
            date_str = datetime.now().strftime("%Y-%m-%d")
            added = 0
            for lg in data.get("leagues", []):
                for m in lg["matches"]:
                    if m.get("time") != "Encerrado":
                        continue
                    event_id = str(m.get("event_id") or "")
                    if not event_id:
                        continue
                    if os.path.exists(_forca_backup_path(date_str, event_id)):
                        continue
                    with _forca_backup_queued_lock:
                        if event_id in _forca_backup_queued:
                            continue
                        _forca_backup_queued.add(event_id)
                    _forca_backup_queue.put((date_str, m))
                    added += 1
            if added:
                print(f"[forca-backup] {added} jogo(s) finalizado(s) novo(s) enfileirado(s)")
        except Exception as e:
            print(f"[forca-backup] Erro escaneando jogos finalizados: {e}")
        time.sleep(_FORCA_BACKUP_SCAN_INTERVAL)


def _painel_fetch_matches(force=False, date_str=None):
    """date_str: "YYYY-MM-DD" opcional — qualquer dia navegável pelo calendário
    do BetExplorer. None/"" = dia atual (comportamento igual ao botão "Hoje")."""
    now = time.time()
    cache_key = date_str or "today"
    with _painel_lock:
        cached = _painel_cache.get(cache_key)
        if not force and cached is not None and (now - cached["ts"]) < _PAINEL_CACHE_TTL:
            return cached["data"]

        date_params = None
        if date_str:
            try:
                y, mo, d = date_str.split("-")
                date_params = {"year": y, "month": mo, "day": d}
            except ValueError:
                date_params = None

        try:
            html_1x2 = _be_fetch_bettype_html("1x2", date_params)
            html_ou  = _be_fetch_bettype_html("ou", date_params)
            matches_1x2, leagues_order = _be_parse_bettype(html_1x2)
            matches_ou, _ = _be_parse_bettype(html_ou)
        except Exception as e:
            # BetExplorer fora do ar/bloqueando (429) — em vez de deixar a página vazia
            # com erro, serve o último resultado que já funcionou (mesmo vencido), com
            # um aviso de que os dados podem estar desatualizados. Só mostra erro puro
            # se nunca conseguimos buscar nada pra esse dia ainda.
            if cached is not None:
                stale = dict(cached["data"])
                stale["stale"] = True
                stale["stale_error"] = str(e)
                return stale
            return {"error": str(e), "leagues": [], "updated_at": now, "date": date_str}

        leagues_map = {}
        for lg in leagues_order:
            leagues_map.setdefault(lg["key"], dict(lg, matches=[]))

        for event_id, m in matches_1x2.items():
            lg = leagues_map.get(m["league_key"])
            if lg is None:
                continue
            ou = matches_ou.get(event_id)
            odds = m["odds"] + [None] * (3 - len(m["odds"]))
            ou_odds = (ou["odds"] if ou else []) + [None, None]
            lg["matches"].append({
                "event_id": event_id,
                "time": m["status_text"],
                "home": m["home"], "away": m["away"],
                "home_logo": m.get("home_logo"), "away_logo": m.get("away_logo"),
                "score_home": m["score_home"], "score_away": m["score_away"],
                "match_url": (BETEXPLORER_BASE + m["match_url"]) if m["match_url"] else None,
                "odd_1": odds[0], "odd_x": odds[1], "odd_2": odds[2],
                "ou_line": ou["line"] if ou else None,
                "odd_over": ou_odds[0], "odd_under": ou_odds[1],
                "ts": m["ts"],
            })

        leagues = [lg for lg in leagues_map.values() if lg["matches"]]
        for lg in leagues:
            lg["matches"].sort(key=lambda x: x["ts"])

        data = {"leagues": leagues, "updated_at": now, "date": date_str}
        _painel_cache[cache_key] = {"ts": now, "data": data}
        return data


@app.route("/api/painel/matches")
def api_painel_matches():
    force = request.args.get("force") == "1"
    date_str = request.args.get("date") or None  # "YYYY-MM-DD"
    return jsonify(_painel_fetch_matches_nowgoal(force=force, date_str=date_str))


# ── Widget de análise — "Últimos resultados" de cada time (BetExplorer) ────────
# Essa seção do BetExplorer só é montada via JS depois que a página carrega (o
# endpoint /gres/ajax/match-content.php exige um token "ts" gerado no client,
# sem padrão fixo pra reproduzir com um simples requests.get). Por isso usamos o
# Playwright (headless) UMA VEZ por partida só pra "ler" da página já renderizada
# o token de torneio ("par") e o ID de cada time — depois disso, trocar entre
# 5/10/15/todos os resultados ou "só esse torneio"/"todos os torneios" é um
# requests.get direto em /res/ajax/team-matches.php (rápido, sem precisar mais
# de browser), então o cache do contexto vale a pena mesmo custando ~2-4s a mais
# na primeira vez que alguém abre a análise de um jogo.
_be_context_cache = {}   # match_url -> {"ts":, "data": {"par":, "home":{"id","name"}, "away":{...}}}
_be_context_lock = threading.Lock()
_BE_CONTEXT_TTL = 6 * 3600
_be_playwright_semaphore = threading.Semaphore(2)  # evita várias janelas headless simultâneas

# Abrir o widget de análise dispara 3 chamadas em paralelo (últimos resultados casa/
# fora + confronto direto) que TODAS precisam do mesmo contexto do jogo. Sem isso,
# as 3 viam o cache vazio ao mesmo tempo e cada uma abria seu próprio Chromium
# headless pra carregar a MESMA página — 3 sessões simultâneas na mesma URL, uma
# assinatura bem óbvia de bot pro BetExplorer. Esse lock por URL garante que só a
# primeira chamada realmente busca; as outras esperam e reaproveitam o resultado.
_be_context_inflight = {}   # match_url -> threading.Lock (só existe enquanto a busca está em andamento)
_be_context_inflight_guard = threading.Lock()


def _be_fetch_match_context(match_url, _attempt=1):
    with _be_context_lock:
        cached = _be_context_cache.get(match_url)
        if cached and (time.time() - cached["ts"]) < _BE_CONTEXT_TTL:
            return cached["data"]

    with _be_context_inflight_guard:
        lock = _be_context_inflight.get(match_url)
        is_leader = lock is None
        if is_leader:
            lock = threading.Lock()
            lock.acquire()
            _be_context_inflight[match_url] = lock

    if not is_leader:
        lock.acquire()  # espera o líder terminar
        lock.release()
        with _be_context_lock:
            cached = _be_context_cache.get(match_url)
        if cached:
            return cached["data"]
        raise RuntimeError("Não foi possível carregar os últimos resultados dessa partida.")

    try:
        return _be_fetch_match_context_uncached(match_url)
    finally:
        with _be_context_inflight_guard:
            _be_context_inflight.pop(match_url, None)
        lock.release()


def _be_fetch_match_context_uncached(match_url, _attempt=1):
    with _be_context_lock:
        cached = _be_context_cache.get(match_url)
        if cached and (time.time() - cached["ts"]) < _BE_CONTEXT_TTL:
            return cached["data"]

    from playwright.sync_api import sync_playwright

    par, teams = None, []
    try:
        with _be_playwright_semaphore:
            with sync_playwright() as p:
                browser = p.chromium.launch(
                    headless=True,
                    args=["--disable-blink-features=AutomationControlled"],
                )
                try:
                    page = browser.new_page(
                        user_agent=BETEXPLORER_HEADERS["User-Agent"],
                        viewport={"width": 1280, "height": 900},
                        locale="pt-BR",
                    )
                    page.goto(match_url, wait_until="domcontentloaded", timeout=30000)
                    try:
                        page.wait_for_load_state("networkidle", timeout=8000)
                    except Exception:
                        pass  # segue mesmo se não ficar 100% ocioso — o seletor abaixo é o que realmente importa
                    page.wait_for_selector("[id^='lm_'][id$='_sel_type']", timeout=20000, state="attached")
                    selects = page.query_selector_all("[id^='lm_'][id$='_sel_type']")
                    headers = page.query_selector_all(".last-results__title .componentHeader")
                    for i, sel in enumerate(selects[:2]):
                        onchange = sel.get_attribute("onchange") or ""
                        m = re.search(r"match_change_team_matches\('([^']*)',\s*'([^']*)',\s*'([^']*)'", onchange)
                        if not m:
                            continue
                        par = m.group(1)
                        name = headers[i].inner_text() if i < len(headers) else ""
                        name = re.sub(r"^.*?:\s*", "", name).strip()
                        teams.append({"id": m.group(3), "name": name})
                finally:
                    browser.close()
    except Exception:
        if _attempt < 2:
            time.sleep(1.5)
            return _be_fetch_match_context_uncached(match_url, _attempt=_attempt + 1)
        raise RuntimeError("O BetExplorer não respondeu a tempo pra carregar os últimos resultados dessa partida. Tente novamente em instantes.")

    if par is None or len(teams) < 2:
        if _attempt < 2:
            time.sleep(1.5)
            return _be_fetch_match_context_uncached(match_url, _attempt=_attempt + 1)
        raise RuntimeError("Não foi possível carregar os últimos resultados dessa partida.")

    data = {"par": par, "home": teams[0], "away": teams[1]}
    with _be_context_lock:
        _be_context_cache[match_url] = {"ts": time.time(), "data": data}
    return data


def _be_parse_team_matches(html):
    soup = BeautifulSoup(html, "html.parser")
    rows = []
    for block in soup.select(".head-to-head__row"):
        date_span = block.select_one(".head-to-head__date .mobileHidden")
        date_text = date_span.get_text(strip=True) if date_span else ""

        home_el = block.select_one(".table-main__participantHome p, .table-main__participantHome div")
        away_el = block.select_one(".table-main__participantAway p, .table-main__participantAway div")
        home_name = home_el.get_text(strip=True) if home_el else ""
        away_name = away_el.get_text(strip=True) if away_el else ""
        home_logo_el = block.select_one(".homeImgMutual")
        away_logo_el = block.select_one(".awayImgMutual")
        home_logo = (BETEXPLORER_BASE + home_logo_el.get("src")) if home_logo_el and home_logo_el.get("src", "").startswith("/") else (home_logo_el.get("src") if home_logo_el else None)
        away_logo = (BETEXPLORER_BASE + away_logo_el.get("src")) if away_logo_el and away_logo_el.get("src", "").startswith("/") else (away_logo_el.get("src") if away_logo_el else None)

        result_div = None
        for d in block.select(".last-results__form-results"):
            if "desktopHidden" not in (d.get("class") or []):
                result_div = d
                break
        result, score_home, score_away = None, None, None
        if result_div:
            for c in (result_div.get("class") or []):
                if c.startswith("last-results__form-results-"):
                    result = c.rsplit("-", 1)[-1]  # W / D / L
            nums = [t for s in result_div.find_all("span") if (t := s.get_text(strip=True)) and t != ":"]
            if len(nums) >= 2:
                score_home, score_away = nums[0], nums[1]

        odds = []
        odds_wrap = block.select_one(".last-results__odds-align")
        if odds_wrap:
            for odd_el in odds_wrap.select(".table-main__odd"):
                span = odd_el.find("span")
                odds.append(span.get("data-odd") if span else None)

        link_tag = block.select_one("a[href]")
        match_href = (BETEXPLORER_BASE + link_tag.get("href")) if link_tag else None

        rows.append({
            "date": date_text, "home": home_name, "away": away_name,
            "home_logo": home_logo, "away_logo": away_logo,
            "score_home": score_home, "score_away": score_away,
            "result": result, "odds": odds, "match_url": match_href,
        })
    return rows


def _be_fetch_team_last_results(match_url, side, count=5, all_tournaments=False):
    ctx = _be_fetch_match_context(match_url)
    team = ctx.get(side)
    if not team:
        raise ValueError("side inválido (use 'home' ou 'away')")
    event_id = match_url.rstrip("/").split("/")[-1]
    params = {
        "par": ctx["par"], "event": event_id, "team": team["id"],
        "type": 2 if all_tournaments else 1, "count": count, "lang": "br",
    }
    r = _be_get(f"{BETEXPLORER_BASE}/res/ajax/team-matches.php", params=params)
    return {"team_name": team["name"], "rows": _be_parse_team_matches(r.text)}


def _be_country_league_path(match_url):
    """Extrai país/liga do caminho da URL do jogo — usado pra montar a URL da
    página de confrontos diretos (mutual-matches), que é por país+liga."""
    m = re.search(r"/football/([^/]+)/([^/]+)/", match_url)
    if not m:
        raise ValueError("Não foi possível identificar a liga a partir da URL do jogo.")
    return m.group(1), m.group(2)


def _be_extract_td_odd(td):
    classes = td.get("class") or []
    colored = "colored" in classes
    val = td.get("data-odd")
    if not val:
        inner = td.find(attrs={"data-odd": True})
        val = inner.get("data-odd") if inner else None
    return val, colored


def _be_parse_h2h(html):
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table", id="js-mutual-table")
    if not table:
        return []
    seasons = []
    current = None
    for tr in table.find_all("tr"):
        if "head-to-head__header" in (tr.get("class") or []):
            th = tr.find("th")
            link = th.find("a") if th else None
            current = {
                "season": link.get_text(strip=True) if link else (th.get_text(strip=True) if th else ""),
                "matches": [],
            }
            seasons.append(current)
            continue
        tds = tr.find_all("td")
        if len(tds) < 7 or current is None:
            continue
        home_name = tds[0].get_text(strip=True)
        away_name = tds[1].get_text(strip=True)
        score_link = tds[2].find("a")
        score_text = score_link.get_text(strip=True) if score_link else tds[2].get_text(strip=True)
        match_href = (BETEXPLORER_BASE + score_link.get("href")) if score_link and score_link.get("href") else None
        score_parts = re.split(r"[:\-]", score_text)
        score_home = score_parts[0].strip() if len(score_parts) == 2 else None
        score_away = score_parts[1].strip() if len(score_parts) == 2 else None
        odds = []
        for td in tds[3:6]:
            val, colored = _be_extract_td_odd(td)
            odds.append({"value": val, "won": colored})
        current["matches"].append({
            "home": home_name, "away": away_name,
            "score_home": score_home, "score_away": score_away,
            "match_url": match_href, "date": tds[6].get_text(strip=True), "odds": odds,
        })
    return seasons


def _be_h2h_summary(seasons, home_name, away_name):
    wins_home = wins_away = draws = 0
    for season in seasons:
        for m in season["matches"]:
            try:
                sh, sa = int(m["score_home"]), int(m["score_away"])
            except (TypeError, ValueError):
                continue
            if sh == sa:
                draws += 1
            else:
                winner = m["home"] if sh > sa else m["away"]
                if winner == home_name:
                    wins_home += 1
                elif winner == away_name:
                    wins_away += 1
    total = wins_home + wins_away + draws
    pct_home = round(100 * wins_home / total) if total else 0
    pct_away = round(100 * wins_away / total) if total else 0
    return {
        "wins_home": wins_home, "wins_away": wins_away, "draws": draws,
        "pct_home": pct_home, "pct_away": pct_away, "pct_draw": 100 - pct_home - pct_away if total else 0,
    }


def _be_fetch_h2h(match_url):
    ctx = _be_fetch_match_context(match_url)
    country, league = _be_country_league_path(match_url)
    r = _be_get(
        f"{BETEXPLORER_BASE}/br/football/{country}/{league}/mutual-matches/",
        params={"home": ctx["home"]["id"], "away": ctx["away"]["id"], "where": 0},
    )
    seasons = _be_parse_h2h(r.text)
    summary = _be_h2h_summary(seasons, ctx["home"]["name"], ctx["away"]["name"])
    return {
        "home_name": ctx["home"]["name"], "away_name": ctx["away"]["name"],
        "summary": summary, "seasons": seasons,
    }


@app.route("/api/painel/h2h")
def api_painel_h2h():
    match_url = request.args.get("match_url", "")
    if not match_url.startswith(BETEXPLORER_BASE):
        return jsonify({"error": "match_url inválido"}), 400
    try:
        data = _be_fetch_h2h(match_url)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return jsonify(data)


# ── Estatísticas de Confronto Direto (aba Análise) — % vitórias/empates/derrotas,
# médias de gols, BTTS, clean sheets, over/under (tempo integral) ──────────────
# Removido de propósito o placar do intervalo (HT): ele não vem na lista de
# confrontos (mutual-matches), só o placar final — pra ter o HT seria preciso
# 1 requisição extra por jogo do filtro na página de cada confronto antigo, o
# que multiplicava bastante o tráfego pro BetExplorer. Ver histórico do commit
# se precisar recuperar essa lógica.


def _be_h2h_stats(match_url, tournament="1", count="5"):
    h2h = _be_fetch_h2h(match_url)
    home_name, away_name = h2h["home_name"], h2h["away_name"]

    seasons = h2h["seasons"]
    if tournament == "1":
        seasons = [s for s in seasons if not re.search(r"copa|cup", s["season"], re.I)]
    flat = [m for s in seasons for m in s["matches"]]  # mais recente primeiro

    wanted = len(flat) if count == "20" else int(count)
    if len(flat) < wanted:
        return {"enough": False, "available": len(flat), "wanted": wanted}
    used = flat[:wanted]

    # Só tempo integral (FT) — o placar do intervalo exigiria 1 requisição extra
    # por jogo do filtro na página de cada confronto antigo, o que multiplicava
    # bastante o tráfego pro BetExplorer (e contribuiu pro bloqueio 429 que a
    # gente teve). Removido de propósito: FT já vem de graça na mesma lista de
    # confrontos, sem nenhuma chamada adicional.
    n = len(used)
    wins = draws = losses = 0
    goals_for_ft = goals_against_ft = 0
    btts_yes = 0
    clean_sheets_ft = 0
    failed_to_score_ft = 0
    over25 = over15 = 0

    for m in used:
        try:
            sh, sa = int(m["score_home"]), int(m["score_away"])
        except (TypeError, ValueError):
            continue
        is_home_team = m["home"] == home_name
        gf = sh if is_home_team else sa
        ga = sa if is_home_team else sh
        goals_for_ft += gf
        goals_against_ft += ga
        if gf > ga:
            wins += 1
        elif gf == ga:
            draws += 1
        else:
            losses += 1
        if gf == 0:
            failed_to_score_ft += 1
        if ga == 0:
            clean_sheets_ft += 1
        if sh > 0 and sa > 0:
            btts_yes += 1
        total_goals = sh + sa
        if total_goals > 2.5:
            over25 += 1
        if total_goals > 1.5:
            over15 += 1

    def pct(x, total):
        return round(100 * x / total) if total else None

    def avg(x, total):
        return round(x / total, 2) if total else None

    # Cada estatística de time é espelhada pro lado oposto: numa H2H, a vitória de
    # um é a derrota do outro, o gol marcado por um é o gol sofrido pelo outro etc.
    # BTTS e Over/Under são do CONFRONTO (não têm "lado"), por isso ficam repetidos
    # dos dois lados.
    pct_wins_ft, pct_draws_ft, pct_losses_ft = pct(wins, n), pct(draws, n), pct(losses, n)
    avg_gf_ft, avg_ga_ft = avg(goals_for_ft, n), avg(goals_against_ft, n)
    pct_clean_ft = pct(clean_sheets_ft, n)
    pct_fts_ft = pct(failed_to_score_ft, n)

    home_stats = {
        "pct_wins_ft": pct_wins_ft, "pct_draws_ft": pct_draws_ft, "pct_losses_ft": pct_losses_ft,
        "avg_goals_for_ft": avg_gf_ft, "avg_goals_against_ft": avg_ga_ft,
        "pct_clean_sheet_ft": pct_clean_ft, "pct_failed_to_score_ft": pct_fts_ft,
    }
    away_stats = {
        "pct_wins_ft": pct_losses_ft, "pct_draws_ft": pct_draws_ft, "pct_losses_ft": pct_wins_ft,
        "avg_goals_for_ft": avg_ga_ft, "avg_goals_against_ft": avg_gf_ft,
        "pct_clean_sheet_ft": pct_fts_ft, "pct_failed_to_score_ft": pct_clean_ft,
    }

    return {
        "enough": True, "matches_used": n,
        "home_name": home_name, "away_name": away_name,
        "home": home_stats, "away": away_stats,
        "pct_btts_yes": pct(btts_yes, n), "pct_btts_no": pct(n - btts_yes, n),
        "pct_over25_ft": pct(over25, n), "pct_under25_ft": pct(n - over25, n),
        "pct_over15_ft": pct(over15, n), "pct_under15_ft": pct(n - over15, n),
    }


@app.route("/api/painel/h2h_stats")
def api_painel_h2h_stats():
    match_url = request.args.get("match_url", "")
    tournament = request.args.get("tournament", "1")
    count = request.args.get("count", "5")
    if not match_url.startswith(BETEXPLORER_BASE):
        return jsonify({"error": "match_url inválido"}), 400
    try:
        data = _be_h2h_stats(match_url, tournament=tournament, count=count)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return jsonify(data)


@app.route("/api/painel/debug_be")
def api_painel_debug_be():
    """Diagnóstico temporário — testa cada etapa do acesso ao BetExplorer
    isoladamente (requests simples vs Playwright) pra descobrir exatamente
    onde está travando em produção, sem depender de logs do Railway."""
    import traceback
    result = {}

    t0 = time.time()
    try:
        r = http_req.get(f"{BETEXPLORER_BASE}/br/", headers=BETEXPLORER_HEADERS, timeout=15)
        result["plain_request"] = {"ok": True, "status": r.status_code, "seconds": round(time.time() - t0, 2), "len": len(r.text)}
    except Exception as e:
        result["plain_request"] = {"ok": False, "error": f"{type(e).__name__}: {e}", "seconds": round(time.time() - t0, 2)}

    match_url = request.args.get(
        "match_url", "https://www.betexplorer.com/br/football/brazil/serie-a-betano/coritiba-palmeiras/zmCty4Ji/"
    )

    t0 = time.time()
    try:
        from playwright.sync_api import sync_playwright
        # mesmo semáforo usado pelo resto do app — sem isso, um pico de chamadas a
        # esse endpoint de diagnóstico podia sozinho ocupar todas as threads do
        # gunicorn (só 1 worker/4 threads) e travar o site inteiro pra todo mundo.
        acquired = _be_playwright_semaphore.acquire(timeout=10)
        if not acquired:
            return jsonify({"error": "Playwright ocupado (semáforo cheio) — tente de novo em instantes."}), 503
        try:
            with sync_playwright() as p:
                t_launch = time.time()
                browser = p.chromium.launch(
                    headless=True, args=["--disable-blink-features=AutomationControlled"], timeout=15000,
                )
                result["playwright_launch"] = {"ok": True, "seconds": round(time.time() - t_launch, 2)}
                try:
                    page = browser.new_page(
                        user_agent=BETEXPLORER_HEADERS["User-Agent"], viewport={"width": 1280, "height": 900}, locale="pt-BR",
                    )
                    t_goto = time.time()
                    page.goto(match_url, wait_until="domcontentloaded", timeout=15000)
                    result["playwright_goto"] = {"ok": True, "seconds": round(time.time() - t_goto, 2), "title": page.title()}

                    t_sel = time.time()
                    try:
                        page.wait_for_selector("[id^='lm_'][id$='_sel_type']", timeout=10000, state="attached")
                        result["playwright_selector"] = {"ok": True, "seconds": round(time.time() - t_sel, 2)}
                    except Exception as e:
                        result["playwright_selector"] = {"ok": False, "error": f"{type(e).__name__}: {e}", "seconds": round(time.time() - t_sel, 2)}
                        # salva um pedaço do HTML pra ver se veio página de bloqueio/captcha
                        html = page.content()
                        result["page_snippet"] = html[:1500]
                        result["page_length"] = len(html)
                finally:
                    browser.close()
        except Exception as e:
            result["playwright_error"] = f"{type(e).__name__}: {e}"
            result["playwright_traceback"] = traceback.format_exc()
        finally:
            _be_playwright_semaphore.release()
    except Exception as e:
        result["playwright_error"] = f"{type(e).__name__}: {e}"
        result["playwright_traceback"] = traceback.format_exc()
    result["playwright_total_seconds"] = round(time.time() - t0, 2)

    return jsonify(result)


# ── Widget de análise — Classificações / Forma / Over-Under / HT-FT / Marcadores
# Diferente do "últimos resultados" (que precisa de Playwright pro token "ts" do
# JOGO), o "ts" da TABELA/liga já vem embutido direto no HTML estático da página
# do torneio (ex: /br/football/brazil/serie-a-betano/) — então dá pra pegar com
# um requests.get simples, sem precisar de browser headless.
_be_tournament_ts_cache = {}   # (country, league) -> {"ts":, "value": token}
_be_tournament_ts_lock = threading.Lock()
_BE_TOURNAMENT_TS_TTL = 6 * 3600
_BE_STANDINGS_CACHE_TTL = 300   # 5min — tabela de classificação não muda a cada minuto
_be_standings_cache = {}
_be_standings_lock = threading.Lock()


def _be_fetch_tournament_ts(country, league):
    key = (country, league)
    with _be_tournament_ts_lock:
        cached = _be_tournament_ts_cache.get(key)
        if cached and (time.time() - cached["ts"]) < _BE_TOURNAMENT_TS_TTL:
            return cached["value"]

    r = _be_get(f"{BETEXPLORER_BASE}/br/football/{country}/{league}/")
    m = re.search(r"standings/\?table=table&table_sub=&ts=([^&\"]+)", r.text)
    if not m:
        raise RuntimeError("Não foi possível encontrar o token de classificação dessa liga.")
    token = m.group(1)
    with _be_tournament_ts_lock:
        _be_tournament_ts_cache[key] = {"ts": time.time(), "value": token}
    return token


def _be_parse_standings_table(html):
    """Retorna {variant: {"columns":[...], "rows":[...]}}. `variant` é o sufixo
    do id do box (ex.: "10" pra Forma-10-jogos, "2.5" pra linha do Over/Under),
    ou "default" quando só existe uma tabela (Classificações, HT/FT)."""
    soup = BeautifulSoup(html, "html.parser")
    result = {}
    for box in soup.select("[id^='box-table-type-']"):
        box_id = box.get("id", "")
        m = re.match(r"box-table-type-\d+-(.+)$", box_id)
        variant = m.group(1) if m else "default"
        table = box.find("table")
        if not table:
            continue

        columns = []
        thead = table.find("thead")
        if thead:
            for th in thead.find_all("th"):
                columns.append({
                    "key": th.get("data-type", ""), "label": th.get_text(strip=True), "title": th.get("title", ""),
                })

        rows = []
        tbody = table.find("tbody")
        for tr in (tbody.find_all("tr") if tbody else []):
            tds = tr.find_all("td")
            if not tds:
                continue
            logo_span = tr.select_one(".team-logo, .flag")
            logo_url = None
            if logo_span and logo_span.get("style"):
                mlogo = re.search(r"url\(([^)]+)\)", logo_span["style"])
                if mlogo:
                    src = mlogo.group(1)
                    logo_url = (BETEXPLORER_BASE + src) if src.startswith("/") else src

            cells = {}
            for col, td in zip(columns, tds):
                if col["key"] in ("form", "last_5"):
                    badges = []
                    for a in td.select("a[class*='form-']"):
                        cls = " ".join(a.get("class") or [])
                        mres = re.search(r"form-(w|d|l|s|over|under)\b", cls)
                        badges.append(mres.group(1) if mres else "s")
                    cells[col["key"]] = badges
                else:
                    cells[col["key"]] = td.get_text(strip=True)
            rows.append({"logo": logo_url, "cells": cells})
        result[variant] = {"columns": columns, "rows": rows}
    return result


def _be_fetch_standings(match_url, table, table_sub=""):
    country, league = _be_country_league_path(match_url)
    cache_key = (country, league, table, table_sub)
    now = time.time()
    with _be_standings_lock:
        cached = _be_standings_cache.get(cache_key)
        if cached and (now - cached["ts"]) < _BE_STANDINGS_CACHE_TTL:
            return cached["data"]

    token = _be_fetch_tournament_ts(country, league)
    r = _be_get(
        f"{BETEXPLORER_BASE}/br/football/{country}/{league}/standings/",
        params={"table": table, "table_sub": table_sub, "ts": token, "dcheck": 0, "as-ajax": 1, "l": "br"},
    )
    data = _be_parse_standings_table(r.text)
    with _be_standings_lock:
        _be_standings_cache[cache_key] = {"ts": now, "data": data}
    return data


@app.route("/api/painel/standings")
def api_painel_standings():
    match_url = request.args.get("match_url", "")
    table = request.args.get("table", "table")  # table | form | over_under | ht_ft | top_scorers
    table_sub = request.args.get("sub", "")      # "" | home | away
    if not match_url.startswith(BETEXPLORER_BASE):
        return jsonify({"error": "match_url inválido"}), 400
    if table not in ("table", "form", "over_under", "ht_ft", "top_scorers"):
        return jsonify({"error": "table inválido"}), 400
    try:
        data = _be_fetch_standings(match_url, table, table_sub)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return jsonify(data)


@app.route("/api/painel/last_results")
def api_painel_last_results():
    match_url = request.args.get("match_url", "")
    side = request.args.get("side", "home")
    count = request.args.get("count", "5")
    all_tournaments = request.args.get("all") == "1"
    if not match_url.startswith(BETEXPLORER_BASE):
        return jsonify({"error": "match_url inválido"}), 400
    try:
        data = _be_fetch_team_last_results(match_url, side, count=count, all_tournaments=all_tournaments)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return jsonify(data)


@app.route("/version")
def version():
    return jsonify({"version": APP_VERSION, "ts": datetime.now().isoformat()})

@app.route("/")
def index():
    resp = send_from_directory("static", "index.html")
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"]        = "no-cache"
    resp.headers["Expires"]       = "0"
    return resp


@app.route("/api/matches")
def api_matches():
    matches, source = load_predictions()
    # Retorna lista resumida (sem detalhes) para a listagem
    summary = []
    for m in matches:
        conv_casa, _ = _convicao_score(m, True)
        conv_fora, _ = _convicao_score(m, False)
        summary.append({
            "id": m.get("url_detalhes", "").split("/compare/teams/")[-1],
            "hora": m.get("hora"),
            "liga": m.get("liga"),
            "pais": m.get("pais"),
            "casa": m.get("casa"),
            "fora": m.get("fora"),
            "tip": m.get("tip"),
            "odds_1": m.get("odds_1"),
            "odds_x": m.get("odds_x"),
            "odds_2": m.get("odds_2"),
            "odds_h1": m.get("odds_h1"),
            "odds_hx": m.get("odds_hx"),
            "odds_h2": m.get("odds_h2"),
            "over_1_5": m.get("over_1_5"),
            "over_2_5": m.get("over_2_5"),
            "over_3_5": m.get("over_3_5"),
            "bts": m.get("bts"),
            "ots": m.get("ots"),
            "votos_favor": m.get("votos_favor"),
            "votos_contra": m.get("votos_contra"),
            "tem_detalhes": "detalhes" in m and "erro" not in m.get("detalhes", {}),
            "url_detalhes": m.get("url_detalhes"),
            "data_coleta": m.get("data_coleta"),
            "convicao_casa": conv_casa,
            "convicao_fora": conv_fora,
        })
    return jsonify({"total": len(summary), "source": source, "matches": summary})


@app.route("/api/match/<path:match_id>")
def api_match_detail(match_id):
    matches, _ = load_predictions()
    for m in matches:
        mid = m.get("url_detalhes", "").split("/compare/teams/")[-1]
        if mid == match_id:
            conv_casa, h2h = _convicao_score(m, True)
            conv_fora, _ = _convicao_score(m, False)
            m = dict(m)
            m["convicao_casa"] = conv_casa
            m["convicao_fora"] = conv_fora
            m["h2h_confronto_direto"] = h2h
            return jsonify(m)
    abort(404)


@app.route("/api/status")
def api_status():
    matches, source = load_predictions()
    has_details = sum(1 for m in matches if "detalhes" in m and "erro" not in m.get("detalhes", {}))
    last_updated = matches[0].get("data_coleta") if matches else None
    return jsonify({
        "total_partidas": len(matches),
        "com_detalhes": has_details,
        "source": source,
        "last_updated": last_updated,
    })


# ── SCRAPE REMOTO ────────────────────────────────────────────────────────────
_scrape_state = {
    "running": False,
    "started_at": None,
    "finished_at": None,
    "result": None,   # "ok" | "error"
    "message": "",
    "total": 0,
}
_scrape_lock = threading.Lock()


def _run_scrape_bg(fetch_details: bool = False):
    """Raspa jogos do StatArea em background e salva predictions_full.json."""
    global _scrape_state
    with _scrape_lock:
        _scrape_state.update({"running": True, "started_at": datetime.now().isoformat(),
                              "finished_at": None, "result": None, "message": "Raspando...", "total": 0})
    try:
        from scraper import scrape_all, save_json
        matches = scrape_all(fetch_details=fetch_details, delay=0.5 if not fetch_details else 2.0)
        path = os.path.join(DATA_DIR, "predictions_full.json")
        save_json(matches, path)
        # Sobe para GitHub para persistir no próximo redeploy
        github_storage.push_file_bg(path, "predictions_full.json")
        with _scrape_lock:
            _scrape_state.update({
                "running": False, "finished_at": datetime.now().isoformat(),
                "result": "ok", "message": f"{len(matches)} partidas raspadas com sucesso.",
                "total": len(matches),
            })
        print(f"[scrape] ✓ {len(matches)} partidas salvas em predictions_full.json")
    except Exception as e:
        with _scrape_lock:
            _scrape_state.update({
                "running": False, "finished_at": datetime.now().isoformat(),
                "result": "error", "message": str(e), "total": 0,
            })
        print(f"[scrape] ✗ Erro: {e}")


@app.route("/api/scrape/now", methods=["POST"])
def api_scrape_now():
    """Dispara raspagem remota do StatArea (sem detalhes — rápida).
    Sem autenticação: raspa só dados públicos do StatArea, sem risco."""
    with _scrape_lock:
        if _scrape_state["running"]:
            return jsonify({"ok": False, "running": True,
                            "message": "Raspagem já em andamento — aguarde.", **_scrape_state})

    full = request.args.get("full", "0") == "1"
    threading.Thread(target=_run_scrape_bg, args=(full,), daemon=True).start()
    return jsonify({"ok": True, "running": True, "message": "Raspagem iniciada em background.",
                    "full": full})


@app.route("/api/scrape/status")
def api_scrape_status():
    """Retorna o estado atual da raspagem remota."""
    with _scrape_lock:
        return jsonify(dict(_scrape_state))


BACKTEST_DIR = os.path.join(DATA_DIR, "backtest")


def load_backtest_dates():
    """Retorna lista de datas disponíveis no backtest, ordenadas desc."""
    if not os.path.exists(BACKTEST_DIR):
        return []
    files = glob.glob(os.path.join(BACKTEST_DIR, "????-??-??.json"))
    dates = sorted(
        [os.path.basename(f).replace(".json", "") for f in files],
        reverse=True
    )
    return dates


def load_backtest_day(date_str: str):
    path = os.path.join(BACKTEST_DIR, f"{date_str}.json")
    if not os.path.exists(path):
        return None
    with open(path, encoding="utf-8") as f:
        return json.load(f)


@app.route("/api/backtest/<date_str>")
def api_backtest_day(date_str):
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
        abort(400)
    matches = load_backtest_day(date_str)
    if matches is None:
        abort(404)
    # Calcula estatísticas resumidas do dia
    tips = {"1": 0, "X": 0, "2": 0, "N/D": 0}
    for m in matches:
        t = m.get("tip", "N/D")
        tips[t] = tips.get(t, 0) + 1

    over60      = [m for m in matches if (m.get("odds_1") or 0) >= 60 or (m.get("odds_2") or 0) >= 60]
    bts_high    = [m for m in matches if (m.get("bts") or 0) >= 60]
    over25_high = [m for m in matches if (m.get("over_2_5") or 0) >= 65]
    with_result = [m for m in matches if m.get("resultado")]

    # Acurácia dos TIPs
    decididos  = [m for m in matches if m.get("tip_acertou") is not None]
    acertaram  = [m for m in decididos if m.get("tip_acertou") is True]
    tip_acuracia = round(len(acertaram) / len(decididos) * 100, 1) if decididos else None

    return jsonify({
        "date": date_str,
        "total": len(matches),
        "tips": tips,
        "alta_confianca": len(over60),
        "bts_provavel": len(bts_high),
        "over25_provavel": len(over25_high),
        "com_resultado": len(with_result),
        "tips_acertaram": len(acertaram),
        "tips_decididos": len(decididos),
        "tip_acuracia": tip_acuracia,
        "matches": matches,
    })


@app.route("/api/passadas")
def api_passadas():
    """Retorna todas as partidas passadas (todos os dias de backtest) em uma lista única,
    já com placar final e HT, pra alimentar a aba Partidas Passadas."""
    dates = load_backtest_dates()
    all_matches = []
    for date_str in dates:
        matches = load_backtest_day(date_str) or []
        for m in matches:
            if m.get("resultado"):
                m.setdefault("data_partida", date_str)
                all_matches.append(m)

    all_matches.sort(key=lambda m: (m.get("data_partida") or "", m.get("hora") or ""), reverse=True)

    decididos = [m for m in all_matches if m.get("tip_acertou") is not None]
    acertaram = [m for m in decididos if m.get("tip_acertou") is True]

    return jsonify({
        "total": len(all_matches),
        "dates": len(dates),
        "tip_acuracia": round(len(acertaram) / len(decididos) * 100, 1) if decididos else None,
        "matches": all_matches,
    })


@app.route("/api/patterns")
def api_patterns():
    """Cruza padrões do backtest (probabilidades × placares reais)."""
    dates = load_backtest_dates()

    def range_key(val):
        if val is None: return None
        v = int(val)
        if v >= 80: return "80+"
        if v >= 70: return "70-79"
        if v >= 60: return "60-69"
        if v >= 50: return "50-59"
        return None

    def add(d, key, hit):
        if key not in d:
            d[key] = {"total": 0, "hit": 0}
        d[key]["total"] += 1
        if hit:
            d[key]["hit"] += 1

    tip_pat   = {}   # tip_type → range → {total, hit}
    over15    = {}
    over25    = {}
    over35    = {}
    bts_pat   = {}
    score_freq    = {}   # placar FT real → contagem
    ht_score_freq = {}   # placar HT real → contagem
    momento_mat   = {}   # "CasaMomento|ForaMomento" → {total, hit, casa, fora}
    total         = 0
    total_ht      = 0

    for date in dates:
        matches = load_backtest_day(date)
        if not matches:
            continue
        if isinstance(matches, dict):
            matches = matches.get("matches", [])

        for m in matches:
            if not m.get("resultado"):
                continue
            try:
                parts = re.split(r"[-:]", m["resultado"])
                gc, gf = int(parts[0]), int(parts[1])
            except Exception:
                continue

            tg = gc + gf
            total += 1
            score_freq[f"{gc}-{gf}"] = score_freq.get(f"{gc}-{gf}", 0) + 1

            # HT score
            ht_raw = m.get("ht_score")
            if ht_raw:
                try:
                    ht_parts = re.split(r"[-:]", str(ht_raw))
                    hg, ag = int(ht_parts[0]), int(ht_parts[1])
                    ht_key = f"{hg}-{ag}"
                    ht_score_freq[ht_key] = ht_score_freq.get(ht_key, 0) + 1
                    total_ht += 1
                except Exception:
                    pass

            # Over/Under
            for field, thr, acc in [
                ("over_1_5", 1, over15),
                ("over_2_5", 2, over25),
                ("over_3_5", 3, over35),
            ]:
                rng = range_key(m.get(field))
                if rng:
                    add(acc, rng, tg > thr)

            # BTS
            rng = range_key(m.get("bts"))
            if rng:
                add(bts_pat, rng, gc > 0 and gf > 0)

            # TIP accuracy (só TIPs simples)
            tip = m.get("tip", "")
            acertou = m.get("tip_acertou")
            if acertou is not None and tip in ("1", "2", "X"):
                odds_map = {"1": "odds_1", "2": "odds_2", "X": "odds_x"}
                rng = range_key(m.get(odds_map[tip]))
                if rng:
                    if tip not in tip_pat:
                        tip_pat[tip] = {}
                    add(tip_pat[tip], rng, acertou)

            # Matriz de Momento
            mc = m.get("momento_casa")
            mf = m.get("momento_fora")
            if mc and mf and acertou is not None:
                key = f"{mc}|{mf}"
                if key not in momento_mat:
                    momento_mat[key] = {"total": 0, "hit": 0, "casa": mc, "fora": mf}
                momento_mat[key]["total"] += 1
                if acertou:
                    momento_mat[key]["hit"] += 1

    def with_pct(d):
        return {k: {**v, "pct": round(v["hit"] / v["total"] * 100, 1) if v["total"] else 0}
                for k, v in sorted(d.items())}

    # Adiciona % à matriz de momento
    momento_mat_pct = {
        k: {**v, "pct": round(v["hit"] / v["total"] * 100, 1) if v["total"] else 0}
        for k, v in momento_mat.items()
    }

    return jsonify({
        "total_partidas": total,
        "datas": dates,
        "tip":     {k: with_pct(v) for k, v in tip_pat.items()},
        "over_1_5": with_pct(over15),
        "over_2_5": with_pct(over25),
        "over_3_5": with_pct(over35),
        "bts":        with_pct(bts_pat),
        "score_freq":    score_freq,
        "ht_score_freq": ht_score_freq,
        "total_ht":      total_ht,
        "momento_matrix": momento_mat_pct,
    })


_uniscore_full_cache = {"ts": 0, "live": []}

_UNISCORE_PERIOD_OFFSET = {
    "1st_half": 0, "2nd_half": 45,
    "overtime1": 90, "overtime2": 105,
    "penalties": None, "halftime": None, "break_time": None,
}

def _uniscore_minuto(e):
    """Minuto real da partida, calculado a partir de time.currentPeriodStartTimestamp
    (quando o período atual começou) — o campo 'tempo'/status.description só traz o
    NOME do período (ex: '2nd_half'), não o minuto em si."""
    status_desc = e.get("status", {}).get("description", "")
    if status_desc not in _UNISCORE_PERIOD_OFFSET:
        return None
    offset = _UNISCORE_PERIOD_OFFSET[status_desc]
    if offset is None:
        return None  # intervalo/pênaltis — sem minuto corrido
    start_ts = e.get("time", {}).get("currentPeriodStartTimestamp")
    if not start_ts:
        return None
    elapsed_min = int((time.time() - start_ts) / 60)
    if elapsed_min < 0:
        return None
    return f"{offset + elapsed_min}'"

@app.route("/api/radar/live")
def api_radar_live():
    """Lista TODOS os jogos ao vivo via UniScore (todos os locales + paginação)."""
    # Cache de 90s para o endpoint público (mais curto que o cache interno)
    if time.time() - _uniscore_full_cache["ts"] < 90 and _uniscore_full_cache["live"]:
        live = _uniscore_full_cache["live"]
        return jsonify({"live": live, "total": len(live)})

    live = []
    all_by_id = {}
    for locale in _UNISCORE_LOCALES:
        page = 1
        while True:
            try:
                r = http_req.post(
                    f"{_UNISCORE_API}/sport/football/events/live-v2/locale/{locale}",
                    headers=_UNISCORE_HEADERS,
                    json={"page": page},
                    params={"language": "pt-BR"},
                    timeout=12,
                )
                if r.status_code not in (200, 201):
                    break
                data   = r.json().get("data", {})
                events = data.get("events", [])
                pag    = data.get("pagination", {})
                for e in events:
                    if e.get("status", {}).get("type") != "inprogress":
                        continue
                    eid = e["id"]
                    if eid in all_by_id:
                        continue
                    hs  = e.get("homeScore", {}) or {}
                    aws = e.get("awayScore", {}) or {}
                    all_by_id[eid] = {
                        "id":        eid,
                        "casa":      e.get("homeTeam", {}).get("name", ""),
                        "fora":      e.get("awayTeam", {}).get("name", ""),
                        "liga":      e.get("tournament", {}).get("name", ""),
                        "pais":      e.get("tournament", {}).get("category", {}).get("name", ""),
                        "tempo":     e.get("status", {}).get("description", ""),
                        "minuto":    _uniscore_minuto(e),
                        "golCasaFt": hs.get("current", 0),
                        "golForaFt": aws.get("current", 0),
                        "golCasaHt": hs.get("period1", 0),
                        "golForaHt": aws.get("period1", 0),
                        "cartaoCasa": 0,
                        "cartaoFora": 0,
                    }
                if not pag.get("hasNextPage"):
                    break
                page += 1
                if page > 5:
                    break
            except Exception as e:
                print(f"[live] Erro locale={locale}: {e}")
                break

    live = list(all_by_id.values())
    print(f"[live] {len(live)} jogos ao vivo retornados")

    # Mesmo link direto pra Betfair Exchange / Bolsa de Aposta usado no Painel
    # Principal (ver _find_radar_links) — aqui não temos horário de início (o
    # jogo já tá em andamento), então em caso raro de nome ambíguo fica com o
    # 1º candidato em vez de desempatar por horário.
    try:
        for m in live:
            lb, lba, lr = _find_radar_links(m.get("casa"), m.get("fora"))
            m["link_betfair"] = lb
            m["link_bolsa"] = lba
            m["link_radar"] = lr
    except Exception as e:
        print(f"[radar-links] Erro anexando links (ao vivo): {e}")

    # Reaproveita as mesmas odds 1x2/Over-Under do Painel Principal (NowGoal) —
    # casa (casa, fora) do Ao Vivo (Uniscore) com o jogo equivalente lá pelo
    # nome dos times, mesma técnica do _find_radar_links. Busca a lista uma vez
    # só (não por partida) porque _painel_fetch_matches_nowgoal já cacheia por
    # conta própria, mas repetir a chamada pra cada jogo ainda seria bater no
    # lock/dict à toa dezenas de vezes por request.
    try:
        painel_data = _painel_fetch_matches_nowgoal()
        painel_matches = [pm for lg in painel_data.get("leagues", []) for pm in lg["matches"]]
    except Exception as e:
        print(f"[painel-odds] Erro buscando odds do Painel Principal: {e}")
        painel_matches = []
    for m in live:
        pm = next((p for p in painel_matches
                   if _name_match(m.get("casa") or "", p.get("home") or "")
                   and _name_match(m.get("fora") or "", p.get("away") or "")), None)
        m["odd_1"] = pm.get("odd_1") if pm else None
        m["odd_x"] = pm.get("odd_x") if pm else None
        m["odd_2"] = pm.get("odd_2") if pm else None
        m["odd_over"] = pm.get("odd_over") if pm else None
        m["odd_under"] = pm.get("odd_under") if pm else None

    # Se a fetch retornou 0 jogos mas o cache anterior tem dados recentes (< 5 min),
    # mantém o cache antigo para evitar sidebar vazia por falha temporária da API
    if not live and _uniscore_full_cache["live"] and (time.time() - _uniscore_full_cache["ts"] < 300):
        print(f"[live] API retornou 0 jogos — mantendo cache anterior com {len(_uniscore_full_cache['live'])} jogos")
        return jsonify({"live": _uniscore_full_cache["live"], "total": len(_uniscore_full_cache["live"]), "stale": True})

    _uniscore_full_cache["ts"]   = time.time()
    _uniscore_full_cache["live"] = live
    return jsonify({"live": live, "total": len(live), "stale": False})


# ── Cache simples de momentum em memória (evita abrir browser repetidamente) ──
_momentum_cache = {}
_momentum_lock  = threading.Lock()   # proteção para acesso concorrente

# ── Cache de shotmap ao vivo: acumula durante o jogo para não perder ao FT ──
_SHOTMAP_CACHE_FILE = os.path.join(DATA_DIR, ".shotmap_cache.json")
_shotmap_lock       = threading.Lock()

def _load_shotmap_cache() -> dict:
    """Carrega cache de shotmap do disco (sobrevive a restarts)."""
    try:
        if os.path.exists(_SHOTMAP_CACHE_FILE):
            with open(_SHOTMAP_CACHE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            print(f"[shotmap] Cache restaurado: {len(data)} jogo(s)")
            return data
    except Exception as e:
        print(f"[shotmap] Erro ao carregar cache: {e}")
    return {}

def _save_shotmap_cache(cache: dict):
    """Persiste cache de shotmap no disco."""
    try:
        with open(_SHOTMAP_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False)
    except Exception as e:
        print(f"[shotmap] Erro ao salvar cache: {e}")

_shotmap_live_cache = _load_shotmap_cache()


def _fetch_live_matches_for_monitor():
    """Busca lista de jogos ao vivo via UniScore para o monitor de fundo."""
    matches = _get_uniscore_live_matches()
    return [
        {"id": m["id"], "casa": m["home"], "fora": m["away"], "liga": ""}
        for m in matches
    ]


_sofa_live_cache = {"ts": 0, "events": []}
_sofa_live_lock  = threading.Lock()

import unicodedata

def _norm(s: str) -> str:
    """Normaliza string: minúsculo, sem acento, sem caracteres especiais."""
    s = s.lower().strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s

# Sufixos genéricos a ignorar na comparação de nomes de times
_TEAM_SUFFIXES = {
    "fc", "cf", "ac", "sc", "bc", "bk", "sk", "fk", "nk", "rk",
    "united", "utd", "city", "town", "rovers", "wanderers",
    "sporting", "sport", "club", "atletico", "atletico", "deportivo",
    "real", "de", "do", "da", "dos", "las", "los", "el",
}

def _strip_suffixes(words: set) -> set:
    """Remove palavras genéricas de um conjunto de tokens."""
    return {w for w in words if w not in _TEAM_SUFFIXES}

def _name_match(a: str, b: str) -> bool:
    """Verifica se dois nomes de times batem (fuzzy aprimorado)."""
    na, nb = _norm(a), _norm(b)

    # 1. Igualdade exata
    if na == nb:
        return True

    # 2. Substring direta
    if na in nb or nb in na:
        return True

    # 3. Palavras com >= 3 chars em comum (anterior era >= 4)
    words_a = {w for w in na.split() if len(w) >= 3}
    words_b = {w for w in nb.split() if len(w) >= 3}
    if words_a & words_b:
        return True

    # 4. Palavras sem sufixos genéricos — evita falso positivo por "FC"/"Sporting"
    core_a = _strip_suffixes(words_a)
    core_b = _strip_suffixes(words_b)
    if core_a and core_b and core_a & core_b:
        return True

    # 5. Nomes curtos (≤ 4 chars): exige igualdade exata entre os tokens curtos
    short_a = {w for w in na.split() if len(w) <= 4}
    short_b = {w for w in nb.split() if len(w) <= 4}
    if short_a and short_b and short_a == short_b and len(short_a) >= 1:
        return True

    return False


def _get_sofa_live_events():
    """Busca todos os jogos de futebol ao vivo do api.sofascore.com. Cache 2min."""
    with _sofa_live_lock:
        if time.time() - _sofa_live_cache["ts"] < 120:
            return _sofa_live_cache["events"]
    try:
        s = http_req.Session()
        s.headers.update(_SOFA_HEADERS)
        r = s.get("https://api.sofascore.com/api/v1/sport/football/events/live", timeout=12)
        if r.status_code == 200:
            events = r.json().get("events", [])
            parsed = [
                {
                    "id":   e["id"],
                    "home": e.get("homeTeam", {}).get("name", ""),
                    "away": e.get("awayTeam", {}).get("name", ""),
                }
                for e in events
            ]
            print(f"[sofa-live] {len(parsed)} jogos ao vivo")
            with _sofa_live_lock:
                _sofa_live_cache["ts"]     = time.time()
                _sofa_live_cache["events"] = parsed
            return parsed
    except Exception as e:
        print(f"[sofa-live] Erro: {e}")
    return []


def _find_sofa_event_id(casa: str, fora: str):
    """Encontra o ID correto do SofaScore pelo nome dos times."""
    if not casa or not fora:
        return None
    events = _get_sofa_live_events()
    for ev in events:
        if _name_match(casa, ev["home"]) and _name_match(fora, ev["away"]):
            print(f"[sofa-live] Match: {ev['home']} vs {ev['away']} id={ev['id']}")
            return ev["id"]
    return None

_SOFA_HEADERS = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Referer":         "https://www.sofascore.com/",
    "Origin":          "https://www.sofascore.com",
    "Accept":          "application/json, text/plain, */*",
    "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
    "Cache-Control":   "no-cache",
}


_fotmob_live_cache    = {"ts": 0, "matches": []}
_fotmob_live_lock     = threading.Lock()

_UNISCORE_HEADERS = {
    "User-Agent":   "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Origin":       "https://uniscore.com",
    "Referer":      "https://uniscore.com/pt-BR/football",
    "Accept":       "application/json, text/plain, */*",
    "Content-Type": "application/json",
}
_UNISCORE_API     = "https://api.unik8s.com/api/v2"
_uniscore_cache   = {"ts": 0, "matches": []}
_uniscore_lock    = threading.Lock()


_UNISCORE_LOCALES = ["BR", "EU", "AS", "AF", "NA", "SA", "OC"]

def _get_uniscore_live_matches():
    """Busca TODAS as partidas ao vivo do UniScore (todos os locales + paginação).
    Cache de 2 minutos. Retorna lista de {id, homeId, awayId, home, away}."""
    with _uniscore_lock:
        if time.time() - _uniscore_cache["ts"] < 120:
            return _uniscore_cache["matches"]

    all_by_id = {}
    for locale in _UNISCORE_LOCALES:
        page = 1
        while True:
            try:
                r = http_req.post(
                    f"{_UNISCORE_API}/sport/football/events/live-v2/locale/{locale}",
                    headers=_UNISCORE_HEADERS,
                    json={"page": page},
                    params={"language": "pt-BR"},
                    timeout=12,
                )
                if r.status_code not in (200, 201):
                    break
                data      = r.json().get("data", {})
                events    = data.get("events", [])
                pag       = data.get("pagination", {})
                for e in events:
                    if e.get("status", {}).get("type") == "inprogress":
                        eid = e["id"]
                        if eid not in all_by_id:
                            all_by_id[eid] = {
                                "id":     eid,
                                "homeId": e.get("homeTeam", {}).get("id", ""),
                                "awayId": e.get("awayTeam", {}).get("id", ""),
                                "home":   e.get("homeTeam", {}).get("name", ""),
                                "away":   e.get("awayTeam", {}).get("name", ""),
                            }
                if not pag.get("hasNextPage"):
                    break
                page += 1
                if page > 5:   # safety cap
                    break
            except Exception as e:
                print(f"[uniscore] Erro live locale={locale} page={page}: {e}")
                break

    matches = list(all_by_id.values())
    print(f"[uniscore] {len(matches)} partidas ao vivo (todos os locales)")
    with _uniscore_lock:
        _uniscore_cache["ts"]      = time.time()
        _uniscore_cache["matches"] = matches
    return matches


def _find_uniscore_id(casa, fora):
    """Encontra ID UniScore pelo nome dos times (fuzzy com normalização).
    Retorna dict {id, homeId, awayId} ou None."""
    if not casa or not fora:
        return None
    matches = _get_uniscore_live_matches()
    for m in matches:
        if _name_match(casa, m["home"]) and _name_match(fora, m["away"]):
            # try/except só no print: nomes com caracteres fora do cp1252 (ş, Č
            # etc) derrubavam a requisição inteira com UnicodeEncodeError no
            # console do Windows local — é só diagnóstico, não pode quebrar o
            # endpoint por causa disso.
            try:
                print(f"[uniscore] Match: {m['home']} vs {m['away']} id={m['id']}")
            except UnicodeEncodeError:
                pass
            return {"id": m["id"], "homeId": m["homeId"], "awayId": m["awayId"]}
    if matches:
        sample = [(m["home"], m["away"]) for m in matches[:5]]
        try:
            print(f"[uniscore] Sem match p/ '{casa}' vs '{fora}'. Amostra: {sample}")
        except UnicodeEncodeError:
            pass
    return None


def _uniscore_stats_to_flat(stats_list):
    """Converte lista de períodos UniScore para dict plano por período.
    Retorna {"ALL": {stat: {home,away,homeValue,awayValue}}, "1ST": {...}, "2ND": {...}}"""
    result = {}
    for period_data in (stats_list or []):
        period = period_data.get("period", "ALL")
        flat   = {}
        for group in period_data.get("groups", []):
            for item in group.get("statisticsItems", []):
                name = item.get("name")
                if name and name not in flat:
                    flat[name] = {
                        "home":      item.get("home", "0"),
                        "away":      item.get("away", "0"),
                        "homeValue": item.get("homeValue", 0),
                        "awayValue": item.get("awayValue", 0),
                    }
        result[period] = flat
    return result


def _fetch_uniscore_graph(uni_match):
    """Busca graphPoints + estatísticas por período do UniScore.
    uni_match = {id, homeId, awayId}"""
    uniscore_id = uni_match["id"]
    home_id     = uni_match.get("homeId", "")
    away_id     = uni_match.get("awayId", "")

    # Graph (momentum)
    r = http_req.get(
        f"{_UNISCORE_API}/football/event/{uniscore_id}/graph",
        headers=_UNISCORE_HEADERS, timeout=12,
    )
    r.raise_for_status()
    pts = r.json().get("data", {}).get("graphPoints", [])

    # Incidents (gols)
    goals = []
    try:
        ri = http_req.get(
            f"{_UNISCORE_API}/football/event/{uniscore_id}/incidents",
            headers=_UNISCORE_HEADERS, timeout=12,
        )
        if ri.status_code == 200:
            for inc in ri.json().get("data", {}).get("incidents", []):
                if inc.get("incidentType") == "goal":
                    player     = inc.get("player") or inc.get("scorer") or {}
                    added_time = inc.get("addedTime") or 0
                    goals.append({
                        "minute":     inc.get("time", 0),
                        "addedTime":  added_time,
                        "team":       "home" if inc.get("isHome") else "away",
                        "player":     player.get("shortName") or player.get("name") or "",
                        "ownGoal":    inc.get("incidentClass") == "ownGoal",
                    })
    except Exception:
        pass

    # Estatísticas por período (Todos / 1º / 2º)
    statistics_periods = {}
    if home_id and away_id:
        try:
            rs = http_req.get(
                f"{_UNISCORE_API}/football/event/{uniscore_id}/home/{home_id}/away/{away_id}/statistics",
                headers=_UNISCORE_HEADERS, timeout=12,
            )
            if rs.status_code == 200:
                stats_list = rs.json().get("data", {}).get("statistics", [])
                statistics_periods = _uniscore_stats_to_flat(stats_list)
                print(f"[uniscore] Estatísticas: {list(statistics_periods.keys())}")
        except Exception as es:
            print(f"[uniscore] Stats falhou: {es}")

    # Shotmap
    shotmap = []
    try:
        rsm = http_req.get(
            f"{_UNISCORE_API}/football/event/{uniscore_id}/shotmap",
            headers=_UNISCORE_HEADERS, timeout=12,
        )
        if rsm.status_code == 200:
            raw_shots = rsm.json().get("data", {}).get("shotmap", [])
            shotmap = [
                {
                    "id":        s.get("id"),
                    "minute":    s.get("time", 0),
                    "isHome":    s.get("isHome", True),
                    "shotType":  s.get("shotType", "miss"),
                    "bodyPart":  s.get("bodyPart", ""),
                    "situation": s.get("situation", ""),
                    "player":    s.get("player", {}).get("shortName", ""),
                    "x":         s.get("playerCoordinates", {}).get("x", 0),
                    "y":         s.get("playerCoordinates", {}).get("y", 0),
                }
                for s in raw_shots
            ]
            print(f"[uniscore] Shotmap: {len(shotmap)} chutes")
    except Exception as es:
        print(f"[uniscore] Shotmap falhou: {es}")

    # Status (FT?) + placar oficial do UniScore
    finished = False
    score_h  = None
    score_a  = None
    try:
        re = http_req.get(
            f"{_UNISCORE_API}/football/event/{uniscore_id}",
            headers=_UNISCORE_HEADERS,
            params={"language": "pt-BR"}, timeout=10,
        )
        if re.status_code == 200:
            ev_data  = re.json().get("data", {}).get("event", {})
            status   = ev_data.get("status", {})
            finished = status.get("type") == "finished"
            hs = ev_data.get("homeScore", {}) or {}
            as_ = ev_data.get("awayScore", {}) or {}
            score_h    = hs.get("current")
            score_a    = as_.get("current")
            score_ht_h = hs.get("period1")
            score_ht_a = as_.get("period1")
    except Exception:
        pass

    return {
        "graphPoints":        pts,
        "goals":              goals,
        "finished":           finished,
        "score_h":            score_h,
        "score_a":            score_a,
        "score_ht_h":         score_ht_h,
        "score_ht_a":         score_ht_a,
        "statistics":         statistics_periods.get("ALL", {}),
        "statistics_periods": statistics_periods,
        "shotmap":            shotmap,
        "source":             "uniscore",
    }


def _get_fotmob_live_matches():
    """Busca partidas ao vivo do FotMob. Cache de 2 minutos."""
    with _fotmob_live_lock:
        if time.time() - _fotmob_live_cache["ts"] < 120:
            return _fotmob_live_cache["matches"]
    try:
        today_str = datetime.now().strftime("%Y%m%d")
        r = http_req.get(
            "https://www.fotmob.com/api/matches",
            headers=FOTMOB_HEADERS,
            params={"date": today_str},
            timeout=12
        )
        r.raise_for_status()
        data = r.json()
        matches = []
        for league in data.get("leagues", []):
            for m in league.get("matches", []):
                st = m.get("status", {})
                if st.get("started") and not st.get("finished"):
                    home_name = m.get("home", {}).get("name", "").lower()
                    away_name = m.get("away", {}).get("name", "").lower()
                    matches.append({
                        "id":   str(m.get("id", "")),
                        "home": home_name,
                        "away": away_name,
                    })
        print(f"[fotmob] {len(matches)} partidas ao vivo encontradas")
        with _fotmob_live_lock:
            _fotmob_live_cache["ts"]      = time.time()
            _fotmob_live_cache["matches"] = matches
        return matches
    except Exception as e:
        print(f"[fotmob] Erro ao buscar live: {e}")
        return []


def _find_fotmob_id(casa, fora):
    """Encontra o ID do FotMob pelo nome dos times (busca fuzzy)."""
    if not casa or not fora:
        return None
    matches = _get_fotmob_live_matches()
    casa_l  = casa.lower()
    fora_l  = fora.lower()
    # Tenta match exato primeiro, depois substring
    for m in matches:
        if (casa_l in m["home"] or m["home"] in casa_l) and \
           (fora_l in m["away"] or m["away"] in fora_l):
            print(f"[fotmob] Match encontrado: {m['home']} vs {m['away']} (id={m['id']})")
            return m["id"]
    # Log para depuração quando não encontra
    if matches:
        sample = [(m["home"], m["away"]) for m in matches[:5]]
        print(f"[fotmob] Nenhum match para '{casa_l}' vs '{fora_l}'. Amostra: {sample}")
    else:
        print(f"[fotmob] Lista de partidas vazia ao buscar '{casa_l}' vs '{fora_l}'")
    return None


def _fetch_fotmob_momentum(fotmob_id):
    """Busca momentum do FotMob via API direta e converte para formato graphPoints."""
    r = http_req.get(
        "https://www.fotmob.com/api/matchDetails",
        params={"matchId": fotmob_id},
        headers=FOTMOB_HEADERS,
        timeout=15
    )
    r.raise_for_status()
    raw      = r.json()
    mom      = raw.get("content", {}).get("matchFacts", {}).get("momentum", {})
    mom_data = mom.get("main", {}).get("data", [])

    # Converte para formato graphPoints (compatível com o frontend)
    # FotMob: [{minute, value}] onde value > 0 = home, < 0 = away
    graph_points = []
    for pt in mom_data:
        val = pt.get("value", 0)
        minute = pt.get("minute", pt.get("min", 0))
        graph_points.append({
            "minute":    minute,
            "homeValue": max(0, val),
            "awayValue": min(0, val),
        })

    # Extrai gols dos incidents
    incidents_raw = raw.get("content", {}).get("matchFacts", {}).get("events", {})
    goals = []
    for ev in incidents_raw.get("events", []):
        if ev.get("type") in ("goal", "ownGoal"):
            goals.append({
                "minute": ev.get("time", 0),
                "team":   "home" if ev.get("isHome") else "away",
            })

    # Detecta FT
    status   = raw.get("header", {}).get("status", {})
    finished = status.get("finished", False)

    # Stats
    stats_raw = raw.get("content", {}).get("matchFacts", {}).get("stats", {})
    stats_out = []
    for block in stats_raw.get("stats", []):
        for stat in block.get("stats", []):
            vals = stat.get("stats", [])
            if len(vals) >= 2:
                stats_out.append({
                    "title": stat.get("title", ""),
                    "home":  str(vals[0]),
                    "away":  str(vals[1]),
                })

    return {
        "graphPoints": graph_points,
        "goals":       goals,
        "finished":    finished,
        "statistics":  stats_out,
        "source":      "fotmob",
    }


def _pressure_summary(graph_points: list) -> dict:
    """Calcula métricas de pressão/dominância a partir dos graphPoints.
    Valor > 0 = home dominant, < 0 = away dominant.
    """
    if not graph_points:
        return {}
    vals = [p.get("value", 0) for p in graph_points]
    minutes = [p.get("minute", 0) for p in graph_points]
    max_min = max(minutes) if minutes else 90
    half = max_min / 2

    h1 = [v for p, v in zip(graph_points, vals) if p.get("minute", 0) <= half]
    h2 = [v for p, v in zip(graph_points, vals) if p.get("minute", 0) > half]

    def avg(lst): return round(sum(lst) / len(lst), 3) if lst else 0.0

    home_dom = sum(1 for v in vals if v > 0)
    swings = sum(1 for i in range(1, len(vals)) if (vals[i] > 0) != (vals[i-1] > 0))

    return {
        "overall_avg":        avg(vals),
        "h1_avg":             avg(h1),
        "h2_avg":             avg(h2),
        "home_dominance_pct": round(home_dom / len(vals) * 100, 1) if vals else 0.0,
        "max_home":           round(max((v for v in vals if v > 0), default=0.0), 3),
        "max_away":           round(abs(min((v for v in vals if v < 0), default=0.0)), 3),
        "momentum_swings":    swings,
        "total_points":       len(vals),
    }


def _calc_xg(stats_flat: dict) -> dict:
    """Estima xG usando TODOS os campos de estatísticas disponíveis.

    Prioridade:
      1. Campo xG direto do UniScore/SofaScore
      2. Fórmula enriquecida com todos os stats do painel ESTATÍSTICAS

    Pesos baseados em probabilidades de conversão da literatura de analytics:
      shots_on_target  ≈ 0.33  (1 em 3 chutes no alvo vira gol)
      big_chances      ≈ 0.38  (grandes chances têm alta conversão)
      shots_inside_box ≈ 0.09  (chutes dentro da área não no alvo)
      shots_outside    ≈ 0.025 (chutes de fora da área)
      corners          ≈ 0.026 (escanteios geram perigo de área)
      touches_in_box   ≈ 0.008 (toques na área → proximidade de gol)
      final_third      ≈ 0.004 (passes/entradas no terço final → pressão)
      freekicks        ≈ 0.012 (cobranças de falta em posição perigosa)
      saves (oponente) → proxy de chutes no alvo quando shots_on_target = 0
    """
    # Nomes alternativos: UniScore usa Title Case com espaços,
    # código interno usa snake_case — tentamos ambos
    _ALIASES = {
        "shots_on_target":     ["Shots on Target", "shotsOnTarget", "Shots On Target"],
        "shots_inside_box":    ["Shots Inside Box", "shots_inside_box"],
        "shots_outside_box":   ["Shots Outside Box", "shots_outside_box"],
        "big_chances":         ["Big Chances", "bigChancesCreated", "Big Chances Created"],
        "corner_kicks":        ["Corner Kicks", "cornerKicks", "Corners"],
        "touches_in_box":      ["Touches in Box", "Touches In Box", "touches_in_box"],
        "pass_in_final_third": ["Passes in Final Third", "Pass in Final Third", "pass_in_final_third"],
        "final_third_entries": ["Final Third Entries", "final_third_entries"],
        "saves":               ["Saves", "Goalkeeper Saves", "saves"],
        "freekicks":           ["Free Kicks", "Freekicks", "freekicks"],
        "shots":               ["Total Shots", "Shots", "totalShots", "shots"],
    }

    def _get(canonical, fallback=0.0):
        """Busca stat tentando snake_case + aliases UniScore."""
        keys_to_try = [canonical] + _ALIASES.get(canonical, [])
        for k in keys_to_try:
            item = stats_flat.get(k)
            if item and isinstance(item, dict):
                hv = item.get("homeValue")
                av = item.get("awayValue")
                # homeValue pode ser 0 legítimo — só pula se for None
                if hv is None: hv = item.get("home", 0)
                if av is None: av = item.get("away", 0)
                try:
                    h = float(str(hv).replace("%", "").strip() or 0)
                    a = float(str(av).replace("%", "").strip() or 0)
                    return h, a
                except Exception:
                    pass
        return fallback, fallback

    # ── 1. xG direto ─────────────────────────────────────────────────────────
    for key in ("Expected Goals", "xG", "expected_goals", "Expected goals"):
        item = stats_flat.get(key)
        if item and isinstance(item, dict):
            hv = item.get("homeValue") if item.get("homeValue") is not None else item.get("home", 0)
            av = item.get("awayValue") if item.get("awayValue") is not None else item.get("away", 0)
            try:
                return {"home": round(float(hv), 2), "away": round(float(av), 2), "source": "direct"}
            except Exception:
                pass

    # ── 2. Fórmula enriquecida com todos os stats ─────────────────────────────
    ontar_h,   ontar_a   = _get("shots_on_target")
    inside_h,  inside_a  = _get("shots_inside_box")
    outside_h, outside_a = _get("shots_outside_box")
    big_h,     big_a     = _get("big_chances")
    corners_h, corners_a = _get("corner_kicks")
    touches_h, touches_a = _get("touches_in_box")
    fp3_h,     fp3_a     = _get("pass_in_final_third")
    fte_h,     fte_a     = _get("final_third_entries")
    saves_h,   saves_a   = _get("saves")
    free_h,    free_a    = _get("freekicks")
    total_h,   total_a   = _get("shots")

    # Se shots_on_target = 0 mas saves do adversário está disponível,
    # usa saves como proxy (saves_adversario ≈ shots_on_target_proprio)
    eff_ontar_h = ontar_h if ontar_h > 0 else saves_a
    eff_ontar_a = ontar_a if ontar_a > 0 else saves_h

    # Se shots_inside_box = 0 mas total de chutes disponível, estima 60% dentro
    if inside_h == 0 and total_h > 0:
        inside_h = total_h * 0.60
    if inside_a == 0 and total_a > 0:
        inside_a = total_a * 0.60
    if outside_h == 0 and total_h > 0:
        outside_h = total_h * 0.40
    if outside_a == 0 and total_a > 0:
        outside_a = total_a * 0.40

    # Contribuição de pressão posicional (final third + passes finais)
    press_h = fp3_h + fte_h
    press_a = fp3_a + fte_a

    xg_h = round(
        0.33  * eff_ontar_h    # chutes no alvo (maior peso)
      + 0.38  * big_h          # grandes chances
      + 0.09  * inside_h       # chutes dentro da área (não no alvo)
      + 0.025 * outside_h      # chutes fora da área
      + 0.026 * corners_h      # escanteios → perigo de área
      + 0.008 * touches_h      # toques na área adversária
      + 0.004 * press_h        # pressão no terço final
      + 0.012 * free_h,        # cobranças de falta perigosas
    2)

    xg_a = round(
        0.33  * eff_ontar_a
      + 0.38  * big_a
      + 0.09  * inside_a
      + 0.025 * outside_a
      + 0.026 * corners_a
      + 0.008 * touches_a
      + 0.004 * press_a
      + 0.012 * free_a,
    2)

    # Quais campos contribuíram
    fields_used = []
    if eff_ontar_h > 0 or eff_ontar_a > 0:   fields_used.append("shots_on_target")
    if big_h > 0 or big_a > 0:                fields_used.append("big_chances")
    if inside_h > 0 or inside_a > 0:          fields_used.append("shots_inside_box")
    if outside_h > 0 or outside_a > 0:        fields_used.append("shots_outside_box")
    if corners_h > 0 or corners_a > 0:        fields_used.append("corners")
    if touches_h > 0 or touches_a > 0:        fields_used.append("touches_in_box")
    if press_h > 0 or press_a > 0:            fields_used.append("final_third")
    if free_h > 0 or free_a > 0:              fields_used.append("freekicks")

    if xg_h == 0.0 and xg_a == 0.0:
        return {}

    return {
        "home":         xg_h,
        "away":         xg_a,
        "source":       "estimated",
        "fields_used":  fields_used,
        "n_fields":     len(fields_used),
    }


def _extract_score(goals: list) -> dict:
    """Conta gols da lista de incidents para obter o placar final."""
    home = sum(1 for g in goals if g.get("team") == "home")
    away = sum(1 for g in goals if g.get("team") == "away")
    return {"home": home, "away": away}


def _build_save_payload(
    event_id, casa, fora, liga,
    graph_points, goals, stats_flat, stats_periods,
    opening_odds, source, shotmap=None
) -> dict:
    """Monta o payload completo para salvar no momentum_history."""
    today = datetime.now().strftime("%Y-%m-%d")
    return {
        # Identificação
        "event_id":  event_id,
        "date":      today,
        "saved_at":  datetime.now().isoformat(),
        "source":    source,
        "casa":      casa,
        "fora":      fora,
        "liga":      liga,
        # Dados brutos
        "graphPoints":        graph_points,
        "goals":              goals,
        "score":              _extract_score(goals),
        # Mapa de chutes
        "shotmap":            shotmap or [],
        # Estatísticas
        "statistics":         stats_flat,
        "statistics_periods": stats_periods,
        # Indicadores derivados
        "pressure_summary":   _pressure_summary(graph_points),
        "xg":                 _calc_xg(stats_flat),
        # Odds
        "opening_odds":       opening_odds,
    }


def _fetch_sofa_direct(event_id):
    """Busca dados do SofaScore via requests direto (sem Playwright).
    Tenta api.sofascore.com (sem Cloudflare) primeiro, fallback para www.
    Retorna (graph, incidents, statistics) ou levanta exceção."""
    # api.sofascore.com não tem Cloudflare — funciona de IPs de datacenter
    for base_url in [
        f"https://api.sofascore.com/api/v1/event/{event_id}",
        f"https://www.sofascore.com/api/v1/event/{event_id}",
    ]:
        try:
            s = http_req.Session()
            s.headers.update(_SOFA_HEADERS)
            graph      = s.get(f"{base_url}/graph",      timeout=12)
            incidents  = s.get(f"{base_url}/incidents",  timeout=12)
            statistics = s.get(f"{base_url}/statistics", timeout=12)
            # Verifica se retornou dados válidos (não bloqueio Cloudflare)
            g = graph.json()
            pts = g.get("graphPoints", [])
            print(f"[sofa] {base_url.split('/')[2]}: {len(pts)} graphPoints")
            if pts:
                return g, incidents.json(), statistics.json()
        except Exception as e:
            print(f"[sofa] Falhou {base_url.split('/')[2]}: {e}")
            continue
    raise RuntimeError(f"SofaScore: nenhuma fonte retornou graphPoints para event {event_id}")


def _fetch_sofa_playwright(event_id):
    """Fallback: busca dados via Playwright (headless Chromium)."""
    from playwright.sync_api import sync_playwright
    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage",
                  "--disable-blink-features=AutomationControlled"]
        )
        ctx  = browser.new_context(
            user_agent=_SOFA_HEADERS["User-Agent"],
            locale="pt-BR", timezone_id="America/Sao_Paulo",
        )
        page = ctx.new_page()
        page.add_init_script("Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")
        page.goto("https://www.sofascore.com/", timeout=30000, wait_until="domcontentloaded")
        page.wait_for_timeout(2000)
        result = page.evaluate(f"""async () => {{
            try {{
                const [rG, rI, rS] = await Promise.all([
                    fetch('/api/v1/event/{event_id}/graph'),
                    fetch('/api/v1/event/{event_id}/incidents'),
                    fetch('/api/v1/event/{event_id}/statistics')
                ]);
                return {{
                    status: 200,
                    graph:      rG.ok ? await rG.json() : {{}},
                    incidents:  rI.ok ? await rI.json() : {{}},
                    statistics: rS.ok ? await rS.json() : {{}}
                }};
            }} catch(e) {{ return {{error: e.message}}; }}
        }}""")
        browser.close()
    if result.get("status") != 200:
        raise RuntimeError(f"Playwright sem dados: {result.get('error')}")
    return result["graph"], result["incidents"], result["statistics"]


def _process_momentum(event_id, casa="", fora="", liga=""):
    """Busca momentum exclusivamente via UniScore (busca por nome de time).
    Cache de 90s para evitar chamadas repetidas.
    """
    global _pattern_tips_cache, _odds_patterns_cache, _stats_patterns_cache
    # Verifica cache primeiro
    with _momentum_lock:
        cached = _momentum_cache.get(event_id)
        if cached and time.time() - cached["ts"] < 90:
            return cached["data"]

    print(f"[momentum] Buscando '{casa}' vs '{fora}' via UniScore...")

    # ── Único source: UniScore (busca por nome) ───────────────────────────
    uni_match = _find_uniscore_id(casa, fora)
    if not uni_match:
        print(f"[momentum] UniScore: partida não encontrada para '{casa}' vs '{fora}'")
        with _momentum_lock:
            _momentum_cache[event_id] = {"ts": time.time(), "data": None}
        return None

    try:
        udata = _fetch_uniscore_graph(uni_match)
        pts   = udata.get("graphPoints", [])
        print(f"[momentum] UniScore OK: {len(pts)} graphPoints, finished={udata.get('finished')}")

        # ── Calcula xG e pressure_summary para o dado ao vivo ────────────
        # (normalmente só calculados no save; precisamos aqui para os indicadores)
        stats_live = udata.get("statistics", {})
        xg_live    = _calc_xg(stats_live) if stats_live else {}
        ps_live    = _pressure_summary(pts) if pts else {}

        data = {**udata, "saved": False,
                "xg": xg_live, "pressure_summary": ps_live}

        # ── Acumula shotmap ao vivo no cache separado ─────────────────────
        # O endpoint de shotmap só funciona durante o jogo; ao FT fica vazio.
        # Guardamos no disco para sobreviver a restarts/redeploys do Railway.
        live_shots = udata.get("shotmap", [])
        # DIAGNÓSTICO TEMPORÁRIO — investigando por que shotmap_history/ tem tão
        # poucas partidas salvas (12) comparado a momentum_history/ (2600+).
        # Hipótese: o Uniscore só manda chute-a-chute pra uma fatia das ligas
        # cobertas. Sem custo/chamada nova — só loga o que já veio na busca de
        # momentum, que já roda de qualquer jeito. Remover depois de confirmar
        # o padrão em alguns dias de log.
        if not udata.get("finished"):
            ja_teve_chutes = event_id in _shotmap_live_cache
            print(f"[shotmap-diag] liga='{liga}' | {casa} x {fora} | chutes_agora={len(live_shots)} | ja_teve_antes={ja_teve_chutes}")
        if live_shots:
            with _shotmap_lock:
                prev = _shotmap_live_cache.get(event_id, [])
                is_new_event = event_id not in _shotmap_live_cache
                if len(live_shots) != len(prev):   # só escreve se mudou
                    _shotmap_live_cache[event_id] = live_shots
                    _save_shotmap_cache(_shotmap_live_cache)
                    # Push pro GitHub quando é novo evento (sobrevive a restart mid-game)
                    if is_new_event:
                        github_storage.push_file_bg(
                            _SHOTMAP_CACHE_FILE, ".shotmap_cache.json"
                        )

        # ── Auto-save quando a partida termina ───────────────────────────
        if udata.get("finished"):
            today     = datetime.now().strftime("%Y-%m-%d")
            save_file = os.path.join(MOMENTUM_DIR, f"{today}_{event_id}.json")
            if not os.path.exists(save_file):
                # Odds de abertura
                opening_odds = {}
                try:
                    uni_odds_map = _uni_odds_today().get(uni_match["id"], {})
                    if uni_odds_map:
                        opening_odds = {
                            "h":        uni_odds_map.get("h"),
                            "x":        uni_odds_map.get("x"),
                            "a":        uni_odds_map.get("a"),
                            "ou_line":  uni_odds_map.get("ou_line"),
                            "ou_over":  uni_odds_map.get("ou_over"),
                            "ou_under": uni_odds_map.get("ou_under"),
                        }
                except Exception:
                    pass

                # Usa shotmap acumulado durante o jogo (o endpoint de FT fica vazio)
                with _shotmap_lock:
                    best_shotmap = _shotmap_live_cache.get(event_id) or udata.get("shotmap", [])

                payload = _build_save_payload(
                    event_id=event_id,
                    casa=casa, fora=fora, liga=liga,
                    graph_points=pts,
                    goals=udata.get("goals", []),
                    stats_flat=udata.get("statistics", {}),
                    stats_periods=udata.get("statistics_periods", {}),
                    opening_odds=opening_odds,
                    source="uniscore",
                    shotmap=best_shotmap,
                )
                with open(save_file, "w", encoding="utf-8") as f:
                    json.dump(payload, f, ensure_ascii=False, indent=2)
                data["saved"] = True
                sm_count = len(best_shotmap)
                print(f"[momentum] Salvo: {save_file} | shotmap={sm_count} chutes")

                # ── Salva shotmap separadamente em shotmap_history/ ───────────
                if best_shotmap:
                    score_raw = payload.get("score", {})
                    sm_payload = {
                        "event_id":  event_id,
                        "date":      today,
                        "casa":      casa,
                        "fora":      fora,
                        "liga":      liga,
                        "score":     score_raw,
                        "total_shots": sm_count,
                        "shotmap":   best_shotmap,
                    }
                    sm_file = os.path.join(SHOTMAP_DIR, f"{today}_{event_id}.json")
                    with open(sm_file, "w", encoding="utf-8") as f:
                        json.dump(sm_payload, f, ensure_ascii=False, indent=2)
                    print(f"[shotmap] Salvo separado: {sm_file}")
                    github_storage.push_file_bg(sm_file, f"shotmap_history/{today}_{event_id}.json")

                # Limpa cache de shotmap (memória + disco)
                with _shotmap_lock:
                    _shotmap_live_cache.pop(event_id, None)
                    _save_shotmap_cache(_shotmap_live_cache)
                github_storage.push_file_bg(save_file, f"momentum_history/{today}_{event_id}.json")
                _pattern_tips_cache  = {"ts": 0, "data": None}
                _odds_patterns_cache = {"ts": 0, "data": None}
                _stats_patterns_cache = {"ts": 0, "data": None}
                threading.Thread(target=_rebuild_analysis_cache, daemon=True).start()
            else:
                data["saved"] = True

        with _momentum_lock:
            _momentum_cache[event_id] = {"ts": time.time(), "data": data}
        return data

    except Exception as e:
        print(f"[momentum] Erro UniScore event {event_id}: {e}")
        return None


# ── Monitor de fundo: verifica jogos ao vivo a cada 5 min e salva os encerrados ──
def _background_monitor():
    """Thread daemon que varre os jogos ao vivo (UniScore) e salva quando FT."""
    print("[monitor] Thread de monitoramento iniciada.")
    time.sleep(60)  # Aguarda Flask subir
    while True:
        try:
            today     = datetime.now().strftime("%Y-%m-%d")
            live_list = _fetch_live_matches_for_monitor()
            pendentes = [
                m for m in live_list
                if not os.path.exists(
                    os.path.join(MOMENTUM_DIR, f"{today}_{m['id']}.json")
                )
            ]
            if pendentes:
                print(f"[monitor] {len(live_list)} ao vivo, {len(pendentes)} ainda não salvos — verificando...")
                for m in pendentes:
                    data = _process_momentum(m["id"], m["casa"], m["fora"], m["liga"])
                    if data and data.get("finished"):
                        print(f"[monitor] ✓ Encerrado e salvo: {m['casa']} x {m['fora']}")
                    time.sleep(2)
            else:
                print(f"[monitor] {len(live_list)} ao vivo, todos já salvos ou sem jogos.")
        except Exception as e:
            print(f"[monitor] Erro geral: {e}")
        time.sleep(300)


# Inicia a thread de monitoramento (daemon = morre junto com o Flask)
threading.Thread(target=_background_monitor, daemon=True, name="MomentumMonitor").start()

# Restaura dados do GitHub ao iniciar (backtest + momentum_history + shotmap_history + cache).
# _github_sync_done é usado pelo loop de watchlist pra esperar essa restauração
# terminar antes do 1º ciclo — sem isso, o 1º ciclo (que roda quase instantaneamente)
# criava um banco novo VAZIO local e o dava PUSH pro GitHub antes dessa restauração
# (mais lenta, sequencial) conseguir baixar a versão de verdade, apagando o backup bom.
_github_sync_done = threading.Event()


def _github_sync_on_startup_then_flag():
    try:
        github_storage.sync_on_startup(MOMENTUM_DIR, BACKTEST_DIR, DATA_DIR, SHOTMAP_DIR)
        github_storage.pull_directory("forca_history", FORCA_HISTORY_DIR)
    finally:
        _github_sync_done.set()


threading.Thread(
    target=_github_sync_on_startup_then_flag,
    daemon=True,
    name="GitHubSync"
).start()

# Mapa de Sugestões: padrão do dia + lista de sugestões travada — sem isso, cada
# redeploy no Railway apagava a memória e escolhia um padrão novo do zero no meio
# do dia (mesmo problema do backtest2.db, resolvido do mesmo jeito: sempre baixa
# a versão mais recente do GitHub, sobrescrevendo qualquer coisa local).
threading.Thread(
    target=github_storage.pull_directory,
    args=("mapa_cache", MAPA_CACHE_DIR),
    daemon=True,
    name="GitHubSyncMapa"
).start()

# Backup de Força — estava totalmente pronto desde a sessão em que foi desenhado,
# mas os threads nunca tinham sido iniciados (ficou parado, pasta forca_history/
# vazia). Ativado agora a pedido do usuário, pra alimentar odds pré-jogo no Replay.
threading.Thread(target=_forca_backup_worker, daemon=True, name="ForcaBackupWorker").start()
threading.Thread(target=_forca_backup_scan_loop, daemon=True, name="ForcaBackupScan").start()


@app.route("/api/radar/momentum/<event_id>")
def api_radar_momentum(event_id):
    """Busca dados de Attack Momentum do SofaScore via Playwright.
    Query params opcionais: casa, fora, liga — usados ao salvar histórico.
    """
    from flask import request as flask_req

    casa = flask_req.args.get("casa", "")
    fora = flask_req.args.get("fora", "")
    liga = flask_req.args.get("liga", "")

    data = _process_momentum(event_id, casa, fora, liga)
    if data is None:
        return jsonify({"error": "Sem dados do SofaScore"}), 503
    return jsonify(data)


@app.route("/api/radar/shotmap/<event_id>")
def api_shotmap(event_id):
    """Retorna mapa de chutes via UniScore para uma partida.
    Query params: casa, fora (usados para encontrar o ID no UniScore).
    Tenta primeiro no arquivo salvo, depois busca ao vivo."""
    from flask import request as flask_req
    casa = flask_req.args.get("casa", "")
    fora = flask_req.args.get("fora", "")

    # 1. Tenta arquivo já salvo no momentum_history
    today = datetime.now().strftime("%Y-%m-%d")
    save_file = os.path.join(MOMENTUM_DIR, f"{today}_{event_id}.json")
    if os.path.exists(save_file):
        try:
            with open(save_file, encoding="utf-8") as f:
                saved = json.load(f)
            shots = saved.get("shotmap", [])
            if shots:
                return jsonify({"ok": True, "shots": shots, "source": "saved"})
        except Exception:
            pass

    # 2. Busca ao vivo via UniScore (match por nome)
    uni_match = _find_uniscore_id(casa, fora)
    if not uni_match:
        # Tenta usando o event_id diretamente (IDs são compatíveis)
        uni_match = {"id": event_id, "homeId": "", "awayId": ""}

    try:
        rsm = http_req.get(
            f"{_UNISCORE_API}/football/event/{uni_match['id']}/shotmap",
            headers=_UNISCORE_HEADERS, timeout=12,
        )
        if rsm.status_code == 200:
            raw = rsm.json().get("data", {}).get("shotmap", [])
            shots = [
                {
                    "id":        s.get("id"),
                    "minute":    s.get("time", 0),
                    "isHome":    s.get("isHome", True),
                    "shotType":  s.get("shotType", "miss"),
                    "bodyPart":  s.get("bodyPart", ""),
                    "situation": s.get("situation", ""),
                    "player":    s.get("player", {}).get("shortName", ""),
                    "x":         s.get("playerCoordinates", {}).get("x", 0),
                    "y":         s.get("playerCoordinates", {}).get("y", 0),
                }
                for s in raw
            ]
            return jsonify({"ok": True, "shots": shots, "source": "live"})
    except Exception as e:
        print(f"[shotmap] Erro: {e}")

    return jsonify({"ok": False, "shots": [], "source": "none"}), 200


@app.route("/api/momentum/history")
def api_momentum_history():
    """Lista partidas com momentum salvo (mais recentes primeiro).
    Sem 'q': retorna só as 200 mais recentes (browse rápido). Com 'q': varre
    todo o histórico salvo procurando o texto no time/liga (usado pela busca
    da aba Replay)."""
    q = (request.args.get("q") or "").strip().lower()
    files = sorted(glob.glob(os.path.join(MOMENTUM_DIR, "*.json")), reverse=True)
    if not q:
        files = files[:200]
    matches = []
    for fpath in files:
        try:
            with open(fpath, encoding="utf-8") as f:
                d = json.load(f)
            casa = d.get("casa", "")
            fora = d.get("fora", "")
            liga = d.get("liga", "")
            if q and q not in casa.lower() and q not in fora.lower() and q not in liga.lower():
                continue
            pts   = d.get("graphPoints", [])
            total = len(pts)
            matches.append({
                "event_id":  d.get("event_id"),
                "date":      d.get("date"),
                "saved_at":  d.get("saved_at"),
                "casa":      casa,
                "fora":      fora,
                "liga":      liga,
                "goals":     d.get("goals", []),
                "points":    total,
                "filename":  os.path.basename(fpath),
            })
            if q and len(matches) >= 300:
                break
        except Exception:
            pass
    return jsonify({"matches": matches, "total": len(matches)})


def _momentum_detect_chance_spikes(points, chance_pct=0.8, over_pct=0.6):
    """Porta da parte 'Grande Chance' de _live2DetectPressureMarkers (JS) — pico
    pontual (1-2 pontos) de pressão >= chance_pct do máximo daquele jogo, fora de
    qualquer janela sustentada (Momento Over, >=4 pontos >= over_pct do máximo).
    chance_pct/over_pct parametrizados pra permitir a calibração por grid search."""
    n = len(points)
    if n < 4:
        return []
    max_val = max((abs(p.get("value") or 0) for p in points), default=1) or 1
    over_thresh    = max_val * over_pct
    chance_thresh  = max_val * chance_pct

    over_windows = []
    i = 0
    while i < n:
        if abs(points[i].get("value") or 0) >= over_thresh:
            j = i
            while j < n and abs(points[j].get("value") or 0) >= over_thresh:
                j += 1
            if j - i >= 4:
                over_windows.append((i, j - 1))
            i = j
        else:
            i += 1

    def in_over_window(idx):
        return any(s <= idx <= e for s, e in over_windows)

    spikes = []
    i = 0
    while i < n:
        v = abs(points[i].get("value") or 0)
        if v >= chance_thresh and not in_over_window(i):
            j = i
            while j < n and abs(points[j].get("value") or 0) >= chance_thresh and not in_over_window(j):
                j += 1
            if j - i <= 2:
                spikes.append({
                    "minute": int(points[i].get("minute") or 0),
                    "is_home": (points[i].get("value") or 0) >= 0,
                })
            i = j
        else:
            i += 1
    return spikes


def _momentum_detect_over_under_windows(points, over_pct=0.6, under_pct=0.15):
    """Porta das janelas sustentadas 'Momento Over'/'Momento Under' de
    _live2DetectPressureMarkers (JS) — >=4 pontos consecutivos acima (over) ou
    abaixo (under) do limiar, em % do pico daquele jogo. Ao contrário do 'Grande
    Chance', essas janelas não têm time associado no marcador (o Ao Vivo também
    não mostra time nelas) — o sinal é "pressão sustentada" ou "sem pressão",
    não "pressão de X"."""
    n = len(points)
    if n < 4:
        return [], []
    max_val = max((abs(p.get("value") or 0) for p in points), default=1) or 1
    over_thresh  = max_val * over_pct
    under_thresh = max_val * under_pct

    def find_windows(cond):
        windows = []
        i = 0
        while i < n:
            if cond(points[i]):
                j = i
                while j < n and cond(points[j]):
                    j += 1
                if j - i >= 4:
                    windows.append((i, j - 1))
                i = j
            else:
                i += 1
        return windows

    over_idx  = find_windows(lambda p: abs(p.get("value") or 0) >= over_thresh)
    under_idx = find_windows(lambda p: abs(p.get("value") or 0) <= under_thresh)

    def to_markers(idx_windows):
        out = []
        for s, e in idx_windows:
            m_start = points[s].get("minute") or 0
            m_end   = points[e].get("minute") or 0
            out.append({"minute": (m_start + m_end) / 2, "end_minute": m_end})
        return out

    return to_markers(over_idx), to_markers(under_idx)


_MOMENTUM_ALL_MATCHES_CACHE = {"ts": 0, "data": None}
_MOMENTUM_ALL_MATCHES_TTL = 10 * 60  # 10min — evita reler ~2500 arquivos do disco
# a cada request, já que o único consumidor hoje (Grande Chance) também tem seu
# próprio cache de 30min por cima disso

def _momentum_load_all_matches():
    """Carrega (graphPoints, goals) de todo o momentum_history — cacheado por
    10min, pra não reler os ~2500 arquivos do disco a cada request (recarregar
    tudo repetidamente já chegou a derrubar o worker do Railway por estourar o
    timeout do gunicorn com 1 worker só)."""
    now = time.time()
    if _MOMENTUM_ALL_MATCHES_CACHE["data"] is not None and (now - _MOMENTUM_ALL_MATCHES_CACHE["ts"]) < _MOMENTUM_ALL_MATCHES_TTL:
        return _MOMENTUM_ALL_MATCHES_CACHE["data"]

    matches = []
    for fpath in glob.glob(os.path.join(MOMENTUM_DIR, "*.json")):
        try:
            with open(fpath, encoding="utf-8") as f:
                d = json.load(f)
        except Exception:
            continue
        points = d.get("graphPoints") or []
        goals  = d.get("goals") or []
        if len(points) >= 4:
            matches.append((points, goals))

    _MOMENTUM_ALL_MATCHES_CACHE["data"] = matches
    _MOMENTUM_ALL_MATCHES_CACHE["ts"] = now
    return matches


# ── SCANNER — etapa 2: motor de similaridade ao vivo ────────────────────────
# Dado o estado atual de uma partida (placar + minuto), busca no
# momentum_history partidas passadas que tiveram o MESMO placar num minuto
# parecido, e mede — daquele ponto em diante — a chance de cada time marcar
# de novo em cada janela de tempo. Os 3 tipos de scalping que o usuário
# descreveu (Under no placar limite, placar exato limite, contra uma equipe)
# são todos derivados dos mesmos 2 sinais binários por analog (casa marcou?
# fora marcou?):
#   - "sem_mais_gols"   (nem casa nem fora marcou) → Under no placar atual+0.5
#   - "casa_nao_marca"  → placar exato limite pro lado de cima (ex: 2-1→3-1
#     fica de fora) OU scalping contra a casa
#   - "fora_nao_marca"  → idem pro lado de baixo (ex: 2-1→2-2 fica de fora)
#     OU scalping contra o visitante
# Não decide um mercado/janela fixos — testa uma grade de janelas e devolve
# QUALQUER combinação que bata o limiar de segurança, deixando o "padrão
# dinâmico" (ideia do usuário) emergir dos dados em vez de forçar uma regra.
_SCANNER_MINUTE_TOLERANCE = 3     # cohort: mesmo placar dentro de ±3min do minuto atual
_SCANNER_WINDOWS = (3, 5, 7, 10)  # janelas de scalping testadas, em minutos
_SCANNER_MIN_SAMPLE = 20          # amostra mínima pro percentual valer alguma coisa
_SCANNER_SAFE_THRESHOLD = 0.95    # confirmado com o usuário: ≥95% = "seguro"


def _scanner_score_at_minute(goals, minute):
    """Reconstrói o placar (h, a) num minuto específico, a partir da lista de
    gols {minute, team}. Mesma técnica já usada em outros pontos do projeto
    (ex: reconstrução de placar por minuto nas metodologias antigas)."""
    h = sum(1 for g in goals if g.get("team") == "home" and (g.get("minute") or 0) <= minute)
    a = sum(1 for g in goals if g.get("team") == "away" and (g.get("minute") or 0) <= minute)
    return h, a


def _scanner_find_cohort(all_matches, target_h, target_a, target_minute):
    """Varre momentum_history procurando partidas que tiveram o placar
    (target_h, target_a) em algum minuto dentro de ±_SCANNER_MINUTE_TOLERANCE
    de target_minute. Cada partida entra no máximo 1x (o primeiro minuto em
    que bateu), guardando esse minuto (pra medir 'dali pra frente' na
    linha do tempo DELA, não na da partida atual) + o minuto máximo
    rastreado (pra saber se ela viu o suficiente pra contar em cada janela)."""
    cohort = []
    target_minute = int(target_minute)
    lo = max(0, target_minute - _SCANNER_MINUTE_TOLERANCE)
    hi = target_minute + _SCANNER_MINUTE_TOLERANCE
    for points, goals in all_matches:
        # minute às vezes vem fracionário (ex: 45.5, ponto de intervalo) — arredonda
        # pra baixo, range() exige int.
        max_minute = int(max((p.get("minute") or 0) for p in points)) if points else 0
        if max_minute < lo:
            continue  # não rastreou nem até o começo da janela de tolerância
        for m in range(lo, min(hi, max_minute) + 1):
            h, a = _scanner_score_at_minute(goals, m)
            if h == target_h and a == target_a:
                cohort.append((goals, m, max_minute))
                break
    return cohort


def _scanner_evaluate_cohort(cohort):
    """Pra cada janela candidata, mede — entre as partidas do cohort que
    foram rastreadas até o fim da janela — a fração que NÃO teve gol de cada
    time (e de nenhum dos dois) dali em diante. Só devolve janelas com
    amostra suficiente."""
    out = {}
    for w in _SCANNER_WINDOWS:
        total = casa_safe = fora_safe = ambos_safe = 0
        for goals, m2, max_minute in cohort:
            if max_minute < m2 + w:
                continue  # não deu pra observar a janela inteira nessa partida
            total += 1
            casa_marcou = any(g.get("team") == "home" and m2 < (g.get("minute") or 0) <= m2 + w for g in goals)
            fora_marcou = any(g.get("team") == "away" and m2 < (g.get("minute") or 0) <= m2 + w for g in goals)
            if not casa_marcou:
                casa_safe += 1
            if not fora_marcou:
                fora_safe += 1
            if not casa_marcou and not fora_marcou:
                ambos_safe += 1
        if total >= _SCANNER_MIN_SAMPLE:
            out[w] = {
                "total": total,
                "casa_nao_marca": round(casa_safe / total, 4),
                "fora_nao_marca": round(fora_safe / total, 4),
                "sem_mais_gols":  round(ambos_safe / total, 4),
            }
    return out


def _scanner_build_alerts(evaluated, placar_h, placar_a):
    """Converte a grade janela→taxas num punhado de alertas prontos pra UI —
    só os que bateram o limiar de ≥95%, com o texto já explicando o mercado
    (placar exato / under / contra o time), ordenados por janela (os mais
    imediatos primeiro).

    Por pedido do usuário (2026-08-23): só Under por enquanto — validar esse
    mercado sozinho antes de abrir pros outros dois (placar exato / contra um
    time). _scanner_evaluate_cohort já calcula os 3 sinais (é barato, mesmo
    laço), só não surgem como alerta ainda — pra reativar, é só descomentar
    as 2 linhas abaixo."""
    labels = {
        "sem_mais_gols":  f"Under {placar_h + placar_a}.5 gols (sem mais nenhum gol)",
        # "casa_nao_marca": f"Lay {placar_h + 1}x{placar_a} / contra a Casa marcar de novo",
        # "fora_nao_marca": f"Lay {placar_h}x{placar_a + 1} / contra o Visitante marcar de novo",
    }
    alerts = []
    for w in sorted(evaluated.keys()):
        stats = evaluated[w]
        for signal, label in labels.items():
            rate = stats[signal]
            if rate >= _SCANNER_SAFE_THRESHOLD:
                alerts.append({
                    "window_min": w,
                    "signal": signal,
                    "label": label,
                    "rate": rate,
                    "sample": stats["total"],
                })
    return alerts


@app.route("/api/scanner/analyze")
def api_scanner_analyze():
    """Motor de similaridade do Scanner: pega o estado ao vivo atual (via
    _process_momentum, o mesmo dado que já alimenta Ao Vivo/momentum) e busca
    no momentum_history partidas com o mesmo placar num minuto parecido,
    devolvendo qualquer combinação janela+mercado que bateu ≥95% de segurança
    na amostra encontrada. Chamado periodicamente pelo frontend só pros jogos
    marcados como 'watch' (não pra todos os jogos ao vivo — custo)."""
    event_id = request.args.get("event_id", "")
    casa = request.args.get("casa", "")
    fora = request.args.get("fora", "")
    liga = request.args.get("liga", "")
    if not event_id or not casa or not fora:
        return jsonify({"error": "faltam parâmetros (event_id, casa, fora)"}), 400

    data = _process_momentum(event_id, casa, fora, liga)
    if not data or not data.get("graphPoints"):
        return jsonify({"error": "sem dados ao vivo pra essa partida ainda"}), 404

    points = data["graphPoints"]
    goals = data.get("goals") or []
    minuto_atual = int(max((p.get("minute") or 0) for p in points)) if points else 0
    # Prefere o placar OFICIAL do UniScore (score_h/score_a, vem direto do
    # status da partida) em vez de reconstruir contando a lista de gols —
    # essa lista pode perder gol (achado testando: Barracas Central 1x0
    # aparecia como 0x0 na análise porque o gol de pênalti não batia o
    # filtro incidentType=="goal" de _fetch_uniscore_graph). Só cai pro
    # cálculo via goals se o placar oficial não vier.
    if data.get("score_h") is not None and data.get("score_a") is not None:
        placar_h, placar_a = int(data["score_h"]), int(data["score_a"])
    else:
        placar_h, placar_a = _scanner_score_at_minute(goals, minuto_atual)

    all_matches = _momentum_load_all_matches()
    cohort = _scanner_find_cohort(all_matches, placar_h, placar_a, minuto_atual)
    evaluated = _scanner_evaluate_cohort(cohort)
    alerts = _scanner_build_alerts(evaluated, placar_h, placar_a)

    return jsonify({
        "minuto": minuto_atual,
        "placar": f"{placar_h}-{placar_a}",
        "cohort_size": len(cohort),
        # Total de partidas na base (momentum_history) — o frontend usa isso
        # pra mostrar "N jogos parecidos (X% da base)", dando noção de quão
        # raro/comum é esse cenário, não só o número bruto.
        "total_base": len(all_matches),
        "alerts": alerts,
        # graphPoints/goals já vieram nessa mesma busca (_process_momentum) —
        # devolve pro frontend desenhar o gráfico de pressão nos jogos
        # monitorados sem precisar de uma requisição extra.
        "graphPoints": points,
        "goals": goals,
    })


def _momentum_eval_pattern(matches, chance_pct, over_pct, window_min):
    """Roda a detecção de picos com um limiar/janela específicos em cima da base
    já carregada e mede o lift real (taxa de acerto vs taxa-base) dessa combinação."""
    total_spikes = hits = total_goals = total_minutes = 0
    for points, goals in matches:
        total_minutes += max((p.get("minute") or 0) for p in points)
        total_goals   += len(goals)
        for sp in _momentum_detect_chance_spikes(points, chance_pct, over_pct):
            total_spikes += 1
            team = "home" if sp["is_home"] else "away"
            saiu_gol = any(
                g.get("team") == team
                and 0 <= (g.get("minute") or 0) - sp["minute"] <= window_min
                for g in goals
            )
            if saiu_gol:
                hits += 1

    hit_rate = (hits / total_spikes) if total_spikes else 0.0
    # Taxa-base: probabilidade de sair gol de UM time específico numa janela aleatória
    # do mesmo tamanho — aproximação a partir da taxa média de gols/minuto da base,
    # dividida por 2 (metade das vezes o gol é do time "certo" por acaso)
    gols_por_minuto = (total_goals / total_minutes) if total_minutes else 0.0
    baseline_rate = (gols_por_minuto * window_min) / 2
    lift = (hit_rate / baseline_rate) if baseline_rate else 0.0

    return {
        "chance_pct":    chance_pct,
        "over_pct":      over_pct,
        "window_min":    window_min,
        "total_spikes":  total_spikes,
        "hits":          hits,
        "hit_rate":      round(hit_rate, 4),
        "baseline_rate": round(baseline_rate, 4),
        "lift":          round(lift, 3),
    }


_CHANCE_PATTERN_CACHE = {"ts": 0, "data": None}
_CHANCE_PATTERN_TTL = 30 * 60  # 30 minutos — reescanear/recalibrar a base inteira a cada request seria lento
_CHANCE_PATTERN_MIN_SAMPLE = 30
_CHANCE_PATTERN_MIN_LIFT = 1.3
# Grid de combinações testadas — quanto mais a base cresce, mais confiável fica a
# escolha da melhor combinação (thresholds mais "esquisitos"/específicos só ganham
# quando o tamanho da amostra sustenta, senão MIN_SAMPLE já descarta)
_CHANCE_PATTERN_THRESH_GRID = [0.60, 0.65, 0.70, 0.75, 0.80, 0.85, 0.90]
_CHANCE_PATTERN_WINDOW_GRID = [5, 8, 10, 12, 15]
# Limiar do "Momento Over" (janela sustentada) que delimita o que NÃO conta como pico
# pontual — antes ficava fixo em 0.6; agora também entra na busca, só descartando
# combinações onde over_pct >= chance_pct (não faria sentido: a janela sustentada
# "engoliria" o próprio pico antes dele contar como Grande Chance)
_CHANCE_PATTERN_OVER_GRID = [0.45, 0.50, 0.55, 0.60, 0.65, 0.70]

@app.route("/api/momentum/chance_pattern_stats")
def api_momentum_chance_pattern_stats():
    """Calibra dinamicamente o indicador 'Grande Chance' (Ao Vivo 2) contra toda a
    base de jogos salvos em momentum_history: testa uma grade de combinações de
    limiar de pico (60%-90% do máximo), limiar da janela sustentada que separa o pico
    (45%-70%) e janela pós-pico (5-15min), e escolhe a combinação com o maior lift
    real (taxa de acerto vs taxa-base) entre as que têm amostra suficiente. Conforme
    mais jogos vão sendo salvos, essa escolha muda sozinha — não é um limiar fixo,
    é recalculado a cada 30min em cima da base atual."""
    now = time.time()
    if _CHANCE_PATTERN_CACHE["data"] and (now - _CHANCE_PATTERN_CACHE["ts"]) < _CHANCE_PATTERN_TTL:
        return jsonify(_CHANCE_PATTERN_CACHE["data"])

    matches = _momentum_load_all_matches()

    resultados = [
        _momentum_eval_pattern(matches, chance_pct, over_pct, window_min)
        for chance_pct in _CHANCE_PATTERN_THRESH_GRID
        for over_pct in _CHANCE_PATTERN_OVER_GRID
        for window_min in _CHANCE_PATTERN_WINDOW_GRID
        if over_pct < chance_pct
    ]

    candidatos = [r for r in resultados if r["total_spikes"] >= _CHANCE_PATTERN_MIN_SAMPLE]
    melhor = max(candidatos, key=lambda r: r["lift"]) if candidatos else None
    valido = bool(melhor and melhor["lift"] >= _CHANCE_PATTERN_MIN_LIFT)

    data = {
        "matches_used":  len(matches),
        "grid_tested":   len(resultados),
        "melhor":        melhor,
        "valido":        valido,
        # Campos usados pelo frontend pra desenhar o marcador com o limiar calibrado
        "chance_pct":    melhor["chance_pct"] if valido else 0.8,
        "over_pct":      melhor["over_pct"] if valido else 0.6,
        "window_min":    melhor["window_min"] if valido else 10,
        "lift":          melhor["lift"] if melhor else 0.0,
    }
    _CHANCE_PATTERN_CACHE["data"] = data
    _CHANCE_PATTERN_CACHE["ts"]   = now
    return jsonify(data)


_SIGNAL_STATS_MINUTE_CACHE = {}     # (tipo, time, window_min, bucket) -> {"ts":, "data":}
_SIGNAL_STATS_MINUTE_TTL = 10 * 60  # 10min — mesmo TTL do cache da base (_momentum_load_all_matches)

@app.route("/api/momentum/signal_stats_by_minute")
def api_momentum_signal_stats_by_minute():
    """Estatística de um tipo de sinal (Grande Chance / Momento Over / Momento
    Under) quebrada por faixa de minuto (buckets de N minutos, padrão 5) contra
    toda a base salva em momentum_history — sem combinar tipos de sinal entre si
    (decisão deliberada: combinações têm risco real de comparações múltiplas /
    achar padrão que não é real). Cada faixa já vem com as duas métricas juntas:

    - hits_gol/rate_gol: saiu gol dentro da janela pós-sinal (window_min).
    - hits_placar/rate_placar: o placar do momento exato do sinal (contando só
      os gols até ali) se manteve como placar FINAL da partida.

    Vêm as duas de uma vez pra poder mostrar o detalhe de uma faixa (ex: ao
    clicar numa barra do gráfico) sem precisar de uma segunda consulta."""
    tipo   = request.args.get("tipo", "chance")
    time_f = request.args.get("time", "")
    try:
        window_min = max(1, min(30, int(request.args.get("window", 10))))
    except ValueError:
        window_min = 10
    try:
        bucket = max(1, min(15, int(request.args.get("bucket", 5))))
    except ValueError:
        bucket = 5

    if tipo not in ("chance", "over", "under"):
        return jsonify({"error": "tipo inválido"}), 400

    cache_key = (tipo, time_f, window_min, bucket)
    now = time.time()
    cached = _SIGNAL_STATS_MINUTE_CACHE.get(cache_key)
    if cached and (now - cached["ts"]) < _SIGNAL_STATS_MINUTE_TTL:
        return jsonify(cached["data"])

    matches = _momentum_load_all_matches()
    cfg = _CHANCE_PATTERN_CACHE["data"] or {}
    chance_pct = cfg.get("chance_pct", 0.8)
    over_pct   = cfg.get("over_pct", 0.6)

    def placar_se_manteve(goals, ref_minute):
        return not any((g.get("minute") or 0) > ref_minute for g in goals)

    buckets = {}  # minute_start -> {"total":, "hits_gol":, "hits_placar":}
    def add(minute, hit_gol, hit_placar):
        b_start = (int(minute) // bucket) * bucket
        b = buckets.setdefault(b_start, {"total": 0, "hits_gol": 0, "hits_placar": 0})
        b["total"] += 1
        if hit_gol:
            b["hits_gol"] += 1
        if hit_placar:
            b["hits_placar"] += 1

    for points, goals in matches:
        if tipo == "chance":
            for sp in _momentum_detect_chance_spikes(points, chance_pct, over_pct):
                team = "home" if sp["is_home"] else "away"
                if time_f and team != time_f:
                    continue
                hit_gol = any(g.get("team") == team and 0 <= (g.get("minute") or 0) - sp["minute"] <= window_min for g in goals)
                hit_placar = placar_se_manteve(goals, sp["minute"])
                add(sp["minute"], hit_gol, hit_placar)
        else:
            over_mk, under_mk = _momentum_detect_over_under_windows(points, over_pct)
            for mk in (over_mk if tipo == "over" else under_mk):
                saiu_gol = any(0 <= (g.get("minute") or 0) - mk["end_minute"] <= window_min for g in goals)
                hit_gol = saiu_gol if tipo == "over" else not saiu_gol
                hit_placar = placar_se_manteve(goals, mk["end_minute"])
                add(mk["minute"], hit_gol, hit_placar)

    result = [
        {
            "minute_start": b_start, "minute_end": b_start + bucket - 1,
            "total": b["total"],
            "hits_gol": b["hits_gol"], "rate_gol": round(b["hits_gol"] / b["total"], 4) if b["total"] else 0.0,
            "hits_placar": b["hits_placar"], "rate_placar": round(b["hits_placar"] / b["total"], 4) if b["total"] else 0.0,
        }
        for b_start, b in sorted(buckets.items())
    ]

    data = {
        "tipo": tipo, "time": time_f or "qualquer",
        "window_min": window_min, "bucket_size": bucket, "matches_used": len(matches),
        "buckets": result,
    }
    _SIGNAL_STATS_MINUTE_CACHE[cache_key] = {"ts": now, "data": data}
    return jsonify(data)


@app.route("/api/momentum/history/<event_id>")
def api_momentum_history_match(event_id):
    """Retorna dados completos de um evento salvo."""
    files = glob.glob(os.path.join(MOMENTUM_DIR, f"*_{event_id}.json"))
    if not files:
        return jsonify({"error": "Não encontrado"}), 404
    with open(files[0], encoding="utf-8") as f:
        return jsonify(json.load(f))


def _minute_range(minute):
    """Classifica o minuto em faixa de jogo."""
    if minute <= 30:  return "early"   # 0-30
    if minute <= 45:  return "ht"      # 31-45 (+ acrésc. 1T)
    if minute <= 70:  return "second"  # 46-70
    return "late"                       # 71-90+


def _goal_situation(team, sh_before, sa_before):
    """Situação do time que marcou, no momento antes do gol."""
    if sh_before == sa_before:
        return "drawing"
    if team == "home":
        return "leading" if sh_before > sa_before else "trailing"
    else:
        return "leading" if sa_before > sh_before else "trailing"


@app.route("/api/momentum/patterns")
def api_momentum_patterns():
    """Extrai padrões pré-gol de todos os históricos salvos.
    Cada padrão inclui minute_range e situation para filtragem contextual.
    """
    WINDOW = 8   # graphPoints antes do gol a capturar
    patterns = []
    files = glob.glob(os.path.join(MOMENTUM_DIR, "*.json"))
    for fpath in files:
        try:
            with open(fpath, encoding="utf-8") as f:
                data = json.load(f)
            pts   = sorted(data.get("graphPoints", []), key=lambda p: p.get("minute", 0))
            goals = sorted(data.get("goals", []), key=lambda g: g.get("minute", 0))
            if not pts or not goals:
                continue
            for idx, goal in enumerate(goals):
                gmin = goal.get("minute", 0)
                team = goal.get("team", "home")

                # Score antes deste gol (conta gols anteriores)
                sh = sum(1 for g in goals[:idx] if g.get("team") == "home")
                sa = sum(1 for g in goals[:idx] if g.get("team") == "away")

                before = [p for p in pts if p.get("minute", 0) < gmin]
                if len(before) < 3:
                    continue
                window = before[-WINDOW:]
                values = [p.get("value", 0) for p in window]
                patterns.append({
                    "team":         team,
                    "values":       values,
                    "goal_min":     gmin,
                    "match":        f"{data.get('casa','?')} x {data.get('fora','?')}",
                    "date":         data.get("date", ""),
                })
        except Exception:
            continue
    return jsonify({"patterns": patterns, "total": len(patterns), "window": WINDOW})


@app.route("/api/shotmap/history")
def api_shotmap_history():
    """Lista todos os shotmaps salvos em shotmap_history/."""
    files = sorted(glob.glob(os.path.join(SHOTMAP_DIR, "*.json")), reverse=True)
    result = []
    for fpath in files:
        try:
            with open(fpath, encoding="utf-8") as f:
                d = json.load(f)
            result.append({
                "event_id":    d.get("event_id"),
                "date":        d.get("date"),
                "casa":        d.get("casa"),
                "fora":        d.get("fora"),
                "liga":        d.get("liga"),
                "score":       d.get("score", {}),
                "total_shots": d.get("total_shots", len(d.get("shotmap", []))),
                "file":        os.path.basename(fpath),
            })
        except Exception:
            continue
    return jsonify({"total": len(result), "matches": result})


@app.route("/api/shotmap/history/<event_id>")
def api_shotmap_history_detail(event_id):
    """Retorna shotmap completo de uma partida específica."""
    files = glob.glob(os.path.join(SHOTMAP_DIR, f"*_{event_id}.json"))
    if not files:
        return jsonify({"error": "not found"}), 404
    try:
        with open(files[0], encoding="utf-8") as f:
            return jsonify(json.load(f))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/forca/history/match")
def api_forca_history_match():
    """Busca odds pré-jogo/força salvos no Backup de Força casando por time+data.
    forca_history vem do NowGoal (event_id próprio dele) enquanto momentum/shotmap
    vêm de outra fonte — não dá pra confiar que o event_id bata entre os dois, então
    casa pelo nome dos times (mesmo critério fuzzy usado em _find_radar_links)."""
    casa = request.args.get("casa", "")
    fora = request.args.get("fora", "")
    date_str = request.args.get("date", "")
    if not casa or not fora or not date_str:
        return jsonify({"error": "casa, fora e date são obrigatórios"}), 400
    files = glob.glob(os.path.join(FORCA_HISTORY_DIR, f"{date_str}_*.json"))
    for fpath in files:
        try:
            with open(fpath, encoding="utf-8") as f:
                d = json.load(f)
            if _name_match(casa, d.get("home", "")) and _name_match(fora, d.get("away", "")):
                return jsonify(d)
        except Exception:
            continue
    return jsonify({"error": "not found"}), 404


@app.route("/api/shotmap/patterns")
def api_shotmap_patterns():
    """Agrega padrões de todos os shotmaps do momentum_history."""
    # Coleta todos os chutes de todos os arquivos
    all_shots = []
    match_count = 0
    seen_ids = set()

    for fpath in sorted(glob.glob(os.path.join(MOMENTUM_DIR, "*.json"))):
        try:
            with open(fpath, encoding="utf-8") as f:
                d = json.load(f)
            shots = d.get("shotmap", [])
            if not shots:
                continue
            event_id = d.get("event_id", "")
            if event_id in seen_ids:
                continue
            seen_ids.add(event_id)
            match_count += 1
            score = d.get("score", {})
            home_g = score.get("home", 0) or 0
            away_g = score.get("away", 0) or 0
            if home_g > away_g:
                match_result = "casa"
            elif away_g > home_g:
                match_result = "vis"
            else:
                match_result = "emp"
            for s in shots:
                all_shots.append({**s, "match_result": match_result})
        except Exception:
            pass

    # Também lê do shotmap_history (arquivos que podem não estar no momentum)
    for fpath in sorted(glob.glob(os.path.join(SHOTMAP_DIR, "*.json"))):
        try:
            with open(fpath, encoding="utf-8") as f:
                d = json.load(f)
            event_id = d.get("event_id", "")
            if event_id in seen_ids:
                continue
            seen_ids.add(event_id)
            shots = d.get("shotmap", [])
            if not shots:
                continue
            match_count += 1
            score = d.get("score", {})
            home_g = score.get("home", 0) or 0
            away_g = score.get("away", 0) or 0
            if home_g > away_g:
                match_result = "casa"
            elif away_g > home_g:
                match_result = "vis"
            else:
                match_result = "emp"
            for s in shots:
                all_shots.append({**s, "match_result": match_result})
        except Exception:
            pass

    if not all_shots:
        return jsonify({"total_shots": 0, "total_matches": 0})

    total = len(all_shots)

    # ── Por desfecho (shotType) ──────────────────────────────────────────
    from collections import defaultdict
    outcome_counts = defaultdict(int)
    for s in all_shots:
        outcome_counts[s.get("shotType", "unknown")] += 1

    # ── Por parte do corpo ───────────────────────────────────────────────
    body_counts = defaultdict(int)
    body_goals  = defaultdict(int)
    for s in all_shots:
        bp = s.get("bodyPart", "unknown")
        body_counts[bp] += 1
        if s.get("shotType") == "goal":
            body_goals[bp] += 1

    # ── Por situação ─────────────────────────────────────────────────────
    sit_counts = defaultdict(int)
    sit_goals  = defaultdict(int)
    for s in all_shots:
        st = s.get("situation", "unknown")
        sit_counts[st] += 1
        if s.get("shotType") == "goal":
            sit_goals[st] += 1

    # ── Casa vs Visitante ────────────────────────────────────────────────
    home_shots = [s for s in all_shots if s.get("isHome")]
    away_shots = [s for s in all_shots if not s.get("isHome")]
    home_goals = sum(1 for s in home_shots if s.get("shotType") == "goal")
    away_goals = sum(1 for s in away_shots if s.get("shotType") == "goal")

    # ── Por zona (x = distância do gol) ─────────────────────────────────
    def classify_zone(x):
        x = x or 0
        if x < 12:
            return "area_pequena"
        elif x < 32:
            return "area_grande"
        elif x < 55:
            return "meia_distancia"
        else:
            return "longa_distancia"

    zone_shots = defaultdict(int)
    zone_goals = defaultdict(int)
    for s in all_shots:
        z = classify_zone(s.get("x", 50))
        zone_shots[z] += 1
        if s.get("shotType") == "goal":
            zone_goals[z] += 1

    zone_labels = [
        ("area_pequena",    "Área Pequena",   "x < 12m"),
        ("area_grande",     "Área Grande",    "12-32m"),
        ("meia_distancia",  "Meia Distância", "32-55m"),
        ("longa_distancia", "Longa Distância","55m+"),
    ]
    by_zone = []
    for key, label, desc in zone_labels:
        n = zone_shots[key]
        g = zone_goals[key]
        by_zone.append({
            "key": key, "label": label, "desc": desc,
            "shots": n, "goals": g,
            "goal_pct": round(g / n * 100, 1) if n else 0,
        })

    # ── Por período (minuto) ─────────────────────────────────────────────
    buckets_def = [
        ("1–15", 1, 15), ("16–30", 16, 30), ("31–45", 31, 45),
        ("46–60", 46, 60), ("61–75", 61, 75), ("76–90+", 76, 200),
    ]
    by_minute = []
    for label, lo, hi in buckets_def:
        sl = [s for s in all_shots if lo <= (s.get("minute") or 0) <= hi]
        gl = [s for s in sl if s.get("shotType") == "goal"]
        by_minute.append({
            "label": label, "shots": len(sl), "goals": len(gl),
            "goal_pct": round(len(gl) / len(sl) * 100, 1) if sl else 0,
        })

    # ── Por resultado do jogo ────────────────────────────────────────────
    result_shots = defaultdict(int)
    result_goals = defaultdict(int)
    for s in all_shots:
        r = s.get("match_result", "emp")
        result_shots[r] += 1
        if s.get("shotType") == "goal":
            result_goals[r] += 1

    def _pct(a, b):
        return round(a / b * 100, 1) if b else 0

    return jsonify({
        "total_shots":   total,
        "total_matches": match_count,
        "by_outcome": dict(outcome_counts),
        "by_body_part": {
            bp: {"shots": body_counts[bp], "goals": body_goals[bp],
                 "goal_pct": _pct(body_goals[bp], body_counts[bp])}
            for bp in body_counts
        },
        "by_situation": {
            st: {"shots": sit_counts[st], "goals": sit_goals[st],
                 "goal_pct": _pct(sit_goals[st], sit_counts[st])}
            for st in sit_counts
        },
        "home_away": {
            "home": {"total": len(home_shots), "goals": home_goals,
                     "goal_pct": _pct(home_goals, len(home_shots))},
            "away": {"total": len(away_shots), "goals": away_goals,
                     "goal_pct": _pct(away_goals, len(away_shots))},
        },
        "by_zone":   by_zone,
        "by_minute": by_minute,
        "by_result": {
            r: {"shots": result_shots[r], "goals": result_goals[r],
                "goal_pct": _pct(result_goals[r], result_shots[r])}
            for r in result_shots
        },
    })


@app.route("/api/momentum/export")
def api_momentum_export():
    """Gera planilha Excel com os dados das partidas ao vivo salvas."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file
    import io

    # Carrega todos os arquivos salvos
    files = sorted(glob.glob(os.path.join(MOMENTUM_DIR, "*.json")), reverse=True)
    matches_data = []
    for fpath in files:
        try:
            with open(fpath, encoding="utf-8") as f:
                matches_data.append(json.load(f))
        except Exception:
            pass

    wb = Workbook()

    # ── Estilos ───────────────────────────────────────────────────────────
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", fgColor="1E3A5F")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c_align   = Alignment(horizontal="center", vertical="center")
    fill_green = PatternFill("solid", fgColor="C6EFCE")
    fill_red   = PatternFill("solid", fgColor="FFC7CE")
    fill_alt   = PatternFill("solid", fgColor="F0F4FF")
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def _hcell(ws, row, col, value, width=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = hdr_align; c.border = thin
        if width:
            ws.column_dimensions[get_column_letter(col)].width = width
        return c

    def _dcell(ws, row, col, value, fill=None):
        c = ws.cell(row=row, column=col, value=value)
        c.alignment = c_align; c.border = thin
        if fill: c.fill = fill
        return c

    # ── Sheet 1: Resumo ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumo"
    ws1.row_dimensions[1].height = 32

    hdrs = [
        ("Data",         12), ("Liga",         22), ("Casa",         20),
        ("Fora",         20), ("Placar FT",     10), ("Placar HT",     10),
        ("Gols Casa FT",  10), ("Gols Fora FT", 10), ("Gols Casa HT", 10),
        ("Gols Fora HT", 10), ("Total Gols",    10), ("Minutos\nSalvos", 10),
    ]
    for col, (label, width) in enumerate(hdrs, 1):
        _hcell(ws1, 1, col, label, width)

    for ri, d in enumerate(matches_data, 2):
        goals = d.get("goals", [])
        pts   = d.get("graphPoints", [])
        gc_ft = sum(1 for g in goals if g.get("team") == "home")
        gf_ft = sum(1 for g in goals if g.get("team") == "away")
        gc_ht = sum(1 for g in goals if g.get("team") == "home" and g.get("minute", 0) <= 45)
        gf_ht = sum(1 for g in goals if g.get("team") == "away" and g.get("minute", 0) <= 45)
        row_fill = fill_alt if ri % 2 == 0 else None
        _dcell(ws1, ri, 1, d.get("date", ""), row_fill)
        _dcell(ws1, ri, 2, d.get("liga", ""), row_fill)
        _dcell(ws1, ri, 3, d.get("casa", ""), row_fill)
        _dcell(ws1, ri, 4, d.get("fora", ""), row_fill)
        _dcell(ws1, ri, 5, f"{gc_ft}:{gf_ft}", fill_green if gc_ft + gf_ft > 0 else row_fill)
        _dcell(ws1, ri, 6, f"{gc_ht}:{gf_ht}", row_fill)
        _dcell(ws1, ri, 7, gc_ft, row_fill)
        _dcell(ws1, ri, 8, gf_ft, row_fill)
        _dcell(ws1, ri, 9, gc_ht, row_fill)
        _dcell(ws1, ri, 10, gf_ht, row_fill)
        _dcell(ws1, ri, 11, gc_ft + gf_ft,
               fill_green if gc_ft + gf_ft >= 3 else (fill_red if gc_ft + gf_ft == 0 else row_fill))
        _dcell(ws1, ri, 12, len(pts), row_fill)

    ws1.freeze_panes = "A2"

    # ── Sheet 2: Gols por Minuto ──────────────────────────────────────────
    ws2 = wb.create_sheet("Gols")
    ws2.row_dimensions[1].height = 40

    _hcell(ws2, 1, 1, "Minuto", 8)
    _hcell(ws2, 1, 2, "Partida", 26)
    _hcell(ws2, 1, 3, "Data", 12)
    _hcell(ws2, 1, 4, "Liga", 20)
    _hcell(ws2, 1, 5, "Time", 12)
    _hcell(ws2, 1, 6, "Acréscimo", 10)
    _hcell(ws2, 1, 7, "Placar Após", 12)

    ri = 2
    for d in matches_data:
        goals = sorted(d.get("goals", []), key=lambda g: g.get("minute", 0))
        partida = f"{d.get('casa','?')} x {d.get('fora','?')}"
        sh, sa = 0, 0
        for g in goals:
            team = g.get("team", "")
            if team == "home": sh += 1
            else: sa += 1
            row_fill = fill_green if team == "home" else fill_red
            _dcell(ws2, ri, 1, g.get("minute", ""), row_fill)
            _dcell(ws2, ri, 2, partida, row_fill)
            _dcell(ws2, ri, 3, d.get("date", ""), row_fill)
            _dcell(ws2, ri, 4, d.get("liga", ""), row_fill)
            _dcell(ws2, ri, 5, d.get("casa", "") if team == "home" else d.get("fora", ""), row_fill)
            _dcell(ws2, ri, 6, g.get("addedTime", 0) or 0, row_fill)
            _dcell(ws2, ri, 7, f"{sh}:{sa}", row_fill)
            ri += 1
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Momentum (matrix minuto × partida) ───────────────────────
    ws3 = wb.create_sheet("Momentum")
    ws3.row_dimensions[1].height = 44

    _hcell(ws3, 1, 1, "Min", 5)
    for ci, d in enumerate(matches_data, 2):
        label = f"{d.get('casa','?')} x {d.get('fora','?')}\n{d.get('date','')}"
        _hcell(ws3, 1, ci, label, 18)
        ws3.column_dimensions[get_column_letter(ci)].width = 8

    for rmi, minute in enumerate(range(1, 91), 2):
        c = ws3.cell(row=rmi, column=1, value=minute)
        c.font = Font(bold=True, size=9)
        c.alignment = c_align
        for ci, d in enumerate(matches_data, 2):
            # Monta mapa tolerando int e float como chave
            pts_map = {}
            for p in d.get("graphPoints", []):
                pts_map[int(p["minute"])] = p["value"]
            val = pts_map.get(minute)      # None quando realmente ausente
            cell = ws3.cell(row=rmi, column=ci, value=val if val is not None else "")
            cell.alignment = c_align
            cell.border = thin
            cell.font = Font(size=8)
            if val is not None:
                if val > 0:   cell.fill = fill_green
                elif val < 0: cell.fill = fill_red
    ws3.freeze_panes = "B2"

    # ── Sheet 4: Dados_ML (formato largo para modelo — 1 linha por partida) ─
    ws4 = wb.create_sheet("Dados_ML")
    ws4.row_dimensions[1].height = 28

    # Colunas fixas de contexto
    ctx_cols = [
        ("date", 12), ("liga", 20), ("casa", 18), ("fora", 18),
        ("gc_ft", 8), ("gf_ft", 8), ("gc_ht", 8), ("gf_ht", 8),
        ("total_gols", 10), ("tem_gol", 8),
    ]
    # + colunas de minuto 1..90
    min_cols = [f"min_{m}" for m in range(1, 91)]

    all_cols = ctx_cols + [(c, 7) for c in min_cols]
    for ci, (label, width) in enumerate(all_cols, 1):
        _hcell(ws4, 1, ci, label, width)

    for ri, d in enumerate(matches_data, 2):
        goals = d.get("goals", [])
        pts   = d.get("graphPoints", [])
        gc_ft = sum(1 for g in goals if g.get("team") == "home")
        gf_ft = sum(1 for g in goals if g.get("team") == "away")
        gc_ht = sum(1 for g in goals if g.get("team") == "home" and g.get("minute", 0) <= 45)
        gf_ht = sum(1 for g in goals if g.get("team") == "away" and g.get("minute", 0) <= 45)
        total = gc_ft + gf_ft
        row_fill = fill_alt if ri % 2 == 0 else None

        ctx_vals = [
            d.get("date", ""), d.get("liga", ""), d.get("casa", ""), d.get("fora", ""),
            gc_ft, gf_ft, gc_ht, gf_ht, total, 1 if total > 0 else 0,
        ]
        for ci, val in enumerate(ctx_vals, 1):
            _dcell(ws4, ri, ci, val, row_fill)

        # Mapa minuto → valor (int chave)
        pts_map = {}
        for p in pts:
            pts_map[int(p["minute"])] = p["value"]

        for mi, minute in enumerate(range(1, 91), len(ctx_cols) + 1):
            val = pts_map.get(minute, 0)   # 0 quando minuto ausente
            cell = ws4.cell(row=ri, column=mi, value=val)
            cell.alignment = c_align
            cell.border = thin
            cell.font = Font(size=8)
            if val > 0:   cell.fill = fill_green
            elif val < 0: cell.fill = fill_red
            elif row_fill: cell.fill = row_fill

    ws4.freeze_panes = "A2"

    # ── Envia o arquivo ───────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"partidas_ao_vivo_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


ANALYSIS_CACHE_FILE = os.path.join(DATA_DIR, "momentum_analysis.json")


def _compute_analysis():
    """Computa padrões de gol com janela ótima INDIVIDUAL por categoria (3-15 min)."""
    import math, random as _rand

    NONE_GAP  = 12
    NONE_STEP = 3
    CANDIDATES = list(range(3, 16))

    # Pré-carrega todos os arquivos uma única vez
    files_list = sorted(glob.glob(os.path.join(MOMENTUM_DIR, "*.json")))
    all_data = []
    for fpath in files_list:
        try:
            with open(fpath, encoding="utf-8") as f:
                d = json.load(f)
            pts   = sorted(d.get("graphPoints", []), key=lambda p: float(p.get("minute", 0)))
            goals = sorted(d.get("goals",       []), key=lambda g: float(g.get("minute", 0)))
            if len(pts) < 5:
                continue
            all_data.append({
                "pt_list":   [(float(p["minute"]), p["value"]) for p in pts],
                "goal_list": [(float(g["minute"]), g["team"])  for g in goals],
            })
        except Exception:
            continue

    total_matches = len(all_data)
    total_goals   = sum(len(d["goal_list"]) for d in all_data)

    def feats(w):
        tail  = sum(w[-4:]) / max(len(w), 1)
        trend = w[-1] - w[0]
        peak  = max(w, key=abs)
        return tail + 0.3 * trend + 0.2 * peak

    # ── Pré-extrai todas as janelas para todos os W candidatos ────────────
    wins_cache = {}
    for W in CANDIDATES:
        hw, aw, nw, ht_w, st_w = [], [], [], [], []
        for d in all_data:
            pt_list   = d["pt_list"]
            goal_list = d["goal_list"]
            if len(pt_list) < W + 2:
                continue
            goal_mins = [gm for gm, _ in goal_list]
            for gmin, team in goal_list:
                before = [(m, v) for m, v in pt_list if m < gmin]
                if len(before) < W:
                    continue
                win = [v for _, v in before[-W:]]
                (hw if team == "home" else aw).append(win)
                (ht_w if gmin <= 45 else st_w).append(win)
            for i in range(W, len(pt_list) - 1, NONE_STEP):
                m_now = pt_list[i][0]
                if any(abs(gm - m_now) <= NONE_GAP for gm in goal_mins):
                    continue
                nw.append([pt_list[j][1] for j in range(i - W, i)])
        wins_cache[W] = (hw, aw, nw, ht_w, st_w)

    # ── Busca janela+threshold ótimos para uma categoria específica ───────
    def find_best_for(pos_fn, neg_fn, direction="pos"):
        """
        direction: "pos"  → score > T prediz positivo (gol casa)
                   "neg"  → score < -T prediz positivo (gol visitante)
                   "any"  → |score| > T prediz positivo (qualquer gol / 1T / 2T)
                   "none" → |score| < T prediz positivo (sem gol)
        """
        best_W = CANDIDATES[0]; best_acc = 0.0; best_T = 8
        for W in CANDIDATES:
            pos = pos_fn(W)
            neg = neg_fn(W)
            if len(pos) < 4:
                continue
            neg_bal = _rand.sample(neg, min(len(neg), max(len(pos), 1)))
            labeled = [(w, True) for w in pos] + [(w, False) for w in neg_bal]
            for T in range(1, 60):
                if direction == "pos":
                    correct = sum(1 for w, lbl in labeled if (feats(w) > T) == lbl)
                elif direction == "neg":
                    correct = sum(1 for w, lbl in labeled if (feats(w) < -T) == lbl)
                elif direction == "any":
                    correct = sum(1 for w, lbl in labeled if (abs(feats(w)) > T) == lbl)
                else:  # none
                    correct = sum(1 for w, lbl in labeled if (abs(feats(w)) <= T) == lbl)
                bal = correct / max(len(labeled), 1)
                if bal > best_acc:
                    best_acc = bal; best_T = T; best_W = W
        return best_W, best_T, round(best_acc * 100, 1)

    # Janela ótima individual por categoria
    home_W, home_T, home_acc = find_best_for(
        lambda W: wins_cache[W][0], lambda W: wins_cache[W][2], "pos")
    away_W, away_T, away_acc = find_best_for(
        lambda W: wins_cache[W][1], lambda W: wins_cache[W][2], "neg")
    any_W,  any_T,  any_acc  = find_best_for(
        lambda W: wins_cache[W][0] + wins_cache[W][1], lambda W: wins_cache[W][2], "any")
    none_W, none_T, none_acc = find_best_for(
        lambda W: wins_cache[W][2], lambda W: wins_cache[W][0] + wins_cache[W][1], "none")
    ht_W,   ht_T,   ht_acc   = find_best_for(
        lambda W: wins_cache[W][3], lambda W: wins_cache[W][2], "any")
    st_W,   st_T,   st_acc   = find_best_for(
        lambda W: wins_cache[W][4], lambda W: wins_cache[W][2], "any")

    # Janela global (para o modelo geral e prob)
    window_scores = {}
    best_window = 8; best_acc_overall = 0.0; best_T_overall = 8
    for W in CANDIDATES:
        hw, aw, nw, _, _ = wins_cache[W]
        goal_n = len(hw) + len(aw)
        if goal_n < 4:
            window_scores[W] = 0.0; continue
        none_bal = _rand.sample(nw, min(len(nw), max(goal_n, 1)))
        labeled  = ([(w,"home") for w in hw] + [(w,"away") for w in aw] +
                    [(w,"none") for w in none_bal])
        best_bal = 0.0; T_found = 8
        for T in range(1, 60):
            cnt = {"home":[0,0],"away":[0,0],"none":[0,0]}
            for w, lbl in labeled:
                score = feats(w)
                pred  = "home" if score > T else ("away" if score < -T else "none")
                cnt[lbl][1] += 1; cnt[lbl][0] += int(pred == lbl)
            recalls = [cnt[k][0]/cnt[k][1] for k in cnt if cnt[k][1]]
            bal = sum(recalls)/len(recalls) if recalls else 0
            if bal > best_bal:
                best_bal = bal; T_found = T
        window_scores[W] = round(best_bal * 100, 1)
        if best_bal > best_acc_overall:
            best_acc_overall = best_bal; best_window = W; best_T_overall = T_found

    # ── win_stats com W dinâmico por categoria ────────────────────────────
    def win_stats(wins, W):
        n = len(wins)
        if not n:
            return {"avg": [0.0]*W, "std": [0.0]*W, "n": 0, "tail_mean": 0.0}
        avg = [sum(w[i] for w in wins) / n for i in range(W)]
        std = [math.sqrt(sum((w[i]-avg[i])**2 for w in wins)/max(n-1,1)) for i in range(W)]
        tail_mean = sum(sum(w[-4:])/max(len(w),1) for w in wins) / n
        return {"avg": [round(v,1) for v in avg],
                "std": [round(v,1) for v in std],
                "n": n, "tail_mean": round(tail_mean, 1)}

    home_wins = wins_cache[home_W][0]
    away_wins = wins_cache[away_W][1]
    any_wins  = wins_cache[any_W][0]  + wins_cache[any_W][1]
    none_wins = wins_cache[none_W][2]
    ht_wins   = wins_cache[ht_W][3]
    st_wins   = wins_cache[st_W][4]

    # ── Probabilidade: P(gol nos próx. LOOKAHEAD min | sinal X) ──────────
    LOOKAHEAD = 10
    W = best_window
    prob = {"home": [0,0], "away": [0,0], "any": [0,0], "none": [0,0]}
    for d in all_data:
        pt_list   = d["pt_list"]
        goal_list = d["goal_list"]
        if len(pt_list) < W + 2:
            continue
        for i in range(W, len(pt_list)):
            m_now = pt_list[i][0]
            win   = [pt_list[j][1] for j in range(i - W, i)]
            score = feats(win)
            if score > best_T_overall:               sig = "home"
            elif score < -best_T_overall:            sig = "away"
            elif abs(score) > best_T_overall * 0.6: sig = "any"
            else:                                    sig = "none"
            goals_ahead = [(gm, gt) for gm, gt in goal_list
                           if gm > m_now and gm <= m_now + LOOKAHEAD]
            prob[sig][1] += 1
            if sig == "home"  and any(gt=="home"  for _,gt in goals_ahead): prob[sig][0] += 1
            elif sig == "away" and any(gt=="away" for _,gt in goals_ahead): prob[sig][0] += 1
            elif sig == "any"  and goals_ahead:                             prob[sig][0] += 1
            elif sig == "none" and not goals_ahead:                         prob[sig][0] += 1

    def sp(a, b): return round(a/b*100, 1) if b else None

    per_cat = {
        "home": {"window": home_W, "threshold": home_T, "accuracy": home_acc},
        "away": {"window": away_W, "threshold": away_T, "accuracy": away_acc},
        "any":  {"window": any_W,  "threshold": any_T,  "accuracy": any_acc},
        "none": {"window": none_W, "threshold": none_T, "accuracy": none_acc},
        "ht":   {"window": ht_W,   "threshold": ht_T,   "accuracy": ht_acc},
        "st":   {"window": st_W,   "threshold": st_T,   "accuracy": st_acc},
    }

    return {
        "patterns": {
            "home": win_stats(home_wins, home_W),
            "away": win_stats(away_wins, away_W),
            "any":  win_stats(any_wins,  any_W),
            "none": win_stats(none_wins, none_W),
            "ht":   win_stats(ht_wins,   ht_W),
            "st":   win_stats(st_wins,   st_W),
        },
        "per_category":  per_cat,
        "threshold":     best_T_overall,
        "accuracy":      round(best_acc_overall * 100, 1),
        "acc_home":      home_acc,
        "acc_away":      away_acc,
        "acc_none":      none_acc,
        "prob_home":     sp(*prob["home"]),
        "prob_away":     sp(*prob["away"]),
        "prob_any":      sp(*prob["any"]),
        "prob_none":     sp(*prob["none"]),
        "prob_lookahead": LOOKAHEAD,
        "total_matches": total_matches,
        "total_goals":   total_goals,
        "window":        best_window,
        "window_scores": window_scores,
        "computed_at":   datetime.now().isoformat(),
    }


def _rebuild_analysis_cache():
    """Reconstrói o cache de análise e salva em disco. Chamado automaticamente."""
    try:
        result = _compute_analysis()
        with open(ANALYSIS_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f"[analysis] Cache atualizado — {result['total_matches']} partidas, "
              f"acurácia {result['accuracy']}%")
    except Exception as e:
        print(f"[analysis] Erro ao reconstruir cache: {e}")


@app.route("/api/momentum/analysis")
def api_momentum_analysis():
    """Retorna análise de padrões de gol. Usa cache em disco; recalcula se ausente."""
    # Tenta ler cache
    if os.path.exists(ANALYSIS_CACHE_FILE):
        try:
            with open(ANALYSIS_CACHE_FILE, encoding="utf-8") as f:
                return jsonify(json.load(f))
        except Exception:
            pass
    # Cache ausente/corrompido: calcula na hora e salva
    result = _compute_analysis()
    try:
        with open(ANALYSIS_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
    except Exception:
        pass
    return jsonify(result)


@app.route("/api/momentum/similar", methods=["POST"])
def api_momentum_similar():
    """Busca partidas salvas com padrão de momentum similar ao atual.

    Usa uma comparação leve (soma/tendência da janela, não o vetor ponto-a-ponto
    completo) pra ficar rápido mesmo varrendo milhares de jogos salvos — comparar
    cada posição de cada arquivo com distância euclidiana completa era o gargalo
    que deixava a aba Análise lenta."""
    try:
        body    = request.get_json(force=True) or {}
        pts_raw = body.get("points", [])
        W       = int(body.get("window", 8))
        LOOKAHEAD = int(body.get("lookahead", 10))

        if len(pts_raw) < W:
            return jsonify({"similar": [], "total": 0, "goal_home": 0, "goal_away": 0, "goal_none": 0})

        cur_vals    = [float(p.get("value", 0)) for p in pts_raw[-W:]]
        cur_signal  = sum(cur_vals) / W                       # tendência média da janela
        cur_swing   = max(cur_vals) - min(cur_vals) if cur_vals else 0  # volatilidade

        STRIDE = 2  # varre de 2 em 2 minutos em vez de todo minuto — ~2x mais rápido

        similar = []
        for d in _get_momentum_files_cached():
            try:
                pt_list = d["pt_list"]
                goals   = d["goals"]
                if len(pt_list) < W + 2:
                    continue

                best_dist = float("inf")
                best_outcome = "none"
                best_min = 0

                for i in range(W, len(pt_list), STRIDE):
                    win = [pt_list[j][1] for j in range(i - W, i)]
                    win_signal = sum(win) / W
                    win_swing  = max(win) - min(win) if win else 0
                    dist = abs(win_signal - cur_signal) + 0.3 * abs(win_swing - cur_swing)
                    if dist < best_dist:
                        best_dist = dist
                        m_now = pt_list[i][0]
                        ahead = [(float(g.get("minute", 0)), g.get("team", ""))
                                 for g in goals
                                 if float(g.get("minute", 0)) > m_now
                                 and float(g.get("minute", 0)) <= m_now + LOOKAHEAD]
                        if any(t == "home" for _, t in ahead):
                            best_outcome = "home"
                        elif any(t == "away" for _, t in ahead):
                            best_outcome = "away"
                        elif ahead:
                            best_outcome = "any"
                        else:
                            best_outcome = "none"
                        best_min = int(m_now)

                similar.append({
                    "casa":     d.get("casa", "—"),
                    "fora":     d.get("fora", "—"),
                    "liga":     d.get("liga", ""),
                    "date":     d.get("date", ""),
                    "outcome":  best_outcome,
                    "distance": round(best_dist, 3),
                    "minute":   best_min,
                })
            except Exception:
                continue

        similar.sort(key=lambda x: x["distance"])
        total = len(similar)

        # Top 5 para exibição na lista
        top_display = similar[:5]

        # Estatísticas em TODAS as partidas com distância <= limiar (sem tamanho fixo) —
        # ou seja, pega quantas forem realmente parecidas com o padrão atual, nem mais nem menos.
        DIST_THRESHOLD = float(body.get("dist_threshold", 8))
        MIN_SAMPLE = 10  # piso pra evitar % instável com amostra minúscula
        top_stat = [s for s in similar if s["distance"] <= DIST_THRESHOLD]
        if len(top_stat) < MIN_SAMPLE:
            top_stat = similar[:MIN_SAMPLE]
        STAT_N = len(top_stat)

        return jsonify({
            "similar":    top_display,
            "total":      total,
            "stat_n":     STAT_N,
            "goal_home":  sum(1 for s in top_stat if s["outcome"] == "home"),
            "goal_away":  sum(1 for s in top_stat if s["outcome"] == "away"),
            "goal_none":  sum(1 for s in top_stat if s["outcome"] == "none"),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/momentum/similar-scores", methods=["POST"])
def api_momentum_similar_scores():
    """Identifica o placar atual (casa/visitante) e busca no histórico partidas que
    tiveram esse MESMO placar em algum momento do jogo, informando os placares finais
    mais comuns entre elas. Sem análise de gráfico — só o placar, bem mais rápido."""
    try:
        body     = request.get_json(force=True) or {}
        cur_casa = int(body.get("cur_casa", 0))
        cur_fora = int(body.get("cur_fora", 0))

        counts = {}
        total_checked = 0
        for d in _get_momentum_files_cached():
            try:
                goals = sorted(d["goals"], key=lambda g: float(g.get("minute", 0)))
                score_final = d.get("score") or {}
                fh, fa = score_final.get("home"), score_final.get("away")
                if fh is None or fa is None:
                    continue
                total_checked += 1

                # Recria o placar minuto a minuto pra ver se em algum ponto bateu com o atual
                h = a = 0
                hit = (cur_casa == 0 and cur_fora == 0)  # todo jogo começa 0-0
                for g in goals:
                    if g.get("team") == "home": h += 1
                    elif g.get("team") == "away": a += 1
                    if h == cur_casa and a == cur_fora:
                        hit = True
                        break
                    if h > cur_casa or a > cur_fora:
                        break  # passou do placar atual sem bater — não teve esse momento

                if hit:
                    key = f"{fh}-{fa}"
                    counts[key] = counts.get(key, 0) + 1
            except Exception:
                continue

        stat_n = sum(counts.values())
        if stat_n == 0:
            return jsonify({"scores": [], "total": total_checked, "stat_n": 0})

        scores = sorted(
            [{"score": k, "count": v, "pct": round(v / stat_n * 100)} for k, v in counts.items()],
            key=lambda x: -x["count"]
        )[:4]

        return jsonify({"scores": scores, "total": total_checked, "stat_n": stat_n})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def _shotmap_feature_vector(shots):
    """Converte uma lista de chutes num vetor de 10 posições: contagem por zona
    (Área Pequena / Área Grande / Meia Distância / Longa Distância) e chutes no
    alvo, separado por casa/visitante. Mesmas zonas usadas no mapa de chutes visual."""
    zones = {"casa": [0, 0, 0, 0], "fora": [0, 0, 0, 0]}
    on_target = {"casa": 0, "fora": 0}
    for s in shots:
        side = "casa" if s.get("isHome") else "fora"
        x = float(s.get("x", 50))
        if x <= 12:
            zi = 0
        elif x <= 32:
            zi = 1
        elif x <= 45:
            zi = 2
        else:
            zi = 3
        zones[side][zi] += 1
        if s.get("shotType") in ("goal", "save"):
            on_target[side] += 1
    vec = zones["casa"] + zones["fora"] + [on_target["casa"], on_target["fora"]]
    total = sum(abs(v) for v in vec) or 1
    return [v / total for v in vec]


@app.route("/api/shotmap/similar", methods=["POST"])
def api_shotmap_similar():
    """Busca partidas salvas com padrão de mapa de chutes parecido e informa
    a % de jogos em que casa/visitante marcou gol nesse padrão."""
    try:
        body  = request.get_json(force=True) or {}
        shots = body.get("shots", [])
        if not shots:
            return jsonify({"similar": [], "total": 0, "goal_home": 0, "goal_away": 0, "goal_none": 0})

        cur_vec = _shotmap_feature_vector(shots)

        similar = []
        for d in _get_momentum_files_cached():
            try:
                hist_shots = d["shotmap"]
                if not hist_shots:
                    continue
                score = d.get("score", {}) or {}
                gh, ga = score.get("home"), score.get("away")
                if gh is None or ga is None:
                    continue

                hist_vec = _shotmap_feature_vector(hist_shots)
                dist = sum((a - b) ** 2 for a, b in zip(cur_vec, hist_vec)) ** 0.5

                outcome = "home" if gh > 0 and ga == 0 else \
                          "away" if ga > 0 and gh == 0 else \
                          "both" if gh > 0 and ga > 0 else "none"

                similar.append({
                    "casa": d.get("casa", "—"), "fora": d.get("fora", "—"),
                    "liga": d.get("liga", ""), "date": d.get("date", ""),
                    "outcome": outcome, "distance": round(dist, 3),
                    "placar": f"{gh}-{ga}",
                    "home_scored": gh > 0, "away_scored": ga > 0,
                })
            except Exception:
                continue

        similar.sort(key=lambda x: x["distance"])
        total = len(similar)
        STAT_N = min(30, total)
        top_stat = similar[:STAT_N]

        return jsonify({
            "similar":   similar[:5],
            "total":     total,
            "stat_n":    STAT_N,
            "goal_home": sum(1 for s in top_stat if s["home_scored"]),
            "goal_away": sum(1 for s in top_stat if s["away_scored"]),
            "goal_none": sum(1 for s in top_stat if not s["home_scored"] and not s["away_scored"]),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Padrões de Estatísticas ──────────────────────────────────────────────────

# Estatísticas que queremos rastrear — suporta keys SofaScore (camelCase) e UniScore (snake_case)
_STAT_LABELS = {
    # ── Posse & Ataque ──────────────────────────────────────────────────────
    "ball_possession":        "Posse de Bola (%)",
    "ballPossession":         "Posse de Bola (%)",
    "shots":                  "Total de Chutes",
    "totalShots":             "Total de Chutes",
    "shots_on_target":        "Chutes no Alvo",
    "shotsOnTarget":          "Chutes no Alvo",
    "blocked_shots":          "Chutes Bloqueados",
    "shots_inside_box":       "Chutes Dentro da Área",
    "shots_outside_box":      "Chutes Fora da Área",
    "touches_in_box":         "Toques na Área Adversária",
    "big_chances":            "Grandes Chances",
    "bigChancesCreated":      "Grandes Chances",
    "corner_kicks":           "Escanteios",
    "cornerKicks":            "Escanteios",
    "freekicks":              "Cobranças de Falta",
    # ── Passes ──────────────────────────────────────────────────────────────
    "passes":                 "Total de Passes",
    "totalPasses":            "Total de Passes",
    "pass_in_final_third":    "Passes no Terço Final",
    "final_third_entries":    "Entradas no Último Terço",
    "long_balls":             "Lançamentos Longos",
    "crosses_accuracy":       "Cruzamentos",
    "throw_in":               "Arremessos Laterais",
    # ── Duelos & Dribles ────────────────────────────────────────────────────
    "duels":                  "Duelos Totais",
    "ground_duels":           "Duelos no Chão",
    "aerial_duels":           "Duelos Aéreos",
    "dribble":                "Dribles",
    "dispossessed":           "Perda de Posse (Dribble)",
    # ── Defesa ──────────────────────────────────────────────────────────────
    "tackles":                "Desarmes",
    "tacklesWon":             "Desarmes",
    "interceptions":          "Intercepções",
    "recoveries":             "Recuperações",
    "clearances":             "Rebotes/Afastamentos",
    # ── Goleiro ─────────────────────────────────────────────────────────────
    "saves":                  "Defesas (GK)",
    "goal_kicks":             "Tiros de Meta",
    # ── Disciplina ──────────────────────────────────────────────────────────
    "fouls":                  "Faltas Cometidas",
    "foulsCommitted":         "Faltas Cometidas",
    "was_fouled":             "Sofreu Falta",
    "yellow_cards":           "Cartões Amarelos",
    "yellowCards":            "Cartões Amarelos",
    # ── Perda de Posse ──────────────────────────────────────────────────────
    "poss_losts":             "Perda de Posse Total",
    # ── Indicadores derivados (calculados no save) ───────────────────────
    "xg":                     "xG (Gols Esperados)",
    "pressure_home_dom_pct":  "Dominância de Pressão (%)",
    "pressure_overall_avg":   "Pressão Média (saldo)",
    "pressure_momentum_swings": "Trocas de Dominância",
}

_stats_patterns_cache = {"ts": 0, "data": None}


def _parse_stat_val(raw):
    """Converte string (ex: '57%', '3') ou número para float. None se inválido."""
    if raw is None:
        return None
    try:
        return float(str(raw).replace("%", "").strip())
    except (ValueError, TypeError):
        return None


def _extract_stats(statistics_raw):
    """Extrai {key: (home_val, away_val)} do período ALL.
    Aceita dois formatos:
      - Novo (flat dict): {"ball_possession": {"homeValue": 55, "awayValue": 45, ...}, ...}
      - Antigo (lista SofaScore): [{"period":"ALL","groups":[{"statisticsItems":[...]}]}]
    """
    result = {}

    # ── Formato novo: dict flat ────────────────────────────────────────────
    if isinstance(statistics_raw, dict):
        for key, item in statistics_raw.items():
            if not isinstance(item, dict):
                continue
            hv = _parse_stat_val(item.get("homeValue") or item.get("home"))
            av = _parse_stat_val(item.get("awayValue") or item.get("away"))
            if hv is not None or av is not None:
                result[key] = (hv, av)
        return result

    # ── Formato antigo: lista com period/groups/statisticsItems ───────────
    if isinstance(statistics_raw, list):
        for period_data in statistics_raw:
            if period_data.get("period") != "ALL":
                continue
            for group in period_data.get("groups", []):
                for item in group.get("statisticsItems", []):
                    key = item.get("key", "")
                    if not key:
                        continue
                    hv = _parse_stat_val(item.get("homeValue") or item.get("home"))
                    av = _parse_stat_val(item.get("awayValue") or item.get("away"))
                    if hv is not None or av is not None:
                        result[key] = (hv, av)
            break  # só período ALL

    return result


@app.route("/api/momentum/stats-patterns")
def api_stats_patterns():
    """Analisa padrões das estatísticas finais das partidas salvas,
    agrupando por resultado (Casa V., Empate, Vis. V., Over/Under 2.5, BTS)."""
    global _stats_patterns_cache
    if (time.time() - _stats_patterns_cache["ts"] < 1800
            and _stats_patterns_cache["data"] is not None):
        return jsonify(_stats_patterns_cache["data"])

    OUTCOME_KEYS = ["casaV", "emp", "visV", "o25", "u25", "btts", "nbtts"]
    # {outcome: {stat_key_h|_a: [values]}}
    buckets = {oc: {} for oc in OUTCOME_KEYS}
    total = 0

    for fpath in glob.glob(os.path.join(MOMENTUM_DIR, "*.json")):
        try:
            with open(fpath, encoding="utf-8") as f:
                d = json.load(f)

            # ── Determina placar ──────────────────────────────────────────
            # Prefere score salvo, fallback para contar goals
            score = d.get("score", {})
            if score and "home" in score and "away" in score:
                gh = int(score["home"])
                ga = int(score["away"])
            else:
                goals = d.get("goals", [])
                gh = sum(1 for g in goals if g.get("team") == "home")
                ga = sum(1 for g in goals if g.get("team") == "away")
            tot = gh + ga

            # ── Estatísticas base ─────────────────────────────────────────
            stats_raw = d.get("statistics", [])
            stat_vals = _extract_stats(stats_raw) if stats_raw else {}

            # ── xG como métricas extras ───────────────────────────────────
            xg = d.get("xg", {})
            if xg and xg.get("home") is not None:
                stat_vals["xg"] = (float(xg["home"]), float(xg["away"]))

            # ── Pressure summary como métricas extras ─────────────────────
            ps = d.get("pressure_summary", {})
            if ps:
                if ps.get("home_dominance_pct") is not None:
                    stat_vals["pressure_home_dom_pct"] = (float(ps["home_dominance_pct"]), 100.0 - float(ps["home_dominance_pct"]))
                if ps.get("overall_avg") is not None:
                    stat_vals["pressure_overall_avg"] = (float(ps["overall_avg"]), None)
                if ps.get("momentum_swings") is not None:
                    stat_vals["pressure_momentum_swings"] = (float(ps["momentum_swings"]), None)

            if not stat_vals:
                continue

            active = set()
            if gh > ga:   active.add("casaV")
            elif ga > gh: active.add("visV")
            else:         active.add("emp")
            active.add("o25" if tot > 2 else "u25")
            active.add("btts" if gh >= 1 and ga >= 1 else "nbtts")

            for oc in active:
                for key, (hv, av) in stat_vals.items():
                    if hv is not None:
                        buckets[oc].setdefault(key + "_h", []).append(hv)
                    if av is not None:
                        buckets[oc].setdefault(key + "_a", []).append(av)
            total += 1
        except Exception:
            continue

    # Calcula médias e contagens
    outcomes_result = {}
    for oc, stats in buckets.items():
        if not stats:
            outcomes_result[oc] = {"n": 0}
            continue
        entry = {"n": 0}
        for k, vals in stats.items():
            if vals:
                entry[k] = round(sum(vals) / len(vals), 1)
                entry["n"] = max(entry["n"], len(vals))
        outcomes_result[oc] = entry

    # Detecta quais chaves de stats existem em pelo menos 1 outcome
    all_keys = set()
    for oc_data in outcomes_result.values():
        all_keys.update(k for k in oc_data if k not in ("n",))
    stat_keys_found = sorted(all_keys)

    result = {
        "total":        total,
        "outcomes":     outcomes_result,
        "stat_keys":    stat_keys_found,
        "stat_labels":  _STAT_LABELS,
        "computed_at":  datetime.now().isoformat(),
    }
    _stats_patterns_cache = {"ts": time.time(), "data": result}
    return jsonify(result)


# ── Correlação de Odds de Abertura ──────────────────────────────────────────

_ODDS_BUCKETS_H = [
    ("1.01–1.30", 1.01, 1.30),
    ("1.31–1.60", 1.31, 1.60),
    ("1.61–2.00", 1.61, 2.00),
    ("2.01–3.00", 2.01, 3.00),
    ("3.01–5.00", 3.01, 5.00),
    (">5.00",     5.01, 99.0),
]

_odds_patterns_cache = {"ts": 0, "data": None}

@app.route("/api/momentum/odds-patterns")
def api_odds_patterns():
    """Correlaciona faixas de odd de abertura (Casa 1X2) com resultados reais.
    Retorna edge vs. probabilidade implícita por faixa."""
    global _odds_patterns_cache
    if (time.time() - _odds_patterns_cache["ts"] < 1800
            and _odds_patterns_cache["data"] is not None):
        return jsonify(_odds_patterns_cache["data"])

    # {label: {n, casaV, emp, visV, o25, u25, sum_imp_h, sum_imp_x, sum_imp_a}}
    bkts = {
        lbl: {"n": 0, "casaV": 0, "emp": 0, "visV": 0,
              "o25": 0, "u25": 0,
              "sum_imp_h": 0.0, "sum_imp_x": 0.0, "sum_imp_a": 0.0}
        for lbl, _, _ in _ODDS_BUCKETS_H
    }
    total = 0

    for fpath in glob.glob(os.path.join(MOMENTUM_DIR, "*.json")):
        try:
            with open(fpath, encoding="utf-8") as f:
                d = json.load(f)
            oo = d.get("opening_odds", {})
            if not oo or not oo.get("h"):
                continue
            try:
                h_odd = float(oo["h"])
                x_odd = float(oo.get("x") or 0)
                a_odd = float(oo.get("a") or 0)
            except (ValueError, TypeError):
                continue
            if h_odd <= 0:
                continue

            goals = d.get("goals", [])
            gc = sum(1 for g in goals if g.get("team") == "home")
            gf = sum(1 for g in goals if g.get("team") == "away")
            outcome = "casaV" if gc > gf else "visV" if gf > gc else "emp"
            is_o25  = 1 if gc + gf > 2 else 0

            for lbl, lo, hi in _ODDS_BUCKETS_H:
                if lo <= h_odd <= hi:
                    b = bkts[lbl]
                    b["n"]       += 1
                    b[outcome]   += 1
                    b["o25"]     += is_o25
                    b["u25"]     += 1 - is_o25
                    b["sum_imp_h"] += (1 / h_odd * 100) if h_odd > 0 else 0
                    b["sum_imp_x"] += (1 / x_odd * 100) if x_odd > 0 else 0
                    b["sum_imp_a"] += (1 / a_odd * 100) if a_odd > 0 else 0
                    break
            total += 1
        except Exception:
            continue

    buckets_out = []
    for lbl, _, _ in _ODDS_BUCKETS_H:
        b = bkts[lbl]
        n = b["n"]
        if n == 0:
            continue
        imp_h = round(b["sum_imp_h"] / n, 1)
        imp_x = round(b["sum_imp_x"] / n, 1)
        imp_a = round(b["sum_imp_a"] / n, 1)
        buckets_out.append({
            "label":     lbl,
            "n":         n,
            "casaV":     b["casaV"],
            "emp":       b["emp"],
            "visV":      b["visV"],
            "casaV_pct": round(b["casaV"] / n * 100),
            "emp_pct":   round(b["emp"]   / n * 100),
            "visV_pct":  round(b["visV"]  / n * 100),
            "o25_pct":   round(b["o25"]   / n * 100),
            "u25_pct":   round(b["u25"]   / n * 100),
            "imp_h":     imp_h,
            "imp_x":     imp_x,
            "imp_a":     imp_a,
            "edge_h":    round(b["casaV"] / n * 100 - imp_h, 1),
            "edge_x":    round(b["emp"]   / n * 100 - imp_x, 1),
            "edge_a":    round(b["visV"]  / n * 100 - imp_a, 1),
        })

    result = {"total": total, "buckets": buckets_out, "computed_at": datetime.now().isoformat()}
    _odds_patterns_cache = {"ts": time.time(), "data": result}
    return jsonify(result)


# ── Indicadores ao vivo ──────────────────────────────────────────────────────

def _profile_similarity(current: dict, profile: dict) -> float | None:
    """Calcula similaridade 0-100 entre stats atuais e perfil histórico de um resultado.
    current = {key: (home_val, away_val)}
    profile = {key_h: avg, key_a: avg}  (formato de outcomes_result)
    """
    diffs = []
    for key, vals in current.items():
        hv, av = vals if isinstance(vals, tuple) else (vals, None)
        h_hist = profile.get(key + "_h")
        a_hist = profile.get(key + "_a")
        if h_hist is not None and hv is not None and h_hist > 0:
            diffs.append(abs(float(hv) - float(h_hist)) / float(h_hist))
        if a_hist is not None and av is not None and a_hist > 0:
            diffs.append(abs(float(av) - float(a_hist)) / float(a_hist))
    if not diffs:
        return None
    avg_diff = sum(diffs) / len(diffs)
    return round(1 / (1 + avg_diff) * 100, 1)


def _momentum_contrib(outcome: str, ps: dict) -> float:
    """Contribuição do momentum para um resultado (0-100)."""
    dom = ps.get("home_dominance_pct", 50.0)
    if outcome == "casaV":
        return round(min(dom, 100), 1)
    if outcome == "visV":
        return round(min(100 - dom, 100), 1)
    if outcome == "emp":
        return round(max(0, 100 - abs(dom - 50) * 2), 1)
    if outcome == "o25":
        swings = ps.get("momentum_swings", 0)
        return round(min(50 + swings * 2, 100), 1)
    if outcome == "u25":
        swings = ps.get("momentum_swings", 0)
        return round(max(0, 50 - swings * 2), 1)
    if outcome == "btts":
        return round(min(50 + abs(ps.get("overall_avg", 0)) * 30, 100), 1)
    return 50.0


def _xg_contrib(outcome: str, xg: dict) -> float | None:
    """Contribuição do xG para um resultado (0-100)."""
    if not xg or xg.get("home") is None:
        return None
    h, a = float(xg.get("home", 0) or 0), float(xg.get("away", 0) or 0)
    diff = h - a
    # sigmoid suavizada: 0 diff → 50, +2 diff → ~85, -2 diff → ~15
    import math
    sig = lambda x: 1 / (1 + math.exp(-x * 0.8))
    if outcome == "casaV":
        return round(sig(diff) * 100, 1)
    if outcome == "visV":
        return round(sig(-diff) * 100, 1)
    if outcome == "emp":
        return round((1 - abs(diff) / max(h + a + 0.01, 1)) * 100, 1)
    if outcome == "o25":
        total = h + a
        return round(min(total / 3 * 100, 100), 1)
    if outcome == "u25":
        total = h + a
        return round(max(0, (1 - total / 3) * 100), 1)
    if outcome == "btts":
        return round(min(h, 1) * min(a, 1) * 100, 1)
    return 50.0


@app.route("/api/momentum/indicators/<event_id>")
def api_indicators(event_id):
    """Indicadores dinâmicos ao vivo: perfil histórico + momentum + xG + value de odds."""
    from flask import request as flask_req
    casa = flask_req.args.get("casa", "")
    fora = flask_req.args.get("fora", "")

    # ── Dados atuais da partida ───────────────────────────────────────────
    with _momentum_lock:
        cached = _momentum_cache.get(event_id)
    mdata = (cached or {}).get("data") or {}

    if not mdata:
        # Tenta buscar se não estiver em cache
        mdata = _process_momentum(event_id, casa, fora) or {}

    stats_raw   = mdata.get("statistics", {})
    current     = _extract_stats(stats_raw) if stats_raw else {}
    ps          = mdata.get("pressure_summary", {})
    xg          = mdata.get("xg", {})
    open_odds   = mdata.get("opening_odds", {})
    goals       = mdata.get("goals", [])
    score_h     = sum(1 for g in goals if g.get("team") == "home")
    score_a     = sum(1 for g in goals if g.get("team") == "away")

    # Adiciona momentum e xG como pseudo-stats para o perfil
    if ps.get("home_dominance_pct") is not None:
        current["pressure_home_dom_pct"] = (ps["home_dominance_pct"], 100 - ps["home_dominance_pct"])
    if xg.get("home") is not None:
        current["xg"] = (float(xg["home"]), float(xg["away"]))

    # ── Padrões históricos ────────────────────────────────────────────────
    global _stats_patterns_cache, _odds_patterns_cache
    if _stats_patterns_cache["data"] is None:
        with app.test_request_context():
            api_stats_patterns()
    hist_data     = _stats_patterns_cache.get("data") or {}
    outcomes_hist = hist_data.get("outcomes", {})
    total_hist    = hist_data.get("total", 0)

    if _odds_patterns_cache["data"] is None:
        with app.test_request_context():
            api_odds_patterns()
    odds_data    = _odds_patterns_cache.get("data") or {}
    odds_buckets = odds_data.get("buckets", [])

    # ── Value Score (edge de odds) ────────────────────────────────────────
    value = {}
    if open_odds and open_odds.get("h"):
        try:
            h_odd = float(open_odds["h"])
            for b in odds_buckets:
                lbl = b["label"]
                # converte label "1.20-1.40" para faixa numérica
                parts = lbl.replace(">","").split("-")
                lo = float(parts[0])
                hi = float(parts[1]) if len(parts) > 1 else 99.0
                if lo <= h_odd <= hi:
                    value = {
                        "label":    lbl,
                        "n":        b["n"],
                        "edge_h":   b.get("edge_h", 0),
                        "edge_x":   b.get("edge_x", 0),
                        "edge_a":   b.get("edge_a", 0),
                        "casaV_pct": b.get("casaV_pct", 0),
                        "emp_pct":   b.get("emp_pct", 0),
                        "visV_pct":  b.get("visV_pct", 0),
                        "o25_pct":   b.get("o25_pct", 0),
                    }
                    break
        except Exception:
            pass

    # ── Peso dinâmico por volume de dados ────────────────────────────────
    # Com poucos dados, momentum/xG pesam mais; com muitos, perfil pesa mais
    w_profile  = min(0.6, max(0.15, total_hist / 200))   # 0.15 → 0.60
    w_momentum = 0.25
    w_xg       = max(0.15, 0.40 - w_profile * 0.4)
    w_value    = 1 - w_profile - w_momentum - w_xg
    w_value    = max(0, min(w_value, 0.20))

    # ── Calcula indicador composto por resultado ──────────────────────────
    OUTCOMES = {
        "casaV": "Casa Vence",
        "emp":   "Empate",
        "visV":  "Visitante",
        "o25":   "Over 2.5",
        "u25":   "Under 2.5",
        "btts":  "Ambos Marcam",
        "nbtts": "Não BTTS",
    }

    indicators = {}
    for oc, label in OUTCOMES.items():
        oc_hist = outcomes_hist.get(oc, {})
        n       = oc_hist.get("n", 0)

        # 1. Perfil
        profile_sim = _profile_similarity(current, oc_hist) if (n >= 3 and current) else None

        # 2. Momentum
        mom_score = _momentum_contrib(oc, ps) if ps else None

        # 3. xG
        xg_score = _xg_contrib(oc, xg) if xg else None

        # 4. Value edge normalizado 0-100 (edge de -20% a +20%)
        val_score = None
        if value:
            edge = value.get(f"edge_{oc[:4]}", value.get("edge_h" if oc == "casaV" else
                             "edge_x" if oc == "emp" else
                             "edge_a" if oc == "visV" else None))
            if edge is not None:
                val_score = round(min(100, max(0, 50 + edge * 2.5)), 1)

        # Composição ponderada com os componentes disponíveis
        components = []
        weights_used = []
        if profile_sim is not None:
            components.append(profile_sim * w_profile)
            weights_used.append(w_profile)
        if mom_score is not None:
            components.append(mom_score * w_momentum)
            weights_used.append(w_momentum)
        if xg_score is not None:
            components.append(xg_score * w_xg)
            weights_used.append(w_xg)
        if val_score is not None and value.get("n", 0) >= 5:
            components.append(val_score * w_value)
            weights_used.append(w_value)

        if components:
            w_sum   = sum(weights_used)
            composite = round(sum(components) / w_sum, 1)
        else:
            composite = None

        indicators[oc] = {
            "label":       label,
            "composite":   composite,
            "profile":     profile_sim,
            "momentum":    mom_score,
            "xg":          xg_score,
            "value_edge":  value.get("edge_h" if oc=="casaV" else "edge_x" if oc=="emp" else "edge_a", None) if value else None,
            "n_hist":      n,
        }

    return jsonify({
        "ok":            True,
        "event_id":      event_id,
        "total_hist":    total_hist,
        "weights":       {"profile": round(w_profile,2), "momentum": round(w_momentum,2),
                          "xg": round(w_xg,2), "value": round(w_value,2)},
        "score":         {"home": score_h, "away": score_a},
        "indicators":    indicators,
        "pressure":      ps,
        "xg":            xg,
        "value":         value,
        "has_live_data": bool(current),
    })


# ── Pattern Tips — TIPs dinâmicos por sequência de sinais ───────────────────

_pattern_tips_cache = {"ts": 0, "data": None, "n_files": 0}
_PTIPS_TTL = 3600   # 1 hora (mas invalida se chegar novo arquivo)

@app.route("/api/momentum/pattern-tips")
def api_momentum_pattern_tips():
    """Analisa sequências de sinais históricos (C/V/G/N) e correlaciona com
    resultados finais de todos os mercados definidos.
    Re-aprende automaticamente sempre que um novo arquivo é salvo."""
    global _pattern_tips_cache
    current_n = len(glob.glob(os.path.join(MOMENTUM_DIR, "*.json")))
    cache_valid = (
        _pattern_tips_cache["data"] is not None
        and time.time() - _pattern_tips_cache["ts"] < _PTIPS_TTL
        and _pattern_tips_cache["n_files"] == current_n   # re-aprende se novo arquivo
    )
    if cache_valid:
        return jsonify(_pattern_tips_cache["data"])

    T = 8    # threshold fixo (equivalente ao global padrão)
    W = 8    # janela de pontos

    def _feats(w):
        if not w:
            return 0.0
        tail  = sum(w[-4:]) / max(len(w), 1)
        trend = (w[-1] - w[0]) if len(w) > 1 else 0.0
        peak  = max(w, key=abs)
        return tail + 0.3 * trend + 0.2 * peak

    def _sig(score):
        if score > T:              return 'C'   # Casa domina
        if score < -T:             return 'V'   # Visitante domina
        if abs(score) > T * 0.6:   return 'G'   # Possível gol
        return 'N'                              # Sem padrão

    # patterns_data[pat][mkt] = [total, count]
    patterns_data = {}
    files         = glob.glob(os.path.join(MOMENTUM_DIR, "*.json"))
    total_files   = 0

    for fpath in files:
        try:
            with open(fpath, encoding="utf-8") as f:
                d = json.load(f)
            pts   = sorted(d.get("graphPoints", []),
                           key=lambda p: float(p.get("minute", 0)))
            goals = d.get("goals", [])
            if len(pts) < W + 2:
                continue

            # ── Extrai sequência de sinais (transições) ───────────────────
            sig_seq = []
            last_sig = None
            for i in range(W, len(pts)):
                win   = [float(pts[j].get("value", 0)) for j in range(i - W, i)]
                score = _feats(win)
                sig   = _sig(score)
                if sig != last_sig:
                    sig_seq.append(sig)
                    last_sig = sig

            if len(sig_seq) < 2:
                continue

            # ── Calcula resultados do jogo ────────────────────────────────
            gh  = sum(1 for g in goals if g.get("team") == "home")
            ga  = sum(1 for g in goals if g.get("team") == "away")
            gh_ht = sum(1 for g in goals
                        if g.get("team") == "home"
                        and float(g.get("minute", 0)) <= 45)
            ga_ht = sum(1 for g in goals
                        if g.get("team") == "away"
                        and float(g.get("minute", 0)) <= 45)
            tot    = gh + ga
            tot_ht = gh_ht + ga_ht

            outcomes = {
                "casaV":  int(gh > ga),
                "visV":   int(ga > gh),
                "emp":    int(gh == ga),
                "o25":    int(tot > 2),
                "u25":    int(tot <= 2),
                "o15":    int(tot > 1),
                "u15":    int(tot <= 1),
                "btts":   int(gh >= 1 and ga >= 1),
                "nbtts":  int(not (gh >= 1 and ga >= 1)),
                "o05ht":  int(tot_ht >= 1),
                "u05ht":  int(tot_ht == 0),
                "u15ht":  int(tot_ht <= 1),
                "1x":     int(gh >= ga),
                "x2":     int(ga >= gh),
            }
            # Placares corretos (capped at 3)
            for h in range(4):
                for a in range(4):
                    outcomes[f"cs_{h}-{a}"] = int(gh == h and ga == a)

            # ── Gera sub-padrões de comprimento 2, 3, 4 ──────────────────
            for length in (2, 3, 4):
                for i in range(len(sig_seq) - length + 1):
                    pat = "".join(sig_seq[i : i + length])
                    pd  = patterns_data.setdefault(pat, {})
                    for mkt, val in outcomes.items():
                        rec = pd.setdefault(mkt, [0, 0])
                        rec[0] += 1
                        rec[1] += val

            total_files += 1
        except Exception:
            continue

    # ── Filtra padrões com n ≥ 15 amostras ───────────────────────────────
    MIN_N = 15
    result_patterns = {}
    for pat, mkts in patterns_data.items():
        pat_res = {}
        for mkt, (total, count) in mkts.items():
            if total >= MIN_N:
                pat_res[mkt] = {"rate": round(count / total * 100, 1), "n": total}
        if pat_res:
            result_patterns[pat] = pat_res

    result = {
        "patterns":     result_patterns,
        "total_files":  total_files,
        "n_files":      current_n,
        "computed_at":  datetime.now().isoformat(),
    }
    _pattern_tips_cache = {"ts": time.time(), "data": result, "n_files": current_n}
    return jsonify(result)


# ── Uniscore / unik8s ────────────────────────────────────────────────────────

UNISCORE_HEADERS = {
    "Origin":       "https://uniscore.com",
    "Referer":      "https://uniscore.com/pt-BR/",
    "User-Agent":   "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept":       "application/json, text/plain, */*",
    "Accept-Language": "pt-BR,pt;q=0.9",
    "Content-Type": "application/json",
}
UNISCORE_BASE = "https://api.unik8s.com/api/v2"

# Cache da lista de partidas do dia (30s TTL)
_uni_events_cache = {"ts": 0, "data": []}

# Cache de odds do dia (60s TTL)
_uni_odds_cache = {"ts": 0, "data": {}}

def _uni_odds_today():
    """Retorna dict {eventId: {h, x, a, ah_line, ah_h, ah_a, ou_line, ou_over, ou_under}} (cache 60s)."""
    global _uni_odds_cache
    if time.time() - _uni_odds_cache["ts"] < 60 and _uni_odds_cache["data"]:
        return _uni_odds_cache["data"]
    try:
        date = datetime.now().strftime("%Y-%m-%d")
        r = http_req.get(
            f"{UNISCORE_BASE}/sport/football/odds/8/{date}/offset/180",
            headers=UNISCORE_HEADERS, timeout=8
        )
        raw = r.json().get("data", {}).get("odds", "")
        odds_map = {}
        for entry in raw.split("!"):
            parts = entry.split("^")
            if len(parts) < 5:
                continue
            eid = parts[0]
            def _parse(seg):
                v = seg.split(":")
                return v if len(v) >= 3 else ["", "", ""]
            ah  = _parse(parts[2])   # Asian Handicap: line, home, away
            fx2 = _parse(parts[3])   # 1X2: home, draw, away
            ou  = _parse(parts[4])   # Over/Under em HK odds → converte para decimal (+1)
            def hk2dec(v):
                try:
                    f = float(v)
                    return str(round(f + 1, 2)) if f < 1.0 else v
                except: return v
            if fx2[0]:
                odds_map[eid] = {
                    "h": fx2[0], "x": fx2[1], "a": fx2[2],
                    "ah_line": ah[0], "ah_h": ah[1], "ah_a": ah[2],
                    "ou_line":  ou[0],
                    "ou_over":  hk2dec(ou[1]),
                    "ou_under": hk2dec(ou[2]),
                }
        _uni_odds_cache = {"ts": time.time(), "data": odds_map}
        return odds_map
    except Exception:
        return {}

def _uni_events_today():
    """Retorna lista de eventos de futebol do dia — TODOS os locales + paginação,
    combinando os jogos AO VIVO (live-v2) com os AINDA NÃO COMEÇADOS (scheduled-events).
    O endpoint de scheduled-events sozinho não inclui partidas já em andamento, por isso
    a combinação — senão a maioria dos jogos ao vivo não é encontrada pra enriquecimento.
    Todas as chamadas (locale × fonte) rodam em PARALELO — sequencial chegava a travar
    dezenas de segundos numa única busca. Cache de 30s."""
    from concurrent.futures import ThreadPoolExecutor
    global _uni_events_cache
    if time.time() - _uni_events_cache["ts"] < 30 and _uni_events_cache["data"]:
        return _uni_events_cache["data"]
    today = datetime.now().strftime("%Y-%m-%d")
    all_by_id = {}
    lock = threading.Lock()

    def _paginate_locale(url_fmt, locale):
        page = 1
        while True:
            try:
                r = http_req.post(url_fmt.format(locale=locale), json={"page": page},
                                   headers=UNISCORE_HEADERS, timeout=6)
                data   = r.json().get("data", {})
                events = data.get("events", [])
                pag    = data.get("pagination", {})
                with lock:
                    for ev in events:
                        eid = ev.get("id")
                        if eid and eid not in all_by_id:
                            all_by_id[eid] = ev
                if not pag.get("hasNextPage"):
                    break
                page += 1
                if page > 4:   # safety cap (mais baixo — já roda em paralelo por locale)
                    break
            except Exception:
                break

    urls = [
        f"{UNISCORE_BASE}/sport/football/events/live-v2/locale/{{locale}}",
        f"{UNISCORE_BASE}/sport/football/scheduled-events-pagination-v2/{today}/locale/{{locale}}/type/all?language=pt-BR",
    ]
    tasks = [(url_fmt, locale) for url_fmt in urls for locale in _UNISCORE_LOCALES]
    with ThreadPoolExecutor(max_workers=len(tasks)) as ex:
        futures = [ex.submit(_paginate_locale, url_fmt, locale) for url_fmt, locale in tasks]
        for fut in futures:
            try:
                fut.result(timeout=15)
            except Exception:
                pass

    events = list(all_by_id.values())
    _uni_events_cache = {"ts": time.time(), "data": events}
    print(f"[uniscore] {len(events)} eventos de hoje (ao vivo + agendados, todos os locales, em paralelo)")
    return events


# Seleções nacionais mudam MUITO de nome entre idiomas (ex: "Brasil"/"Brazil",
# "Alemanha"/"Germany") — ao contrário de clubes, que costumam ser parecidos. O Uniscore
# às vezes devolve o nome em inglês pra um evento e em português pra outro no mesmo
# request (mistura de locale), então sem esse mapa a busca por seleção falha silenciosamente.
_COUNTRY_ALIASES = {
    "brasil": "brazil", "alemanha": "germany", "espanha": "spain", "franca": "france",
    "inglaterra": "england", "italia": "italy", "holanda": "netherlands",
    "paises baixos": "netherlands", "belgica": "belgium", "suica": "switzerland",
    "suecia": "sweden", "noruega": "norway", "dinamarca": "denmark", "polonia": "poland",
    "austria": "austria", "escocia": "scotland", "irlanda": "ireland",
    "irlanda do norte": "northern ireland", "pais de gales": "wales", "gales": "wales",
    "russia": "russia", "ucrania": "ukraine", "sercia": "serbia", "servia": "serbia",
    "croacia": "croatia", "romenia": "romania", "grecia": "greece", "turquia": "turkey",
    "portugal": "portugal", "mexico": "mexico", "estados unidos": "united states",
    "eua": "united states", "canada": "canada", "argentina": "argentina",
    "uruguai": "uruguay", "paraguai": "paraguay", "chile": "chile", "colombia": "colombia",
    "equador": "ecuador", "peru": "peru", "venezuela": "venezuela", "bolivia": "bolivia",
    "japao": "japan", "coreia do sul": "south korea", "coreia do norte": "north korea",
    "china": "china", "australia": "australia", "arabia saudita": "saudi arabia",
    "ira": "iran", "iraque": "iraq", "egito": "egypt", "marrocos": "morocco",
    "argelia": "algeria", "tunisia": "tunisia", "nigeria": "nigeria", "senegal": "senegal",
    "camaroes": "cameroon", "gana": "ghana", "africa do sul": "south africa",
    "costa do marfim": "ivory coast", "cabo verde": "cape verde", "nova zelandia": "new zealand",
}


# ── Memória de apelidos Uniscore — toda vez que a busca fuzzy abaixo acha um evento com
# confiança boa, grava aqui o nome EXATO que o Uniscore usa pra aquele confronto
# (StatArea "Brasil"/"Noruega" → Uniscore "Brazil"/"Norway"). Da próxima vez que a mesma
# dupla de times aparecer, usamos esse nome direto (sem depender do dicionário de países
# nem da pontuação fuzzy), então o sistema "aprende" qualquer confronto que já resolveu,
# não só seleções que estão no dicionário manual. Persiste em disco entre reinícios. ──
_UNI_NAME_ALIASES_PATH = os.path.join(DATA_DIR, "uni_name_aliases.json")
_uni_name_aliases_cache = None

def _load_uni_name_aliases():
    global _uni_name_aliases_cache
    if _uni_name_aliases_cache is not None:
        return _uni_name_aliases_cache
    try:
        with open(_UNI_NAME_ALIASES_PATH, encoding="utf-8") as f:
            _uni_name_aliases_cache = json.load(f)
    except Exception:
        _uni_name_aliases_cache = {}
    return _uni_name_aliases_cache

def _save_uni_name_alias(casa, fora, uni_home, uni_away):
    aliases = _load_uni_name_aliases()
    key = f"{casa.strip().lower()}|{fora.strip().lower()}"
    if aliases.get(key) == [uni_home, uni_away]:
        return  # já está salvo, não regrava toda vez
    aliases[key] = [uni_home, uni_away]
    try:
        with open(_UNI_NAME_ALIASES_PATH, "w", encoding="utf-8") as f:
            json.dump(aliases, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def _uni_find(casa, fora):
    """Encontra evento Uniscore pelo nome dos times (fuzzy com remoção de acentos)."""
    import unicodedata

    def norm(s):
        """Normaliza: minúsculo, sem acentos, sem pontuação, com apelidos de seleção traduzidos."""
        s = s.lower()
        s = unicodedata.normalize("NFD", s)
        s = "".join(c for c in s if unicodedata.category(c) != "Mn")  # remove diacritics
        s = re.sub(r"[^a-z0-9 ]", " ", s)
        s = re.sub(r"\s+", " ", s).strip()
        return _COUNTRY_ALIASES.get(s, s)

    events = _uni_events_today()

    # 1. Já resolvemos essa dupla antes? Usa o nome exato do Uniscore que aprendemos,
    # sem precisar de fuzzy nem do dicionário de países.
    aliases = _load_uni_name_aliases()
    key = f"{casa.strip().lower()}|{fora.strip().lower()}"
    known = aliases.get(key)
    if known:
        uni_home, uni_away = known
        for ev in events:
            if ev.get("homeTeam", {}).get("name") == uni_home and ev.get("awayTeam", {}).get("name") == uni_away:
                return ev

    def words(s):
        """Conjunto de palavras significativas (>= 3 letras)."""
        return {w for w in norm(s).split() if len(w) >= 3}

    def score_team(query, candidate):
        """Pontuação de similaridade entre dois nomes de time."""
        qn = norm(query)
        cn = norm(candidate)
        # Prefixo dos primeiros 5 chars
        prefix_match = qn[:5] == cn[:5] and len(qn) >= 4
        # Interseção de palavras
        qw = words(query)
        cw = words(candidate)
        common = qw & cw
        word_score = len(common) / max(len(qw), 1)
        # Substring bidirecional
        substr = (qn[:8] in cn) or (cn[:8] in qn)
        return (3 if prefix_match else 0) + (word_score * 2) + (1 if substr else 0)

    best = None
    best_score = 0.0
    for ev in events:
        hn = ev.get("homeTeam", {}).get("name", "")
        an = ev.get("awayTeam", {}).get("name", "")
        sc = score_team(casa, hn) + score_team(fora, an)
        if sc > best_score:
            best_score = sc
            best = ev
    # Threshold mínimo: pelo menos um nome com score >= 2
    if best_score < 2.0:
        return None

    # 2. Achou com boa confiança — grava o apelido pra próxima vez nem precisar de fuzzy
    if best_score >= 3.0:
        _save_uni_name_alias(casa, fora, best.get("homeTeam", {}).get("name"), best.get("awayTeam", {}).get("name"))

    return best

def _uni_enrich_one(casa, fora):
    """Busca e retorna dados Uniscore para um par casa/fora. Retorna dict."""
    ev = _uni_find(casa, fora)
    if not ev:
        return {"found": False, "casa": casa, "fora": fora}

    eid = ev.get("id")
    home_tid = ev.get("homeTeam", {}).get("id", "")
    away_tid = ev.get("awayTeam", {}).get("id", "")

    # Odds do dia (cache 60s)
    odds = _uni_odds_today().get(eid, {})

    result = {
        "found":    True,
        "id":       eid,
        "home_tid": home_tid,
        "away_tid": away_tid,
        "casa":   ev.get("homeTeam", {}).get("name"),
        "fora":   ev.get("awayTeam", {}).get("name"),
        "status": ev.get("status", {}).get("description"),
        "minuto": ev.get("time", {}).get("current"),
        "placar": {
            "casa": ev.get("homeScore", {}).get("current"),
            "fora": ev.get("awayScore", {}).get("current"),
        },
        "stats": {
            "escanteios_casa": ev.get("homeCornerKicks", 0),
            "escanteios_fora": ev.get("awayCornerKicks", 0),
            "chutes_casa":     ev.get("homeShotOnTarget", 0),
            "chutes_fora":     ev.get("awayShotOnTarget", 0),
            "amarelos_casa":   ev.get("homeYellowCards", 0),
            "amarelos_fora":   ev.get("awayYellowCards", 0),
            "vermelhos_casa":  ev.get("homeRedCards", 0),
            "vermelhos_fora":  ev.get("awayRedCards", 0),
        },
        "odds": odds,  # {h, x, a, ah_line, ah_h, ah_a, ou_line, ou_over, ou_under}
    }

    def _fetch_details():
        try:
            r = http_req.get(f"{UNISCORE_BASE}/football/event/{eid}?language=pt-BR",
                             headers=UNISCORE_HEADERS, timeout=6)
            d = r.json().get("data", {}).get("event", {})
            return {
                "clima":   d.get("environment"),
                "arbitro": d.get("referee", {}).get("name") if isinstance(d.get("referee"), dict) else d.get("referee"),
                "estadio": d.get("venue", {}).get("name") if isinstance(d.get("venue"), dict) else None,
            }
        except Exception:
            return {}

    def _fetch_incidents():
        try:
            r = http_req.get(f"{UNISCORE_BASE}/football/event/{eid}/incidents?language=pt-BR",
                             headers=UNISCORE_HEADERS, timeout=6)
            incs = r.json().get("data", {}).get("incidents", [])
            return {"incidents": [
                {
                    "type":    i.get("incidentType"),
                    "class":   i.get("incidentClass"),
                    "minute":  i.get("time"),
                    "player":  i.get("player", {}).get("name"),
                    "assist":  i.get("assist1", {}).get("name") if i.get("assist1") else None,
                    "isHome":  i.get("isHome"),
                    "score_h": i.get("homeScore"),
                    "score_a": i.get("awayScore"),
                }
                for i in (incs or [])
            ]}
        except Exception:
            return {"incidents": []}

    def _fetch_form():
        try:
            r = http_req.get(f"{UNISCORE_BASE}/football/event/{eid}/recent-form?language=pt-BR",
                             headers=UNISCORE_HEADERS, timeout=6)
            d = r.json().get("data", {})
            def parse_form(matches):
                out = []
                for m in (matches or []):
                    hs = m.get("homeScore", {})
                    as_ = m.get("awayScore", {})
                    out.append({
                        "home":   m.get("homeTeam", {}).get("name"),
                        "away":   m.get("awayTeam", {}).get("name"),
                        "score":  f"{hs.get('current',0)}-{as_.get('current',0)}",
                        "status": m.get("status", {}).get("type"),
                    })
                return out
            return {
                "forma_casa": parse_form(d.get("home", {}).get("latest_matches", [])),
                "forma_fora": parse_form(d.get("away", {}).get("latest_matches", [])),
            }
        except Exception:
            return {"forma_casa": [], "forma_fora": []}

    def _fetch_top_players():
        try:
            r = http_req.get(f"{UNISCORE_BASE}/sport/football/events/{eid}/top-players?language=pt-BR",
                             headers=UNISCORE_HEADERS, timeout=6)
            d = r.json().get("data", {})
            def parse_player(p):
                if not p: return None
                return {
                    "name":   p.get("name"),
                    "pos":    p.get("position"),
                    "rating": p.get("rating"),
                    "attrs":  p.get("attributes", {}),
                }
            return {
                "top_casa": parse_player(d.get("home_player")),
                "top_fora": parse_player(d.get("away_player")),
            }
        except Exception:
            return {"top_casa": None, "top_fora": None}

    # Estatísticas detalhadas (chutes, posse, passes, duelos...)
    # Mapeamento: nome original UniScore → chave snake_case usada no frontend
    _UNI_STAT_NORM = {
        "Ball Possession":          "ball_possession",
        "Total Shots":              "shots",
        "Shots":                    "shots",
        "Shots on Target":          "shots_on_target",
        "Blocked Shots":            "blocked_shots",
        "Shots Inside Box":         "shots_inside_box",
        "Shots Outside Box":        "shots_outside_box",
        "Touches In Box":           "touches_in_box",
        "Touches in Box":           "touches_in_box",
        "Big Chances":              "big_chances",
        "Big Chances Created":      "big_chances",
        "Corner Kicks":             "corner_kicks",
        "Corners":                  "corner_kicks",
        "Free Kicks":               "freekicks",
        "Passes":                   "passes",
        "Total Passes":             "passes",
        "Accurate Passes":          "passes",
        "Passes Accurate":          "passes",
        "Passes in Final Third":    "pass_in_final_third",
        "Pass in Final Third":      "pass_in_final_third",
        "Final Third Entries":      "final_third_entries",
        "Long Balls":               "long_balls",
        "Crosses":                  "crosses_accuracy",
        "Crosses Accurate":         "crosses_accuracy",
        "Duels":                    "duels",
        "Ground Duels":             "ground_duels",
        "Aerial Duels":             "aerial_duels",
        "Dribbles":                 "dribble",
        "Successful Dribbles":      "dribble",
        "Tackles":                  "tackles",
        "Tackles Won":              "tackles",
        "Interceptions":            "interceptions",
        "Recoveries":               "recoveries",
        "Clearances":               "clearances",
        "Saves":                    "saves",
        "Goalkeeper Saves":         "saves",
        "Goal Kicks":               "goal_kicks",
        "Yellow Cards":             "yellow_cards",
        "Fouls":                    "fouls",
        "Fouls Committed":          "fouls",
        "Was Fouled":               "was_fouled",
        "Possession Losses":        "poss_losts",
        "Total Possession Losses":  "poss_losts",
        "Expected Goals":           "expected_goals",
    }

    def _build_stat_map(period_data):
        """Constrói dict com chave original + alias snake_case."""
        stat_map = {}
        for grp in (period_data or {}).get("groups", []):
            for item in grp.get("statisticsItems", []):
                if not isinstance(item, dict):
                    continue
                name = item.get("name") or item.get("fields") or ""
                if not name:
                    continue
                stat_map[name] = item                        # chave original
                snake = _UNI_STAT_NORM.get(name)
                if snake and snake not in stat_map:
                    stat_map[snake] = item                   # alias snake_case
        return stat_map

    def _fetch_statistics():
        try:
            url_stats = f"{UNISCORE_BASE}/football/event/{eid}/home/{home_tid}/away/{away_tid}/statistics"
            r = http_req.get(url_stats, headers=UNISCORE_HEADERS, timeout=6)
            periods = r.json().get("data", {}).get("statistics", [])
            stat_periods_out = {}
            for pd in periods:
                period_key = pd.get("period", "ALL")
                stat_periods_out[period_key] = _build_stat_map(pd)
            all_pd = next((p for p in periods if p.get("period") == "ALL"), periods[0] if periods else None)
            return {
                "statistics":         _build_stat_map(all_pd) if all_pd else {},
                "statistics_periods": stat_periods_out,
            }
        except Exception:
            return {"statistics": {}, "statistics_periods": {}}

    def _fetch_lineups():
        try:
            r = http_req.get(f"{UNISCORE_BASE}/football/event/{eid}/lineups?language=pt-BR",
                             headers=UNISCORE_HEADERS, timeout=6)
            ld = r.json().get("data", {})
            def parse_side(side):
                sd = ld.get(side, {})
                def pp(p):
                    pl = p.get("player", {})
                    return {
                        "name":    pl.get("fullName") or pl.get("name"),
                        "number":  p.get("shirtNumber"),
                        "pos":     p.get("position"),
                        "captain": p.get("captain", False),
                        "rating":  p.get("rating"),
                        "order":   p.get("counterOrder"),
                    }
                players = sd.get("players", [])
                titulares = [pp(p) for p in players if not p.get("substitute", False)]
                titulares.sort(key=lambda p: p.get("order") or 99)
                return {
                    "formation":  sd.get("formation", ""),
                    "confirmed":  ld.get("confirmed", False),
                    "titulares":  titulares,
                    "reservas":   [pp(p) for p in players if p.get("substitute", False)],
                }
            return {"lineup_casa": parse_side("home"), "lineup_fora": parse_side("away")}
        except Exception:
            return {"lineup_casa": None, "lineup_fora": None}

    def _fetch_graph():
        try:
            r = http_req.get(f"{UNISCORE_BASE}/football/event/{eid}/graph",
                             headers=UNISCORE_HEADERS, timeout=6)
            pts = r.json().get("data", {}).get("graphPoints", [])
            return {"graph": [{"m": p.get("minute"), "v": p.get("value")} for p in pts]}
        except Exception:
            return {"graph": []}

    # Todas as chamadas acima são independentes — rodam em paralelo em vez de uma
    # atrás da outra, senão uma única partida podia levar 8-12s pra carregar.
    from concurrent.futures import ThreadPoolExecutor as _TPE
    fetchers = [_fetch_details, _fetch_incidents, _fetch_form, _fetch_top_players,
                _fetch_statistics, _fetch_lineups, _fetch_graph]
    with _TPE(max_workers=len(fetchers)) as ex:
        for fut in [ex.submit(fn) for fn in fetchers]:
            try:
                result.update(fut.result(timeout=10))
            except Exception:
                pass

    # Odds ao vivo (mesma fonte das odds de abertura — o feed do dia já reflete
    # a movimentação de mercado durante o jogo, então serve pra odds ao vivo também)
    try:
        odds_today = _uni_odds_today()
        od = odds_today.get(eid)
        if od and od.get("h"):
            result["live_odds"] = {
                "h": od["h"], "x": od["x"], "a": od["a"],
                "ou_line": od.get("ou_line", ""),
                "ou_over": od.get("ou_over", ""),
                "ou_under": od.get("ou_under", ""),
                "changed": True,
            }
    except Exception:
        pass

    return result


# ── Aba "CD" (Sequências do time / confronto direto) + "Dados" do Uniscore ────
# Descoberto interceptando as chamadas reais do site uniscore.com (devtools):
# team-streaks = exatamente o card "Sequências do time" / "Sequências de
# confrontos diretos". "Dados" (Mais De 2.5, BTTS, Gols/Jogo...) não tem
# endpoint próprio — o site calcula na hora a partir do /recent-form de cada
# time, então aqui replicamos o mesmo cálculo (mesmo estilo já usado na aba
# Leitura do Painel Principal).
_UNI_STREAK_NAMED_LABELS = {
    "first_to_score": "Primeiro a marcar",
    "both_team_scoring": "Ambas as equipes marcando",
    "without_clean_sheet": "Sem goleiro sem sofrer gol",
    "no_wins": "Sem vitórias",
    "no_losses": "Sem derrotas",
    "no_draws": "Sem empates",
    "clean_sheet": "Sem sofrer gol",
    "failed_to_score": "Não marcou",
}
_UNI_STREAK_STAT_LABELS = {
    "more_goals": "Mais de {n} gols",
    "less_goals": "Menos de {n} gols",
    "more_corners": "Mais de {n} escanteios",
    "less_corners": "Menos de {n} escanteios",
    "more_cards": "Mais de {n} cartões",
    "less_cards": "Menos de {n} cartões",
}


def _uni_streak_label(name):
    if name in _UNI_STREAK_NAMED_LABELS:
        return _UNI_STREAK_NAMED_LABELS[name]
    m = re.match(r"(more|less)_(goals|corners|cards)_([\d.]+)", name or "")
    if m:
        kind, stat, n = m.groups()
        base = _UNI_STREAK_STAT_LABELS.get(f"{kind}_{stat}")
        if base:
            return base.format(n=n)
    return (name or "").replace("_", " ").capitalize()


def _uni_fetch_team_streaks(eid, home_tid, away_tid, start_ts):
    try:
        r = http_req.get(
            f"{UNISCORE_BASE}/football/event/{eid}/home/{home_tid}/away/{away_tid}/start-time/{start_ts}/team-streaks?language=pt-BR",
            headers=UNISCORE_HEADERS, timeout=6)
        d = r.json().get("data", {})
        def parse(items):
            out = []
            for it in (items or []):
                out.append({
                    "label": _uni_streak_label(it.get("name")),
                    "team": "Casa" if it.get("team") == "Home" else "Fora",
                    "value": it.get("value"),
                })
            return out
        return {"geral": parse(d.get("general")), "confronto_direto": parse(d.get("head2head"))}
    except Exception:
        return {"geral": [], "confronto_direto": []}


def _uni_fetch_analytics(eid):
    """Aba 'Dados' de verdade — achado interceptando a rede do uniscore.com. Vem
    pronto do próprio Uniscore (fonte deles é season-long, tipo FootyStats), não é
    cálculo nosso — por isso os números batem exatamente com o app deles."""
    try:
        r = http_req.get(f"{UNISCORE_BASE}/football/event/{eid}/analytics?language=pt-BR",
                         headers=UNISCORE_HEADERS, timeout=6)
        d = r.json().get("data", r.json())
        return {
            "home_ppg": d.get("home_ppg"), "away_ppg": d.get("away_ppg"),
            "over25_pct": d.get("o25_potential"), "over15_pct": d.get("o15_potential"),
            "btts_pct": d.get("btts_potential"), "gols_jogo": d.get("avg_potential"),
            "escanteios": d.get("corners_potential"), "cartoes": d.get("cards_potential"),
        }
    except Exception:
        return None


def _uni_cd_dados_one(casa, fora):
    ev = _uni_find(casa, fora)
    if not ev:
        return {"found": False}
    eid = ev.get("id")
    home_tid = ev.get("homeTeam", {}).get("id", "")
    away_tid = ev.get("awayTeam", {}).get("id", "")
    start_ts = ev.get("startTimestamp", 0)

    streaks = _uni_fetch_team_streaks(eid, home_tid, away_tid, start_ts)
    analytics = _uni_fetch_analytics(eid)

    return {
        "found": True, "casa": ev.get("homeTeam", {}).get("name"), "fora": ev.get("awayTeam", {}).get("name"),
        "streaks": streaks, "dados": analytics,
    }


@app.route("/api/uniscore/cd_dados")
def api_uniscore_cd_dados():
    casa = request.args.get("casa", "")
    fora = request.args.get("fora", "")
    if not casa or not fora:
        return jsonify({"error": "casa e fora obrigatórios"}), 400
    return jsonify(_uni_cd_dados_one(casa, fora))


@app.route("/api/uniscore/enrich")
def api_uniscore_enrich():
    """Retorna dados enriquecidos do Uniscore para uma partida ao vivo."""
    casa = request.args.get("casa", "")
    fora = request.args.get("fora", "")
    if not casa or not fora:
        return jsonify({"error": "casa e fora obrigatórios"}), 400
    return jsonify(_uni_enrich_one(casa, fora))


@app.route("/api/uniscore/live-all", methods=["POST"])
def api_uniscore_live_all():
    """Recebe lista [{casa, fora}] e retorna enriquecimento paralelo de todas."""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    body = request.get_json(force=True, silent=True) or {}
    partidas = body.get("partidas", [])
    if not partidas:
        return jsonify([])

    # Pré-aquece cache de eventos (1 chamada para todas)
    _uni_events_today()

    results = [None] * len(partidas)
    with ThreadPoolExecutor(max_workers=16) as ex:
        fut_map = {
            ex.submit(_uni_enrich_one, p.get("casa", ""), p.get("fora", "")): i
            for i, p in enumerate(partidas)
        }
        for fut in as_completed(fut_map):
            idx = fut_map[fut]
            try:
                results[idx] = fut.result()
            except Exception:
                results[idx] = {"found": False,
                                "casa": partidas[idx].get("casa", ""),
                                "fora": partidas[idx].get("fora", "")}

    return jsonify(results)


@app.route("/api/fotmob/live")
def api_fotmob_live():
    """Busca jogos ao vivo diretamente via FotMob /api/matches (sem bloqueio)."""
    try:
        today_str = datetime.now().strftime("%Y%m%d")
        r = http_req.get(
            "https://www.fotmob.com/api/matches",
            headers=FOTMOB_HEADERS,
            params={"date": today_str},
            timeout=12
        )
        r.raise_for_status()
        data = r.json()

        live = []
        for league in data.get("leagues", []):
            for match in league.get("matches", []):
                status   = match.get("status", {})
                started  = status.get("started", False)
                finished = status.get("finished", False)
                if not started or finished:
                    continue
                home = match.get("home", {})
                away = match.get("away", {})
                live.append({
                    "id":        str(match.get("id", "")),
                    "home":      home.get("name", ""),
                    "away":      away.get("name", ""),
                    "homeScore": home.get("score"),
                    "awayScore": away.get("score"),
                    "minute":    status.get("liveTime", {}).get("short", ""),
                    "league":    league.get("name", ""),
                    "country":   league.get("ccode", ""),
                })
        return jsonify({"live": live, "total": len(live)})
    except Exception as e:
        return jsonify({"error": str(e), "live": [], "total": 0}), 503


@app.route("/api/fotmob/match/<match_id>")
def api_fotmob_match_detail(match_id):
    """Busca detalhes completos via Playwright Chrome (bypassa CF — inclui momentum)."""
    from playwright.sync_api import sync_playwright

    result = {}
    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(
                headless=True,
                args=["--no-sandbox", "--disable-dev-shm-usage"]
            )
            ctx  = browser.new_context(locale="pt-BR")
            page = ctx.new_page()

            def on_resp(resp):
                if "matchDetails" in resp.url and not result:
                    try:
                        result["data"] = resp.json()
                    except Exception:
                        pass

            page.on("response", on_resp)
            page.goto(
                f"https://www.fotmob.com/match/{match_id}",
                wait_until="domcontentloaded",
                timeout=25000
            )
            # Aguarda a chamada assíncrona do matchDetails
            for _ in range(30):
                if result:
                    break
                page.wait_for_timeout(300)
            browser.close()

        if not result:
            return jsonify({"error": "matchDetails não capturado"}), 503

        raw = result["data"]

        # ── Header ──
        teams  = raw.get("header", {}).get("teams", [])
        status = raw.get("header", {}).get("status", {})
        home_t = teams[0] if len(teams) > 0 else {}
        away_t = teams[1] if len(teams) > 1 else {}
        gen    = raw.get("general", {})

        # ── Momentum ──
        momentum = raw.get("content", {}).get("matchFacts", {}).get("momentum", {})
        mom_data = momentum.get("main", {}).get("data", [])

        # ── Stats ──
        stats_raw = raw.get("content", {}).get("matchFacts", {}).get("stats", {})
        stats_out = []
        for block in stats_raw.get("stats", []):
            for stat in block.get("stats", []):
                vals = stat.get("stats", [])
                if len(vals) >= 2:
                    stats_out.append({
                        "title": stat.get("title", ""),
                        "home":  str(vals[0]),
                        "away":  str(vals[1]),
                    })

        # ── Escalações ──
        lineup_raw = raw.get("content", {}).get("lineup", {}).get("lineup", [])
        lineup_out = {"home": [], "away": []}
        if lineup_raw:
            block = lineup_raw[0]
            players = block.get("players", [[], []])
            for i, side in enumerate(["home", "away"]):
                for p in players[i] if i < len(players) else []:
                    lineup_out[side].append({
                        "name":    p.get("name", {}).get("lastName") or p.get("name", {}).get("fullName", ""),
                        "shirt":   p.get("shirt", ""),
                        "pos":     p.get("position", ""),
                        "starter": True,
                    })

        # ── Odds ──
        odds_out = {}
        for market in raw.get("content", {}).get("odds", {}).get("parser", []):
            mname = market.get("name", "")
            odds  = market.get("odds", [{}])[0] if market.get("odds") else {}
            if "1x2" in mname.lower() or "match" in mname.lower():
                odds_out = {
                    "home": odds.get("homeOdds") or odds.get("1"),
                    "draw": odds.get("drawOdds") or odds.get("X"),
                    "away": odds.get("awayOdds") or odds.get("2"),
                }
                break

        return jsonify({
            "home":       home_t.get("name", ""),
            "away":       away_t.get("name", ""),
            "homeScore":  home_t.get("score"),
            "awayScore":  away_t.get("score"),
            "minute":     status.get("liveTime", {}).get("short", ""),
            "league":     gen.get("leagueName", ""),
            "momentum":   mom_data,
            "stats":      stats_out,
            "lineup":     lineup_out,
            "odds":       odds_out,
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 503


# ── ODDSPEDIA (Ao Vivo 2 — experimental) ──────────────────────────────────────
ODDSPEDIA_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept-Language": "pt-BR,pt;q=0.9",
}

@app.route("/api/oddspedia/debug/<path:slug>")
def api_oddspedia_debug(slug):
    """DEBUG: abre a página da partida no Oddspedia via Playwright e intercepta TODAS
    as chamadas de rede que parecem trazer odds/dados de partida — usado só pra
    descobrir a URL real da API antes de fazer o scraper de verdade. slug = ex:
    'br/futebol/brasil-noruega-1982539' (sem o domínio, como aparece no link do site)."""
    from playwright.sync_api import sync_playwright

    captured = []
    KEYWORDS = ("odds", "bookmaker", "api", "market", "event", "match")

    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True, args=["--no-sandbox", "--disable-dev-shm-usage"])
            ctx  = browser.new_context(locale="pt-BR", user_agent=ODDSPEDIA_HEADERS["User-Agent"])
            page = ctx.new_page()

            def on_resp(resp):
                url = resp.url
                low = url.lower()
                if "oddspedia.com" not in low:
                    return
                if "/_nuxt/" in low or low.endswith((".js", ".css", ".woff2", ".png", ".svg", ".ico")):
                    return
                if not any(k in low for k in KEYWORDS):
                    return
                entry = {"url": url, "status": resp.status}
                try:
                    ct = resp.headers.get("content-type", "")
                    if "json" in ct:
                        body = resp.json()
                        entry["body_preview"] = json.dumps(body, ensure_ascii=False)[:1500]
                except Exception:
                    pass
                captured.append(entry)

            page.on("response", on_resp)
            page.goto(f"https://oddspedia.com/{slug}", wait_until="domcontentloaded", timeout=25000)
            page.wait_for_timeout(6000)
            browser.close()

        return jsonify({"slug": slug, "total_capturado": len(captured), "chamadas": captured})
    except Exception as e:
        return jsonify({"error": str(e)}), 503


# ── FLASHSCORE/SOCCERWAY (Ao Vivo 2 — experimental) ───────────────────────────
# O Soccerway roda em cima da mesma infraestrutura do Flashscore/Livesport. O feed
# devolve texto num formato próprio (blocos separados por "~", campos "chave÷valor"
# separados por "¬") — sem JSON, mas simples de parsear. Sem Cloudflare, só precisa
# de um Referer válido e um header x-fsign (não parece ser validado com rigor).
FS_BASE = "https://global.flashscore.ninja/2051/x/feed"
FS_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Referer": "https://br.soccerway.com/",
    "x-fsign": "SW9D1eZo",
}

def _fs_get(path):
    r = http_req.get(f"{FS_BASE}/{path}", headers=FS_HEADERS, timeout=10)
    r.raise_for_status()
    return r.text

def _fs_blocks(text):
    return [b for b in text.split("~") if b.strip()]

def _fs_kv(block):
    d = {}
    for part in block.split("¬"):
        if "÷" in part:
            k, _, v = part.partition("÷")
            d[k] = v
    return d

_fs_live_cache = {"ts": 0, "data": []}

def _fs_all_matches():
    """Lista de TODAS as partidas do dia no feed (agendadas + ao vivo + encerradas),
    com liga/país (o feed lista um bloco "ZA" (liga) seguido dos jogos "AA" daquela
    liga, em ordem — então vamos guardando a liga atual enquanto percorremos)."""
    global _fs_live_cache
    if time.time() - _fs_live_cache["ts"] < 30 and _fs_live_cache["data"]:
        return _fs_live_cache["data"]
    try:
        text = _fs_get("f_1_0_-3_pt-br_1")
    except Exception:
        return _fs_live_cache["data"]
    matches = []
    liga_atual, pais_atual = "", ""
    for block in _fs_blocks(text):
        if block.startswith("ZA"):
            d = _fs_kv(block)
            liga_atual = d.get("ZA", "")
            pais_atual = d.get("ZY", "")
            continue
        if not block.startswith("AA"):
            continue
        d = _fs_kv(block)
        if not d.get("AE") or not d.get("AF"):
            continue
        # Escudos: link direto pro CDN de imagens da própria fonte (campos "OA"/"OB" do
        # feed) — não baixa/armazena nada aqui, só monta a URL e o navegador do usuário
        # carrega direto de lá quando renderizar o <img>
        escudo_casa = f"https://www.flashscore.com/res/image/data/{d['OA']}" if d.get("OA") else None
        escudo_fora = f"https://www.flashscore.com/res/image/data/{d['OB']}" if d.get("OB") else None
        matches.append({
            "id":          d.get("AA"),
            "home":        d.get("AE"),
            "away":        d.get("AF"),
            "home_score":  d.get("AG"),
            "away_score":  d.get("AH"),
            "status":      d.get("AB"),   # 1=agendado, 2=ao vivo, 3=encerrado (aproximado)
            "kickoff_ts":  d.get("AD"),
            "liga":        liga_atual,
            "pais":        pais_atual,
            "escudo_casa": escudo_casa,
            "escudo_fora": escudo_fora,
        })
    _fs_live_cache = {"ts": time.time(), "data": matches}
    return matches

# Mantém o nome antigo funcionando (usado pelo _fs_find_match) — mesmo dado, só outro nome
def _fs_live_matches():
    return _fs_all_matches()

def _fs_find_match(casa, fora):
    """Acha o jogo no feed pelo nome dos times (fuzzy, mesma técnica do _uni_find)."""
    import unicodedata

    def norm(s):
        s = (s or "").lower()
        s = unicodedata.normalize("NFD", s)
        s = "".join(c for c in s if unicodedata.category(c) != "Mn")
        s = re.sub(r"[^a-z0-9 ]", " ", s)
        return re.sub(r"\s+", " ", s).strip()

    def side_score(query, candidate):
        s = 0
        if len(query) >= 4 and candidate[:5] == query[:5]: s += 2
        if query[:6] and (query[:6] in candidate or candidate[:6] in query): s += 1
        return s

    nc, nf = norm(casa), norm(fora)
    best, best_score = None, 0
    for m in _fs_live_matches():
        h, a = norm(m["home"]), norm(m["away"])
        # Testa nos dois sentidos (casa/fora direto E invertido) — o mesmo jogo
        # às vezes vem com mandante/visitante trocado entre a fonte do Ao Vivo
        # (NowGoal) e a do Flashscore (usada só aqui pra tabela/H2H), então uma
        # comparação só na ordem "direta" perdia esses jogos.
        sc_direto = side_score(nc, h) and side_score(nf, a) and (side_score(nc, h) + side_score(nf, a))
        sc_invertido = side_score(nc, a) and side_score(nf, h) and (side_score(nc, a) + side_score(nf, h))
        score = max(sc_direto or 0, sc_invertido or 0)
        if score == 0:
            continue
        if score > best_score:
            best_score = score
            best = m
    if best_score < 4:
        return None
    return best

def _fs_match_stats(event_id):
    """Estatísticas da partida (posse, chutes, xG, etc.), agrupadas por seção."""
    text = _fs_get(f"df_st_1_{event_id}")
    sections = []
    current = None
    for block in _fs_blocks(text):
        d = _fs_kv(block)
        if "SF" in d and "SG" not in d:
            current = {"title": d["SF"], "rows": []}
            sections.append(current)
        elif "SG" in d:
            if current is None:
                current = {"title": "Geral", "rows": []}
                sections.append(current)
            current["rows"].append({"label": d.get("SG", ""), "home": d.get("SH", ""), "away": d.get("SI", "")})
    return sections

def _fs_h2h(event_id):
    """Últimos jogos de cada time + confrontos diretos, agrupados por aba (campo KA:
    'Total' / 'Time A - Casa' / 'Time B - Fora') e por seção dentro da aba (campo KB)."""
    text = _fs_get(f"df_hh_1_{event_id}")
    RESULT_LABEL = {"w": "V", "d": "E", "l": "D"}
    tabs = []
    tab = None
    section = None
    for block in _fs_blocks(text):
        d = _fs_kv(block)
        if "KA" in d:
            tab = {"label": d["KA"], "sections": []}
            tabs.append(tab)
            section = None
            continue
        if "KB" in d:
            section = {"title": d["KB"], "rows": []}
            if tab is None:
                tab = {"label": "Total", "sections": []}
                tabs.append(tab)
            tab["sections"].append(section)
            continue
        if "KC" not in d or "KJ" not in d:
            continue
        if d.get("KP") == event_id:
            continue  # é a própria partida ainda não disputada, não um jogo passado
        if section is None:
            if tab is None:
                tab = {"label": "Total", "sections": []}
                tabs.append(tab)
            section = {"title": "Confrontos", "rows": []}
            tab["sections"].append(section)
        home = (d.get("KJ") or "").lstrip("*")
        away = (d.get("KK") or "").lstrip("*")
        section["rows"].append({
            "id":     d.get("KP", ""),
            "date":   d.get("KC", ""),
            "league": d.get("KF", ""),
            "pais":   d.get("KH", ""),
            "home":   home,
            "away":   away,
            "score":  d.get("KL", ""),
            "result": RESULT_LABEL.get(d.get("WIS", ""), ""),
        })
    return tabs

def _fs_half_time(event_id):
    """Placar do 1º tempo de uma partida já disputada (feed df_sui_, blocos 'AC': o
    primeiro bloco 'AC' é sempre o 1º tempo, com IG/IH = gols de casa/fora nesse tempo)."""
    text = _fs_get(f"df_sui_1_{event_id}")
    for block in _fs_blocks(text):
        d = _fs_kv(block)
        if "AC" in d and "1" in d["AC"]:
            return {"home": d.get("IG", "0"), "away": d.get("IH", "0")}
    return None

def _fs_goal_minutes(event_id):
    """Minutos de cada gol da partida, separados por time (casa/fora), usando o feed
    df_sui_ (resumo/timeline). Cada bloco de gol (IK='Gol') traz INX/IOX = placar da
    casa/fora IMEDIATAMENTE APÓS aquele gol — comparando com o placar anterior dá pra
    saber de qual time foi o gol."""
    text = _fs_get(f"df_sui_1_{event_id}")

    def _parse_minute(raw):
        # "26'" -> 26 ; "45+2'" -> 47 (soma o acréscimo)
        raw = (raw or "").rstrip("'")
        if "+" in raw:
            try:
                base, extra = raw.split("+")
                return int(base) + int(extra)
            except ValueError:
                return None
        try:
            return int(raw)
        except ValueError:
            return None

    home_minutes, away_minutes = [], []
    prev_home, prev_away = 0, 0
    for block in _fs_blocks(text):
        d = _fs_kv(block)
        if d.get("IK") != "Gol":
            continue
        try:
            cur_home, cur_away = int(d.get("INX", prev_home)), int(d.get("IOX", prev_away))
        except ValueError:
            continue
        minute = _parse_minute(d.get("IB"))
        if minute is not None:
            if cur_home > prev_home:
                home_minutes.append(minute)
            elif cur_away > prev_away:
                away_minutes.append(minute)
        prev_home, prev_away = cur_home, cur_away
    return {"home": home_minutes, "away": away_minutes}

def _fs_goal_events(event_id):
    """Igual a _fs_goal_minutes, mas guarda também quem marcou/tomou cartão — o feed
    já traz o nome do jogador em 'IF' (ex: 'Loupatty E.') em cada bloco de evento, só
    não era usado até agora. 'ICT' indica pênalti/gol contra quando presente (testado
    em jogos reais — normalmente vem vazio, então não força uma tag quando não vier).
    Pra gol, o lado (casa/fora) é derivado comparando INX/IOX com o placar anterior;
    pra cartão não tem placar pra comparar, então usa 'IA' direto (1=casa, 2=fora,
    confirmado testando contra gols onde os dois métodos batem)."""
    text = _fs_get(f"df_sui_1_{event_id}")

    def _parse_minute(raw):
        raw = (raw or "").rstrip("'")
        if "+" in raw:
            try:
                base, extra = raw.split("+")
                return int(base) + int(extra)
            except ValueError:
                return None
        try:
            return int(raw)
        except ValueError:
            return None

    events = []
    prev_home, prev_away = 0, 0
    for block in _fs_blocks(text):
        d = _fs_kv(block)
        ik = d.get("IK")
        minute = _parse_minute(d.get("IB"))
        if ik == "Gol":
            try:
                cur_home, cur_away = int(d.get("INX", prev_home)), int(d.get("IOX", prev_away))
            except ValueError:
                continue
            is_home = cur_home > prev_home
            if minute is not None and (is_home or cur_away > prev_away):
                events.append({
                    "type": "gol",
                    "minute": minute,
                    "minute_label": d.get("IB", ""),
                    "player": d.get("IF", ""),
                    "isHome": is_home,
                    "note": d.get("ICT", ""),  # ex: pênalti/gol contra, quando o feed manda
                })
            prev_home, prev_away = cur_home, cur_away
        elif ik in ("Cartão Vermelho", "Cartão Amarelo") and minute is not None:
            events.append({
                "type": "cartao_vermelho" if ik == "Cartão Vermelho" else "cartao_amarelo",
                "minute": minute,
                "minute_label": d.get("IB", ""),
                "player": d.get("IF", ""),
                "isHome": d.get("IA") == "1",
                "note": "",
            })
    return events

def _fs_standings(event_id, home_name="", away_name=""):
    """Tabela de classificação da liga da partida (posição, pontos, V/E/D, saldo)."""
    text = _fs_get(f"df_tl_1_{event_id}")
    rows = []

    def _norm(s):
        return (s or "").strip().lower()
    nh, na = _norm(home_name), _norm(away_name)

    for block in _fs_blocks(text):
        d = _fs_kv(block)
        if "TR" not in d or "TN" not in d:
            continue
        nome = d.get("TN", "")
        rows.append({
            "pos":     d.get("TR", ""),
            "team":    nome,
            "jogos":   d.get("TM", ""),
            "vitorias":  d.get("TW", ""),
            "empates":   d.get("TDR", ""),
            "derrotas":  d.get("TL", ""),
            "gols":    d.get("TG", ""),
            "pontos":  d.get("TP", ""),
            "destacado": _norm(nome) in (nh, na),
        })
    return rows

def _fs_odds(event_id, bookmaker_id=574, bet_type="HOME_DRAW_AWAY", bet_scope="FULL_TIME"):
    """Odds 1X2 de uma casa de apostas específica (bookmaker_id) via GraphQL da lsapp.eu.
    Timeout curto (4s) — em ligas menores, essa API às vezes trava ao invés de retornar
    404 rápido; com timeout de 10s e fallback de 3 casas, um jogo lento sozinho podia
    travar até 30s, e multiplicado por centenas de jogos no ranking do Backtest 2 isso
    inflava o tempo total demais."""
    url = ("https://global.ds.lsapp.eu/odds/pq_graphql"
           f"?_hash=ope2&eventId={event_id}&bookmakerId={bookmaker_id}"
           f"&betType={bet_type}&betScope={bet_scope}")
    r = http_req.get(url, headers=FS_HEADERS, timeout=4)
    r.raise_for_status()
    return r.json().get("data", {}).get("findPrematchOddsForBookmaker")

@app.route("/api/flashscore/today")
def api_flashscore_today():
    """Todas as partidas do dia no Soccerway/Flashscore (agendadas + ao vivo + encerradas),
    com liga/país — usado pela aba experimental 'Hoje 2'."""
    matches = _fs_all_matches()
    return jsonify({"total": len(matches), "matches": matches})

@app.route("/api/flashscore/match")
def api_flashscore_match():
    """Busca id/placar/status do jogo pelo nome dos times (casa/fora), sem estatísticas."""
    casa = request.args.get("casa", "")
    fora = request.args.get("fora", "")
    m = _fs_find_match(casa, fora)
    if not m:
        return jsonify({"found": False}), 404
    return jsonify({"found": True, **m})

@app.route("/api/flashscore/stats")
def api_flashscore_stats():
    """Estatísticas completas da partida (xG, posse, chutes, passes, defesa, etc.)."""
    casa = request.args.get("casa", "")
    fora = request.args.get("fora", "")
    m = _fs_find_match(casa, fora)
    if not m:
        return jsonify({"found": False}), 404
    try:
        sections = _fs_match_stats(m["id"])
    except Exception as e:
        return jsonify({"found": True, "match": m, "error": str(e)}), 503
    return jsonify({"found": True, "match": m, "sections": sections})

@app.route("/api/flashscore/odds")
def api_flashscore_odds():
    """Odds 1X2 (uma ou mais casas de apostas) da partida, pelo nome dos times."""
    casa = request.args.get("casa", "")
    fora = request.args.get("fora", "")
    m = _fs_find_match(casa, fora)
    if not m:
        return jsonify({"found": False}), 404
    # Algumas casas conhecidas nesse feed (bet365=16, Betano.br=574, ...) — tenta a
    # primeira que responder com dados válidos, sem travar em uma só
    BOOKMAKERS = [(16, "bet365"), (574, "Betano.br"), (49, "Tipsport")]
    for bid, name in BOOKMAKERS:
        try:
            odds = _fs_odds(m["id"], bookmaker_id=bid)
            if odds:
                return jsonify({"found": True, "match": m, "bookmaker": name, "odds": odds})
        except Exception:
            continue
    return jsonify({"found": True, "match": m, "odds": None})

FS_ODDS_MARKETS = [
    ("1x2", "HOME_DRAW_AWAY"),
    ("over_under", "OVER_UNDER"),
    ("ambos_marcam", "BOTH_TEAMS_TO_SCORE"),
    ("dupla_chance", "DOUBLE_CHANCE"),
    ("handicap_asiatico", "ASIAN_HANDICAP"),
    ("placar_exato", "CORRECT_SCORE"),
]

FS_ODDS_BOOKMAKERS = [(16, "bet365"), (574, "Betano.br"), (49, "Tipsport")]

from concurrent.futures import ThreadPoolExecutor

# Pools separados (mercados dentro de 1 jogo vs. vários jogos ao mesmo tempo) pra evitar
# que um pool único fique com todos os workers presos esperando sub-tarefas dele mesmo.
_fs_market_pool = ThreadPoolExecutor(max_workers=30)
_fs_event_pool = ThreadPoolExecutor(max_workers=12)
# Pools dedicados só pro "jogos que se encaixam na metodologia" (Backtest 2) — sem
# isso, essa busca rápida (poucos jogos) ficava presa na fila atrás do cálculo pesado
# do ranking (centenas de lookups de odds), que usa _fs_event_pool E _fs_market_pool
# por até 1 minuto. Isolando os dois níveis (evento e mercado), a busca continua
# rápida mesmo com um recálculo de ranking rodando ao mesmo tempo.
_bt2_matches_pool = ThreadPoolExecutor(max_workers=10)
_bt2_matches_market_pool = ThreadPoolExecutor(max_workers=20)

# Em dias com muitos jogos (200+), buscar odds de TODOS os agendados demora
# minutos (até 3 tentativas de casa de apostas x timeout por jogo, dividido
# entre poucos workers). Limita aos próximos N jogos por horário — cobre o
# que realmente dá pra apostar em breve, sem travar a busca.
_BT2_MATCHES_MAX_CANDIDATOS = 60


def _bt2_matches_candidatos(include_finished=False):
    """Jogos de hoje candidatos pros filtros (Metodologias/Parâmetros/Análise) —
    por padrão só agendados ("1"), já que os filtros dependem de odds. Passando
    include_finished=True, também inclui encerrados ("3") — as odds continuam
    disponíveis na Flashscore mesmo depois do jogo acabar (campo "opening" é a
    odd de abertura, "value" fica com o último valor antes do jogo travar)."""
    statuses = ("1", "3") if include_finished else ("1",)
    candidatos = [m for m in _fs_all_matches() if m.get("status") in statuses]
    candidatos.sort(key=lambda m: m.get("kickoff_ts") or "")
    return candidatos[:_BT2_MATCHES_MAX_CANDIDATOS]


_TODAY2_ODDS_SNAPSHOT_TTL = 300  # 5min — odds mudam pouco em poucos minutos
_today2_odds_snapshot_cache = {"ts": 0, "data": None}


def _today2_odds_snapshot(force=False):
    """Busca as odds de TODOS os mercados de cada jogo candidato de hoje (agendado
    ou encerrado) UMA VEZ SÓ, cacheada 5min — usada por vários consumidores (Filtro
    de Metodologias, Filtro de Parâmetros, odd média por bucket do Backtest CS).
    Sem esse cache compartilhado, cada checkbox marcado no Filtro de Metodologias
    disparava seu PRÓPRIO fetch completo (~70s, casa de aposta x mercado x jogo) —
    com várias metodologias marcadas ao mesmo tempo, elas competiam pelo mesmo pool
    de threads e o filtro parecia simplesmente não funcionar (ainda calculando
    minutos depois, sem indicação de carregamento)."""
    now = time.time()
    if not force and _today2_odds_snapshot_cache["data"] is not None and \
            (now - _today2_odds_snapshot_cache["ts"]) < _TODAY2_ODDS_SNAPSHOT_TTL:
        return _today2_odds_snapshot_cache["data"]

    candidatos = _bt2_matches_candidatos(include_finished=True)

    def _fetch_one(m):
        try:
            _, markets = _fs_odds_all_markets_any_bookmaker(m["id"], pool=_bt2_matches_market_pool)
        except Exception:
            markets = {}
        return m, markets

    snapshot = list(_bt2_matches_pool.map(_fetch_one, candidatos))
    _today2_odds_snapshot_cache["ts"] = now
    _today2_odds_snapshot_cache["data"] = snapshot
    return snapshot

def _fs_odds_has_data(key, odds):
    """A API às vezes retorna um objeto 'válido' mas vazio (ex: Placar Exato com
    items:[] quando essa casa não tem esse mercado pra esse jogo) — sem isso, o código
    tratava como 'achei dados' e parava de tentar outras casas de apostas."""
    if not odds:
        return False
    if key in ("over_under", "handicap_asiatico"):
        return bool(odds.get("opportunities"))
    if key == "placar_exato":
        return bool(odds.get("items"))
    if key == "1x2":
        return bool(odds.get("home") and odds.get("draw") and odds.get("away"))
    if key == "ambos_marcam":
        return bool(odds.get("yes") and odds.get("no"))
    if key == "dupla_chance":
        return bool(odds.get("homeOrDraw") and odds.get("homeOrAway") and odds.get("drawOrAway"))
    return True

def _fs_odds_all_markets(event_id, bookmaker_id=16, pool=None, markets_wanted=None):
    """Busca os mercados confirmados pra um event_id numa casa específica, em paralelo
    (uma requisição por mercado ao mesmo tempo). Retorna markets_dict (pode vir vazio).
    Aceita um pool alternativo (default: _fs_market_pool) pra isolar chamadas que não
    podem ficar presas atrás de cálculos pesados que também usam o pool padrão.
    markets_wanted (lista de keys, ex: ["1x2"]) restringe aos mercados pedidos em vez
    dos 6 confirmados — usado por quem só precisa de 1, evita chamadas à toa."""
    pool = pool or _fs_market_pool
    items = FS_ODDS_MARKETS if markets_wanted is None else [i for i in FS_ODDS_MARKETS if i[0] in markets_wanted]

    def _fetch_one(item):
        key, bet_type = item
        try:
            return key, _fs_odds(event_id, bookmaker_id=bookmaker_id, bet_type=bet_type)
        except Exception:
            return key, None
    markets = {}
    for key, odds in pool.map(_fetch_one, items):
        if _fs_odds_has_data(key, odds):
            markets[key] = odds
    return markets

def _fs_odds_all_markets_any_bookmaker(event_id, pool=None, markets_wanted=None):
    """Tenta bet365/Betano/Tipsport nessa ordem, MISTURANDO casas por mercado — a
    Betano.br, por exemplo, não expõe 'Ambos Marcam' nessa API pra praticamente
    nenhuma partida (testado e confirmado, não é bug de parsing: a API retorna
    null mesmo), então usar só a 1ª casa que respondeu deixava esse mercado
    faltando toda vez que o bet365 não tinha dados e caía pro Betano. Agora
    completa os mercados que faltarem com a próxima casa da lista. O nome
    retornado é da 1ª casa que contribuiu com algo (a mais completa via de
    regra), pros mercados que vieram de outra casa não tem atribuição individual
    no retorno — é uma simplificação aceitável já que o objetivo é preencher
    lacunas, não misturar odds de mercados que uma mesma casa já tem.
    markets_wanted restringe quais dos 6 mercados são buscados (default: todos) —
    quem só precisa de "1x2" (ex: Backtest CS) evita 5/6 das chamadas à toa."""
    bookmaker_name = None
    markets = {}
    wanted_keys = markets_wanted if markets_wanted is not None else [k for k, _ in FS_ODDS_MARKETS]
    for bid, name in FS_ODDS_BOOKMAKERS:
        faltando = [k for k in wanted_keys if k not in markets]
        if not faltando:
            break
        casa_markets = _fs_odds_all_markets(event_id, bookmaker_id=bid, pool=pool, markets_wanted=faltando)
        if casa_markets and bookmaker_name is None:
            bookmaker_name = name
        for k, v in casa_markets.items():
            markets.setdefault(k, v)
    return bookmaker_name, markets

@app.route("/api/flashscore/odds_all")
def api_flashscore_odds_all():
    """Todos os mercados de odds confirmados (1X2, Acima/Abaixo, Ambos Marcam,
    Dupla Chance, Handicap Asiático, Placar Exato) de uma casa de apostas, pelo
    nome dos times. Usado apenas na aba experimental 'Hoje 2'."""
    casa = request.args.get("casa", "")
    fora = request.args.get("fora", "")
    m = _fs_find_match(casa, fora)
    if not m:
        return jsonify({"found": False}), 404
    bookmaker_name, markets = _fs_odds_all_markets_any_bookmaker(m["id"])
    return jsonify({"found": True, "match": m, "bookmaker": bookmaker_name, "markets": markets})

@app.route("/api/flashscore/odds_all_batch")
def api_flashscore_odds_all_batch():
    """Odds (todos os mercados) de várias partidas JÁ DISPUTADAS de uma vez, direto
    pelo event_id de cada uma (sem precisar buscar por nome de time) — usado pela aba
    'Profit' pra calcular o profit histórico usando as odds de cada jogo passado do H2H.
    Tenta bet365/Betano/Tipsport (várias ligas menores só têm odds em uma delas) e
    busca todos os ids em paralelo. Limitado a 20 ids por chamada pra não sobrecarregar
    o feed."""
    ids = [i for i in request.args.get("ids", "").split(",") if i][:20]

    def _fetch_one(event_id):
        try:
            return event_id, _fs_odds_all_markets_any_bookmaker(event_id)
        except Exception:
            return event_id, (None, {})

    result = {}
    for event_id, (bookmaker_name, markets) in _fs_event_pool.map(_fetch_one, ids):
        if markets:
            result[event_id] = {"bookmaker": bookmaker_name, "markets": markets}
    return jsonify({"results": result})

@app.route("/api/flashscore/h2h")
def api_flashscore_h2h():
    """Últimos jogos de cada time + confrontos diretos, pelo nome dos times."""
    casa = request.args.get("casa", "")
    fora = request.args.get("fora", "")
    m = _fs_find_match(casa, fora)
    if not m:
        return jsonify({"found": False}), 404
    try:
        tabs = _fs_h2h(m["id"])
    except Exception as e:
        return jsonify({"found": True, "match": m, "error": str(e)}), 503
    return jsonify({"found": True, "match": m, "tabs": tabs})

@app.route("/api/flashscore/half_time_batch")
def api_flashscore_half_time_batch():
    """Placar do 1º tempo de várias partidas já disputadas de uma vez (usado para
    mostrar 'gols no 1º tempo' nos jogos passados listados no H2H). Limitado a 15
    ids por chamada pra não sobrecarregar o feed.
    Buscado em PARALELO (_fs_event_pool), igual odds_all_batch/goal_minutes_batch —
    antes era um `for` sequencial, um jogo de cada vez, e isso sozinho já explicava
    boa parte da lentidão ao abrir qualquer partida (a aba Jogo depende desse
    endpoint pra quase todo jogo do histórico)."""
    ids = [i for i in request.args.get("ids", "").split(",") if i][:15]

    def _fetch_one(event_id):
        try:
            return event_id, _fs_half_time(event_id)
        except Exception:
            return event_id, None

    result = {}
    for event_id, ht in _fs_event_pool.map(_fetch_one, ids):
        if ht:
            result[event_id] = ht
    return jsonify({"results": result})

@app.route("/api/flashscore/goal_minutes_batch")
def api_flashscore_goal_minutes_batch():
    """Minutos dos gols (por time) de várias partidas já disputadas de uma vez — usado
    pra calcular 'tempo sem marcar/sofrer gol' na aba Jogo. Busca em paralelo, limitado
    a 20 ids por chamada."""
    ids = [i for i in request.args.get("ids", "").split(",") if i][:20]

    def _fetch_one(event_id):
        try:
            return event_id, _fs_goal_minutes(event_id)
        except Exception:
            return event_id, None

    result = {}
    for event_id, gm in _fs_event_pool.map(_fetch_one, ids):
        if gm is not None:
            result[event_id] = gm
    return jsonify({"results": result})

@app.route("/api/flashscore/goal_events")
def api_flashscore_goal_events():
    """Quem marcou cada gol (e cartões vermelho/amarelo) e em que minuto, pra
    partidas já encerradas — usado no Hoje 2 pra mostrar os artilheiros/cartões
    (nome + minuto) igual ao placar final."""
    casa = request.args.get("casa", "")
    fora = request.args.get("fora", "")
    m = _fs_find_match(casa, fora)
    if not m:
        return jsonify({"found": False}), 404
    try:
        events = _fs_goal_events(m["id"])
    except Exception as e:
        return jsonify({"found": True, "match": m, "error": str(e)}), 503
    return jsonify({"found": True, "match": m, "events": events})

@app.route("/api/flashscore/standings")
def api_flashscore_standings():
    """Tabela de classificação da liga da partida, pelo nome dos times."""
    casa = request.args.get("casa", "")
    fora = request.args.get("fora", "")
    m = _fs_find_match(casa, fora)
    if not m:
        return jsonify({"found": False}), 404
    try:
        rows = _fs_standings(m["id"], home_name=m.get("home", ""), away_name=m.get("away", ""))
    except Exception as e:
        return jsonify({"found": True, "match": m, "error": str(e)}), 503
    return jsonify({"found": True, "match": m, "rows": rows})



# ── BACKTEST 2 — ranking global de metodologias ──────────────────────────────
# Roda a MESMA simulação de lucro da aba "Profit" (Hoje 2), mas agregando o
# histórico de TODOS os jogos de hoje de uma vez, pra ter amostra estatística
# grande o suficiente pra rankear mercado+seleção do mais pro menos lucrativo.
_bt2_cache = {"ts": 0, "data": None}
_BT2_CACHE_TTL = 30 * 60  # 30 minutos

_BT2_MARKET_LABELS = {
    "1x2": "1X2", "over_under": "Acima/Abaixo", "ambos_marcam": "Ambos Marcam",
    "dupla_chance": "Dupla Chance", "handicap_asiatico": "Handicap Asiático",
}
_BT2_MIN_SAMPLE = 15
_BT2_MAX_HISTORICAL_IDS = 400

# Faixas de odd — usadas pra segmentar o ranking por odd em vez de só juntar tudo
# (ex: "Casa" no 1X2 pode ter ROI bem diferente com odd 1.50 vs odd 3.50). A odd de
# cada aposta já é gravada em bt2_bets, então isso é calculado na hora da leitura,
# sem precisar de coluna nova nem migração de schema.
_BT2_ODD_RANGES = [
    (0.0, 1.5,  "1.01–1.50"),
    (1.5, 2.0,  "1.51–2.00"),
    (2.0, 3.0,  "2.01–3.00"),
    (3.0, 5.0,  "3.01–5.00"),
    (5.0, 9e9,  "5.01+"),
]
_BT2_MIN_SAMPLE_ODD_RANGE = 8  # amostra menor exigida pra cada faixa (o corte reparte a amostra total)

def _bt2_odd_range_label(odd):
    if not odd:
        return None
    for lo, hi, label in _BT2_ODD_RANGES:
        if lo < odd <= hi:
            return label
    return None


def _bt2_parse_score(s):
    m = re.match(r"^(\d+):(\d+)$", s or "")
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


def _bt2_team_rows_from_section(tabs, team_home):
    """Porta do _profitTeamRowsFromSection (JS) pra Python: pega a seção "Últimos
    jogos: <time>" (não confrontos) do time casa/fora dentro da aba 'Total' (idx 0)
    e extrai {id,h,a,date} de cada jogo com placar válido (até 30 por time)."""
    if not tabs:
        return []
    tab = tabs[0]
    sections = [s for s in tab["sections"] if "confront" not in (s.get("title") or "").lower()]
    sec = sections[0] if team_home else (sections[1] if len(sections) > 1 else (sections[0] if sections else None))
    if not sec:
        return []
    rows = []
    for r in sec["rows"][:30]:
        sc = _bt2_parse_score(r.get("score"))
        if not sc:
            continue
        rows.append({"id": r.get("id"), "h": sc[0], "a": sc[1], "date": r.get("date"),
                     "home": r.get("home"), "away": r.get("away")})
    return rows


def _bt2_market_selections(market_key, o, h, a):
    """Porta do _profitMarketSelections (JS) pra Python — mesma lógica, mesmos 6
    mercados, mesmas condições de vitória por seleção."""
    if not o:
        return []
    total = h + a

    def _f(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    if market_key == "1x2":
        return [
            {"label": "Casa", "odd": _f((o.get("home") or {}).get("value")), "won": h > a},
            {"label": "Empate", "odd": _f((o.get("draw") or {}).get("value")), "won": h == a},
            {"label": "Fora", "odd": _f((o.get("away") or {}).get("value")), "won": a > h},
        ]
    if market_key == "ambos_marcam":
        btts = h >= 1 and a >= 1
        return [
            {"label": "Sim", "odd": _f((o.get("yes") or {}).get("value")), "won": btts},
            {"label": "Não", "odd": _f((o.get("no") or {}).get("value")), "won": not btts},
        ]
    if market_key == "dupla_chance":
        return [
            {"label": "1X", "odd": _f((o.get("homeOrDraw") or {}).get("value")), "won": h >= a},
            {"label": "12", "odd": _f((o.get("homeOrAway") or {}).get("value")), "won": h != a},
            {"label": "X2", "odd": _f((o.get("drawOrAway") or {}).get("value")), "won": a >= h},
        ]
    if market_key == "over_under":
        sels = []
        for op in o.get("opportunities") or []:
            line = _f((op.get("handicap") or {}).get("value"))
            if line is None or total == line:
                continue
            sels.append({"label": f"Acima {line}", "odd": _f((op.get("over") or {}).get("value")), "won": total > line})
            sels.append({"label": f"Abaixo {line}", "odd": _f((op.get("under") or {}).get("value")), "won": total < line})
        return sels
    if market_key == "handicap_asiatico":
        sels = []
        for op in o.get("opportunities") or []:
            line = _f((op.get("handicap") or {}).get("value"))
            if line is None:
                continue
            adj_home = h + line
            if adj_home == a:
                continue
            sign = "+" if line >= 0 else ""
            sign2 = "+" if -line >= 0 else ""
            sels.append({"label": f"Casa ({sign}{line})", "odd": _f((op.get("home") or {}).get("value")), "won": adj_home > a})
            sels.append({"label": f"Fora ({sign2}{-line})", "odd": _f((op.get("away") or {}).get("value")), "won": adj_home < a})
        return sels
    return []


# ── BACKTEST 2 — persistência acumulada em SQLite ────────────────────────────
# Além do snapshot "últimos 30 jogos" (em memória, recalculado a cada 30min),
# guardamos cada aposta histórica individual num SQLite pra construir um ranking
# ACUMULADO que cresce com o tempo (não se limita ao histórico dos jogos de hoje).
_BT2_DB_PATH = os.path.join(DATA_DIR, "backtest2.db")
_bt2_db_lock = threading.Lock()


def _bt2_push_db_bg():
    """Envia o backtest2.db pro GitHub (branch 'data') em segundo plano — sem isso,
    o histórico acumulado (Backtest 2 e Backtest CS, que moram no mesmo arquivo)
    se perdia a cada redeploy no Railway, porque o disco local não sobrevive entre
    deploys. Restaurado de volta em github_storage.sync_on_startup."""
    github_storage.push_file_bg(_BT2_DB_PATH, "backtest2.db")


def _bt2_db_conn():
    conn = sqlite3.connect(_BT2_DB_PATH, timeout=30)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS bt2_bets (
            event_id TEXT NOT NULL,
            market_key TEXT NOT NULL,
            selection TEXT NOT NULL,
            odd REAL NOT NULL,
            won INTEGER NOT NULL,
            profit REAL NOT NULL,
            match_date TEXT,
            inserted_at TEXT NOT NULL,
            PRIMARY KEY (event_id, market_key, selection)
        )
    """)
    return conn


def _bt2_persist_bets(rows):
    """Grava linhas [(event_id, market_key, selection, odd, won, profit, match_date), ...]
    no SQLite via INSERT OR IGNORE (chave event_id+market_key+selection evita
    duplicar apostas já processadas em execuções anteriores)."""
    if not rows:
        return
    now = datetime.utcnow().isoformat() + "Z"
    with _bt2_db_lock:
        conn = _bt2_db_conn()
        try:
            conn.executemany(
                """INSERT OR IGNORE INTO bt2_bets
                   (event_id, market_key, selection, odd, won, profit, match_date, inserted_at)
                   VALUES (?,?,?,?,?,?,?,?)""",
                [(eid, mk, sel, odd, 1 if won else 0, profit, date, now) for (eid, mk, sel, odd, won, profit, date) in rows],
            )
            conn.commit()
        finally:
            conn.close()
    _bt2_push_db_bg()


def _bt2_compute_variance(profits):
    mean = sum(profits) / len(profits)
    variance = sum((p - mean) ** 2 for p in profits) / len(profits)
    stddev = variance ** 0.5
    cv = (stddev / abs(mean)) if mean != 0 else None
    return variance, stddev, cv


def _bt2_max_streaks(dates, wons):
    """Maior sequência de vitórias seguidas (max green) e maior sequência de
    derrotas seguidas (max red), na ordem cronológica das apostas (por
    match_date) — não é drawdown em unidades, é contagem de apostas."""
    pairs = sorted(zip(dates, wons), key=lambda x: (x[0] or ""))
    max_win = max_loss = cur_win = cur_loss = 0
    for _, won in pairs:
        if won:
            cur_win += 1
            cur_loss = 0
        else:
            cur_loss += 1
            cur_win = 0
        max_win = max(max_win, cur_win)
        max_loss = max(max_loss, cur_loss)
    return max_win, max_loss


def _bt2_odd_ranges_breakdown(odds, wons, profits):
    """A partir das odds/won/profit de cada aposta de uma seleção, agrupa por faixa
    de odd (ver _BT2_ODD_RANGES) e devolve a lista de faixas com amostra suficiente,
    ordenada na mesma ordem das faixas (menor odd primeiro)."""
    by_range = {}
    for odd, won, profit in zip(odds, wons, profits):
        label = _bt2_odd_range_label(odd)
        if not label:
            continue
        r = by_range.setdefault(label, {"bets": 0, "wins": 0, "profit": 0.0})
        r["bets"] += 1
        if won:
            r["wins"] += 1
        r["profit"] += profit

    order = [label for _, _, label in _BT2_ODD_RANGES]
    out = []
    for label in order:
        r = by_range.get(label)
        if not r or r["bets"] < _BT2_MIN_SAMPLE_ODD_RANGE:
            continue
        out.append({
            "faixa":   label,
            "bets":    r["bets"],
            "wins":    r["wins"],
            "winrate": round(r["wins"] / r["bets"], 4),
            "profit":  round(r["profit"], 4),
            "roi":     round(r["profit"] / r["bets"] * 100, 2),
        })
    return out


def _bt2_compute_ranking_acumulado():
    """Ranking acumulado: lê TODAS as apostas já persistidas em bt2_bets (não só as
    dos jogos de hoje) e agrega por mercado+seleção, com timeline cumulativa por
    match_date. Sem cache — leitura no SQLite é barata."""
    with _bt2_db_lock:
        conn = _bt2_db_conn()
        try:
            cur = conn.execute(
                "SELECT market_key, selection, odd, won, profit, match_date FROM bt2_bets ORDER BY match_date ASC"
            )
            all_rows = cur.fetchall()
            sample_matches_used = conn.execute("SELECT COUNT(DISTINCT event_id) FROM bt2_bets").fetchone()[0]
        finally:
            conn.close()

    agg = {}
    for market_key, selection, odd, won, profit, match_date in all_rows:
        bucket = agg.setdefault(market_key, {})
        r = bucket.setdefault(selection, {"bets": 0, "wins": 0, "profit": 0.0, "profits": [], "dates": [], "odds": [], "wons": []})
        r["bets"] += 1
        if won:
            r["wins"] += 1
        r["profit"] += profit
        r["profits"].append(profit)
        r["dates"].append(match_date or "")
        r["odds"].append(odd)
        r["wons"].append(won)

    methodologies = []
    for market_key, bucket in agg.items():
        market_label = _BT2_MARKET_LABELS.get(market_key, market_key)
        for selection, r in bucket.items():
            if r["bets"] < _BT2_MIN_SAMPLE:
                continue
            profits = r["profits"]
            variance, stddev, cv = _bt2_compute_variance(profits)
            max_win_streak, max_loss_streak = _bt2_max_streaks(r["dates"], r["wons"])

            pairs = sorted(zip(r["dates"], profits), key=lambda x: (x[0] or ""))
            cum = 0.0
            timeline = []
            for date, p in pairs:
                cum += p
                timeline.append({"date": date, "cumulative_profit": round(cum, 4)})

            methodologies.append({
                "market": market_key,
                "market_label": market_label,
                "selection": selection,
                "bets": r["bets"],
                "wins": r["wins"],
                "winrate": round(r["wins"] / r["bets"], 4),
                "profit": round(r["profit"], 4),
                "roi": round(r["profit"] / r["bets"] * 100, 2),
                "odd_ranges": _bt2_odd_ranges_breakdown(r["odds"], r["wons"], profits),
                "variance": round(variance, 4),
                "stddev": round(stddev, 4),
                "cv": round(cv, 4) if cv is not None else None,
                "avg_odd": round(sum(r["odds"]) / len(r["odds"]), 2) if r["odds"] else None,
                "max_win_streak": max_win_streak,
                "max_loss_streak": max_loss_streak,
                "timeline": timeline,
            })

    methodologies.sort(key=lambda x: x["profit"], reverse=True)

    return {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "sample_matches_used": sample_matches_used,
        "methodologies": methodologies,
    }


def _bt2_compute_ranking():
    """Algoritmo completo: coleta histórico de todos os jogos de hoje, busca odds em
    lote, agrega bets/wins/profit por mercado+seleção e rankeia por lucro."""
    today_matches = _fs_all_matches()

    # 1) Coleta ids históricos (até 400 distintos) percorrendo o H2H de cada jogo de hoje
    historical_by_id = {}
    for m in today_matches:
        if len(historical_by_id) >= _BT2_MAX_HISTORICAL_IDS:
            break
        try:
            tabs = _fs_h2h(m["id"])
        except Exception:
            continue
        for team_home in (True, False):
            for row in _bt2_team_rows_from_section(tabs, team_home):
                if not row["id"] or row["id"] in historical_by_id:
                    continue
                historical_by_id[row["id"]] = row
                if len(historical_by_id) >= _BT2_MAX_HISTORICAL_IDS:
                    break

    hist_ids = list(historical_by_id.keys())

    # 2) Busca odds de todos os jogos históricos em paralelo (reaproveita o pool existente)
    def _fetch_odds(event_id):
        try:
            return event_id, _fs_odds_all_markets_any_bookmaker(event_id)
        except Exception:
            return event_id, (None, {})

    odds_by_id = {}
    for event_id, (bookmaker_name, markets) in _fs_event_pool.map(_fetch_odds, hist_ids):
        if markets:
            odds_by_id[event_id] = markets

    # 3) Agrega globalmente por mercado+seleção
    agg = {}  # market_key -> { label -> {bets,wins,profit,profits:[],dates:[]} }
    sample_matches_used = 0
    persist_rows = []  # (event_id, market_key, selection, odd, won, profit, match_date) p/ SQLite
    for event_id, markets in odds_by_id.items():
        row = historical_by_id.get(event_id)
        if not row:
            continue
        sample_matches_used += 1
        for market_key, market_label in _BT2_MARKET_LABELS.items():
            sels = _bt2_market_selections(market_key, markets.get(market_key), row["h"], row["a"])
            for sel in sels:
                if not sel["odd"]:
                    continue
                bucket = agg.setdefault(market_key, {})
                r = bucket.setdefault(sel["label"], {"bets": 0, "wins": 0, "profit": 0.0, "profits": [], "dates": [], "odds": [], "wons": []})
                p = (sel["odd"] - 1) if sel["won"] else -1
                r["bets"] += 1
                if sel["won"]:
                    r["wins"] += 1
                r["profit"] += p
                r["profits"].append(p)
                r["dates"].append(row.get("date") or "")
                r["odds"].append(sel["odd"])
                r["wons"].append(sel["won"])
                persist_rows.append((str(event_id), market_key, sel["label"], sel["odd"], sel["won"], p, row.get("date") or ""))

    # Persiste cada aposta individual no SQLite (modo acumulado), sem afetar o
    # snapshot "últimos 30" retornado abaixo — é só um efeito colateral de gravação.
    try:
        _bt2_persist_bets(persist_rows)
    except Exception:
        pass

    # 4) Monta lista final com variância/roi/timeline, filtrando amostra mínima
    methodologies = []
    for market_key, bucket in agg.items():
        market_label = _BT2_MARKET_LABELS.get(market_key, market_key)
        for selection, r in bucket.items():
            if r["bets"] < _BT2_MIN_SAMPLE:
                continue
            profits = r["profits"]
            mean = sum(profits) / len(profits)
            variance = sum((p - mean) ** 2 for p in profits) / len(profits)
            stddev = variance ** 0.5
            cv = (stddev / abs(mean)) if mean != 0 else None
            max_win_streak, max_loss_streak = _bt2_max_streaks(r["dates"], r["wons"])

            # timeline: cumulativo ordenado por data ascendente (data no formato timestamp
            # unix em string, mesmo campo "KC" usado no resto do H2H)
            pairs = sorted(zip(r["dates"], profits), key=lambda x: (x[0] or ""))
            cum = 0.0
            timeline = []
            for date, p in pairs:
                cum += p
                timeline.append({"date": date, "cumulative_profit": round(cum, 4)})

            methodologies.append({
                "market": market_key,
                "market_label": market_label,
                "selection": selection,
                "bets": r["bets"],
                "wins": r["wins"],
                "winrate": round(r["wins"] / r["bets"], 4),
                "profit": round(r["profit"], 4),
                "roi": round(r["profit"] / r["bets"] * 100, 2),
                "odd_ranges": _bt2_odd_ranges_breakdown(r["odds"], r["wons"], profits),
                "variance": round(variance, 4),
                "stddev": round(stddev, 4),
                "cv": round(cv, 4) if cv is not None else None,
                "avg_odd": round(sum(r["odds"]) / len(r["odds"]), 2) if r["odds"] else None,
                "max_win_streak": max_win_streak,
                "max_loss_streak": max_loss_streak,
                "timeline": timeline,
            })

    methodologies.sort(key=lambda x: x["profit"], reverse=True)

    return {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "sample_matches_used": sample_matches_used,
        "methodologies": methodologies,
    }


@app.route("/api/backtest2/ranking")
def api_backtest2_ranking():
    """Ranking global de metodologias (mercado+seleção) por lucro, agregando a
    simulação de profit em cima do histórico de TODOS os jogos de hoje (não só um
    jogo). Cálculo caro (centenas de lookups de odds), então cacheia em memória por
    30 minutos — use ?refresh=1 pra forçar recálculo."""
    mode = request.args.get("mode") or "recente"
    if mode == "acumulado":
        return jsonify(_bt2_compute_ranking_acumulado())
    force = request.args.get("refresh") == "1"
    if not force and _bt2_cache["data"] is not None and (time.time() - _bt2_cache["ts"]) < _BT2_CACHE_TTL:
        return jsonify(_bt2_cache["data"])
    data = _bt2_compute_ranking()
    _bt2_cache["ts"] = time.time()
    _bt2_cache["data"] = data
    return jsonify(data)


@app.route("/api/backtest2/ranking_acumulado")
def api_backtest2_ranking_acumulado():
    """Ranking ACUMULADO: agrega TODAS as apostas históricas já persistidas em
    bt2_bets (cresce a cada vez que /api/backtest2/ranking roda), não só o
    histórico dos jogos de hoje. Sem cache — leitura SQLite é barata."""
    return jsonify(_bt2_compute_ranking_acumulado())


def _bt2_market_selection_labels(market_key, o):
    """Como _bt2_market_selections, mas SEM depender do placar final (jogo ainda não
    aconteceu) — não filtra linhas 'push' nem calcula 'won', só lista as seleções e
    odds disponíveis nesse mercado pra poder casar com o rótulo de uma metodologia."""
    if not o:
        return []

    def _f(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    if market_key == "1x2":
        return [
            {"label": "Casa", "odd": _f((o.get("home") or {}).get("value"))},
            {"label": "Empate", "odd": _f((o.get("draw") or {}).get("value"))},
            {"label": "Fora", "odd": _f((o.get("away") or {}).get("value"))},
        ]
    if market_key == "ambos_marcam":
        return [
            {"label": "Sim", "odd": _f((o.get("yes") or {}).get("value"))},
            {"label": "Não", "odd": _f((o.get("no") or {}).get("value"))},
        ]
    if market_key == "dupla_chance":
        return [
            {"label": "1X", "odd": _f((o.get("homeOrDraw") or {}).get("value"))},
            {"label": "12", "odd": _f((o.get("homeOrAway") or {}).get("value"))},
            {"label": "X2", "odd": _f((o.get("drawOrAway") or {}).get("value"))},
        ]
    if market_key == "over_under":
        sels = []
        for op in o.get("opportunities") or []:
            line = _f((op.get("handicap") or {}).get("value"))
            if line is None:
                continue
            sels.append({"label": f"Acima {line}", "odd": _f((op.get("over") or {}).get("value"))})
            sels.append({"label": f"Abaixo {line}", "odd": _f((op.get("under") or {}).get("value"))})
        return sels
    if market_key == "handicap_asiatico":
        sels = []
        for op in o.get("opportunities") or []:
            line = _f((op.get("handicap") or {}).get("value"))
            if line is None:
                continue
            sign = "+" if line >= 0 else ""
            sign2 = "+" if -line >= 0 else ""
            sels.append({"label": f"Casa ({sign}{line})", "odd": _f((op.get("home") or {}).get("value"))})
            sels.append({"label": f"Fora ({sign2}{-line})", "odd": _f((op.get("away") or {}).get("value"))})
        return sels
    return []


@app.route("/api/backtest2/matches_for_methodology")
def api_backtest2_matches_for_methodology():
    """Jogos de HOJE (agendados ou já encerrados) onde a metodologia selecionada
    (mercado+seleção) está disponível nas odds — pra saber em quais partidas reais
    dá pra aplicar aquele sinal hoje, incluindo jogos que já encerraram (útil pra
    conferir quais bateram a metodologia mesmo depois do apito final)."""
    market_key = request.args.get("market", "")
    selection = request.args.get("selection", "")
    if not market_key or not selection:
        return jsonify({"matches": []}), 400

    resultado = []
    for m, markets in _today2_odds_snapshot():
        sels = _bt2_market_selection_labels(market_key, markets.get(market_key))
        for sel in sels:
            if sel["label"] == selection and sel["odd"]:
                resultado.append({
                    "id": m["id"], "home": m["home"], "away": m["away"],
                    "liga": m.get("liga", ""), "pais": m.get("pais", ""),
                    "kickoff_ts": m.get("kickoff_ts"), "odd": sel["odd"],
                })
                break
    resultado.sort(key=lambda x: x.get("kickoff_ts") or "")
    return jsonify({"matches": resultado})


# ── BACKTEST CS — ranking global dos 19 buckets de "placar exato" por TAXA DE
# ACERTO ────────────────────────────────────────────────────────────────────
# Irmã do Backtest 2, mas em vez de rankear mercado+seleção por LUCRO, rankeia
# os 19 buckets fixos de placar exato (0-0 .. 3-3 + 3 "qualquer outro") por
# quantas vezes o placar final de um jogo histórico realmente caiu ali. Não
# precisa de odds pro cálculo do ranking em si (o placar já vem no H2H), só
# reaproveita a coleta de jogos históricos já feita pro Backtest 2 — por isso é
# bem mais rápido. Odds só entram na hora de achar "jogos de hoje que se
# encaixam" num bucket específico.
_btcs_cache = {"ts": 0, "data": None}
_BTCS_CACHE_TTL = 30 * 60  # 30 minutos
_BTCS_MAX_HISTORICAL_IDS = _BT2_MAX_HISTORICAL_IDS

# Amostra mínima pra considerar o ranking como um todo minimamente confiável.
# Diferente do Backtest 2 (onde cada mercado+seleção tem sua própria amostra
# independente), aqui todos os 19 buckets compartilham a MESMA amostra total
# (cada jogo histórico contribui pra exatamente um bucket), então o gate
# relevante é sobre o TOTAL coletado, não por bucket. Ainda assim, exigimos
# pelo menos 1 acerto (hits >= 1) pra listar um bucket, senão buckets raros
# (tipo "3-3") apareceriam com 0% poluindo a tabela sem sinal nenhum.
_BTCS_MIN_TOTAL_SAMPLE = 15
_BTCS_MIN_HITS = 0  # 0 inclui buckets que NUNCA ocorreram — é justamente o que o
                     # filtro "⚡ Placares mais improváveis" precisa mostrar

# Porta de PLACAR_EXATO_BUCKETS_ORDER (JS, static/index.html) — mesma ordem.
PLACAR_EXATO_BUCKETS_ORDER = [
    "0-0", "0-1", "0-2", "0-3", "1-0", "1-1", "1-2", "1-3",
    "2-0", "2-1", "2-2", "2-3", "3-0", "3-1", "3-2", "3-3",
    "Qualquer outra vitória em casa", "Qualquer Outra Vitória de Visitante", "Qualquer outro empate",
]


def _btcs_bucket_key(h, a):
    """Porta exata de _placarExatoBucketKey (JS) pra Python."""
    if h <= 3 and a <= 3:
        return f"{h}-{a}"
    if h > a:
        return "Qualquer outra vitória em casa"
    if a > h:
        return "Qualquer Outra Vitória de Visitante"
    return "Qualquer outro empate"


# ── BACKTEST CS — persistência acumulada em SQLite (mesmo arquivo do Backtest 2,
# tabela própria) ─────────────────────────────────────────────────────────────
def _btcs_db_conn():
    conn = sqlite3.connect(_BT2_DB_PATH, timeout=30)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS btcs_results (
            event_id TEXT PRIMARY KEY,
            bucket TEXT NOT NULL,
            match_date TEXT,
            inserted_at TEXT NOT NULL
        )
    """)
    # Linhas por TIME (não só por jogo) — guarda o suficiente pra recalcular as
    # variáveis dos Padrões (média de gols, mandante etc.) em cima de TODO o
    # histórico já visto, não só os últimos 30 jogos do dia. Mesmo jogo pode
    # aparecer 2x (uma por time), então a chave inclui o time.
    conn.execute("""
        CREATE TABLE IF NOT EXISTS btcs_pattern_rows (
            event_id TEXT NOT NULL,
            team_key TEXT NOT NULL,
            team_name TEXT NOT NULL,
            h INTEGER NOT NULL,
            a INTEGER NOT NULL,
            home_name TEXT,
            match_date TEXT,
            inserted_at TEXT NOT NULL,
            PRIMARY KEY (event_id, team_key)
        )
    """)
    # Migração pra bases já criadas antes da coluna "odd" existir — a odd 1X2 do
    # próprio time naquele jogo histórico, usada como variável de padrão (força do
    # time conforme o mercado precificava, não só o resultado em si)
    try:
        conn.execute("ALTER TABLE btcs_pattern_rows ADD COLUMN odd REAL")
    except sqlite3.OperationalError:
        pass  # coluna já existe
    # Migração pra base de "opp_odd" (odd do ADVERSÁRIO naquele jogo histórico) —
    # precisa das duas odds (própria + adversário) pra replicar as métricas da
    # aba Jogo (Valor do Gol/Ponto/Saldo, Custo do Gol 2.0) como variáveis de padrão
    try:
        conn.execute("ALTER TABLE btcs_pattern_rows ADD COLUMN opp_odd REAL")
    except sqlite3.OperationalError:
        pass  # coluna já existe
    # Migração pra base de "away_name" — precisamos dos DOIS times do jogo
    # histórico (não só o mandante) pra casar com o backtest/ do StatArea
    # (fonte diferente, sem ID em comum, casamento é por nome+data — ver
    # _statarea_lookup). Sem essa coluna só dava pra recuperar o nome do
    # adversário quando o time da seção era o visitante daquele jogo.
    try:
        conn.execute("ALTER TABLE btcs_pattern_rows ADD COLUMN away_name TEXT")
    except sqlite3.OperationalError:
        pass  # coluna já existe
    return conn


def _btcs_persist_results(rows):
    """Grava [(event_id, bucket, match_date), ...] via INSERT OR IGNORE — chave
    é só event_id, então o mesmo jogo histórico reaparecendo em janelas futuras
    de 'últimos 30' não é contado duas vezes no acumulado."""
    if not rows:
        return
    now = datetime.utcnow().isoformat() + "Z"
    with _bt2_db_lock:
        conn = _btcs_db_conn()
        try:
            conn.executemany(
                """INSERT OR IGNORE INTO btcs_results
                   (event_id, bucket, match_date, inserted_at)
                   VALUES (?,?,?,?)""",
                [(eid, bucket, date, now) for (eid, bucket, date) in rows],
            )
            conn.commit()
        finally:
            conn.close()
    _bt2_push_db_bg()


def _btcs_persist_pattern_rows(rows):
    """Grava [(event_id, team_key, team_name, h, a, home_name, away_name, match_date, odd, opp_odd), ...]
    via upsert — chave é (event_id, team_key), então o mesmo jogo reaparecendo em
    janelas futuras de 'últimos 30' não duplica pro mesmo time. Em conflito,
    COALESCE mantém odd/opp_odd/away_name já salvos se a nova rodada vier vazia
    nesses campos, mas PREENCHE se a linha antiga não tinha esse dado ainda —
    sem isso, linhas salvas antes de um campo existir ficavam pra sempre sem
    esse dado (INSERT OR IGNORE nunca atualiza linha já existente)."""
    if not rows:
        return
    now = datetime.utcnow().isoformat() + "Z"
    with _bt2_db_lock:
        conn = _btcs_db_conn()
        try:
            conn.executemany(
                """INSERT INTO btcs_pattern_rows
                   (event_id, team_key, team_name, h, a, home_name, away_name, match_date, inserted_at, odd, opp_odd)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?)
                   ON CONFLICT(event_id, team_key) DO UPDATE SET
                     odd = COALESCE(btcs_pattern_rows.odd, excluded.odd),
                     opp_odd = COALESCE(btcs_pattern_rows.opp_odd, excluded.opp_odd),
                     away_name = COALESCE(btcs_pattern_rows.away_name, excluded.away_name)""",
                [(eid, tk, tn, h, a, home, away, date, now, odd, opp_odd)
                 for (eid, tk, tn, h, a, home, away, date, odd, opp_odd) in rows],
            )
            conn.commit()
        finally:
            conn.close()
    _bt2_push_db_bg()


def _btcs_build_methodologies(bucket_dates):
    """A partir de {bucket: [dates ordenáveis...]} (uma entrada por acerto,
    já na ordem de ocorrência) e do total_sample, monta a lista final ordenada
    por winrate desc, com timeline de hit count acumulado."""
    total_sample = sum(len(v) for v in bucket_dates.values())
    methodologies = []
    if total_sample >= _BTCS_MIN_TOTAL_SAMPLE:
        for bucket in PLACAR_EXATO_BUCKETS_ORDER:
            dates = sorted(bucket_dates.get(bucket, []), key=lambda d: d or "")
            hits = len(dates)
            if hits < _BTCS_MIN_HITS:
                continue
            cum = 0
            timeline = []
            for date in dates:
                cum += 1
                timeline.append({"date": date, "cumulative_hits": cum})
            methodologies.append({
                "bucket": bucket,
                "bets": total_sample,
                "wins": hits,
                "winrate": round(hits / total_sample, 4),
                "timeline": timeline,
            })
    methodologies.sort(key=lambda x: x["winrate"], reverse=True)
    return total_sample, methodologies


def _btcs_compute_ranking():
    """Coleta o histórico de todos os jogos de hoje (mesma lógica de
    _bt2_compute_ranking, sem precisar de odds) e rankeia os 19 buckets de
    placar exato por taxa de acerto."""
    today_matches = _fs_all_matches()

    historical_by_id = {}
    for m in today_matches:
        if len(historical_by_id) >= _BTCS_MAX_HISTORICAL_IDS:
            break
        try:
            tabs = _fs_h2h(m["id"])
        except Exception:
            continue
        for team_home in (True, False):
            for row in _bt2_team_rows_from_section(tabs, team_home):
                if not row["id"] or row["id"] in historical_by_id:
                    continue
                historical_by_id[row["id"]] = row
                if len(historical_by_id) >= _BTCS_MAX_HISTORICAL_IDS:
                    break

    bucket_dates = {}
    persist_rows = []
    for event_id, row in historical_by_id.items():
        bucket = _btcs_bucket_key(row["h"], row["a"])
        bucket_dates.setdefault(bucket, []).append(row.get("date") or "")
        persist_rows.append((str(event_id), bucket, row.get("date") or ""))

    try:
        _btcs_persist_results(persist_rows)
    except Exception:
        pass

    total_sample, methodologies = _btcs_build_methodologies(bucket_dates)

    return {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "sample_matches_used": total_sample,
        "methodologies": methodologies,
    }


def _btcs_compute_ranking_acumulado():
    """Ranking acumulado: lê TODAS as linhas já persistidas em btcs_results
    (dedupadas por event_id) e reaplica a mesma agregação por bucket."""
    with _bt2_db_lock:
        conn = _btcs_db_conn()
        try:
            cur = conn.execute("SELECT bucket, match_date FROM btcs_results ORDER BY match_date ASC")
            all_rows = cur.fetchall()
        finally:
            conn.close()

    bucket_dates = {}
    for bucket, match_date in all_rows:
        bucket_dates.setdefault(bucket, []).append(match_date or "")

    total_sample, methodologies = _btcs_build_methodologies(bucket_dates)

    return {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "sample_matches_used": total_sample,
        "methodologies": methodologies,
    }


# ── BACKUP DIÁRIO DO BACKTEST (Tradicional + CS) ────────────────────────────
# A aba Backtest foi removida do frontend (usuário não usa mais por enquanto,
# mas quer voltar quando tiver mais volume de dados histórico) — sem a aba,
# ninguém mais dispara o cálculo caro (_bt2_compute_ranking/_btcs_compute_ranking,
# que são os únicos jeitos de bt2_bets/btcs_results crescerem). Essa thread roda
# esse cálculo sozinha 1x por dia, sem depender de ninguém clicar em nada, e sobe
# o backtest2.db pro GitHub — mesmo padrão de persistência já usado pelo resto
# do app (github_storage), pra sobreviver a reinícios/redeploys do Railway.
def _background_backtest_daily_backup():
    while True:
        try:
            print("[backtest-backup] Rodando simulação diária do Backtest (Tradicional + CS)...")
            _bt2_compute_ranking()
            _btcs_compute_ranking()
            # _btcs_compute_patterns() é a ÚNICA função que popula btcs_pattern_rows
            # (usada pela aba Padrões do Placar Exato) — sem chamar ela aqui, essa
            # tabela nunca crescia: o frontend só bate em mode=acumulado (leitura
            # pura, de propósito, pra não travar o worker) e mais ninguém disparava
            # o modo caro que persiste as linhas. Resultado: "amostra: 0 times, 0
            # jogos" pra sempre, mesmo com Backtest Tradicional e Placar Exato
            # (ranking) acumulando normalmente.
            _btcs_compute_patterns()
            github_storage.push_file_bg(_BT2_DB_PATH, "backtest2.db")
            print("[backtest-backup] Concluído — backtest2.db atualizado e enviado pro GitHub.")
        except Exception as e:
            print(f"[backtest-backup] Erro: {e}")
        time.sleep(24 * 3600)


threading.Thread(target=_background_backtest_daily_backup, daemon=True, name="BacktestDailyBackup").start()


# ── BACKTEST CS — PADRÕES: em que condições dos times a taxa de acerto de um
# bucket de placar exato melhora vs a taxa geral (baseline) ─────────────────
# Reusa a MESMA coleta de histórico do ranking, mas mantendo o agrupamento POR
# TIME (o ranking achata tudo num dict global de event_id -> row; aqui
# precisamos saber a QUEM cada jogo pertence pra calcular médias/percentuais
# por time e então segmentar). Não busca odds — só o H2H já dá tudo que
# precisa, por isso é rápido como o ranking.
_btcs_patterns_cache = {"ts": 0, "data": None}

_BTCS_PATTERN_VARIABLES = {
    "avg_scored": "Média de gols marcados",
    "avg_conceded": "Média de gols sofridos",
    "pct_scored": "Chance de marcar gol",
    "pct_conceded": "Chance de sofrer gol",
    "pct_over25": "Over/Under 2.5",
    "pct_ambas_marcam": "Ambas Marcam",
    "saldo": "Saldo de Gols",
    "mandante": "Mandante",
    "odd": "Odd 1X2 (força do time)",
    # Row-level, só contam nos jogos com odd PRÓPRIA e do ADVERSÁRIO conhecidas
    # (opp_odd) — mesmas fórmulas da aba Jogo (Aulas 09/10 do curso: pondera o
    # resultado bruto pela força do adversário, medida pela odd dele).
    "valor_ponto": "Valor do Ponto",
    "valor_gol": "Valor do Gol",
    "valor_saldo": "Valor do Saldo",
    "custo_gol2": "Custo do Gol 2.0",
    # Row-level, só conta nos jogos onde achamos o par no backtest/ do StatArea
    # (casamento por time+data, fonte diferente do FlashScore — ver
    # _statarea_lookup) — probabilidade que o PRÓPRIO StatArea deu pro time
    # daquele jogo vencer, testando se a previsão deles é sinal de verdade
    # pro placar exato ou não.
    "previsao_statarea": "Previsão StatArea (vitória do time)",
}

_BTCS_SALDO_RANGES = [(-0.3, "<-0.3"), (0.3, "-0.3–0.3"), (None, "≥0.3")]
_BTCS_VALOR_PONTO_RANGES = [(0.3, "<0.3"), (0.6, "0.3–0.6"), (1.0, "0.6–1.0"), (None, "≥1.0")]
_BTCS_VALOR_GOL_RANGES = [(0.15, "<0.15"), (0.3, "0.15–0.3"), (0.5, "0.3–0.5"), (None, "≥0.5")]
_BTCS_VALOR_SALDO_RANGES = [(-0.05, "<-0.05"), (0.05, "-0.05–0.05"), (None, "≥0.05")]
_BTCS_CUSTO_GOL2_RANGES = [(0.5, "<0.5"), (0.7, "0.5–0.7"), (0.9, "0.7–0.9"), (None, "≥0.9")]

_BTCS_PATTERN_MIN_SEGMENT_GAMES = 30
_BTCS_PATTERN_MIN_SEGMENT_HITS = 3

# Pares de variáveis combinadas (ex: "média de gols marcados" + "mandante") têm
# amostra naturalmente menor que uma variável isolada — exige menos jogos/acertos
# pra não descartar tudo, mas ainda o suficiente pra não virar coincidência.
_BTCS_COMBO_MIN_SEGMENT_GAMES = 20
_BTCS_COMBO_MIN_SEGMENT_HITS = 2
_BTCS_PATTERN_MIN_LIFT = 1.3
_BTCS_PATTERN_TOP_N = 30

# "Padrões que reduzem a chance" — o oposto: condições em que um placar fica
# AINDA MENOS provável que o normal (útil pra apostar CONTRA aquele placar
# com mais confiança). Aqui não exigimos um mínimo de acertos no segmento
# (0 acertos é o caso ideal), só amostra suficiente pra confiar no número.
_BTCS_PATTERN_MAX_LIFT_AVOID = 0.7
_BTCS_PATTERN_AVOID_TOP_N = 30

# Faixas de segmentação: lista de (limite_superior_exclusivo, label), em ordem
# crescente, o último item deve ter limite None (>= o penúltimo limite).
_BTCS_AVG_RANGES = [(1.5, "<1.5 gols/jogo"), (2.5, "1.5–2.5 gols/jogo"),
                     (3.5, "2.5–3.5 gols/jogo"), (None, "≥3.5 gols/jogo")]
_BTCS_PCT_RANGES = [(0.70, "<70%"), (0.90, "70–90%"), (None, "≥90%")]
_BTCS_OVER_RANGES = [(0.40, "<40%"), (0.60, "40–60%"), (None, "≥60%")]
# Diferente de _BTCS_PCT_RANGES (usada pra taxas medidas sobre 30 jogos, que
# tendem a ficar altas) — probabilidade de VITÓRIA pra um jogo só raramente
# passa de 70%, então as faixas são mais granuladas na base.
_BTCS_STATAREA_PROB_RANGES = [(0.30, "<30%"), (0.50, "30–50%"), (0.70, "50–70%"), (None, "≥70%")]


def _btcs_norm_name(s):
    """Normaliza nome de time pra comparação (minúsculo, sem acento, só
    alfanumérico+espaço) — mesma técnica do norm() interno de _fs_find_match."""
    import unicodedata
    s = (s or "").lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


# ── Cruzamento com backtest/ do StatArea — fonte independente do FlashScore,
# sem ID em comum, então o casamento é por time+data (fuzzy, mesma técnica do
# _name_match já usado em momentum_history↔forca_history). Quando não acha um
# par confiável (nome muito diferente, StatArea não cobriu aquele jogo etc.),
# retorna None — não força casamento errado, só deixa aquele jogo de fora da
# variável nova (ver [[reference_trading_metodologias_esportivas]] pro mesmo
# princípio aplicado antes).
_statarea_backtest_index_cache = {"ts": 0, "data": None}
_STATAREA_INDEX_TTL = 6 * 3600  # 6h — só ~100 arquivos, recarregar é barato, mas não precisa toda hora


def _statarea_backtest_index():
    """Carrega todo backtest/*.json (previsões diárias do StatArea, já com
    resultado real anexado) num índice {data_partida: [linhas]} pra consulta
    rápida em memória."""
    now = time.time()
    cache = _statarea_backtest_index_cache
    if cache["data"] is not None and (now - cache["ts"]) < _STATAREA_INDEX_TTL:
        return cache["data"]
    index = {}
    for path in glob.glob(os.path.join(BACKTEST_DIR, "*.json")):
        try:
            with open(path, encoding="utf-8") as f:
                rows = json.load(f)
        except Exception:
            continue
        if not isinstance(rows, list):
            continue
        for r in rows:
            date = r.get("data_partida")
            if not date:
                continue
            index.setdefault(date, []).append(r)
    cache["ts"] = now
    cache["data"] = index
    return index


def _statarea_lookup(ts, home_name, away_name):
    """Acha a linha do backtest/ (StatArea) correspondente a um jogo histórico
    do FlashScore, dado seu timestamp (campo KC, unix UTC) e os dois nomes de
    time. Tenta a data exata +-1 dia (KC é UTC; o StatArea grava a data local
    do jogo, então partidas perto da meia-noite podem cair no dia seguinte/
    anterior). Retorna a linha (dict com tip/odds_1/odds_x/odds_2/etc) ou None."""
    if not ts or not home_name or not away_name:
        return None
    try:
        base = datetime.utcfromtimestamp(int(ts))
    except (TypeError, ValueError, OSError):
        return None
    index = _statarea_backtest_index()
    for delta in (0, -1, 1):
        date_str = (base + timedelta(days=delta)).strftime("%Y-%m-%d")
        for cand in index.get(date_str, []):
            if _name_match(home_name, cand.get("casa", "")) and _name_match(away_name, cand.get("fora", "")):
                return cand
    return None


def _statarea_own_win_prob(statarea_row, is_home):
    """Probabilidade (0-1) que o StatArea deu pra vitória do time da
    perspectiva atual (odds_1 se ele jogou em casa, odds_2 se jogou fora) —
    apesar do nome do campo ('odds_1'), é uma probabilidade em % que o
    StatArea calcula, não uma odd decimal de casa de aposta."""
    if not statarea_row:
        return None
    val = statarea_row.get("odds_1" if is_home else "odds_2")
    if val is None:
        return None
    try:
        return float(val) / 100.0
    except (TypeError, ValueError):
        return None


def _btcs_segment_range(value, ranges):
    for hi, label in ranges:
        if hi is None or value < hi:
            return label
    return ranges[-1][1]


def _btcs_build_patterns_from_teams(teams):
    """Núcleo compartilhado entre o modo 'últimos 30' e o 'acumulado': recebe
    {team_key: {"name":..., "rows":[{id,h,a,date,home}, ...]}} já montado (de
    onde vier — H2H de hoje ou SQLite acumulado) e calcula variáveis team-level
    (média de gols marcados/sofridos, chance de marcar/sofrer, over/under 2.5)
    + uma variável row-level (mandante), medindo o LIFT da taxa de acerto de
    cada bucket de placar exato dentro de cada segmento vs a taxa geral."""
    # Baseline: taxa de acerto geral de cada bucket sobre TODOS os jogos
    # coletados (mesma amostra usada pra formar os segmentos abaixo).
    baseline_bucket_counts = {}
    total_games = 0
    for entry in teams.values():
        for row in entry["rows"]:
            bucket = _btcs_bucket_key(row["h"], row["a"])
            baseline_bucket_counts[bucket] = baseline_bucket_counts.get(bucket, 0) + 1
            total_games += 1
    baseline_rate = {b: (c / total_games if total_games else 0.0) for b, c in baseline_bucket_counts.items()}

    # {variable: {segment: {bucket: hits, "_total": n_jogos_no_segmento}}}
    seg_data = {v: {} for v in _BTCS_PATTERN_VARIABLES}
    # {(var1,var2): {"segA + segB": {bucket: hits, "_total": n}}} — pares de
    # variáveis combinadas (ex: "2.5-3.5 gols/jogo" + "Jogando em casa")
    combo_seg_data = {}

    def _accum(variable, segment, bucket):
        seg = seg_data[variable].setdefault(segment, {"_total": 0})
        seg["_total"] += 1
        seg[bucket] = seg.get(bucket, 0) + 1

    def _accum_combo(var_combo, label_combo, bucket):
        seg = combo_seg_data.setdefault(var_combo, {})
        combo_label = " + ".join(label_combo)
        entry = seg.setdefault(combo_label, {"_total": 0})
        entry["_total"] += 1
        entry[bucket] = entry.get(bucket, 0) + 1

    _TEAM_LEVEL_VARS = ("avg_scored", "avg_conceded", "pct_scored", "pct_conceded", "pct_over25",
                         "pct_ambas_marcam", "saldo")
    # valor_ponto/valor_gol/valor_saldo/custo_gol2 são row-level como "odd" (só
    # contam nos jogos com odd própria E do adversário conhecidas — ver
    # _btcs_row_valor_segs). previsao_statarea também é row-level, só conta nos
    # jogos em que achamos o par no backtest/ do StatArea (ver _statarea_lookup).
    _ALL_PATTERN_VARS = _TEAM_LEVEL_VARS + ("mandante", "odd", "valor_ponto", "valor_gol", "valor_saldo",
                                             "custo_gol2", "previsao_statarea")
    # Combinações de 2 a 3 variáveis ao mesmo tempo (ex: média de gols marcados
    # + chance de sofrer gol + mandante). Com 12 variáveis disponíveis agora,
    # combos de tamanho 4 explodiriam o tempo de cálculo (C(12,4) = 495 vs
    # C(12,3) = 220) sem ganho relevante de sinal — cortado por segurança de
    # performance (mesma lição aprendida com o travamento do fetch de odds do
    # Mapa de Sugestões).
    _COMBO_SIZES = (2, 3)
    _VAR_COMBOS = [c for size in _COMBO_SIZES for c in itertools.combinations(_ALL_PATTERN_VARS, size)]

    def _btcs_row_valor_segs(own, opp, own_odd, opp_odd):
        """Segmentos das 4 variáveis dependentes de odd, ou None se faltar
        odd própria ou do adversário nesse jogo específico."""
        if not own_odd or not opp_odd:
            return None
        own_prob = 1.0 / own_odd
        opp_prob = 1.0 / opp_odd
        pontos = 3 if own > opp else (1 if own == opp else 0)
        saldo_row = own - opp
        return {
            "valor_ponto": _btcs_segment_range(pontos * opp_prob, _BTCS_VALOR_PONTO_RANGES),
            "valor_gol": _btcs_segment_range(own * opp_prob, _BTCS_VALOR_GOL_RANGES),
            "valor_saldo": _btcs_segment_range(saldo_row * opp_prob, _BTCS_VALOR_SALDO_RANGES),
            "custo_gol2": _btcs_segment_range((own / 2) + (own_prob / 2), _BTCS_CUSTO_GOL2_RANGES),
        }

    for entry in teams.values():
        rows = entry["rows"]
        n = len(rows)
        if n == 0:
            continue
        team_key = _btcs_norm_name(entry["name"])
        scored_total = conceded_total = 0
        games_scored = games_conceded = games_over25 = games_ambas = 0
        row_is_home = []
        for row in rows:
            row_home_key = _btcs_norm_name(row.get("home"))
            # Se por algum motivo o nome não bate com nenhum lado (dado
            # incompleto), assume casa como padrão conservador.
            is_home = (row_home_key == team_key) if row_home_key and team_key else True
            own = row["h"] if is_home else row["a"]
            opp = row["a"] if is_home else row["h"]
            scored_total += own
            conceded_total += opp
            if own >= 1:
                games_scored += 1
            if opp >= 1:
                games_conceded += 1
            if row["h"] + row["a"] >= 3:
                games_over25 += 1
            if own >= 1 and opp >= 1:
                games_ambas += 1
            row_is_home.append(is_home)

        team_segs = {
            "avg_scored": _btcs_segment_range(scored_total / n, _BTCS_AVG_RANGES),
            "avg_conceded": _btcs_segment_range(conceded_total / n, _BTCS_AVG_RANGES),
            "pct_scored": _btcs_segment_range(games_scored / n, _BTCS_PCT_RANGES),
            "pct_conceded": _btcs_segment_range(games_conceded / n, _BTCS_PCT_RANGES),
            "pct_over25": _btcs_segment_range(games_over25 / n, _BTCS_OVER_RANGES),
            "pct_ambas_marcam": _btcs_segment_range(games_ambas / n, _BTCS_PCT_RANGES),
            "saldo": _btcs_segment_range((scored_total - conceded_total) / n, _BTCS_SALDO_RANGES),
        }

        for row in rows:
            bucket = _btcs_bucket_key(row["h"], row["a"])
            for v in _TEAM_LEVEL_VARS:
                _accum(v, team_segs[v], bucket)

        # Mandante, Odd e as 4 variáveis de "valor" são row-level: os jogos de
        # um time podem se dividir entre segmentos diferentes jogo a jogo,
        # diferente das variáveis acima que atribuem o time inteiro a UM
        # segmento. Monta o perfil row-level UMA VEZ e reaproveita tanto pro
        # acúmulo individual quanto pras combinações abaixo.
        row_valor_segs = []
        row_statarea_segs = []
        for row, is_home in zip(rows, row_is_home):
            bucket = _btcs_bucket_key(row["h"], row["a"])
            segment = "Jogando em casa" if is_home else "Jogando fora"
            _accum("mandante", segment, bucket)
            odd_label = _bt2_odd_range_label(row.get("odd"))
            if odd_label:
                _accum("odd", odd_label, bucket)
            own = row["h"] if is_home else row["a"]
            opp = row["a"] if is_home else row["h"]
            valor_segs = _btcs_row_valor_segs(own, opp, row.get("odd"), row.get("opp_odd"))
            if valor_segs:
                for v, seg in valor_segs.items():
                    _accum(v, seg, bucket)
            row_valor_segs.append(valor_segs)
            sa_row = _statarea_lookup(row.get("date"), row.get("home"), row.get("away"))
            sa_prob = _statarea_own_win_prob(sa_row, is_home)
            sa_label = _btcs_segment_range(sa_prob, _BTCS_STATAREA_PROB_RANGES) if sa_prob is not None else None
            if sa_label:
                _accum("previsao_statarea", sa_label, bucket)
            row_statarea_segs.append(sa_label)

        # Combinações de 2-3 variáveis: monta o "perfil" completo de CADA jogo
        # (segmento de todas as variáveis team-level, que são as mesmas em
        # todos os jogos do time, + mandante/odd/valor/previsão StatArea
        # daquele jogo específico) e acumula em cada combinação de variáveis
        # que existir. Jogos sem odd (própria ou do adversário) ou sem par
        # achado no backtest/ do StatArea só ficam de fora das combinações que
        # incluem essas variáveis.
        for row, is_home, valor_segs, sa_label in zip(rows, row_is_home, row_valor_segs, row_statarea_segs):
            bucket = _btcs_bucket_key(row["h"], row["a"])
            row_segs = dict(team_segs)
            row_segs["mandante"] = "Jogando em casa" if is_home else "Jogando fora"
            odd_label = _bt2_odd_range_label(row.get("odd"))
            if odd_label:
                row_segs["odd"] = odd_label
            if valor_segs:
                row_segs.update(valor_segs)
            if sa_label:
                row_segs["previsao_statarea"] = sa_label
            for var_combo in _VAR_COMBOS:
                if any(v not in row_segs for v in var_combo):
                    continue
                _accum_combo(var_combo, [row_segs[v] for v in var_combo], bucket)

    patterns = []
    patterns_avoid = []
    for variable, segments in seg_data.items():
        variable_label = _BTCS_PATTERN_VARIABLES[variable]
        for segment, counts in segments.items():
            segment_games = counts["_total"]
            if segment_games < _BTCS_PATTERN_MIN_SEGMENT_GAMES:
                continue
            # Buckets que nunca acertaram nesse segmento não aparecem em
            # `counts` (só acumulamos quando há hit) — pra achar os padrões
            # de "reduz a chance" precisamos considerar TODOS os 19 buckets,
            # inclusive os com 0 acertos no segmento.
            for bucket in PLACAR_EXATO_BUCKETS_ORDER:
                hits = counts.get(bucket, 0)
                base = baseline_rate.get(bucket, 0.0)
                if base <= 0:
                    continue  # sem baseline confiável pra comparar, pula
                segment_rate = hits / segment_games
                lift = segment_rate / base
                row = {
                    "variable": variable,
                    "variable_label": variable_label,
                    "segment": segment,
                    "segment_label": segment,
                    "bucket": bucket,
                    "segment_games": segment_games,
                    "segment_hits": hits,
                    "segment_rate": round(segment_rate, 4),
                    "baseline_rate": round(base, 4),
                    "lift": round(lift, 3),
                    "combined": False,
                }
                if hits >= _BTCS_PATTERN_MIN_SEGMENT_HITS and lift >= _BTCS_PATTERN_MIN_LIFT:
                    patterns.append(row)
                elif lift <= _BTCS_PATTERN_MAX_LIFT_AVOID:
                    patterns_avoid.append(row)

    # Mesma lógica acima, mas pros pares de variáveis combinadas — limiares
    # de amostra/acertos mais baixos (_BTCS_COMBO_*) porque cada combinação
    # naturalmente reparte a amostra em fatias menores.
    for var_combo, segments in combo_seg_data.items():
        variable_label = " + ".join(_BTCS_PATTERN_VARIABLES[v] for v in var_combo)
        for combo_label, counts in segments.items():
            segment_games = counts["_total"]
            if segment_games < _BTCS_COMBO_MIN_SEGMENT_GAMES:
                continue
            for bucket in PLACAR_EXATO_BUCKETS_ORDER:
                hits = counts.get(bucket, 0)
                base = baseline_rate.get(bucket, 0.0)
                if base <= 0:
                    continue
                segment_rate = hits / segment_games
                lift = segment_rate / base
                row = {
                    "variable": "+".join(var_combo),
                    "variable_label": variable_label,
                    "segment": combo_label,
                    "segment_label": combo_label,
                    "bucket": bucket,
                    "segment_games": segment_games,
                    "segment_hits": hits,
                    "segment_rate": round(segment_rate, 4),
                    "baseline_rate": round(base, 4),
                    "lift": round(lift, 3),
                    "combined": True,
                    "combo_size": len(var_combo),
                }
                if hits >= _BTCS_COMBO_MIN_SEGMENT_HITS and lift >= _BTCS_PATTERN_MIN_LIFT:
                    patterns.append(row)
                elif lift <= _BTCS_PATTERN_MAX_LIFT_AVOID:
                    patterns_avoid.append(row)

    # "Aumenta a chance": ordena por lift desc; segment_hits como desempate
    # (entre lifts iguais/muito próximos, prefere o padrão com mais evidência).
    patterns.sort(key=lambda p: (p["lift"], p["segment_hits"]), reverse=True)
    patterns = patterns[:_BTCS_PATTERN_TOP_N]

    # "Reduz a chance": ordena por lift ASC (quanto mais perto de 0, mais forte
    # o sinal de que aquele placar não vai sair), segment_games como desempate
    # (entre lifts iguais, prefere o padrão com amostra maior/mais confiável).
    patterns_avoid.sort(key=lambda p: (p["lift"], -p["segment_games"]))

    # Melhor padrão (menor lift) de CADA um dos 19 placares, calculado sobre a
    # lista COMPLETA (antes do corte de top-30 abaixo) — assim cobre todo
    # placar que tiver algum padrão qualificado, não só os que entram no top-30.
    best_avoid_by_bucket = {}
    for p in patterns_avoid:
        cur = best_avoid_by_bucket.get(p["bucket"])
        if cur is None or p["lift"] < cur["lift"]:
            best_avoid_by_bucket[p["bucket"]] = p
    patterns_avoid_best_by_bucket = sorted(best_avoid_by_bucket.values(), key=lambda p: p["lift"])

    patterns_avoid = patterns_avoid[:_BTCS_PATTERN_AVOID_TOP_N]

    return {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "sample_teams_used": len(teams),
        "sample_matches_used": total_games,
        "patterns": patterns,
        "patterns_avoid": patterns_avoid,
        "patterns_avoid_best_by_bucket": patterns_avoid_best_by_bucket,
    }


def _btcs_compute_patterns():
    """Coleta o histórico dos times de hoje (últimos 30 jogos por time,
    mesma fonte do H2H) mantendo o agrupamento POR TIME, monta os padrões via
    _btcs_build_patterns_from_teams, e persiste cada linha (jogo+time) no
    SQLite como efeito colateral — é isso que alimenta o modo 'acumulado'."""
    today_matches = _fs_all_matches()

    teams = {}  # norm(nome) -> {"name": nome original, "rows": [...]}
    all_ids_seen = set()
    total_ids = 0
    for m in today_matches:
        if total_ids >= _BTCS_MAX_HISTORICAL_IDS:
            break
        try:
            tabs = _fs_h2h(m["id"])
        except Exception:
            continue
        for team_home in (True, False):
            if total_ids >= _BTCS_MAX_HISTORICAL_IDS:
                break
            team_name = m["home"] if team_home else m["away"]
            key = _btcs_norm_name(team_name)
            if not key:
                continue
            rows = _bt2_team_rows_from_section(tabs, team_home)
            if not rows:
                continue
            entry = teams.setdefault(key, {"name": team_name, "rows": []})
            existing_ids = {r["id"] for r in entry["rows"]}
            for row in rows:
                if not row["id"] or row["id"] in existing_ids or row["id"] in all_ids_seen:
                    continue
                entry["rows"].append(row)
                existing_ids.add(row["id"])
                all_ids_seen.add(row["id"])
                total_ids += 1
                if total_ids >= _BTCS_MAX_HISTORICAL_IDS:
                    break

    # Busca a odd 1X2 de cada jogo histórico único (mesmo pool/endpoint já usado
    # no Backtest Tradicional) e anexa a odd do PRÓPRIO time (Casa se ele jogou em
    # casa naquele jogo, Fora se jogou fora) em cada linha — alimenta a variável
    # de padrão "odd" (força do time conforme o mercado precificava). Só pede o
    # mercado "1x2" (markets_wanted) — é o único que essa função usa, e pedir os
    # outros 5 à toa (o padrão de _fs_odds_all_markets_any_bookmaker) foi o que
    # deixava isso ~700s pra 400 jogos: medido e confirmado que o gargalo real
    # era aqui, não no fetch de H2H (que sozinho leva só alguns segundos).
    def _fetch_odds_btcs(event_id):
        try:
            return event_id, _fs_odds_all_markets_any_bookmaker(event_id, markets_wanted=["1x2"])
        except Exception:
            return event_id, (None, {})

    odds_by_id = {}
    for event_id, (bookmaker_name, markets) in _fs_event_pool.map(_fetch_odds_btcs, list(all_ids_seen)):
        if markets:
            odds_by_id[event_id] = markets

    for key, entry in teams.items():
        for row in entry["rows"]:
            row_home_key = _btcs_norm_name(row.get("home"))
            is_home = (row_home_key == key) if row_home_key and key else True
            markets = odds_by_id.get(row["id"])
            row["odd"] = None
            row["opp_odd"] = None
            if markets:
                sels = _bt2_market_selections("1x2", markets.get("1x2"), row["h"], row["a"])
                target_label = "Casa" if is_home else "Fora"
                opp_label = "Fora" if is_home else "Casa"
                for sel in sels:
                    if sel["label"] == target_label:
                        row["odd"] = sel["odd"]
                    elif sel["label"] == opp_label:
                        row["opp_odd"] = sel["odd"]

    persist_rows = [
        (row["id"], key, entry["name"], row["h"], row["a"], row.get("home"), row.get("away"),
         row.get("date"), row.get("odd"), row.get("opp_odd"))
        for key, entry in teams.items()
        for row in entry["rows"]
    ]
    try:
        _btcs_persist_pattern_rows(persist_rows)
    except Exception:
        pass

    return _btcs_build_patterns_from_teams(teams)


def _btcs_compute_patterns_acumulado():
    """Mesma lógica de _btcs_compute_patterns, mas monta 'teams' a partir de
    TODO o histórico já persistido em btcs_pattern_rows (não só os times de
    hoje nem os últimos 30 jogos) — cresce a cada vez que o modo 'últimos 30'
    roda. Sem cache — leitura no SQLite é barata."""
    with _bt2_db_lock:
        conn = _btcs_db_conn()
        try:
            cur = conn.execute(
                "SELECT event_id, team_key, team_name, h, a, home_name, away_name, match_date, odd, opp_odd FROM btcs_pattern_rows"
            )
            all_rows = cur.fetchall()
        finally:
            conn.close()

    teams = {}
    for event_id, team_key, team_name, h, a, home_name, away_name, match_date, odd, opp_odd in all_rows:
        entry = teams.setdefault(team_key, {"name": team_name, "rows": []})
        entry["rows"].append({"id": event_id, "h": h, "a": a, "home": home_name, "away": away_name,
                               "date": match_date, "odd": odd, "opp_odd": opp_odd})

    return _btcs_build_patterns_from_teams(teams)


# ── MAPA DE SUGESTÕES — motor genérico de metodologias — pra cada metodologia
# (Lay Casa, Lay Visitante, Casa Vence, Over/Under, Ambas Marcam, Lay Placar X,
# etc.) acha DINAMICAMENTE qual variável (das mesmas usadas na aba Jogo +
# Backtest CS: média/chance de gols, over/under, ambas marcam, saldo) melhor
# prediz o evento-alvo daquela metodologia, testando cada uma isoladamente
# contra o histórico do time relevante (mandante jogando em casa, ou
# visitante jogando fora) — sem limiar fixo escolhido por mim, o sistema
# escolhe sozinho a variável+faixa com maior lift e amostra suficiente, e
# aplica esse padrão validado às partidas de hoje pra gerar as sugestões.
# Cada metodologia tem seu próprio padrão (congelado 1x por dia) e sua
# própria lista de sugestões acumulada — ver notas em _mapa_compute().
_LAYCASA_MIN_SEGMENT_GAMES = 30
_LAYCASA_MIN_LIFT = 1.15
_LAYCASA_TTL = 30 * 60
_LAYCASA_SALDO_RANGES = [(-0.3, "<-0.3"), (0.3, "-0.3–0.3"), (None, "≥0.3")]

# cache de dados (30min) e cache do padrão do dia — ambos keyed por metodologia
_MAPA_CACHE = {}      # metodologia -> {"ts":..., "data":...}
_MAPA_DAY_CACHE = {}  # metodologia -> {"date":..., "melhor":..., "comparativo":..., "sugestoes_by_id":...}


def _mapa_cache_path(metodologia_key):
    return os.path.join(MAPA_CACHE_DIR, f"{metodologia_key}.json")


def _mapa_load_day_cache_from_disk(metodologia_key, today_str):
    """Tenta restaurar o padrão do dia + lista de sugestões travada, salvos por
    uma execução anterior (sobrevive a redeploy/reinício, que apaga a memória).
    Só usa o arquivo se for do dia de hoje — de outra forma, ignora (vira o dia,
    escolhe um padrão novo)."""
    path = _mapa_cache_path(metodologia_key)
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if data.get("date") != today_str:
            return None
        return data
    except Exception as e:
        print(f"[mapa] Erro lendo cache do disco ({metodologia_key}): {e}")
        return None


def _mapa_save_day_cache_to_disk(metodologia_key, day_cache):
    path = _mapa_cache_path(metodologia_key)
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(day_cache, f, ensure_ascii=False)
        github_storage.push_file_bg(path, f"mapa_cache/{metodologia_key}.json")
    except Exception as e:
        print(f"[mapa] Erro salvando cache no disco ({metodologia_key}): {e}")


def _mapa_slug(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "_", s.lower()).strip("_")


def _mapa_variable_labels(side):
    suf = "em casa" if side == "casa" else "fora"
    return {
        "avg_scored": f"Média de gols marcados ({suf})",
        "avg_conceded": f"Média de gols sofridos ({suf})",
        "pct_scored": f"Chance de marcar ({suf})",
        "pct_conceded": f"Chance de sofrer gol ({suf})",
        "pct_over25": f"Over/Under 2.5 ({suf})",
        "pct_ambas_marcam": f"Ambas Marcam ({suf})",
        "saldo": f"Saldo de Gols ({suf})",
    }


_MAPA_METODOLOGIAS = {}


def _mapa_add(key, label, side, target_fn, needs_ht=False):
    _MAPA_METODOLOGIAS[key] = {"label": label, "side": side, "target_fn": target_fn, "needs_ht": needs_ht}


_mapa_add("lay_casa", "Lay Casa (contra o mandante vencer)", "casa", lambda r: r["own"] <= r["opp"])
_mapa_add("lay_visitante", "Lay Visitante (contra o visitante vencer)", "visitante", lambda r: r["own"] <= r["opp"])
_mapa_add("casa_vence", "Casa Vence", "casa", lambda r: r["own"] > r["opp"])
_mapa_add("visitante_vence", "Visitante Vence", "visitante", lambda r: r["own"] > r["opp"])
_mapa_add("empate", "Empate", "casa", lambda r: r["own"] == r["opp"])
_mapa_add("over_15", "Over 1.5 FT", "casa", lambda r: (r["own"] + r["opp"]) >= 2)
_mapa_add("over_25", "Over 2.5 FT", "casa", lambda r: (r["own"] + r["opp"]) >= 3)
_mapa_add("under_15", "Under 1.5 FT", "casa", lambda r: (r["own"] + r["opp"]) <= 1)
_mapa_add("under_25", "Under 2.5 FT", "casa", lambda r: (r["own"] + r["opp"]) <= 2)
# HT precisa de um fetch extra por jogo histórico (placar do intervalo não vem
# junto do H2H) — o target_fn retorna None quando não achou esse dado, e
# _mapa_compute() ignora linhas com None nas contagens (ver needs_ht abaixo).
_mapa_add("over_05ht", "Over 0.5 HT", "casa",
          lambda r: ((r["own_ht"] + r["opp_ht"]) >= 1) if "own_ht" in r else None, needs_ht=True)
_mapa_add("over_15ht", "Over 1.5 HT", "casa",
          lambda r: ((r["own_ht"] + r["opp_ht"]) >= 2) if "own_ht" in r else None, needs_ht=True)
_mapa_add("ambas_sim", "Ambas Marcam Sim", "casa", lambda r: r["own"] >= 1 and r["opp"] >= 1)
_mapa_add("ambas_nao", "Ambas Marcam Não", "casa", lambda r: not (r["own"] >= 1 and r["opp"] >= 1))
for _bucket in PLACAR_EXATO_BUCKETS_ORDER:
    _mapa_add(f"lay_placar_{_mapa_slug(_bucket)}", f"Lay Placar {_bucket}", "casa",
              (lambda bucket: (lambda r: _btcs_bucket_key(r["own"], r["opp"]) != bucket))(_bucket))


def _laycasa_home_rows(tabs):
    """Jogos do mandante jogando EM CASA (aba 'Casa' do H2H, índice 1) — mesma
    lógica de _jogoStatsRowsFromSection(tabs, 1, true) no frontend (aba Jogo).
    'own'/'opp' = gols do próprio mandante / do adversário nesse jogo passado."""
    if not tabs or len(tabs) < 2:
        return []
    tab = tabs[1]
    sections = [s for s in tab.get("sections", []) if "confront" not in (s.get("title") or "").lower()]
    sec = sections[0] if sections else None
    if not sec:
        return []
    rows = []
    for r in sec["rows"][:30]:
        sc = _bt2_parse_score(r.get("score"))
        if not sc:
            continue
        rows.append({"id": r.get("id"), "own": sc[0], "opp": sc[1], "date": r.get("date"),
                     "home": r.get("home"), "away": r.get("away")})
    return rows


def _laycasa_away_rows(tabs):
    """Jogos do visitante jogando FORA (aba 'Fora' do H2H, índice 2) — mesma
    lógica de _jogoStatsRowsFromSection(tabs, 2, false) no frontend (aba Jogo).
    'own'/'opp' = gols do próprio visitante / do adversário (mandante) nesse
    jogo passado — por isso invertido em relação ao placar h:a literal."""
    if not tabs or len(tabs) < 3:
        return []
    tab = tabs[2]
    sections = [s for s in tab.get("sections", []) if "confront" not in (s.get("title") or "").lower()]
    sec = sections[0] if sections else None
    if not sec:
        return []
    rows = []
    for r in sec["rows"][:30]:
        sc = _bt2_parse_score(r.get("score"))
        if not sc:
            continue
        rows.append({"id": r.get("id"), "own": sc[1], "opp": sc[0], "date": r.get("date"),
                     "home": r.get("home"), "away": r.get("away")})
    return rows


def _laycasa_team_metrics(rows):
    """Calcula as variáveis descritivas (sem depender de odds) pra UM time, só
    com os jogos em que ele jogou no papel relevante (casa ou fora). Retorna
    (segments_dict, n) ou (None, 0) se não tiver amostra suficiente pra
    confiar na média."""
    n = len(rows)
    if n < 5:
        return None, 0
    scored = sum(r["own"] for r in rows)
    conceded = sum(r["opp"] for r in rows)
    games_scored = sum(1 for r in rows if r["own"] >= 1)
    games_conceded = sum(1 for r in rows if r["opp"] >= 1)
    games_over25 = sum(1 for r in rows if r["own"] + r["opp"] >= 3)
    ambas_marcaram = sum(1 for r in rows if r["own"] >= 1 and r["opp"] >= 1)

    segs = {
        "avg_scored": _btcs_segment_range(scored / n, _BTCS_AVG_RANGES),
        "avg_conceded": _btcs_segment_range(conceded / n, _BTCS_AVG_RANGES),
        "pct_scored": _btcs_segment_range(games_scored / n, _BTCS_PCT_RANGES),
        "pct_conceded": _btcs_segment_range(games_conceded / n, _BTCS_PCT_RANGES),
        "pct_over25": _btcs_segment_range(games_over25 / n, _BTCS_OVER_RANGES),
        "pct_ambas_marcam": _btcs_segment_range(ambas_marcaram / n, _BTCS_PCT_RANGES),
        "saldo": _btcs_segment_range((scored - conceded) / n, _LAYCASA_SALDO_RANGES),
    }
    return segs, n


def _mapa_compute(metodologia_key):
    cfg = _MAPA_METODOLOGIAS.get(metodologia_key)
    if not cfg:
        return None
    side = cfg["side"]
    target_fn = cfg["target_fn"]
    rows_fn = _laycasa_home_rows if side == "casa" else _laycasa_away_rows
    team_field = "home" if side == "casa" else "away"
    var_labels = _mapa_variable_labels(side)

    cache = _MAPA_CACHE.setdefault(metodologia_key, {"ts": 0, "data": None})
    now = time.time()
    if cache["data"] and (now - cache["ts"]) < _LAYCASA_TTL:
        return cache["data"]

    today_str = datetime.now().strftime("%Y-%m-%d")
    day_cache = _MAPA_DAY_CACHE.setdefault(metodologia_key, {"date": None, "melhor": None, "comparativo": None, "sugestoes_by_id": {}})
    if day_cache["date"] != today_str:
        # virou o dia (ou 1ª execução deste processo) — antes de zerar tudo,
        # tenta restaurar do disco o que uma execução anterior já tinha travado
        # hoje (sobrevive a redeploy/reinício do Railway, que apaga a memória —
        # sem isso, cada deploy escolhia um padrão novo do zero no meio do dia).
        restored = _mapa_load_day_cache_from_disk(metodologia_key, today_str)
        if restored:
            day_cache["date"] = today_str
            day_cache["melhor"] = restored.get("melhor")
            day_cache["comparativo"] = restored.get("comparativo")
            day_cache["sugestoes_by_id"] = restored.get("sugestoes_by_id") or {}
        else:
            # virou o dia (ou primeira execução) — libera escolher um padrão novo e
            # zera a lista acumulada de sugestões
            day_cache["date"] = today_str
            day_cache["melhor"] = None
            day_cache["comparativo"] = None
            day_cache["sugestoes_by_id"] = {}

    # só entra gente NOVA na lista de sugestões na hora em que o padrão do dia é
    # escolhido pela primeira vez (nesta execução ou restaurado do disco já
    # travado) — depois disso a lista trava: só atualiza placar/resultado de
    # quem já está nela, nunca acrescenta partida nova (pedido do usuário, que
    # achava estranho a lista mudar toda vez que clicava em "Recalcular").
    pode_adicionar_novas = day_cache["melhor"] is None

    # Agendados (pra sugerir) + já encerrados hoje (pra mostrar o placar final e
    # se a sugestão teria acertado) — pools separados e cada um com seu próprio
    # teto, senão um dia com muitos jogos já encerrados "engoliria" as vagas dos
    # próximos jogos ainda por vir na lista de candidatos.
    scheduled = sorted((m for m in _fs_all_matches() if m.get("status") == "1"), key=lambda m: m.get("kickoff_ts") or "")
    finished_today = sorted((m for m in _fs_all_matches() if m.get("status") == "3"), key=lambda m: m.get("kickoff_ts") or "")
    candidatos = scheduled[:_BT2_MATCHES_MAX_CANDIDATOS] + finished_today[:_BT2_MATCHES_MAX_CANDIDATOS]
    today_matches = candidatos

    def _fetch_h2h(m):
        try:
            return m["id"], rows_fn(_fs_h2h(m["id"]))
        except Exception:
            return m["id"], []

    rows_by_match = dict(_fs_event_pool.map(_fetch_h2h, candidatos))

    teams = {}  # norm(nome do time relevante) -> {"name":..., "rows":[...]}
    all_ids_seen = set()
    total_ids = 0
    for m in candidatos:
        if total_ids >= _BTCS_MAX_HISTORICAL_IDS:
            break
        rows = rows_by_match.get(m["id"]) or []
        key = _btcs_norm_name(m[team_field])
        if not key or not rows:
            continue
        entry = teams.setdefault(key, {"name": m[team_field], "rows": []})
        existing_ids = {r["id"] for r in entry["rows"]}
        for row in rows:
            if total_ids >= _BTCS_MAX_HISTORICAL_IDS:
                break
            if not row["id"] or row["id"] in existing_ids or row["id"] in all_ids_seen:
                continue
            entry["rows"].append(row)
            existing_ids.add(row["id"])
            all_ids_seen.add(row["id"])
            total_ids += 1

    # Metodologias de 1º tempo (needs_ht) precisam do placar do intervalo de
    # cada jogo histórico, que NÃO vem junto do H2H — busca extra, um feed por
    # jogo, mas rápida em paralelo (~1s pra 20 jogos testado manualmente,
    # bem mais barato que o fetch de odds que trava com centenas de jogos).
    if cfg["needs_ht"] and all_ids_seen:
        ht_by_id = dict(_fs_event_pool.map(lambda eid: (eid, _fs_half_time(eid)), all_ids_seen))
        for entry in teams.values():
            for row in entry["rows"]:
                ht = ht_by_id.get(row["id"])
                if not ht:
                    continue
                try:
                    hs, as_ = int(ht["home"]), int(ht["away"])
                except (TypeError, ValueError, KeyError):
                    continue
                row["own_ht"] = hs if side == "casa" else as_
                row["opp_ht"] = as_ if side == "casa" else hs

    # baseline: taxa do evento-alvo sobre TODOS os jogos coletados — é contra
    # isso que o lift de cada segmento é medido
    baseline_total = baseline_hits = 0
    for entry in teams.values():
        for row in entry["rows"]:
            hit = target_fn(row)
            if hit is None:  # sem dado de HT pra essa linha (needs_ht) — ignora
                continue
            baseline_total += 1
            if hit:
                baseline_hits += 1
    baseline_rate = (baseline_hits / baseline_total) if baseline_total else 0.0

    seg_data = {}       # variável -> faixa -> {"total":n, "hits":n}
    team_profile = {}   # chave normalizada -> {"name":..., "segs":..., "n":...}
    for key, entry in teams.items():
        segs, n = _laycasa_team_metrics(entry["rows"])
        if not segs:
            continue
        team_profile[key] = {"name": entry["name"], "segs": segs, "n": n}
        for var, seg in segs.items():
            bucket = seg_data.setdefault(var, {}).setdefault(seg, {"total": 0, "hits": 0})
            for row in entry["rows"]:
                hit = target_fn(row)
                if hit is None:
                    continue
                bucket["total"] += 1
                if hit:
                    bucket["hits"] += 1

    # acha a MELHOR combinação variável+faixa (maior lift, com amostra
    # suficiente) — é isso que faz o sistema ser dinâmico em vez de eu fixar
    # manualmente um limiar. Só é escolhida UMA VEZ POR DIA (fica congelada em
    # day_cache) — sem isso, cada recálculo (a cada 30min) podia escolher uma
    # variável+faixa diferente e a grade de sugestões ficava trocando de jogos
    # o dia todo, o que não faz sentido pro usuário acompanhar ao longo do dia.
    if day_cache["melhor"] is not None:
        melhor = day_cache["melhor"]
        comparativo = day_cache["comparativo"]
    else:
        melhor = None
        for var, segs in seg_data.items():
            for seg_label, counts in segs.items():
                total = counts["total"]
                if total < _LAYCASA_MIN_SEGMENT_GAMES:
                    continue
                rate = counts["hits"] / total
                lift = (rate / baseline_rate) if baseline_rate else 0.0
                if lift < _LAYCASA_MIN_LIFT:
                    continue
                if melhor is None or lift > melhor["lift"]:
                    melhor = {
                        "variable": var, "variable_label": var_labels.get(var, var),
                        "segment": seg_label, "total": total, "rate": round(rate, 4),
                        "baseline_rate": round(baseline_rate, 4), "lift": round(lift, 3),
                    }

        # gráfico comparativo: taxa de acerto acumulada (geral vs com o filtro
        # validado acima) ao longo de todas as entradas históricas coletadas,
        # ordenadas por data — mesmo padrão do timeline de cumulative_hits usado
        # no Backtest CS, só que aqui é % de acerto acumulado, não contagem bruta
        comparativo = None
        if melhor:
            baseline_entries = []
            filtro_entries = []
            for key, entry in teams.items():
                profile = team_profile.get(key)
                in_segment = bool(profile) and profile["segs"].get(melhor["variable"]) == melhor["segment"]
                for row in entry["rows"]:
                    hit = target_fn(row)
                    if hit is None:
                        continue
                    baseline_entries.append((row.get("date") or "", hit))
                    if in_segment:
                        filtro_entries.append((row.get("date") or "", hit))

            def _cum_hitrate_timeline(entries):
                entries = sorted(entries, key=lambda e: e[0])
                timeline = []
                hits = 0
                for i, (date, hit) in enumerate(entries, start=1):
                    if hit:
                        hits += 1
                    timeline.append({"n": i, "date": date, "taxa_acerto": round(hits / i, 4)})
                return timeline

            comparativo = {
                "geral": _cum_hitrate_timeline(baseline_entries),
                "com_filtro": _cum_hitrate_timeline(filtro_entries),
            }

        day_cache["melhor"] = melhor
        day_cache["comparativo"] = comparativo

    # aplica o padrão validado às partidas de hoje: quais times relevantes têm
    # o PRÓPRIO segmento igual ao segmento validado acima? A lista TRAVA na
    # primeira vez que o padrão do dia é escolhido (pode_adicionar_novas) — só
    # entram partidas novas nesse momento. Depois disso, cada recálculo só
    # ATUALIZA placar/encerrado/acertou de quem já está na lista (nunca some,
    # nunca ganha gente nova) — evita a lista mudando toda vez que o usuário
    # clica em "Recalcular" ou o servidor reinicia no meio do dia.
    sugestoes_by_id = day_cache["sugestoes_by_id"]
    if melhor:
        for m in today_matches:
            key = _btcs_norm_name(m[team_field])
            id_str = str(m["id"])
            ja_esta_na_lista = id_str in sugestoes_by_id
            if not ja_esta_na_lista and not pode_adicionar_novas:
                continue
            profile = team_profile.get(key)
            if not profile:
                continue
            if profile["segs"].get(melhor["variable"]) == melhor["segment"]:
                status = m.get("status")
                item = {
                    "id": m["id"], "home": m["home"], "away": m["away"],
                    "liga": m.get("liga", ""), "pais": m.get("pais", ""),
                    "kickoff_ts": m.get("kickoff_ts"),
                    "amostra_time": profile["n"],
                    "encerrado": status == "3",
                }
                if status == "3":
                    hs, as_ = m.get("home_score"), m.get("away_score")
                    item["home_score"] = hs
                    item["away_score"] = as_
                    if hs is not None and as_ is not None:
                        try:
                            hs_i, as_i = int(hs), int(as_)
                            final_row = {
                                "own": hs_i if side == "casa" else as_i,
                                "opp": as_i if side == "casa" else hs_i,
                            }
                            if cfg["needs_ht"]:
                                ht = _fs_half_time(m["id"])
                                if ht:
                                    hs_ht, as_ht = int(ht["home"]), int(ht["away"])
                                    final_row["own_ht"] = hs_ht if side == "casa" else as_ht
                                    final_row["opp_ht"] = as_ht if side == "casa" else hs_ht
                            item["acertou"] = target_fn(final_row)
                        except (TypeError, ValueError):
                            pass
                sugestoes_by_id[id_str] = item
        _mapa_save_day_cache_to_disk(metodologia_key, day_cache)
    sugestoes = sorted(sugestoes_by_id.values(), key=lambda x: x.get("kickoff_ts") or "")

    data = {
        "metodologia": metodologia_key,
        "metodologia_label": cfg["label"],
        "side": side,
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "matches_used": total_ids,
        "melhor_padrao": melhor,
        "valido": melhor is not None,
        "comparativo": comparativo,
        "sugestoes": sugestoes,
    }
    cache["data"] = data
    cache["ts"] = now
    return data


@app.route("/api/masterlist/metodologias")
def api_masterlist_metodologias():
    """Lista as metodologias disponíveis no Mapa de Sugestões, pra popular o
    dropdown do frontend."""
    return jsonify({"metodologias": [{"key": k, "label": v["label"]} for k, v in _MAPA_METODOLOGIAS.items()]})


@app.route("/api/masterlist/<metodologia>")
def api_masterlist_generic(metodologia):
    """'Mapa de Sugestões' — acha dinamicamente qual variável (das usadas na
    aba Jogo + Backtest CS) melhor prediz o evento-alvo da metodologia
    escolhida, valida contra o histórico, e aplica às partidas de hoje.
    Cacheado 30min (mesmo TTL do Backtest CS)."""
    data = _mapa_compute(metodologia)
    if data is None:
        return jsonify({"error": "metodologia desconhecida"}), 404
    return jsonify(data)


@app.route("/api/masterlist/<metodologia>/atualizar_dia", methods=["POST"])
def api_masterlist_atualizar_dia(metodologia):
    """Reconstrói o padrão do dia + lista de sugestões do ZERO pra essa
    metodologia — usado quando o Mapa fica "preso" mostrando jogos antigos
    (ex: virou o dia mas por algum motivo o cache não acompanhou). Diferente
    do recálculo normal (que só atualiza placar de quem já está na lista, de
    propósito, pra não ficar mudando a lista toda hora), esse endpoint apaga
    tudo e deixa escolher um padrão novo — por isso só permite 1x por dia,
    pra não reintroduzir o mesmo problema que a trava resolveu."""
    if metodologia not in _MAPA_METODOLOGIAS:
        return jsonify({"error": "metodologia desconhecida"}), 404

    today_str = datetime.now().strftime("%Y-%m-%d")
    day_cache = _MAPA_DAY_CACHE.get(metodologia)
    if day_cache and day_cache.get("manual_reset_date") == today_str:
        return jsonify({"error": "essa metodologia já foi atualizada manualmente hoje — só é permitido 1x por dia"}), 429

    _MAPA_DAY_CACHE[metodologia] = {
        "date": today_str, "melhor": None, "comparativo": None,
        "sugestoes_by_id": {}, "manual_reset_date": today_str,
    }
    _MAPA_CACHE[metodologia] = {"ts": 0, "data": None}
    path = _mapa_cache_path(metodologia)
    if os.path.exists(path):
        os.remove(path)

    data = _mapa_compute(metodologia)
    return jsonify(data)


# ── ABA JOGO — MÉDIAS GERAIS — baseline pra comparação: qual a média de cada
# estatística (vitórias, gols, ambas marcam, over/under...) entre TODOS os
# times de hoje, não só o time que o usuário está olhando. Reaproveita
# _laycasa_home_rows/_laycasa_away_rows (mesma extração de linhas do Mapa de
# Sugestões) — um único fetch de H2H por jogo já dá tanto a amostra "casa"
# quanto "fora". Recalculada 1x por dia (não faz sentido recalcular toda hora,
# a média de centenas de times não muda de uma hora pra outra).
_JOGO_MEDIAS_CACHE = {"date": None, "data": None}


def _jogo_fetch_extra(rows, side):
    """Busca minuto dos gols (pra 'minuto médio do 1º gol', gols por faixa de
    minuto e vencedor a cada checkpoint), placar do intervalo (pra 'metade com
    mais gols') e odd 1x2 própria/adversário (pra Valor do Gol/Ponto/Saldo,
    Custo do Gol 2.0) de cada linha, tudo numa passada só em paralelo. Só pede
    o mercado "1x2" (markets_wanted) — mesmo truque que já derrubou o tempo do
    Backtest CS de 706s pra 48s, evitando buscar os outros 5 mercados à toa.
    Sem esses dados a linha simplesmente não entra nas médias que dependem
    deles (ver _jogo_stats_from_rows / _jogo_goal_pattern_from_rows /
    _jogo_goal_diff_from_rows)."""
    def _fetch(row):
        eid = row["id"]
        if not eid:
            return row
        try:
            gm = _fs_goal_minutes(eid)
        except Exception:
            gm = None
        if gm:
            if side == "casa":
                row["gm_own"], row["gm_opp"] = gm.get("home", []), gm.get("away", [])
            else:
                row["gm_own"], row["gm_opp"] = gm.get("away", []), gm.get("home", [])
        try:
            ht = _fs_half_time(eid)
            hs, as_ = int(ht["home"]), int(ht["away"])
            row["ht_own"], row["ht_opp"] = (hs, as_) if side == "casa" else (as_, hs)
        except Exception:
            pass
        try:
            _, markets = _fs_odds_all_markets_any_bookmaker(eid, markets_wanted=["1x2"])
            h, a = (row["own"], row["opp"]) if side == "casa" else (row["opp"], row["own"])
            sels = _bt2_market_selections("1x2", markets.get("1x2"), h, a)
            target_label = "Casa" if side == "casa" else "Fora"
            opp_label = "Fora" if side == "casa" else "Casa"
            for sel in sels:
                if sel["label"] == target_label and sel["odd"]:
                    row["odd"] = sel["odd"]
                elif sel["label"] == opp_label and sel["odd"]:
                    row["opp_odd"] = sel["odd"]
        except Exception:
            pass
        return row

    return list(_fs_event_pool.map(_fetch, rows))


def _jogo_stats_from_rows(rows):
    n = len(rows)
    if n == 0:
        return None
    vit = sum(1 for r in rows if r["own"] > r["opp"])
    emp = sum(1 for r in rows if r["own"] == r["opp"])
    der = n - vit - emp
    marcou = sum(1 for r in rows if r["own"] >= 1)
    sofreu = sum(1 for r in rows if r["opp"] >= 1)
    ambas = sum(1 for r in rows if r["own"] >= 1 and r["opp"] >= 1)
    sem_sofrer = sum(1 for r in rows if r["opp"] == 0)
    sem_marcar = sum(1 for r in rows if r["own"] == 0)
    over25 = sum(1 for r in rows if (r["own"] + r["opp"]) > 2.5)

    def _ou_total(line):
        over = sum(1 for r in rows if (r["own"] + r["opp"]) > line)
        return {"sobre": round(over / n * 100, 1), "sob": round((n - over) / n * 100, 1)}

    def _ou_team(line):
        over = sum(1 for r in rows if r["own"] > line)
        return {"sobre": round(over / n * 100, 1), "sob": round((n - over) / n * 100, 1)}

    # Minuto médio do 1º gol — só nas linhas com dado de minutos de gol
    gm_rows = [r for r in rows if "gm_own" in r]
    tempo_marcado = tempo_sofrido = tempo_partida = None
    if gm_rows:
        def _first_min(mins):
            return min(mins) if mins else 90
        tempo_marcado = round(sum(_first_min(r["gm_own"]) for r in gm_rows) / len(gm_rows), 1)
        tempo_sofrido = round(sum(_first_min(r["gm_opp"]) for r in gm_rows) / len(gm_rows), 1)
        tempo_partida = round(sum(min(_first_min(r["gm_own"]), _first_min(r["gm_opp"])) for r in gm_rows) / len(gm_rows), 1)

    # Valor do Gol/Ponto/Saldo e Custo do Gol 2.0 — só nas linhas com odd
    # própria E do adversário conhecidas, mesmas fórmulas da aba Jogo
    odd_rows = [r for r in rows if r.get("odd") and r.get("opp_odd")]
    valor_gol_marcado = valor_gol_sofrido = custo_gol_marcado = custo_gol_sofrido = valor_saldo = valor_ponto = None
    if odd_rows:
        nn = len(odd_rows)
        vgm = vgs = cgm = cgs = vs = vp = 0.0
        for r in odd_rows:
            own_prob = 1.0 / r["odd"]
            opp_prob = 1.0 / r["opp_odd"]
            pontos = 3 if r["own"] > r["opp"] else (1 if r["own"] == r["opp"] else 0)
            saldo_row = r["own"] - r["opp"]
            vgm += r["own"] * opp_prob
            vgs += r["opp"] * own_prob
            cgm += (r["own"] / 2) + (own_prob / 2)
            cgs += (r["opp"] / 2) + (own_prob / 2)
            vs += saldo_row * opp_prob
            vp += pontos * opp_prob
        valor_gol_marcado, valor_gol_sofrido = round(vgm / nn, 3), round(vgs / nn, 3)
        custo_gol_marcado, custo_gol_sofrido = round(cgm / nn, 3), round(cgs / nn, 3)
        valor_saldo, valor_ponto = round(vs / nn, 3), round(vp / nn, 3)

    return {
        "n": n,
        "vitFT_pct": round(vit / n * 100, 1),
        "empFT_pct": round(emp / n * 100, 1),
        "derFT_pct": round(der / n * 100, 1),
        "mediaMarcados": round(sum(r["own"] for r in rows) / n, 2),
        "mediaSofridos": round(sum(r["opp"] for r in rows) / n, 2),
        "pctMarcar": round(marcou / n * 100, 1),
        "pctSofrer": round(sofreu / n * 100, 1),
        "pctAmbasMarcam": round(ambas / n * 100, 1),
        "semSofrer_pct": round(sem_sofrer / n * 100, 1),
        "semMarcar_pct": round(sem_marcar / n * 100, 1),
        "over25_pct": round(over25 / n * 100, 1),
        "under25_pct": round((n - over25) / n * 100, 1),
        "ou15_total": _ou_total(1.5),
        "ou25_total": _ou_total(2.5),
        "ou15_team": _ou_team(1.5),
        "ou25_team": _ou_team(2.5),
        "ou35_team": _ou_team(3.5),
        "saldo": round(sum(r["own"] - r["opp"] for r in rows) / n, 2),
        "tempoPrimeiroGolMarcado": tempo_marcado,
        "tempoPrimeiroGolSofrido": tempo_sofrido,
        "tempoPrimeiroGolPartida": tempo_partida,
        "valorGolMarcado": valor_gol_marcado,
        "valorGolSofrido": valor_gol_sofrido,
        "custoGol2Marcado": custo_gol_marcado,
        "custoGol2Sofrido": custo_gol_sofrido,
        "valorSaldo": valor_saldo,
        "valorPonto": valor_ponto,
    }


def _jogo_goal_pattern_from_rows(rows):
    """Porta pra Python de _jogoGoalPatternStats (JS) — faixas de gols, ambas
    marcam, par/ímpar, gols por faixa de 15min e minuto do 1º gol por faixa de
    10min — pra alimentar as médias gerais da seção 'Características gerais
    do gol' / 'Gols nos minutos entre' / 'Primeiro gol nas partidas'. Reusa o
    gm_own/gm_opp já anexado por _jogo_fetch_extra, sem fetch adicional."""
    n = len(rows)
    if n == 0:
        return None
    total_goals = [r["own"] + r["opp"] for r in rows]
    ambos = sum(1 for r in rows if r["own"] >= 1 and r["opp"] >= 1)
    apenas_um = sum(1 for r in rows if (r["own"] >= 1) != (r["opp"] >= 1))
    nenhum = sum(1 for r in rows if r["own"] == 0 and r["opp"] == 0)
    impar = sum(1 for g in total_goals if g % 2 == 1)

    gm_rows = [r for r in rows if "gm_own" in r]
    total_com_gm = len(gm_rows)

    def _bucket15(m):
        if m <= 15:
            return 0
        if m <= 30:
            return 1
        if m <= 45:
            return 2
        if m <= 60:
            return 3
        if m <= 75:
            return 4
        return 5

    def _bucket10(m):
        return min(8, max(0, (m - 1) // 10))

    all_buckets = [0] * 6
    team_buckets = [0] * 6
    all_count = team_count = 0
    primeiro_buckets = [0] * 9
    primeiro_team_buckets = [0] * 9
    sem_gol = sem_gol_team = time_primeiro = adversario_primeiro = 0

    for r in gm_rows:
        pro, contra = r["gm_own"], r["gm_opp"]
        for m in pro:
            mm = min(m, 90)
            all_buckets[_bucket15(mm)] += 1
            all_count += 1
            team_buckets[_bucket15(mm)] += 1
            team_count += 1
        for m in contra:
            mm = min(m, 90)
            all_buckets[_bucket15(mm)] += 1
            all_count += 1

        events = sorted([(m, True) for m in pro] + [(m, False) for m in contra])
        if not events:
            sem_gol += 1
            sem_gol_team += 1
            continue
        first_m, first_is_team = events[0]
        primeiro_buckets[_bucket10(min(first_m, 90))] += 1
        if first_is_team:
            time_primeiro += 1
            primeiro_team_buckets[_bucket10(min(first_m, 90))] += 1
        else:
            adversario_primeiro += 1
            sem_gol_team += 1

    def _pct_list(buckets, total):
        return [round(b / total * 100, 1) for b in buckets] if total else None

    return {
        "faixa01_pct": round(sum(1 for g in total_goals if g <= 1) / n * 100, 1),
        "faixa23_pct": round(sum(1 for g in total_goals if 2 <= g <= 3) / n * 100, 1),
        "faixa4mais_pct": round(sum(1 for g in total_goals if g >= 4) / n * 100, 1),
        "ambosMarcam_pct": round(ambos / n * 100, 1),
        "apenasUm_pct": round(apenas_um / n * 100, 1),
        "nenhum_pct": round(nenhum / n * 100, 1),
        "impar_pct": round(impar / n * 100, 1),
        "par_pct": round((n - impar) / n * 100, 1),
        "totalComGm": total_com_gm,
        "allGoalsBuckets_pct": _pct_list(all_buckets, all_count),
        "teamGoalsBuckets_pct": _pct_list(team_buckets, team_count),
        "primeiroGolBuckets_pct": _pct_list(primeiro_buckets, total_com_gm),
        "primeiroGolTeamBuckets_pct": _pct_list(primeiro_team_buckets, total_com_gm),
        "semGol_pct": round(sem_gol / total_com_gm * 100, 1) if total_com_gm else None,
        "semGolTeam_pct": round(sem_gol_team / total_com_gm * 100, 1) if total_com_gm else None,
        "timeMarcouPrimeiro_pct": round(time_primeiro / total_com_gm * 100, 1) if total_com_gm else None,
        "adversarioMarcouPrimeiro_pct": round(adversario_primeiro / total_com_gm * 100, 1) if total_com_gm else None,
    }


def _jogo_winner_at_minute_from_rows(rows):
    """Porta de _jogoWinnerAtMinuteStats (JS) — quem está na frente no placar
    a cada checkpoint (15/30/45/60/75/90min), usando gm_own/gm_opp já
    anexados por _jogo_fetch_extra."""
    gm_rows = [r for r in rows if "gm_own" in r]
    n = len(gm_rows)
    if n == 0:
        return None
    out = {}
    for cp in (15, 30, 45, 60, 75, 90):
        time_c = adv_c = emp_c = 0
        for r in gm_rows:
            gp = sum(1 for m in r["gm_own"] if m <= cp)
            gc = sum(1 for m in r["gm_opp"] if m <= cp)
            if gp > gc:
                time_c += 1
            elif gc > gp:
                adv_c += 1
            else:
                emp_c += 1
        out[str(cp)] = {
            "time_pct": round(time_c / n * 100, 1),
            "adversario_pct": round(adv_c / n * 100, 1),
            "empate_pct": round(emp_c / n * 100, 1),
        }
    return {"n": n, "checkpoints": out}


def _jogo_goal_diff_from_rows(rows):
    """Porta de _jogoGoalDiffStats (JS) — diferença de gols na partida e qual
    metade do jogo teve mais gols (usando ht_own/ht_opp já anexados por
    _jogo_fetch_extra)."""
    n = len(rows)
    if n == 0:
        return None
    diff_arr = [abs(r["own"] - r["opp"]) for r in rows]
    diff01 = sum(1 for d in diff_arr if d <= 1) / n * 100
    diff23 = sum(1 for d in diff_arr if 2 <= d <= 3) / n * 100
    diff4mais = sum(1 for d in diff_arr if d >= 4) / n * 100

    ht_rows = [r for r in rows if "ht_own" in r]
    n_ht = len(ht_rows)
    primeiro_tempo = segunda_metade = gravata = None
    if n_ht:
        pt = sm = gv = 0
        for r in ht_rows:
            gols_ht = r["ht_own"] + r["ht_opp"]
            gols_st = (r["own"] + r["opp"]) - gols_ht
            if gols_ht > gols_st:
                pt += 1
            elif gols_st > gols_ht:
                sm += 1
            else:
                gv += 1
        primeiro_tempo = round(pt / n_ht * 100, 1)
        segunda_metade = round(sm / n_ht * 100, 1)
        gravata = round(gv / n_ht * 100, 1)

    return {
        "diff01_pct": round(diff01, 1),
        "diff23_pct": round(diff23, 1),
        "diff4mais_pct": round(diff4mais, 1),
        "nHt": n_ht,
        "primeiroTempo_pct": primeiro_tempo,
        "segundaMetade_pct": segunda_metade,
        "gravata_pct": gravata,
    }


def _jogo_medias_gerais_compute():
    today_str = datetime.now().strftime("%Y-%m-%d")
    if _JOGO_MEDIAS_CACHE["date"] == today_str and _JOGO_MEDIAS_CACHE["data"] is not None:
        return _JOGO_MEDIAS_CACHE["data"]

    scheduled = sorted((m for m in _fs_all_matches() if m.get("status") == "1"), key=lambda m: m.get("kickoff_ts") or "")
    finished_today = sorted((m for m in _fs_all_matches() if m.get("status") == "3"), key=lambda m: m.get("kickoff_ts") or "")
    candidatos = scheduled[:_BT2_MATCHES_MAX_CANDIDATOS] + finished_today[:_BT2_MATCHES_MAX_CANDIDATOS]

    def _fetch(m):
        try:
            tabs = _fs_h2h(m["id"])
            return _laycasa_home_rows(tabs), _laycasa_away_rows(tabs)
        except Exception:
            return [], []

    casa_rows, fora_rows = [], []
    seen_casa, seen_fora = set(), set()
    for home_rows, away_rows in _fs_event_pool.map(_fetch, candidatos):
        for row in home_rows:
            if row["id"] and row["id"] not in seen_casa and len(casa_rows) < _BTCS_MAX_HISTORICAL_IDS:
                seen_casa.add(row["id"])
                casa_rows.append(row)
        for row in away_rows:
            if row["id"] and row["id"] not in seen_fora and len(fora_rows) < _BTCS_MAX_HISTORICAL_IDS:
                seen_fora.add(row["id"])
                fora_rows.append(row)

    casa_rows = _jogo_fetch_extra(casa_rows, "casa")
    fora_rows = _jogo_fetch_extra(fora_rows, "fora")
    geral_rows = casa_rows + fora_rows

    def _scope_data(rows):
        stats = _jogo_stats_from_rows(rows)
        if stats is None:
            return None
        for extra in (_jogo_goal_pattern_from_rows(rows), _jogo_goal_diff_from_rows(rows)):
            if extra:
                stats.update(extra)
        winner = _jogo_winner_at_minute_from_rows(rows)
        if winner:
            stats["vencedor"] = winner
        return stats

    data = {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "casa": _scope_data(casa_rows),
        "fora": _scope_data(fora_rows),
        "geral": _scope_data(geral_rows),
    }
    _JOGO_MEDIAS_CACHE["date"] = today_str
    _JOGO_MEDIAS_CACHE["data"] = data
    return data


@app.route("/api/jogo/medias_gerais")
def api_jogo_medias_gerais():
    """Médias gerais (Geral/Casa/Fora) das estatísticas da aba Jogo, pra
    comparação — mesmos números que aparecem no card de cada time, só que
    calculados em cima de uma amostra ampla de times de hoje. Cacheado 1x por
    dia."""
    return jsonify(_jogo_medias_gerais_compute())


@app.route("/api/backtestcs/ranking")
def api_backtestcs_ranking():
    """Ranking global dos 19 buckets de placar exato por taxa de acerto,
    agregando o histórico de TODOS os jogos de hoje. Bem mais barato que o
    Backtest 2 (não busca odds), mas ainda cacheia 30min pra evitar recalcular
    o H2H de todo mundo a cada request — use ?refresh=1 pra forçar."""
    mode = request.args.get("mode") or "recente"
    if mode == "acumulado":
        return jsonify(_btcs_compute_ranking_acumulado())
    force = request.args.get("refresh") == "1"
    if not force and _btcs_cache["data"] is not None and (time.time() - _btcs_cache["ts"]) < _BTCS_CACHE_TTL:
        return jsonify(_btcs_cache["data"])
    data = _btcs_compute_ranking()
    _btcs_cache["ts"] = time.time()
    _btcs_cache["data"] = data
    return jsonify(data)


@app.route("/api/backtestcs/ranking_acumulado")
def api_backtestcs_ranking_acumulado():
    """Ranking ACUMULADO: agrega todos os resultados já persistidos em
    btcs_results (cresce a cada vez que /api/backtestcs/ranking roda)."""
    return jsonify(_btcs_compute_ranking_acumulado())


@app.route("/api/backtestcs/patterns")
def api_backtestcs_patterns():
    """Padrões: quais combinações (variável do time, segmento, bucket de placar
    exato) têm taxa de acerto significativamente melhor/pior que a geral.
    ?mode=acumulado usa TODO o histórico já persistido (sem cache, leitura
    SQLite é barata). Modo padrão ('últimos 30') cacheia 30min — ?refresh=1
    força recálculo."""
    mode = request.args.get("mode") or "recente"
    if mode == "acumulado":
        return jsonify(_btcs_compute_patterns_acumulado())
    force = request.args.get("refresh") == "1"
    if not force and _btcs_patterns_cache["data"] is not None and (time.time() - _btcs_patterns_cache["ts"]) < _BTCS_CACHE_TTL:
        return jsonify(_btcs_patterns_cache["data"])
    data = _btcs_compute_patterns()
    _btcs_patterns_cache["ts"] = time.time()
    _btcs_patterns_cache["data"] = data
    return jsonify(data)


def _btcs_bucket_odds(items):
    """Porta de _placarExatoBuckets (JS) pra Python: agrupa os placares crus
    do mercado 'placar_exato' nos 19 buckets fixos, odd combinada = 1/soma das
    probabilidades implícitas dos placares que caem em cada bucket."""
    prob_sum = {}
    for it in items or []:
        m = re.match(r"^(\d+):(\d+)$", it.get("score") or "")
        if not m:
            continue
        try:
            odd = float((it.get("item") or {}).get("value"))
        except (TypeError, ValueError):
            continue
        if not odd:
            continue
        key = _btcs_bucket_key(int(m.group(1)), int(m.group(2)))
        prob_sum[key] = prob_sum.get(key, 0.0) + 1 / odd
    return {key: 1 / prob_sum[key] for key in prob_sum}


_BTCS_BUCKET_ODDS_CACHE_TTL = 300  # 5min — odds mudam pouco em poucos minutos
_btcs_bucket_odds_cache = {"ts": 0, "data": None}


@app.route("/api/backtestcs/bucket_odds_today")
def api_backtestcs_bucket_odds_today():
    """Odd média ATUAL de cada um dos 19 buckets, calculada em cima dos jogos
    de HOJE que ainda não começaram (média entre os jogos que oferecem odd
    pra aquele bucket). Busca as odds de cada jogo de hoje UMA VEZ SÓ (não
    por bucket) e classifica em paralelo — bem mais barato que chamar
    matches_for_bucket bucket por bucket. Cacheado 5min."""
    force = request.args.get("refresh") == "1"
    if not force and _btcs_bucket_odds_cache["data"] is not None and \
            (time.time() - _btcs_bucket_odds_cache["ts"]) < _BTCS_BUCKET_ODDS_CACHE_TTL:
        return jsonify(_btcs_bucket_odds_cache["data"])

    odds_sum = {}
    odds_n = {}
    for m, markets in _today2_odds_snapshot(force=force):
        odds_by_bucket = _btcs_bucket_odds((markets.get("placar_exato") or {}).get("items"))
        for bucket, odd in odds_by_bucket.items():
            odds_sum[bucket] = odds_sum.get(bucket, 0.0) + odd
            odds_n[bucket] = odds_n.get(bucket, 0) + 1

    avg_odds = {b: round(odds_sum[b] / odds_n[b], 2) for b in odds_sum}
    data = {"avg_odds": avg_odds}
    _btcs_bucket_odds_cache["ts"] = time.time()
    _btcs_bucket_odds_cache["data"] = data
    return jsonify(data)


@app.route("/api/backtestcs/matches_for_bucket")
def api_backtestcs_matches_for_bucket():
    """Jogos de HOJE que ainda não começaram onde o bucket de placar exato
    selecionado tem odd disponível no mercado 'placar_exato' atual."""
    bucket = request.args.get("bucket", "")
    if not bucket:
        return jsonify({"matches": []}), 400

    resultado = []
    for m, markets in _today2_odds_snapshot():
        odds_by_bucket = _btcs_bucket_odds((markets.get("placar_exato") or {}).get("items"))
        odd = odds_by_bucket.get(bucket)
        if odd:
            resultado.append({
                "id": m["id"], "home": m["home"], "away": m["away"],
                "liga": m.get("liga", ""), "pais": m.get("pais", ""),
                "kickoff_ts": m.get("kickoff_ts"), "odd": odd,
            })
    resultado.sort(key=lambda x: x.get("kickoff_ts") or "")
    return jsonify({"matches": resultado})


# ── Classificação de jogos de hoje pelas odds atuais (Resultado Final / Gols /
# Ambas Marcam) — usado só pelo painel "Filtrar por Parâmetros" da aba experimental
# 'Hoje 2'. Cacheado 5min (mesmo padrão do _btcs_bucket_odds_cache) pra não
# reconsultar odds a cada clique de checkbox do usuário. ──
_TODAY2_CLASSIFICATION_CACHE_TTL = 300  # 5min — odds mudam pouco em poucos minutos
_today2_classification_cache = {"ts": 0, "data": None}


def _today2_classify_match(markets):
    """Classifica um jogo pelas odds atuais em Resultado Final / Gols / Ambas Marcam.
    Limiares de favoritismo (odd baixa = mais provável): <1.35 super favorito,
    1.35–1.80 favorito, ambos os lados entre 1.80–2.60 sem favorito claro = parelho.
    Fora dessas faixas (ex: um lado <1.80 e outro >2.60) não classifica resultado
    (fica None) pra não forçar rótulo em jogo sem padrão claro."""
    out = {"resultado": None, "favorito_lado": None, "gols": None, "ambas": None,
           "odd_casa": None, "odd_fora": None}

    m1x2 = markets.get("1x2") or {}
    odd_casa = (m1x2.get("home") or {}).get("value")
    odd_fora = (m1x2.get("away") or {}).get("value")
    try:
        odd_casa = float(odd_casa) if odd_casa is not None else None
    except (TypeError, ValueError):
        odd_casa = None
    try:
        odd_fora = float(odd_fora) if odd_fora is not None else None
    except (TypeError, ValueError):
        odd_fora = None
    out["odd_casa"] = odd_casa
    out["odd_fora"] = odd_fora

    if odd_casa and odd_fora:
        menor = min(odd_casa, odd_fora)
        lado = "casa" if odd_casa <= odd_fora else "fora"
        if menor < 1.35:
            out["resultado"] = "super_favorito"
            out["favorito_lado"] = lado
        elif menor <= 1.80:
            out["resultado"] = "favorito"
            out["favorito_lado"] = lado
        elif 1.80 <= odd_casa <= 2.60 and 1.80 <= odd_fora <= 2.60:
            out["resultado"] = "parelho"

    over_under = markets.get("over_under") or {}
    melhor_op = None
    melhor_dist = None
    for op in over_under.get("opportunities") or []:
        try:
            line = float((op.get("handicap") or {}).get("value"))
        except (TypeError, ValueError):
            continue
        dist = abs(line - 2.5)
        if melhor_dist is None or dist < melhor_dist:
            melhor_dist = dist
            melhor_op = op
    if melhor_op:
        try:
            odd_over = float((melhor_op.get("over") or {}).get("value"))
        except (TypeError, ValueError):
            odd_over = None
        try:
            odd_under = float((melhor_op.get("under") or {}).get("value"))
        except (TypeError, ValueError):
            odd_under = None
        if odd_over and odd_under:
            out["gols"] = "over" if odd_over <= odd_under else "under"

    ambos = markets.get("ambos_marcam") or {}
    try:
        odd_sim = float((ambos.get("yes") or {}).get("value"))
    except (TypeError, ValueError):
        odd_sim = None
    try:
        odd_nao = float((ambos.get("no") or {}).get("value"))
    except (TypeError, ValueError):
        odd_nao = None
    if odd_sim and odd_nao:
        out["ambas"] = "sim" if odd_sim <= odd_nao else "nao"

    return out


@app.route("/api/today2/match_classification")
def api_today2_match_classification():
    """Classifica os jogos de HOJE (ainda não iniciados, limitados aos próximos
    _BT2_MATCHES_MAX_CANDIDATOS) pelas odds atuais — Resultado Final, Gols e Ambas
    Marcam — pra alimentar o painel 'Filtrar por Parâmetros' da aba 'Hoje 2'.
    Reaproveita os pools isolados do Backtest 2 (_bt2_matches_pool /
    _bt2_matches_market_pool), não os pools pesados de ranking, pro filtro
    continuar rápido mesmo com um recálculo de ranking rodando."""
    force = request.args.get("refresh") == "1"
    if not force and _today2_classification_cache["data"] is not None and \
            (time.time() - _today2_classification_cache["ts"]) < _TODAY2_CLASSIFICATION_CACHE_TTL:
        return jsonify(_today2_classification_cache["data"])

    classifications = {}
    for m, markets in _today2_odds_snapshot(force=force):
        if not markets:
            continue
        c = _today2_classify_match(markets)
        if c["resultado"] or c["gols"] or c["ambas"]:
            classifications[str(m["id"])] = c

    data = {"classifications": classifications}
    _today2_classification_cache["ts"] = time.time()
    _today2_classification_cache["data"] = data
    return jsonify(data)


@app.route("/api/today2/odds_raw")
def api_today2_odds_raw():
    """Odds cruas (todos os mercados: 1x2, over_under, ambos_marcam, placar_exato
    etc) de todos os jogos de hoje, pro filtro por odds da lista da aba Hoje.
    Reaproveita o MESMO snapshot cacheado que já alimenta o Filtro de
    Metodologias e a classificação de Parâmetros (_today2_odds_snapshot) —
    não dispara nenhuma busca nova, só reexpõe os mercados crus em vez da
    versão já resumida em tiers (favorito/parelho etc)."""
    force = request.args.get("refresh") == "1"
    result = {}
    for m, markets in _today2_odds_snapshot(force=force):
        if markets:
            result[str(m["id"])] = markets
    return jsonify({"markets": result})


@app.route("/api/match/live/<path:match_id>")
def api_match_live(match_id):
    """Busca detalhes ao vivo do StatArea para qualquer partida. Se casa/fora
    forem passados via query string, também recalcula Convicção e Confronto
    Direto em cima desses detalhes recém-buscados (senão ficariam presos aos
    dados incompletos da 1ª resposta de /api/match/<id>)."""
    from scraper import fetch_match_details, create_session
    url = f"https://www.statarea.com/compare/teams/{match_id}"
    try:
        session = create_session()
        details = fetch_match_details(url, session=session)
        resp = {"detalhes": details, "url_detalhes": url}
        casa, fora = request.args.get("casa"), request.args.get("fora")
        if casa and fora:
            m = {"casa": casa, "fora": fora, "detalhes": details}
            resp["convicao_casa"], resp["h2h_confronto_direto"] = _convicao_score(m, True)
            resp["convicao_fora"], _ = _convicao_score(m, False)
        return jsonify(resp)
    except Exception as e:
        abort(503)


@app.route("/code")
@app.route("/code/<path:filepath>")
def code_viewer(filepath=None):
    base = os.path.dirname(__file__)
    files = {
        "app.py": os.path.join(base, "app.py"),
        "scraper.py": os.path.join(base, "scraper.py"),
        "static/index.html": os.path.join(base, "static", "index.html"),
    }
    if filepath and filepath in files:
        path = files[filepath]
        try:
            with open(path, encoding="utf-8") as f:
                content = f.read()
        except Exception as e:
            content = f"Erro ao ler arquivo: {e}"
        ext = filepath.rsplit(".", 1)[-1]
        lang = {"py": "python", "html": "html", "js": "javascript"}.get(ext, "plaintext")
        size = os.path.getsize(path)
        lines = content.count("\n") + 1
        return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<title>Código — {filepath}</title>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/styles/atom-one-dark.min.css">
<script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/highlight.min.js"></script>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ background: #1a1a2e; color: #e0e0e0; font-family: 'Segoe UI', sans-serif; min-height: 100vh; }}
  .topbar {{ background: #16213e; border-bottom: 1px solid #0f3460; padding: 12px 20px; display: flex; align-items: center; gap: 16px; position: sticky; top: 0; z-index: 10; }}
  .topbar a {{ color: #4ecca3; text-decoration: none; font-size: 14px; }}
  .topbar a:hover {{ text-decoration: underline; }}
  .filename {{ font-size: 16px; font-weight: 600; color: #fff; }}
  .meta {{ font-size: 12px; color: #888; margin-left: auto; }}
  .code-wrap {{ padding: 20px; }}
  pre {{ border-radius: 8px; font-size: 13px; line-height: 1.6; overflow-x: auto; }}
  pre code {{ font-family: 'Consolas', 'Monaco', monospace; }}
  .copy-btn {{ background: #0f3460; border: 1px solid #4ecca3; color: #4ecca3; padding: 6px 14px; border-radius: 6px; cursor: pointer; font-size: 13px; }}
  .copy-btn:hover {{ background: #4ecca3; color: #1a1a2e; }}
  .separator {{ color: #444; }}
</style>
</head>
<body>
<div class="topbar">
  <a href="/code">← Arquivos</a>
  <span class="separator">|</span>
  <span class="filename">📄 {filepath}</span>
  <button class="copy-btn" onclick="copyCode()">Copiar</button>
  <span class="meta">{lines} linhas &nbsp;·&nbsp; {size:,} bytes</span>
</div>
<div class="code-wrap">
  <pre><code class="language-{lang}" id="codeblock">{content.replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')}</code></pre>
</div>
<script>
  hljs.highlightAll();
  function copyCode() {{
    navigator.clipboard.writeText(document.getElementById('codeblock').innerText);
    const btn = document.querySelector('.copy-btn');
    btn.textContent = 'Copiado!';
    setTimeout(() => btn.textContent = 'Copiar', 2000);
  }}
</script>
</body>
</html>"""

    # Página índice dos arquivos
    file_info = []
    for name, path in files.items():
        try:
            size = os.path.getsize(path)
            lines = sum(1 for _ in open(path, encoding="utf-8"))
            icon = "🐍" if name.endswith(".py") else "🌐"
        except:
            size, lines, icon = 0, 0, "📄"
        file_info.append((name, size, lines, icon))

    cards = ""
    for name, size, lines, icon in file_info:
        cards += f"""
        <a href="/code/{name}" class="card">
          <div class="card-icon">{icon}</div>
          <div class="card-info">
            <div class="card-name">{name}</div>
            <div class="card-meta">{lines} linhas · {size:,} bytes</div>
          </div>
        </a>"""

    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<title>Visualizador de Código — Gol em Números</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ background: #1a1a2e; color: #e0e0e0; font-family: 'Segoe UI', sans-serif; min-height: 100vh; }}
  .header {{ background: #16213e; border-bottom: 1px solid #0f3460; padding: 20px; text-align: center; }}
  .header h1 {{ font-size: 22px; color: #4ecca3; }}
  .header p {{ font-size: 13px; color: #888; margin-top: 4px; }}
  .grid {{ display: flex; flex-direction: column; gap: 12px; padding: 30px; max-width: 600px; margin: 0 auto; }}
  .card {{ display: flex; align-items: center; gap: 16px; background: #16213e; border: 1px solid #0f3460; border-radius: 10px; padding: 16px 20px; text-decoration: none; color: inherit; transition: border-color .2s, background .2s; }}
  .card:hover {{ border-color: #4ecca3; background: #1e2d50; }}
  .card-icon {{ font-size: 28px; }}
  .card-name {{ font-size: 16px; font-weight: 600; color: #fff; }}
  .card-meta {{ font-size: 12px; color: #888; margin-top: 3px; }}
  .back {{ display: inline-block; margin: 20px 30px 0; color: #4ecca3; text-decoration: none; font-size: 14px; }}
  .back:hover {{ text-decoration: underline; }}
</style>
</head>
<body>
<div class="header">
  <h1>📂 Visualizador de Código</h1>
  <p>Gol em Números — StatArea Dashboard</p>
</div>
<a href="/" class="back">← Voltar ao painel</a>
<div class="grid">{cards}</div>
</body>
</html>"""


@app.route("/api/test-sofa/<event_id>")
def api_test_sofa(event_id):
    """Endpoint de diagnóstico — testa conexão com SofaScore."""
    import sys
    result = {"event_id": event_id, "method": None, "error": None, "data_preview": None}
    try:
        graph, incidents, statistics = _fetch_sofa_direct(event_id)
        pts = graph.get("graphPoints", [])
        result["method"]       = "requests_direto"
        result["graph_points"] = len(pts)
        result["incidents_ok"] = bool(incidents.get("incidents"))
        result["stats_ok"]     = bool(statistics.get("statistics"))
        print(f"[test-sofa] OK via requests: {event_id}, {len(pts)} pts", file=sys.stderr)
    except Exception as e:
        result["error_requests"] = str(e)
        print(f"[test-sofa] requests falhou: {e}", file=sys.stderr)
        try:
            graph, incidents, statistics = _fetch_sofa_playwright(event_id)
            pts = graph.get("graphPoints", [])
            result["method"]       = "playwright"
            result["graph_points"] = len(pts)
            print(f"[test-sofa] OK via playwright: {event_id}", file=sys.stderr)
        except Exception as e2:
            result["error_playwright"] = str(e2)
            print(f"[test-sofa] playwright falhou: {e2}", file=sys.stderr)
    return jsonify(result)


@app.route("/api/list-backup")
def api_list_backup():
    """Lista arquivos de momentum_history disponíveis no servidor."""
    token    = request.args.get("token", "")
    expected = os.environ.get("UPLOAD_TOKEN", "")
    if not expected or token != expected:
        return jsonify({"ok": False, "error": "Token inválido"}), 403

    files = sorted(glob.glob(os.path.join(MOMENTUM_DIR, "*.json")))
    names = [os.path.basename(f) for f in files]
    return jsonify({"ok": True, "files": names})


@app.route("/api/download-backup/<path:filename>")
def api_download_backup(filename):
    """Baixa um arquivo de momentum_history pelo nome."""
    token    = request.args.get("token", "")
    expected = os.environ.get("UPLOAD_TOKEN", "")
    if not expected or token != expected:
        return jsonify({"ok": False, "error": "Token inválido"}), 403

    # Segurança: só permite nomes de arquivo simples
    if "/" in filename or "\\" in filename or not filename.endswith(".json"):
        return jsonify({"ok": False, "error": "Arquivo inválido"}), 400

    return send_from_directory(MOMENTUM_DIR, filename)


@app.route("/api/upload-backup", methods=["POST"])
def api_upload_backup():
    """Recebe arquivos de backtest/momentum/predictions enviados do ambiente local.
    Requer ?token=UPLOAD_TOKEN no Railway Variables.
    """
    token    = request.args.get("token", "")
    expected = os.environ.get("UPLOAD_TOKEN", "")
    if not expected or token != expected:
        return jsonify({"ok": False, "error": "Token inválido"}), 403

    if "file" not in request.files:
        return jsonify({"ok": False, "error": "Nenhum arquivo enviado"}), 400

    f     = request.files["file"]
    fname = f.filename or ""

    # Determina destino pelo nome do arquivo
    if re.match(r'^\d{4}-\d{2}-\d{2}\.json$', fname):
        dest_dir      = BACKTEST_DIR
        remote_prefix = "backtest"
    elif re.match(r'^\d{4}-\d{2}-\d{2}_[\w]+\.json$', fname):
        dest_dir      = MOMENTUM_DIR
        remote_prefix = "momentum_history"
    elif fname in ("predictions_full.json", "predictions.json"):
        dest_dir      = DATA_DIR
        remote_prefix = ""
    else:
        return jsonify({"ok": False, "error": f"Nome de arquivo não reconhecido: {fname}"}), 400

    os.makedirs(dest_dir, exist_ok=True)
    local_path = os.path.join(dest_dir, fname)
    f.save(local_path)

    remote_path = f"{remote_prefix}/{fname}" if remote_prefix else fname
    github_storage.push_file_bg(local_path, remote_path)

    # Invalida caches de análise sempre que chegar arquivo de histórico ou predictions
    if dest_dir in (MOMENTUM_DIR, BACKTEST_DIR) or fname.startswith("predictions"):
        global _pattern_tips_cache, _odds_patterns_cache, _stats_patterns_cache
        _pattern_tips_cache  = {"ts": 0, "data": None}
        _odds_patterns_cache = {"ts": 0, "data": None}
        _stats_patterns_cache = {"ts": 0, "data": None}
        # Reconstrói em background para a próxima requisição já ter os dados prontos
        threading.Thread(target=_rebuild_analysis_cache, daemon=True).start()

    return jsonify({"ok": True, "saved": fname, "path": local_path})


@app.route("/api/upload-backup-bulk", methods=["POST"])
def api_upload_backup_bulk():
    """Recebe múltiplos arquivos de momentum_history em uma só requisição.
    Mais eficiente que chamar /api/upload-backup N vezes.
    Requer ?token=UPLOAD_TOKEN.
    """
    token    = request.args.get("token", "")
    expected = os.environ.get("UPLOAD_TOKEN", "")
    if not expected or token != expected:
        return jsonify({"ok": False, "error": "Token inválido"}), 403

    files = request.files.getlist("files")
    if not files:
        return jsonify({"ok": False, "error": "Nenhum arquivo enviado"}), 400

    saved, skipped, errors = [], [], []
    for f in files:
        fname = f.filename or ""
        if re.match(r'^\d{4}-\d{2}-\d{2}_[\w]+\.json$', fname):
            dest_dir      = MOMENTUM_DIR
            remote_prefix = "momentum_history"
        elif re.match(r'^\d{4}-\d{2}-\d{2}\.json$', fname):
            dest_dir      = BACKTEST_DIR
            remote_prefix = "backtest"
        else:
            skipped.append(fname)
            continue

        os.makedirs(dest_dir, exist_ok=True)
        local_path = os.path.join(dest_dir, fname)
        try:
            f.save(local_path)
            github_storage.push_file_bg(local_path, f"{remote_prefix}/{fname}")
            saved.append(fname)
        except Exception as e:
            errors.append({"file": fname, "error": str(e)})

    if saved:
        global _pattern_tips_cache, _odds_patterns_cache, _stats_patterns_cache
        _pattern_tips_cache  = {"ts": 0, "data": None}
        _odds_patterns_cache = {"ts": 0, "data": None}
        _stats_patterns_cache = {"ts": 0, "data": None}
        threading.Thread(target=_rebuild_analysis_cache, daemon=True).start()

    return jsonify({
        "ok":      True,
        "saved":   len(saved),
        "skipped": len(skipped),
        "errors":  errors,
        "files":   saved,
    })


TG_CONFIG_FILE = os.path.join(DATA_DIR, "telegram_config.json")
TG_DAILY_FILE  = os.path.join(DATA_DIR, "telegram_daily.json")

def _tg_load_config():
    try:
        with open(TG_CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def _tg_send_message(token, chat_id, message):
    r = http_req.post(
        f"https://api.telegram.org/bot{token}/sendMessage",
        json={"chat_id": chat_id, "text": message, "parse_mode": "HTML"},
        timeout=10,
    )
    return r.json()

def _tg_auto_send():
    """Envia as entradas do dia automaticamente via Telegram."""
    cfg = _tg_load_config()
    if not cfg.get("token") or not cfg.get("chat_id"):
        return
    try:
        with open(TG_DAILY_FILE, "r", encoding="utf-8") as f:
            daily = json.load(f)
    except Exception:
        daily = []
    if not daily:
        return
    hoje = datetime.now().strftime("%d/%m/%Y")
    msg  = f"🤖 <b>Gol em Números — {hoje}</b>\n"
    msg += f"📋 <b>{len(daily)} partida{'s' if len(daily)!=1 else ''}</b> nas metodologias configuradas\n\n"
    for item in daily[:25]:
        hora = f" · {item['hora']}" if item.get("hora") else ""
        msg += f"⚽ <b>{item['casa']} × {item['fora']}</b>{hora}\n"
        msg += f"🏆 {item.get('liga','')}\n"
        for fit in item.get("fits", [])[:3]:
            msg += f"  • {fit}\n"
        msg += "\n"
    if len(daily) > 25:
        msg += f"...e mais {len(daily)-25} partidas.\n"
    try:
        _tg_send_message(cfg["token"], cfg["chat_id"], msg)
        print(f"[Telegram] Envio automático: {len(daily)} partidas enviadas às {datetime.now().strftime('%H:%M')}")
    except Exception as e:
        print(f"[Telegram] Erro no envio automático: {e}")

def _tg_scheduler():
    """Background thread: envio diário no horário configurado."""
    sent_today = None
    while True:
        try:
            cfg = _tg_load_config()
            now       = datetime.now()
            today_str = now.strftime("%Y-%m-%d")

            # Envio diário no horário configurado
            if cfg.get("auto_send") and cfg.get("send_time"):
                hm = now.strftime("%H:%M")
                if hm == cfg["send_time"] and sent_today != today_str:
                    sent_today = today_str
                    _tg_auto_send()

        except Exception as e:
            print(f"[Telegram scheduler] {e}")
        time.sleep(30)

# Inicia background thread do scheduler
threading.Thread(target=_tg_scheduler, daemon=True).start()

@app.route("/api/telegram/send", methods=["POST"])
def api_telegram_send():
    """Envia mensagem via Telegram Bot API."""
    data    = request.json or {}
    token   = (data.get("token") or "").strip()
    chat_id = (data.get("chat_id") or "").strip()
    message = (data.get("message") or "").strip()
    if not token or not chat_id or not message:
        return jsonify({"ok": False, "error": "Campos obrigatórios: token, chat_id, message"}), 400
    try:
        return jsonify(_tg_send_message(token, chat_id, message))
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/telegram/config", methods=["GET", "POST"])
def api_telegram_config():
    """Salva ou carrega configuração do Telegram."""
    if request.method == "POST":
        data = request.json or {}
        try:
            with open(TG_CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return jsonify({"ok": True})
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500
    else:
        return jsonify(_tg_load_config())

@app.route("/api/telegram/daily", methods=["POST"])
def api_telegram_daily():
    """Salva a lista pré-computada de partidas do dia para envio automático."""
    data = request.json or {}
    matches = data.get("matches", [])
    try:
        with open(TG_DAILY_FILE, "w", encoding="utf-8") as f:
            json.dump(matches, f, ensure_ascii=False, indent=2)
        return jsonify({"ok": True, "saved": len(matches)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/telegram/send-now", methods=["POST"])
def api_telegram_send_now():
    """Dispara o envio automático imediatamente."""
    _tg_auto_send()
    return jsonify({"ok": True})


# ── Configuração do Lay Placar (persistida no SERVIDOR, não só no navegador) ─
# Antes só ficava salva no localStorage do navegador — cada aparelho/navegador
# tinha que reimportar o mesmo arquivo. Guardando aqui também, importar 1x (de
# qualquer aparelho) já fica disponível em qualquer outro que acessar o site,
# sem reimportar. Sincroniza com o GitHub (mesmo padrão de todo o resto do
# app) pra sobreviver a redeploy no Railway.
LAY_PLACAR_CONFIG_FILE = os.path.join(DATA_DIR, "lay_placar_config.json")


@app.route("/api/lay_placar/config", methods=["GET", "POST"])
def api_lay_placar_config():
    if request.method == "POST":
        data = request.json or {}
        try:
            with open(LAY_PLACAR_CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False)
            github_storage.push_file_bg(LAY_PLACAR_CONFIG_FILE, "lay_placar_config.json")
            return jsonify({"ok": True})
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500
    else:
        if not os.path.exists(LAY_PLACAR_CONFIG_FILE):
            return jsonify({})
        try:
            with open(LAY_PLACAR_CONFIG_FILE, "r", encoding="utf-8") as f:
                return jsonify(json.load(f))
        except Exception:
            return jsonify({})


threading.Thread(target=_tipster_watch_loop, daemon=True, name="TipsterWatch").start()
# Pré-carga de força (força-prefetch) DESATIVADA de novo — mesmo com só 1
# worker + pausa entre partidas, o Playwright rodando quase sem parar em
# segundo plano parece estar competindo por CPU com o resto do site num
# container com recursos limitados (site voltou a ficar lento, 15-40s pra
# responder, depois desse deploy). A coluna "Força" e o filtro continuam
# funcionando normalmente, só que cache-only (sem pré-carga automática).
# threading.Thread(target=_ng_strength_prefetch_filler_loop, daemon=True, name="ForcaPrefetchFiller").start()
# threading.Thread(target=_ng_strength_prefetch_worker, daemon=True, name="ForcaPrefetchWorker").start()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"Servidor rodando em http://localhost:{port}")
    app.run(debug=False, host="0.0.0.0", port=port, use_reloader=False, threaded=True)
